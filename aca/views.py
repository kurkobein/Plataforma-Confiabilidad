import json
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Count, Prefetch, Q
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.dateparse import parse_date, parse_datetime
from django.utils import timezone

from core import models
from core.access import get_accessible_services, get_profile_for_user, get_service_equipment, get_service_permission
from aca.forms import ACAExcelBulkUploadForm, CriticidadDimensionFormSet, CriticidadDimensionInputForm, ServicioACARegistroForm
from aca.services.progress import (
    build_aca_service_progress_summary,
    compute_criticidad_progress,
    filter_criticidades_by_hierarchy,
    get_aca_criticidad_queryset,
    get_aca_progress_dimensions,
    get_descendant_node_ids,
    get_hierarchy_filter_options,
    group_progress_by_hierarchy_level,
)
from core.services.criticality_rules import apply_criticality_rules, dimension_source_values, matrix_rule_config
from core.views import (
    _calc_slug,
    clear_session_upload,
    _catalog_row_boolean,
    _catalog_row_primary_numeric,
    _catalog_row_text,
    _decimal_or_none,
    _dimension_display_value,
    _equipment_items_payload,
    _export_filename,
    _export_pdf_response,
    _export_xlsx_response,
    _format_matrix_decimal,
    _is_generated_matrix_axis_dimension,
    _json_loads_safe,
    _matrix_cell_for_axis_values,
    _matrix_resolution_mode,
    _matrix_result_from_axis_values,
    _matrix_threshold_axis_values,
    _service_equipment_browser_payload,
    _service_equipment_endpoints,
    _service_or_404,
    _strategy_dimensions,
    open_session_upload,
    store_session_upload,
)


# ---------------------------------------------------------------------------
# Helpers del flujo ACA
# ---------------------------------------------------------------------------
ACA_INITIAL_VERSION = Decimal('1.0')


def _existing_aca_equipment_ids(servicio, exclude_ids=None):
    qs = models.Criticidad.objects.filter(
        aca_carga__servicio=servicio,
        equipo_id__isnull=False,
    )
    exclude_ids = [pk for pk in (exclude_ids or []) if pk]
    if exclude_ids:
        qs = qs.exclude(pk__in=exclude_ids)
    return set(qs.values_list('equipo_id', flat=True))


def _duplicate_aca_equipment_message(equipo):
    label = getattr(equipo, 'tag_display', '') or getattr(equipo, 'tag_equipo', '') or str(equipo)
    nombre = getattr(equipo, 'nombre_equipo', '') or ''
    detail = f'{label} - {nombre}'.strip(' -')
    return f'El equipo {detail} ya tiene un registro ACA en este servicio.'


def _next_service_aca_version(servicio):
    latest = models.Carga.objects.filter(servicio=servicio).order_by('-fecha_analisis', '-id').first()
    if not latest or latest.version_carga is None:
        return ACA_INITIAL_VERSION
    try:
        return Decimal(str(latest.version_carga)) + Decimal('0.1')
    except (InvalidOperation, ValueError, TypeError):
        return ACA_INITIAL_VERSION


def _increment_aca_version(value):
    try:
        return Decimal(str(value or ACA_INITIAL_VERSION)) + Decimal('0.1')
    except (InvalidOperation, ValueError, TypeError):
        return ACA_INITIAL_VERSION


def _aca_version_for_save(carga=None, requested_status=None):
    if not carga:
        return ACA_INITIAL_VERSION
    current = carga.version_carga or ACA_INITIAL_VERSION
    if (
        requested_status == models.Carga.STATUS_COMPLETO
        and getattr(carga, 'status', '') == models.Carga.STATUS_COMPLETO
    ):
        return _increment_aca_version(current)
    return current


def _service_family_payload(servicio):
    payload = []
    familias = (
        models.FamiliaEquipo.objects.filter(servicio=servicio, activa=True)
        .prefetch_related('items__equipo')
        .order_by('nombre', 'id')
    )
    for familia in familias:
        equipos = [item.equipo for item in familia.items.all()]
        equipos = sorted(
            [equipo for equipo in equipos if equipo],
            key=lambda equipo: (
                (getattr(equipo, 'tag_display', '') or equipo.tag_equipo or '').casefold(),
                (equipo.nombre_equipo or '').casefold(),
                (equipo.ut or '').casefold(),
            ),
        )
        payload.append({
            'id': familia.pk,
            'nombre': familia.nombre,
            'equipos': _equipment_items_payload(equipos),
        })
    return payload


def _record_attachment_payload(request):
    files = request.FILES.getlist('adjuntos')
    if not files:
        return [], []
    allowed = set(models.RECORD_ATTACHMENT_EXTENSIONS)
    invalid = []
    payload = []
    for uploaded in files:
        extension = (uploaded.name.rsplit('.', 1)[-1] if '.' in uploaded.name else '').lower()
        if extension not in allowed:
            invalid.append(uploaded.name)
            continue
        payload.append({
            'name': uploaded.name,
            'content': uploaded.read(),
        })
    return payload, invalid


def _save_aca_attachments(criticidad, attachment_payload, usuario):
    now = timezone.now()
    for item in attachment_payload:
        adjunto = models.CriticidadAdjunto(
            criticidad=criticidad,
            nombre_original=item['name'],
            creado_en=now,
            usuario=usuario,
        )
        adjunto.archivo.save(item['name'], ContentFile(item['content']), save=False)
        adjunto.save()


def _service_matrix_selector(servicio):
    matriz = None
    if getattr(servicio, 'estrategia_id', None):
        matriz = models.MatrizRiesgo.objects.filter(
            estrategia=servicio.estrategia
        ).select_related(
            'dimension_probabilidad',
            'dimension_probabilidad__dimension',
            'dimension_impacto',
            'dimension_impacto__dimension',
        ).order_by('-fecha_creado', '-id').first()

    if not matriz:
        return {
            'matriz': None,
            'rows': [],
            'columns': [],
            'axis': 'probabilidad',
            'prob_dimension_id': '',
            'impact_dimension_id': '',
            'prob_estrategia_dimension_id': '',
            'impact_estrategia_dimension_id': '',
            'modo_resolucion': models.MatrizRiesgo.RESOLUCION_EXACTA,
            'prob_axis_generated': False,
            'impact_axis_generated': False,
            'prob_axis_label': 'Probabilidad',
            'impact_axis_label': 'Consecuencia',
            'criticality_rules': [],
        }

    prob_levels = list(
        models.NivelProbabilidad.objects.filter(matriz=matriz).order_by('orden_visual', 'id')
    )
    impact_levels = list(
        models.NivelImpacto.objects.filter(matriz=matriz).order_by('orden_visual', 'id')
    )

    cells = {
        (cell.probabilidad_id, cell.impacto_nivel_id): cell
        for cell in models.MatrizRiesgoCelda.objects.filter(matriz=matriz)
        .select_related('probabilidad', 'impacto_nivel')
        .order_by('probabilidad__orden_visual', 'impacto_nivel__orden_visual')
    }

    rows = []
    for impact in impact_levels:
        row = {
            'id': impact.id,
            'label': impact.nombre,
            'value': impact.valor,
            'cells': [],
        }
        for prob in prob_levels:
            cell = cells.get((prob.id, impact.id))
            row['cells'].append({
                'id': cell.id if cell else '',
                'prob_id': prob.id,
                'prob_label': prob.nombre,
                'prob_value': str(prob.valor),
                'impact_id': impact.id,
                'impact_label': impact.nombre,
                'impact_value': str(impact.valor),
                'result': str(cell.resultado_num) if cell and cell.resultado_num is not None else '',
                'classification': cell.clasificacion if cell else '',
                'color': cell.color if cell and cell.color else '#ffffff',
            })
        rows.append(row)

    columns = [
        {'id': prob.id, 'label': prob.nombre, 'value': prob.valor}
        for prob in prob_levels
    ]

    return {
        'matriz': matriz,
        'rows': rows,
        'columns': columns,
        'axis': 'probabilidad',
        'prob_dimension_id': matriz.dimension_probabilidad.dimension_id if matriz.dimension_probabilidad_id else '',
        'impact_dimension_id': matriz.dimension_impacto.dimension_id if matriz.dimension_impacto_id else '',
        'prob_estrategia_dimension_id': matriz.dimension_probabilidad_id or '',
        'impact_estrategia_dimension_id': matriz.dimension_impacto_id or '',
        'modo_resolucion': _matrix_resolution_mode(matriz),
        'prob_axis_generated': _is_generated_matrix_axis_dimension(matriz.dimension_probabilidad, 'probabilidad') if matriz.dimension_probabilidad_id else False,
        'impact_axis_generated': _is_generated_matrix_axis_dimension(matriz.dimension_impacto, 'impacto') if matriz.dimension_impacto_id else False,
        'prob_axis_label': matriz.dimension_probabilidad.dimension.nombre if matriz.dimension_probabilidad_id else 'Probabilidad',
        'impact_axis_label': matriz.dimension_impacto.dimension.nombre if matriz.dimension_impacto_id else 'Consecuencia',
        'criticality_rules': matrix_rule_config(matriz),
    }

def _save_matrix_dimensions(evaluacion, estrategia, selected_cell):
    matriz = models.MatrizRiesgo.objects.filter(
        estrategia=estrategia
    ).select_related(
        'dimension_probabilidad',
        'dimension_probabilidad__dimension',
        'dimension_impacto',
        'dimension_impacto__dimension',
    ).order_by('-fecha_creado', '-id').first()
    if not matriz:
        return

    prob_dim = matriz.dimension_probabilidad
    impact_dim = matriz.dimension_impacto
    existing_dimension_ids = set(
        evaluacion.dimensiones.values_list('dimension_id', flat=True)
    )

    if prob_dim and prob_dim.dimension_id not in existing_dimension_ids:
        models.CriticidadDimension.objects.create(
            criticidad=evaluacion,
            dimension=prob_dim.dimension,
            estrategia_dimension=prob_dim,
            valor_numerico=selected_cell.probabilidad.valor,
            valor_texto=selected_cell.probabilidad.nombre or selected_cell.probabilidad.descripcion or '',
        )

    if impact_dim and impact_dim.dimension_id not in existing_dimension_ids:
        models.CriticidadDimension.objects.create(
            criticidad=evaluacion,
            dimension=impact_dim.dimension,
            estrategia_dimension=impact_dim,
            valor_numerico=selected_cell.impacto_nivel.valor,
            valor_texto=selected_cell.impacto_nivel.nombre or selected_cell.impacto_nivel.descripcion or '',
        )

def _aca_excluded_dimension_ids(estrategia, matriz=None):
    excluded_ids = set()
    matrix_input_dimension_ids = set()

    if matriz:
        if (
            getattr(matriz, 'dimension_probabilidad_id', None)
            and _is_generated_matrix_axis_dimension(matriz.dimension_probabilidad, 'probabilidad')
        ):
            excluded_ids.add(matriz.dimension_probabilidad.dimension_id)
        elif getattr(matriz, 'dimension_probabilidad_id', None):
            matrix_input_dimension_ids.add(matriz.dimension_probabilidad.dimension_id)
        if (
            getattr(matriz, 'dimension_impacto_id', None)
            and _is_generated_matrix_axis_dimension(matriz.dimension_impacto, 'impacto')
        ):
            excluded_ids.add(matriz.dimension_impacto.dimension_id)
        elif getattr(matriz, 'dimension_impacto_id', None):
            matrix_input_dimension_ids.add(matriz.dimension_impacto.dimension_id)

    derived_keywords = [
        'frecuencia normalizada',
        'valor consecuencia total',
        'probabilidad -',
        'impacto -',
    ]

    dims = models.EstrategiaDimension.objects.filter(
        estrategia=estrategia,
        activo=True,
        proceso_uso__in=[
            models.EstrategiaDimension.PROCESO_ACA,
            models.EstrategiaDimension.PROCESO_AMBOS,
        ],
    ).select_related('dimension').prefetch_related('catalogo')

    for item in dims:
        nombre = (item.dimension.nombre or '').strip().lower()
        if _is_generated_matrix_axis_dimension(item):
            excluded_ids.add(item.dimension_id)
            continue
        if item.dimension_id in matrix_input_dimension_ids:
            continue
        try:
            catalogo = item.catalogo
        except models.DimensionCatalogo.DoesNotExist:
            catalogo = None
        is_user_configured = bool(
            (getattr(item.dimension, 'tipo_calculo', '') or '').strip()
            or (catalogo and catalogo.activa)
        )
        if is_user_configured:
            continue
        if any(keyword in nombre for keyword in derived_keywords):
            excluded_ids.add(item.dimension_id)

    return excluded_ids

def _dependency_ref_from_config(config):
    if isinstance(config, str):
        config = _json_loads_safe(config, {})
    if not isinstance(config, dict):
        return ''
    for key in ['dependencia', 'depende_de', 'source', 'fuente', 'campo_fuente', 'dimension_fuente']:
        value = config.get(key)
        if value not in (None, ''):
            return value if isinstance(value, dict) else str(value).strip()
    if config.get('estrategia_dimension_id') or config.get('dimension_id'):
        return config
    return ''

def _source_ref_candidates(source_ref):
    if source_ref in (None, ''):
        return []
    if isinstance(source_ref, dict):
        candidates = []
        estrategia_dimension_id = (
            source_ref.get('estrategia_dimension_id')
            or source_ref.get('estrategiaDimensionId')
            or source_ref.get('ed_id')
        )
        if estrategia_dimension_id not in (None, ''):
            candidates.append(f'ed:{estrategia_dimension_id}')
            candidates.append(f'estrategia_dimension:{estrategia_dimension_id}')
        dimension_id = source_ref.get('dimension_id') or source_ref.get('dimensionId')
        if dimension_id not in (None, ''):
            candidates.append(str(dimension_id))
            candidates.append(f'dim:{dimension_id}')
        for key in ['campo', 'nombre', 'source', 'fuente', 'dependencia', 'depende_de', 'campo_fuente', 'dimension_fuente']:
            value = source_ref.get(key)
            if value not in (None, '') and not isinstance(value, dict):
                candidates.append(str(value).strip())
        return [candidate for candidate in candidates if candidate]
    return [str(source_ref).strip()]

def _dimension_dependency_key(dimension):
    return next(iter(_source_ref_candidates(_dependency_ref_from_config(getattr(dimension, 'config_calculo', None)))), '')

def _remember_source_value(source_values, estrategia_dimension, value):
    decimal_value = _decimal_or_none(value)
    if decimal_value is None:
        return
    dimension = estrategia_dimension.dimension
    campo = _dimension_source_key(estrategia_dimension)
    source_values[f'ed:{estrategia_dimension.id}'] = decimal_value
    source_values[f'estrategia_dimension:{estrategia_dimension.id}'] = decimal_value
    source_values[str(dimension.id)] = decimal_value
    source_values[f'dim:{dimension.id}'] = decimal_value
    source_values[dimension.nombre] = decimal_value
    source_values[_calc_slug(dimension.nombre)] = decimal_value
    if campo:
        source_values[campo] = decimal_value
        source_values[_calc_slug(campo)] = decimal_value

def _source_value(source_values, source_key):
    for candidate in _source_ref_candidates(source_key):
        value = source_values.get(candidate)
        if value is None:
            value = source_values.get(_calc_slug(candidate))
        decimal_value = _decimal_or_none(value)
        if decimal_value is not None:
            return decimal_value
    return None

def _catalog_bound(values, keys):
    for key in keys:
        value = values.get(key) if isinstance(values, dict) else None
        if value not in (None, ''):
            return _decimal_or_none(value)
    return None

def _match_catalog_range_row(catalogo, source_value):
    source_value = _decimal_or_none(source_value)
    if not catalogo or source_value is None:
        return None

    rows = list(catalogo.filas.prefetch_related('celdas__columna').order_by('orden', 'id'))
    row_bounds = []
    lowers = []
    for row in rows:
        values = row.values_map()
        lower = _catalog_bound(values, ['limite_inferior', 'desde', 'min', 'minimo', 'mí­nimo'])
        upper = _catalog_bound(values, ['limite_superior', 'hasta', 'max', 'maximo', 'máximo'])
        row_bounds.append((row, lower, upper))
        if lower is not None:
            lowers.append(lower)

    for row, lower, upper in row_bounds:
        if lower is not None and source_value < lower:
            continue
        if upper is not None and source_value > upper:
            continue
        if upper is not None and source_value == upper and upper in lowers:
            continue
        return row
    return None

def _catalog_primary_numeric_from_values(values):
    for key in ['valor_numerico', 'valor_principal', 'valor', 'nivel', 'puntaje']:
        value = values.get(key) if isinstance(values, dict) else None
        if value not in (None, ''):
            return _decimal_or_none(value)
    if isinstance(values, dict):
        for key, value in values.items():
            if key in {
                'limite_inferior', 'limite_superior', 'desde', 'hasta',
                'min', 'max', 'minimo', 'mí­nimo', 'maximo', 'máximo',
                'valor_secundario',
            }:
                continue
            decimal_value = _decimal_or_none(value)
            if decimal_value is not None:
                return decimal_value
    return None

def _catalog_match_value(values):
    for key in [
        'valor_dependencia',
        'dependencia',
        'depende_de',
        'valor_origen',
        'valor_fuente',
        'source_value',
        'valor_de_entrada',
        'valor_entrada',
        'entrada',
    ]:
        value = values.get(key) if isinstance(values, dict) else None
        if value not in (None, ''):
            decimal_value = _decimal_or_none(value)
            if decimal_value is not None:
                return decimal_value
    return _catalog_primary_numeric_from_values(values)

def _match_catalog_option_row(catalogo, source_value):
    source_value = _decimal_or_none(source_value)
    if not catalogo or source_value is None:
        return None

    rows = catalogo.filas.prefetch_related('celdas__columna').order_by('orden', 'id')
    for row in rows:
        match_value = _catalog_match_value(row.values_map())
        if match_value is not None and match_value == source_value:
            return row
    return None

def _match_catalog_dependency_row(catalogo, source_value):
    if not catalogo:
        return None
    if catalogo.tipo == 'rangos':
        return _match_catalog_range_row(catalogo, source_value)
    if catalogo.tipo == 'opciones':
        return _match_catalog_option_row(catalogo, source_value)
    return None

def _dimension_formset(request, estrategia, initial=None, bind_post=True, exclude_dimension_ids=None):
    initial = initial or []
    if not estrategia:
        return CriticidadDimensionFormSet(
            prefix='dims',
            form_kwargs={'estrategia': None, 'proceso': models.EstrategiaDimension.PROCESO_ACA},
        )

    if not initial:
        dims_qs = (
            models.EstrategiaDimension.objects.filter(
                estrategia=estrategia,
                activo=True,
                proceso_uso__in=[
                    models.EstrategiaDimension.PROCESO_ACA,
                    models.EstrategiaDimension.PROCESO_AMBOS,
                ],
            )
            .select_related('dimension')
            .order_by('orden', 'id')
        )
        if exclude_dimension_ids:
            dims_qs = dims_qs.exclude(dimension_id__in=exclude_dimension_ids)

        dims = list(dims_qs)
        dims.sort(key=lambda dim: (
            2 if _dimension_dependency_key(dim.dimension)
            else 1 if (getattr(dim.dimension, 'tipo_calculo', '') or '').strip()
            else 0,
            dim.orden,
            dim.id,
        ))
        initial = [{'dimension': dim.dimension} for dim in dims]

    if request.method == 'POST' and bind_post:
        return CriticidadDimensionFormSet(
            request.POST,
            prefix='dims',
            form_kwargs={'estrategia': estrategia, 'proceso': models.EstrategiaDimension.PROCESO_ACA},
        )
    return CriticidadDimensionFormSet(
        initial=initial,
        prefix='dims',
        form_kwargs={'estrategia': estrategia, 'proceso': models.EstrategiaDimension.PROCESO_ACA},
    )


def get_aca_bulk_dimensions(estrategia, excluded_dimension_ids=None):
    if not estrategia:
        return []
    dims_qs = (
        models.EstrategiaDimension.objects.filter(
            estrategia=estrategia,
            activo=True,
            proceso_uso__in=[
                models.EstrategiaDimension.PROCESO_ACA,
                models.EstrategiaDimension.PROCESO_AMBOS,
            ],
        )
        .select_related('dimension')
        .order_by('orden', 'id')
    )
    if excluded_dimension_ids:
        dims_qs = dims_qs.exclude(dimension_id__in=excluded_dimension_ids)
    return list(dims_qs)


@login_required
@transaction.atomic
def service_aca_bulk_dimension_order(request, pk):
    servicio, _permission = _service_or_404(request, pk, edit=True)
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Metodo no permitido.'}, status=405)
    if not servicio.estrategia_id:
        return JsonResponse({'ok': False, 'error': 'El servicio no tiene estrategia.'}, status=400)

    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({'ok': False, 'error': 'No se pudo leer el orden.'}, status=400)

    raw_ids = payload.get('dimension_ids') if isinstance(payload, dict) else None
    if not isinstance(raw_ids, list):
        return JsonResponse({'ok': False, 'error': 'Orden de dimensiones invalido.'}, status=400)

    dimension_ids = []
    for raw_id in raw_ids:
        try:
            dimension_id = int(raw_id)
        except (TypeError, ValueError):
            return JsonResponse({'ok': False, 'error': 'La lista contiene una dimension invalida.'}, status=400)
        if dimension_id not in dimension_ids:
            dimension_ids.append(dimension_id)

    strategy_dimensions = list(
        models.EstrategiaDimension.objects
        .filter(estrategia_id=servicio.estrategia_id, activo=True)
        .order_by('orden', 'id')
    )
    allowed_ids = {
        item.pk
        for item in strategy_dimensions
        if item.proceso_uso in (
            models.EstrategiaDimension.PROCESO_ACA,
            models.EstrategiaDimension.PROCESO_AMBOS,
        )
    }
    if len(dimension_ids) < 2 or any(item_id not in allowed_ids for item_id in dimension_ids):
        return JsonResponse({'ok': False, 'error': 'El orden no corresponde a las dimensiones ACA del servicio.'}, status=400)

    movable_ids = set(dimension_ids)
    current_movable_ids = [item.pk for item in strategy_dimensions if item.pk in movable_ids]
    if set(current_movable_ids) != movable_ids:
        return JsonResponse({'ok': False, 'error': 'No se encontraron todas las dimensiones indicadas.'}, status=400)

    reordered_ids = iter(dimension_ids)
    final_ids = [next(reordered_ids) if item.pk in movable_ids else item.pk for item in strategy_dimensions]
    by_id = {item.pk: item for item in strategy_dimensions}
    changed = []
    for order, item_id in enumerate(final_ids, start=1):
        item = by_id[item_id]
        if item.orden != order:
            item.orden = order
            changed.append(item)
    if changed:
        models.EstrategiaDimension.objects.bulk_update(changed, ['orden'])

    return JsonResponse({'ok': True, 'dimension_ids': dimension_ids})


def _bulk_json_safe(value):
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, dict):
        return {key: _bulk_json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_bulk_json_safe(item) for item in value]
    return value


def get_aca_bulk_dimension_payload(estrategia, excluded_dimension_ids=None):
    payload = []
    for estrategia_dimension in get_aca_bulk_dimensions(estrategia, excluded_dimension_ids):
        form = CriticidadDimensionInputForm(
            initial={'dimension': estrategia_dimension.dimension},
            estrategia=estrategia,
            proceso=models.EstrategiaDimension.PROCESO_ACA,
        )
        payload.append(_bulk_json_safe({
            'estrategia_dimension_id': estrategia_dimension.pk,
            'dimension_id': estrategia_dimension.dimension_id,
            'nombre': estrategia_dimension.dimension.nombre,
            'tipo_funcional': estrategia_dimension.dimension.tipo_funcional,
            'tipo_dato': estrategia_dimension.dimension.tipo_dato,
            'input_mode': form.input_mode,
            'option_headers': form.option_headers,
            'option_rows': form.option_rows,
            'catalog_dependency': form.catalog_dependency,
            'catalog_dependency_estrategia_dimension_id': form.catalog_dependency_estrategia_dimension_id,
            'catalog_dependency_dimension_id': form.catalog_dependency_dimension_id,
            'catalog_dependency_campo': form.catalog_dependency_campo,
            'catalog_type': form.catalog_type,
            'allows_empty_comment': form.allows_empty_comment,
            'config_calculo': form.config_calculo,
            'config_calculo_json': form.config_calculo_json,
            'tipo_calculo': form.tipo_calculo,
            'dimension_campo': form.dimension_campo,
            'obligatorio': estrategia_dimension.obligatorio,
        }))
    return payload

def _dimension_formset_initial_from_criticidad(criticidad, estrategia, exclude_dimension_ids=None):
    if not criticidad or not estrategia:
        return []

    existing = {
        item.dimension_id: item
        for item in criticidad.dimensiones.select_related(
            'dimension',
            'catalogo_fila',
            'escala_valor',
        ).all()
    }
    dims_qs = (
        models.EstrategiaDimension.objects.filter(
            estrategia=estrategia,
            activo=True,
            proceso_uso__in=[
                models.EstrategiaDimension.PROCESO_ACA,
                models.EstrategiaDimension.PROCESO_AMBOS,
            ],
        )
        .select_related('dimension')
        .order_by('orden', 'id')
    )
    if exclude_dimension_ids:
        dims_qs = dims_qs.exclude(dimension_id__in=exclude_dimension_ids)

    dims = list(dims_qs)
    dims.sort(key=lambda dim: (
        2 if _dimension_dependency_key(dim.dimension)
        else 1 if (getattr(dim.dimension, 'tipo_calculo', '') or '').strip()
        else 0,
        dim.orden,
        dim.id,
    ))

    initial = []
    for estrategia_dimension in dims:
        item = existing.get(estrategia_dimension.dimension_id)
        payload = {'dimension': estrategia_dimension.dimension}
        if item:
            payload.update({
                'escala_valor': item.escala_valor,
                'catalogo_fila': item.catalogo_fila,
                'valor_numerico': item.valor_numerico,
                'valor_booleano': item.valor_booleano,
                'valor_texto': item.valor_texto,
                'comentario': item.comentario or '',
            })
        initial.append(payload)
    return initial

def _dimension_source_key(estrategia_dimension):
    if not estrategia_dimension:
        return ''
    try:
        catalogo = getattr(estrategia_dimension, 'catalogo', None)
        if catalogo and catalogo.campo:
            return catalogo.campo
    except Exception:
        pass
    return _calc_slug(estrategia_dimension.dimension.nombre)

def _calculation_operation_result(tipo, values):
    if not values:
        return None

    if tipo == 'suma':
        return sum(values, Decimal('0'))
    if tipo == 'resta':
        result = values[0]
        for value in values[1:]:
            result -= value
        return result
    if tipo == 'multiplicacion':
        result = Decimal('1')
        for value in values:
            result *= value
        return result
    if tipo == 'division':
        result = values[0]
        for value in values[1:]:
            if value == 0:
                return None
            result /= value
        return result
    if tipo == 'maximo':
        return max(values)
    if tipo == 'minimo':
        return min(values)
    return None

def _calculation_operand_value(operand, source_values, previous_result=None):
    if isinstance(operand, dict):
        if operand.get('resultado') is True or operand.get('tipo') == 'resultado':
            return previous_result
    else:
        raw = str(operand)
        if raw in {'$resultado', '__resultado__', 'resultado_anterior'}:
            return previous_result

    for candidate in _source_ref_candidates(operand):
        value = source_values.get(candidate)
        if value is None:
            value = source_values.get(_calc_slug(candidate))
        decimal_value = _decimal_or_none(value)
        if decimal_value is not None:
            return decimal_value

    return None

def _calculation_steps(tipo_calculo, config_calculo):
    tipo = (tipo_calculo or '').strip().lower()
    if not tipo:
        return []

    if isinstance(config_calculo, str):
        try:
            config_calculo = json.loads(config_calculo or '{}')
        except Exception:
            config_calculo = {}
    if not isinstance(config_calculo, dict):
        config_calculo = {}

    raw_steps = config_calculo.get('pasos') or config_calculo.get('steps') or []
    if isinstance(raw_steps, list) and raw_steps:
        steps = []
        for raw_step in raw_steps:
            if not isinstance(raw_step, dict):
                continue
            operacion = str(raw_step.get('operacion') or raw_step.get('tipo_calculo') or raw_step.get('operation') or '').strip().lower()
            operandos = raw_step.get('operandos') or raw_step.get('campos') or raw_step.get('sources') or []
            if operacion and isinstance(operandos, list):
                steps.append({
                    'operacion': operacion,
                    'operandos': operandos,
                    'modo': raw_step.get('modo') or ('ponderado' if raw_step.get('ponderado') is True else ''),
                })
        return steps

    operandos = config_calculo.get('operandos') or config_calculo.get('campos') or config_calculo.get('sources') or []
    if not isinstance(operandos, list):
        operandos = []
    return [{'operacion': tipo, 'operandos': operandos}]

def _calculation_operand_weight(operand):
    if not isinstance(operand, dict):
        return None
    return _decimal_or_none(operand.get('peso', operand.get('ponderador', operand.get('weight'))))

def _weighted_calculation_values(operands, resolved_values):
    if not operands or len(operands) != len(resolved_values):
        return None
    weights = [_calculation_operand_weight(operand) for operand in operands]
    if all(weight is None for weight in weights):
        equal_weight = Decimal('1') / Decimal(len(weights))
        weights = [equal_weight] * len(weights)
    elif any(weight is None for weight in weights):
        assigned = sum((weight for weight in weights if weight is not None), Decimal('0'))
        missing_count = sum(1 for weight in weights if weight is None)
        remaining = Decimal('1') - assigned
        if remaining < 0:
            return None
        missing_weight = remaining / Decimal(missing_count)
        weights = [missing_weight if weight is None else weight for weight in weights]
    if any(weight < 0 or weight > 1 for weight in weights):
        return None
    if abs(sum(weights, Decimal('0')) - Decimal('1')) > Decimal('0.0001'):
        return None
    return [value * weight for value, weight in zip(resolved_values, weights)]

def _evaluate_dimension_calculation(tipo_calculo, config_calculo, source_values):
    steps = _calculation_steps(tipo_calculo, config_calculo)
    if not steps:
        return None

    result = None
    for step in steps:
        values = []
        for operand in step['operandos']:
            value = _calculation_operand_value(operand, source_values, result)
            if value is not None:
                values.append(value)

        if step.get('modo') == 'ponderado' and step['operacion'] == 'suma':
            values = _weighted_calculation_values(step['operandos'], values)
            if values is None:
                return None

        result = _calculation_operation_result(step['operacion'], values)
        if result is None:
            return None

    return result

def _extract_dimension_form_value(data, catalogo_fila=None, escala_valor=None):
    valor_numerico = data.get('valor_numerico')
    valor_secundario = None
    valor_booleano = data.get('valor_booleano')
    valor_texto = data.get('valor_texto', '')
    escala_unificada = None

    if escala_valor:
        escala_unificada = escala_valor.escala_unificada
        if valor_numerico in (None, ''):
            valor_numerico = escala_valor.valor_numerico
        if not valor_texto:
            valor_texto = escala_valor.codigo or escala_valor.descripcion or ''

    if catalogo_fila:
        if valor_numerico in (None, ''):
            valor_numerico = _catalog_row_primary_numeric(catalogo_fila)
        if valor_booleano is None:
            valor_booleano = _catalog_row_boolean(catalogo_fila)
        if not valor_texto:
            valor_texto = _catalog_row_text(catalogo_fila)

    return valor_numerico, valor_secundario, valor_booleano, valor_texto, escala_unificada

def _prepare_dimension_items(estrategia, formset):
    prepared = []
    source_values = {}

    for form in formset:
        data = getattr(form, 'cleaned_data', None)
        if not data:
            continue
        dimension = data.get('dimension')
        if not dimension:
            continue

        estrategia_dimension = models.EstrategiaDimension.objects.filter(
            estrategia=estrategia,
            activo=True,
            proceso_uso__in=[
                models.EstrategiaDimension.PROCESO_ACA,
                models.EstrategiaDimension.PROCESO_AMBOS,
            ],
            dimension=dimension,
        ).select_related('dimension').first()
        if not estrategia_dimension:
            continue

        catalogo_fila = data.get('catalogo_fila')
        escala_valor = data.get('escala_valor')
        is_calculated = bool((getattr(dimension, 'tipo_calculo', '') or '').strip())
        dependency_key = _dependency_ref_from_config(getattr(dimension, 'config_calculo', None))
        comentario = str(data.get('comentario') or '').strip()
        if is_calculated or dependency_key:
            comentario = ''
        valor_numerico, valor_secundario, valor_booleano, valor_texto, escala_unificada = _extract_dimension_form_value(
            data,
            catalogo_fila=catalogo_fila,
            escala_valor=escala_valor,
        )

        item = {
            'dimension': dimension,
            'estrategia_dimension': estrategia_dimension,
            'catalogo_fila': catalogo_fila,
            'escala_valor': escala_valor,
            'escala_unificada': escala_unificada,
            'valor_numerico': valor_numerico,
            'valor_secundario': valor_secundario,
            'valor_booleano': valor_booleano,
            'valor_texto': valor_texto or '',
            'comentario': comentario,
            'is_calculated': is_calculated,
            'dependency_key': dependency_key,
        }
        prepared.append(item)

        if not item['is_calculated'] and not item['dependency_key']:
            numeric_value = valor_numerico
            if numeric_value is not None:
                _remember_source_value(source_values, estrategia_dimension, numeric_value)

    for item in prepared:
        if item['is_calculated']:
            value = _evaluate_dimension_calculation(
                item['dimension'].tipo_calculo,
                item['dimension'].config_calculo,
                source_values,
            )
            item['valor_numerico'] = value
            item['valor_texto'] = '' if value is None else str(value)
            if value is not None:
                _remember_source_value(source_values, item['estrategia_dimension'], value)

    for item in prepared:
        if not item['dependency_key']:
            continue

        source_value = _source_value(source_values, item['dependency_key'])
        try:
            catalogo = item['estrategia_dimension'].catalogo
        except models.DimensionCatalogo.DoesNotExist:
            catalogo = None

        matched_row = _match_catalog_dependency_row(catalogo, source_value)
        if not matched_row:
            continue

        item['catalogo_fila'] = matched_row
        valor_numerico, valor_secundario, valor_booleano, valor_texto, escala_unificada = _extract_dimension_form_value(
            {},
            catalogo_fila=matched_row,
            escala_valor=None,
        )
        item['valor_numerico'] = valor_numerico
        item['valor_secundario'] = valor_secundario
        item['valor_booleano'] = valor_booleano
        item['valor_texto'] = valor_texto or ''
        item['escala_unificada'] = escala_unificada

        numeric_value = valor_numerico
        if numeric_value is not None:
            _remember_source_value(source_values, item['estrategia_dimension'], numeric_value)

    return prepared, source_values


def _bulk_dimension_value(raw, *names):
    if not isinstance(raw, dict):
        return None
    for name in names:
        value = raw.get(name)
        if value not in (None, ''):
            return value
    return None


def prepare_bulk_dimension_items(estrategia, row_dimensions, excluded_dimension_ids=None, allow_incomplete=False):
    row_dimensions = row_dimensions if isinstance(row_dimensions, dict) else {}
    prepared = []
    source_values = {}
    errors = []

    for estrategia_dimension in get_aca_bulk_dimensions(estrategia, excluded_dimension_ids):
        dimension = estrategia_dimension.dimension
        raw = row_dimensions.get(str(estrategia_dimension.pk)) or row_dimensions.get(str(dimension.pk)) or {}
        is_calculated = bool((getattr(dimension, 'tipo_calculo', '') or '').strip())
        dependency_key = _dependency_ref_from_config(getattr(dimension, 'config_calculo', None))

        catalogo_fila = None
        escala_valor = None
        if not is_calculated and not dependency_key:
            catalogo_fila_id = _bulk_dimension_value(raw, 'catalogo_fila_id', 'catalogo_fila')
            escala_valor_id = _bulk_dimension_value(raw, 'escala_valor_id', 'escala_valor')
            if catalogo_fila_id:
                catalogo_fila = models.DimensionCatalogoFila.objects.filter(
                    pk=catalogo_fila_id,
                    catalogo__estrategia_dimension=estrategia_dimension,
                ).first()
                if not catalogo_fila:
                    errors.append(f'{dimension.nombre}: opción de catálogo inválida.')
            if escala_valor_id:
                escala_valor = models.EscalaValor.objects.filter(
                    pk=escala_valor_id,
                    estrategia_dimension=estrategia_dimension,
                ).first()
                if not escala_valor:
                    errors.append(f'{dimension.nombre}: valor de escala inválido.')

        valor_booleano = raw.get('valor_booleano') if isinstance(raw, dict) else None
        if isinstance(valor_booleano, str):
            valor_booleano = None if valor_booleano == '' else valor_booleano.lower() in {'true', '1', 'si', 'sí'}

        data = {
            'valor_numerico': _decimal_or_none(_bulk_dimension_value(raw, 'valor_numerico', 'numeric_value')),
            'valor_booleano': valor_booleano,
            'valor_texto': str(_bulk_dimension_value(raw, 'valor_texto', 'text_value', 'display') or ''),
        }
        comentario = str(_bulk_dimension_value(raw, 'comentario', 'comment') or '').strip()
        if is_calculated or dependency_key:
            comentario = ''
        valor_numerico, valor_secundario, valor_booleano, valor_texto, escala_unificada = _extract_dimension_form_value(
            data,
            catalogo_fila=catalogo_fila,
            escala_valor=escala_valor,
        )

        item = {
            'dimension': dimension,
            'estrategia_dimension': estrategia_dimension,
            'catalogo_fila': catalogo_fila,
            'escala_valor': escala_valor,
            'escala_unificada': escala_unificada,
            'valor_numerico': valor_numerico,
            'valor_secundario': valor_secundario,
            'valor_booleano': valor_booleano,
            'valor_texto': valor_texto or '',
            'comentario': comentario,
            'is_calculated': is_calculated,
            'dependency_key': dependency_key,
        }
        prepared.append(item)

        has_value = any([
            item['escala_valor'],
            item['catalogo_fila'],
            item['valor_numerico'] is not None,
            item['valor_secundario'] is not None,
            item['valor_booleano'] is not None,
            item['valor_texto'],
        ])
        if (
            estrategia_dimension.obligatorio
            and not allow_incomplete
            and not is_calculated
            and not dependency_key
            and not has_value
        ):
            errors.append(f'{dimension.nombre}: valor requerido.')

        if not item['is_calculated'] and not item['dependency_key'] and valor_numerico is not None:
            _remember_source_value(source_values, estrategia_dimension, valor_numerico)

    for item in prepared:
        if item['is_calculated']:
            value = _evaluate_dimension_calculation(
                item['dimension'].tipo_calculo,
                item['dimension'].config_calculo,
                source_values,
            )
            item['valor_numerico'] = value
            item['valor_texto'] = '' if value is None else str(value)
            if value is not None:
                _remember_source_value(source_values, item['estrategia_dimension'], value)
            elif item['estrategia_dimension'].obligatorio and not allow_incomplete:
                errors.append(f"{item['dimension'].nombre}: no se pudo calcular.")

    for item in prepared:
        if not item['dependency_key']:
            continue

        source_value = _source_value(source_values, item['dependency_key'])
        try:
            catalogo = item['estrategia_dimension'].catalogo
        except models.DimensionCatalogo.DoesNotExist:
            catalogo = None

        matched_row = _match_catalog_dependency_row(catalogo, source_value)
        if not matched_row:
            if item['estrategia_dimension'].obligatorio and not allow_incomplete:
                errors.append(f"{item['dimension'].nombre}: no se pudo resolver desde la fuente.")
            continue

        item['catalogo_fila'] = matched_row
        valor_numerico, valor_secundario, valor_booleano, valor_texto, escala_unificada = _extract_dimension_form_value(
            {},
            catalogo_fila=matched_row,
            escala_valor=None,
        )
        item['valor_numerico'] = valor_numerico
        item['valor_secundario'] = valor_secundario
        item['valor_booleano'] = valor_booleano
        item['valor_texto'] = valor_texto or ''
        item['escala_unificada'] = escala_unificada
        if valor_numerico is not None:
            _remember_source_value(source_values, item['estrategia_dimension'], valor_numerico)

    return prepared, source_values, errors


def _create_dimension_items(evaluacion, prepared):
    created = []
    for item in prepared:
        if not any([
            item['escala_valor'],
            item['catalogo_fila'],
            item['valor_numerico'] is not None,
            item['valor_secundario'] is not None,
            item['valor_booleano'] is not None,
            item['valor_texto'],
            item.get('comentario'),
        ]):
            continue

        created.append(models.CriticidadDimension.objects.create(
            criticidad=evaluacion,
            dimension=item['dimension'],
            estrategia_dimension=item['estrategia_dimension'],
            escala_valor=item['escala_valor'],
            escala_unificada=item['escala_unificada'],
            catalogo_fila=item['catalogo_fila'],
            valor_numerico=item['valor_numerico'],
            valor_secundario=item['valor_secundario'],
            valor_booleano=item['valor_booleano'],
            valor_texto=item['valor_texto'] or '',
            comentario=item.get('comentario') or '',
        ))
    return created

def _save_dimension_formset(evaluacion, estrategia, formset):
    evaluacion.dimensiones.all().delete()
    prepared, _source_values = _prepare_dimension_items(estrategia, formset)
    return _create_dimension_items(evaluacion, prepared)

def _dimension_record_numeric_value(record):
    if isinstance(record, dict):
        value = record.get('valor_numerico')
    else:
        value = record.valor_numerico
    return _decimal_or_none(value)

def _dimension_record_ids(record):
    if isinstance(record, dict):
        estrategia_dimension = record.get('estrategia_dimension')
        dimension = record.get('dimension')
        return (
            getattr(estrategia_dimension, 'id', None),
            getattr(dimension, 'id', None),
        )
    estrategia_dimension = getattr(record, 'estrategia_dimension', None)
    return (
        getattr(record, 'estrategia_dimension_id', None),
        getattr(record, 'dimension_id', None) or getattr(estrategia_dimension, 'dimension_id', None),
    )

def _matrix_axis_dimension_refs(matriz):
    prob_dim_id = None
    impact_dim_id = None
    if getattr(matriz, 'dimension_probabilidad_id', None):
        prob_dim_id = matriz.dimension_probabilidad.dimension_id
    if getattr(matriz, 'dimension_impacto_id', None):
        impact_dim_id = matriz.dimension_impacto.dimension_id
    return {
        'prob_ed_id': getattr(matriz, 'dimension_probabilidad_id', None),
        'prob_dim_id': prob_dim_id,
        'impact_ed_id': getattr(matriz, 'dimension_impacto_id', None),
        'impact_dim_id': impact_dim_id,
    }


def _add_matrix_axis_source_values(source_values, matriz, prob_val=None, impact_val=None, result_value=None):
    source_values.update({
        'probabilidad': prob_val,
        'frecuencia': prob_val,
        'consecuencia': impact_val,
        'consecuencia_total': impact_val,
        'resultado_matriz': result_value,
    })
    if getattr(matriz, 'dimension_probabilidad_id', None):
        source_values[f'ed:{matriz.dimension_probabilidad_id}'] = prob_val
        source_values[f'estrategia_dimension:{matriz.dimension_probabilidad_id}'] = prob_val
    if getattr(matriz, 'dimension_impacto_id', None):
        source_values[f'ed:{matriz.dimension_impacto_id}'] = impact_val
        source_values[f'estrategia_dimension:{matriz.dimension_impacto_id}'] = impact_val
    return source_values


def _matrix_axis_values_from_records(matriz, records):
    refs = _matrix_axis_dimension_refs(matriz)
    prob_val = None
    impact_val = None

    for record in records:
        value = _dimension_record_numeric_value(record)
        if value is None:
            continue
        estrategia_dimension_id, dimension_id = _dimension_record_ids(record)

        if (
            refs['prob_ed_id'] and estrategia_dimension_id == refs['prob_ed_id']
        ) or (
            refs['prob_dim_id'] and dimension_id == refs['prob_dim_id']
        ):
            prob_val = value

        if (
            refs['impact_ed_id'] and estrategia_dimension_id == refs['impact_ed_id']
        ) or (
            refs['impact_dim_id'] and dimension_id == refs['impact_dim_id']
        ):
            impact_val = value

    return prob_val, impact_val

def _resolve_matrix_cell_from_dimension_records(matriz, records):
    prob_val, impact_val = _matrix_axis_values_from_records(matriz, records)
    return _matrix_cell_for_axis_values(matriz, prob_val, impact_val)


def _matrix_axis_values_complete(matriz, records):
    if not matriz:
        return True
    prob_val, impact_val = _matrix_axis_values_from_records(matriz, records)
    return prob_val is not None and impact_val is not None


def _resolve_complete_matrix_cell_from_dimension_records(matriz, records):
    if not matriz:
        return None
    prob_val, impact_val = _matrix_axis_values_from_records(matriz, records)
    if prob_val is None or impact_val is None:
        return None
    return _matrix_cell_for_axis_values(matriz, prob_val, impact_val)


def _criticidad_progress_is_complete(evaluacion, estrategia):
    progress_dimensions = get_aca_progress_dimensions(estrategia)
    if not progress_dimensions:
        return False
    progress = compute_criticidad_progress(evaluacion, progress_dimensions)
    return progress.get('progress_percent') == Decimal('100.0')


def _sync_aca_carga_status(evaluacion, estrategia, requested_status):
    carga = getattr(evaluacion, 'aca_carga', None)
    if not carga:
        return
    effective_status = models.Carga.STATUS_INCOMPLETO
    if (
        requested_status == models.Carga.STATUS_COMPLETO
        and _criticidad_progress_is_complete(evaluacion, estrategia)
    ):
        effective_status = models.Carga.STATUS_COMPLETO
    if carga.status != effective_status:
        carga.status = effective_status
        carga.actualizado = timezone.now()
        carga.save(update_fields=['status', 'actualizado'])


def _progress_color(progress_percent):
    if progress_percent is None:
        return '#94a3b8'
    try:
        percent = max(0, min(100, float(progress_percent)))
    except (TypeError, ValueError):
        percent = 0
    hue = int(percent * 1.2)
    return f'hsl({hue}, 72%, 43%)'


def _progress_width_css(progress_percent):
    if progress_percent is None:
        return '0'
    try:
        percent = max(0, min(100, float(progress_percent)))
    except (TypeError, ValueError):
        percent = 0
    return f'{percent:.1f}'.rstrip('0').rstrip('.')


def _contrast_text_color(color):
    value = (color or '').strip()
    if not value.startswith('#'):
        return '#111827'
    hex_value = value[1:]
    if len(hex_value) == 3:
        hex_value = ''.join(ch * 2 for ch in hex_value)
    if len(hex_value) != 6:
        return '#111827'
    try:
        red = int(hex_value[0:2], 16)
        green = int(hex_value[2:4], 16)
        blue = int(hex_value[4:6], 16)
    except ValueError:
        return '#111827'
    luminance = (0.299 * red + 0.587 * green + 0.114 * blue) / 255
    return '#ffffff' if luminance < 0.45 else '#111827'


def _matrix_cell_for_criticidad_row(matriz, criticidad, dimension_records, cells_by_classification):
    if not matriz:
        return None

    cell = _resolve_complete_matrix_cell_from_dimension_records(matriz, dimension_records)
    if cell:
        return cell
    return None

def _sync_criticidad_resumen(evaluacion, estrategia):
    dims = list(
        evaluacion.dimensiones.select_related('dimension', 'estrategia_dimension').all()
    )

    matriz = models.MatrizRiesgo.objects.filter(
        estrategia=estrategia
    ).select_related(
        'dimension_probabilidad',
        'dimension_probabilidad__dimension',
        'dimension_impacto',
        'dimension_impacto__dimension',
    ).order_by('-fecha_creado', '-id').first()

    prob_val = None
    impact_val = None
    valor_cons_total = None
    valor_criticidad_equipo = None
    indicador = evaluacion.indicador_criticidad or ''
    criticidad_final = evaluacion.criticidad_final or ''
    celda = None

    # 1) Priorizar las dimensiones especí­ficas que usa la matriz
    if matriz:
        prob_val, impact_val = _matrix_axis_values_from_records(matriz, dims)
        if prob_val is None or impact_val is None:
            incomplete_sources = dimension_source_values(dims)
            _add_matrix_axis_source_values(incomplete_sources, matriz, prob_val, impact_val, None)
            ruled = apply_criticality_rules(
                matriz,
                base_cell=None,
                base_value=None,
                base_classification='',
                source_values=incomplete_sources,
            )
            if ruled['trace'].get('reglas_aplicadas') and ruled['classification']:
                evaluacion.frecuencia_normalizada = prob_val
                evaluacion.valor_cons_total = impact_val
                evaluacion.valor_criticidad_equipo = ruled['value']
                evaluacion.indicador_criticidad = f"[ Regla: {ruled['trace'].get('regla_aplicada') or '-'} ]"[:100]
                evaluacion.criticidad_final = ruled['classification']
                evaluacion.trazabilidad_criticidad_json = json.dumps(ruled['trace'], ensure_ascii=False)
                evaluacion.save(update_fields=[
                    'frecuencia_normalizada',
                    'valor_cons_total',
                    'valor_criticidad_equipo',
                    'indicador_criticidad',
                    'criticidad_final',
                    'trazabilidad_criticidad_json',
                ])
                return
            evaluacion.frecuencia_normalizada = None
            evaluacion.valor_cons_total = None
            evaluacion.valor_criticidad_equipo = None
            evaluacion.indicador_criticidad = ''
            evaluacion.criticidad_final = ''
            evaluacion.trazabilidad_criticidad_json = ''
            evaluacion.save(
                update_fields=[
                    'frecuencia_normalizada',
                    'valor_cons_total',
                    'valor_criticidad_equipo',
                    'indicador_criticidad',
                    'criticidad_final',
                    'trazabilidad_criticidad_json',
                ]
            )
            return

    # 2) Fallback: usar la frecuencia normalizada y/o el valor de consecuencia ya guardado
    if not matriz and prob_val is None and evaluacion.frecuencia_normalizada is not None:
        prob_val = Decimal(evaluacion.frecuencia_normalizada)

    if not matriz and impact_val is None and evaluacion.valor_cons_total is not None:
        impact_val = Decimal(evaluacion.valor_cons_total)

    # 3) Si existe matriz, intentar resolver la celda configurada.
    #    Por defecto es exacta; algunas matrices pueden usar umbral inferior por resultado.
    if matriz and (prob_val is not None or impact_val is not None):
        celda = _matrix_cell_for_axis_values(matriz, prob_val, impact_val)

        if celda:
            if _matrix_resolution_mode(matriz) == models.MatrizRiesgo.RESOLUCION_UMBRAL_RESULTADO:
                effective_prob_val, effective_impact_val = _matrix_threshold_axis_values(matriz, prob_val, impact_val)
                result_value = _matrix_result_from_axis_values(effective_prob_val, effective_impact_val)
                valor_cons_total = effective_impact_val if effective_impact_val is not None else result_value
                if prob_val is None:
                    prob_val = celda.probabilidad.valor
                valor_criticidad_equipo = result_value if result_value is not None else celda.resultado_num
                indicador = f'[ {_format_matrix_decimal(valor_criticidad_equipo)} -> {_format_matrix_decimal(celda.resultado_num)} ]'
            else:
                valor_cons_total = celda.impacto_nivel.valor
                prob_val = celda.probabilidad.valor
                valor_criticidad_equipo = celda.resultado_num
                indicador = f'[ {_format_matrix_decimal(valor_cons_total)} - {_format_matrix_decimal(prob_val)} ]'
            criticidad_final = celda.clasificacion or criticidad_final

    # 4) Fallback final: si no encontró celda, calcular con los niveles ya resueltos
    if valor_cons_total is None:
        valor_cons_total = impact_val

    if valor_criticidad_equipo is None and valor_cons_total is not None and prob_val is not None:
        try:
            valor_criticidad_equipo = Decimal(valor_cons_total) * Decimal(prob_val)
            indicador = f'[ {_format_matrix_decimal(valor_cons_total)} - {_format_matrix_decimal(prob_val)} ]'
        except Exception:
            valor_criticidad_equipo = evaluacion.valor_criticidad_equipo

    trace = {}
    if matriz and valor_criticidad_equipo is not None:
        source_values = dimension_source_values(dims)
        _add_matrix_axis_source_values(source_values, matriz, prob_val, valor_cons_total, valor_criticidad_equipo)
        ruled = apply_criticality_rules(
            matriz,
            base_cell=celda,
            base_value=valor_criticidad_equipo,
            base_classification=criticidad_final,
            source_values=source_values,
        )
        valor_criticidad_equipo = ruled['value']
        criticidad_final = ruled['classification']
        trace = ruled['trace']

    evaluacion.frecuencia_normalizada = prob_val
    evaluacion.valor_cons_total = valor_cons_total
    evaluacion.valor_criticidad_equipo = valor_criticidad_equipo
    evaluacion.indicador_criticidad = indicador
    evaluacion.criticidad_final = criticidad_final
    evaluacion.trazabilidad_criticidad_json = json.dumps(trace, ensure_ascii=False) if trace else ''
    evaluacion.save(
        update_fields=[
            'frecuencia_normalizada',
            'valor_cons_total',
            'valor_criticidad_equipo',
            'indicador_criticidad',
            'criticidad_final',
            'trazabilidad_criticidad_json',
        ]
    )



def _aca_progress_context(servicio, params):
    selected_progress_nivel = (params.get('progress_nivel') or '').strip()
    selected_progress_nodo = (params.get('progress_nodo') or '').strip()
    selected_avance_min, selected_avance_max = _progress_range_from_params(params)
    selected_progress_node_ids = [
        value.strip()
        for value in selected_progress_nodo.replace(';', ',').split(',')
        if value.strip().isdigit()
    ]
    selected_progress_chart_level = (
        params.get('progress_chart_level')
        or params.get('progress_group_level')
        or ''
    ).strip()

    hierarchy_filters = get_hierarchy_filter_options(servicio)
    hierarchy_filters['filter_nodes'] = [
        node for node in hierarchy_filters['nodes']
        if not selected_progress_nivel or str(node['level_id']) == selected_progress_nivel
    ]

    criticidades_queryset = get_aca_criticidad_queryset(servicio).order_by(
        '-aca_carga__fecha_analisis',
        'equipo__tag_equipo',
        'id',
    )
    criticidades_queryset = filter_criticidades_by_hierarchy(
        criticidades_queryset,
        selected_progress_nodo,
    )
    criticidades = list(criticidades_queryset)
    progress_summary = build_aca_service_progress_summary(servicio, criticidades=criticidades)
    progress_by_criticidad_id = progress_summary.get('progress_by_criticidad_id', {})
    criticidades = _filter_criticidades_by_progress_range(
        criticidades,
        progress_by_criticidad_id,
        selected_avance_min,
        selected_avance_max,
    )
    if selected_avance_min != 0 or selected_avance_max != 100:
        progress_summary = build_aca_service_progress_summary(servicio, criticidades=criticidades)
    progress_dimensions = get_aca_progress_dimensions(servicio.estrategia)
    hierarchy_chart_summary = group_progress_by_hierarchy_level(
        criticidades,
        selected_progress_chart_level,
        progress_dimensions,
    )

    return {
        'progress_criticidades': criticidades,
        'progress_summary': progress_summary,
        'hierarchy_filters': hierarchy_filters,
        'hierarchy_chart_summary': hierarchy_chart_summary,
        'selected_progress_nivel': selected_progress_nivel,
        'selected_progress_nodo': selected_progress_nodo,
        'selected_progress_node_ids': selected_progress_node_ids,
        'selected_progress_chart_level': selected_progress_chart_level,
        'selected_avance_min': selected_avance_min,
        'selected_avance_max': selected_avance_max,
    }


def _progress_range_from_params(params):
    def parse(value, fallback):
        try:
            number = int(value)
        except (TypeError, ValueError):
            return fallback
        return max(0, min(100, number))

    min_value = parse(params.get('avance_min'), 0)
    max_value = parse(params.get('avance_max'), 100)
    if min_value > max_value:
        min_value, max_value = max_value, min_value
    return min_value, max_value


def _filter_criticidades_by_progress_range(criticidades, progress_by_criticidad_id, min_value, max_value):
    if min_value <= 0 and max_value >= 100:
        return criticidades
    filtered = []
    for criticidad in criticidades:
        progress = progress_by_criticidad_id.get(criticidad.pk, {})
        progress_percent = progress.get('progress_percent')
        if progress_percent is None:
            continue
        if Decimal(min_value) <= progress_percent <= Decimal(max_value):
            filtered.append(criticidad)
    return filtered


def _aca_excel_sample_rows(importer, service, rows, limit=10):
    sample = []
    existing_equipment_ids = set(
        models.Criticidad.objects.filter(
            aca_carga__servicio=service,
            equipo_id__isnull=False,
        ).values_list('equipo_id', flat=True)
    )
    for row in rows[:limit]:
        tag = str(row.get('tag') or '').strip()
        ut = str(row.get('ut') or '').strip()
        name = str(row.get('equipo_componente') or '').strip()
        equipo = None
        equipment_qs = get_service_equipment(service)
        if tag:
            equipo = equipment_qs.filter(tag_equipo=tag).first() or models.Equipo.objects.filter(tag_equipo=tag).first()
        if not equipo and ut:
            equipo = equipment_qs.filter(ut=ut).first() or models.Equipo.objects.filter(ut=ut).first()
        status = 'Listo'
        if equipo and equipo.pk in existing_equipment_ids:
            status = 'Ya tiene ACA'
        elif not equipo:
            status = 'Equipo no encontrado'
        sample.append({
            'excel_row': row.get('_excel_row'),
            'tag': tag,
            'ut': ut,
            'equipo_excel': name,
            'equipo_sistema': equipo.nombre_equipo if equipo else '',
            'status': status,
        })
    return sample


def _aca_preview_dimension_value(item):
    if not item:
        return ''
    if item.get('valor_booleano') is not None:
        return 'Sí' if item.get('valor_booleano') else 'No'
    if item.get('valor_numerico') is not None:
        value = item.get('valor_numerico')
        try:
            return str(int(value)) if value == int(value) else str(value)
        except Exception:
            return str(value)
    if item.get('valor_secundario') is not None:
        return str(item.get('valor_secundario'))
    if item.get('valor_texto'):
        return str(item.get('valor_texto'))
    escala_valor = item.get('escala_valor')
    if escala_valor:
        return str(escala_valor.valor_numerico or escala_valor.codigo or escala_valor.descripcion or '')
    catalogo_fila = item.get('catalogo_fila')
    if catalogo_fila:
        values = catalogo_fila.values_map()
        for key in ['valor_numerico', 'valor', 'puntaje', 'nivel', 'valor_texto', 'etiqueta', 'nombre', 'descripcion']:
            value = values.get(key)
            if value not in (None, ''):
                return str(value)
        return catalogo_fila.etiqueta or ''
    return ''


def _resolve_aca_preview_equipment(service, row):
    tag = str(row.get('tag') or '').strip()
    ut = str(row.get('ut') or '').strip()
    equipment_qs = get_service_equipment(service)
    equipo = None
    if tag:
        equipo = equipment_qs.filter(tag_equipo=tag).first() or models.Equipo.objects.filter(tag_equipo=tag).first()
    if not equipo and ut:
        equipo = equipment_qs.filter(ut=ut).first() or models.Equipo.objects.filter(ut=ut).first()
    return equipo


def _aca_excel_preview_rows(importer, service, rows, create_missing_equipment=False, limit=100):
    strategy = service.estrategia
    matrix_selector = _service_matrix_selector(service)
    matriz = matrix_selector.get('matriz')
    excluded_dimension_ids = _aca_excluded_dimension_ids(strategy, matriz)
    dimension_by_alias = importer.dimension_map(strategy, excluded_dimension_ids)
    dimension_columns = sorted(
        {
            ed.pk: {
                'id': ed.pk,
                'name': ed.dimension.nombre,
                'order': ed.orden,
            }
            for ed in dimension_by_alias.values()
        }.values(),
        key=lambda item: (item['order'], item['id']),
    )
    dimension_id_set = {item['id'] for item in dimension_columns}
    existing_equipment_ids = set(
        models.Criticidad.objects.filter(
            aca_carga__servicio=service,
            equipo_id__isnull=False,
        ).values_list('equipo_id', flat=True)
    )
    submitted_equipment_ids = set()
    preview_rows = []

    for row in rows:
        if len(preview_rows) >= limit:
            break
        tag = str(row.get('tag') or '').strip()
        ut = str(row.get('ut') or '').strip()
        name = str(row.get('equipo_componente') or '').strip()
        equipo = _resolve_aca_preview_equipment(service, row)
        status_items = []
        will_load = True
        equipment_status = ''

        if equipo:
            equipment_status = equipo.nombre_equipo
            if equipo.pk in existing_equipment_ids:
                status_items.append('Ya tiene ACA')
                will_load = False
            elif equipo.pk in submitted_equipment_ids:
                status_items.append('Repetido en archivo')
                will_load = False
        elif create_missing_equipment:
            equipment_status = 'Se creará'
        else:
            equipment_status = 'Equipo no encontrado'
            status_items.append('Equipo no encontrado')
            will_load = False

        row_dimension_payload = importer.row_dimensions(row, dimension_by_alias)
        prepared, _source_values, errors = prepare_bulk_dimension_items(
            strategy,
            row_dimension_payload,
            excluded_dimension_ids=excluded_dimension_ids,
            allow_incomplete=False,
        )
        if errors:
            status_items.extend(errors)
            will_load = False

        selected_cell = _resolve_complete_matrix_cell_from_dimension_records(matriz, prepared) if matriz else None
        if matriz and not selected_cell:
            status_items.append('Matriz no resuelta')
            will_load = False

        if equipo and will_load:
            submitted_equipment_ids.add(equipo.pk)

        values_by_ed = {
            item['estrategia_dimension'].pk: _aca_preview_dimension_value(item)
            for item in prepared
            if item['estrategia_dimension'].pk in dimension_id_set
        }
        preview_rows.append({
            'excel_row': row.get('_excel_row'),
            'tag': tag,
            'ut': ut,
            'equipo_excel': name,
            'equipo_sistema': equipment_status,
            'dimension_values': [values_by_ed.get(column['id'], '') for column in dimension_columns],
            'criticidad_valor': selected_cell.resultado_num if selected_cell else row.get('criticidad') or '',
            'criticidad_final': selected_cell.clasificacion if selected_cell else row.get('criticidad_nivel') or '',
            'status': 'Se cargará' if will_load else 'Omitida',
            'status_detail': '; '.join(status_items),
        })

    return dimension_columns, preview_rows


def _run_aca_excel_upload(servicio, uploaded_file, sheet_name, create_missing_equipment=False, replace=False, preview=True):
    from openpyxl import load_workbook
    from aca.management.commands.import_aca_excel import Command as ACAImportCommand

    importer = ACAImportCommand()
    uploaded_file.seek(0)
    workbook = load_workbook(BytesIO(uploaded_file.read()), read_only=True, data_only=True)
    resolved_sheet = (sheet_name or '').strip()
    if not resolved_sheet:
        resolved_sheet = 'ACA' if 'ACA' in workbook.sheetnames else workbook.sheetnames[0]
    if resolved_sheet not in workbook.sheetnames:
        raise ValueError(f'El archivo no tiene hoja "{resolved_sheet}". Hojas: {", ".join(workbook.sheetnames)}')

    rows = importer.read_rows(workbook[resolved_sheet])
    if not rows:
        raise ValueError('No se encontraron filas ACA para importar.')

    origin = f'ACA Excel: {uploaded_file.name}'
    sample_rows = _aca_excel_sample_rows(importer, servicio, rows)
    preview_dimension_columns, preview_rows = _aca_excel_preview_rows(
        importer,
        servicio,
        rows,
        create_missing_equipment=create_missing_equipment,
    )
    now = timezone.now()
    with transaction.atomic():
        if replace:
            importer.delete_previous(servicio, origin)
        result = importer.import_rows(
            service=servicio,
            rows=rows,
            origin=origin,
            create_missing_equipment=create_missing_equipment,
            now=now,
        )
        if preview:
            transaction.set_rollback(True)

    workbook.close()
    return {
        'sheet_name': resolved_sheet,
        'origin': origin,
        'rows_read': len(rows),
        'cargas': result.get('cargas', 0),
        'aca': result.get('aca', 0),
        'dimensiones': result.get('dimensiones', 0),
        'catalogo_fila_asignadas': result.get('catalogo_fila_asignadas', 0),
        'equipos_creados': result.get('equipos_creados', 0),
        'omitidas': result.get('omitidas', 0),
        'warnings': result.get('warnings') or [],
        'sample_rows': sample_rows,
        'preview_dimension_columns': preview_dimension_columns,
        'preview_rows': preview_rows,
        'preview_rows_limited': len(rows) > len(preview_rows),
    }


@login_required
def service_aca_excel_upload(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=True)
    if not permission.get('can_edit'):
        raise PermissionDenied('No tienes permisos para cargar registros ACA en este servicio.')

    report = None
    preview_only = True
    upload_session_key = f'aca_excel_upload_file_{servicio.pk}'
    if request.method == 'POST':
        form = ACAExcelBulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data.get('archivo')
            if uploaded_file:
                store_session_upload(request, upload_session_key, uploaded_file, 'aca')
            stored_file, _stored_ref = open_session_upload(request, upload_session_key)
            if not stored_file:
                form.add_error('archivo', 'Selecciona un archivo Excel para continuar.')
            else:
                form.cleaned_data['archivo'] = stored_file
        if form.is_valid():
            preview_only = request.POST.get('action') != 'confirm'
            try:
                report = _run_aca_excel_upload(
                    servicio,
                    form.cleaned_data['archivo'],
                    form.cleaned_data.get('hoja') or '',
                    create_missing_equipment=form.cleaned_data.get('create_missing_equipment'),
                    replace=form.cleaned_data.get('replace'),
                    preview=preview_only,
                )
                if preview_only:
                    messages.info(request, 'Previsualizacion lista. Puedes confirmar la carga sin volver a seleccionar el archivo.')
                else:
                    try:
                        form.cleaned_data['archivo'].close()
                    except Exception:
                        pass
                    clear_session_upload(request, upload_session_key)
                    messages.success(request, f'Carga ACA completada: {report["aca"]} registros creados.')
            except Exception as exc:
                messages.error(request, str(exc))
            finally:
                try:
                    form.cleaned_data['archivo'].close()
                except Exception:
                    pass
    else:
        clear_session_upload(request, upload_session_key)
        form = ACAExcelBulkUploadForm()
    stored_upload = request.session.get(upload_session_key)

    return render(request, 'core/aca/service_aca_excel_upload.html', {
        'service': servicio,
        'permission': permission,
        'form': form,
        'report': report,
        'preview_only': preview_only,
        'stored_upload': stored_upload,
    })


@login_required
def service_aca_excel_template(request, pk):
    servicio, _permission = _service_or_404(request, pk, edit=False)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    matrix_selector = _service_matrix_selector(servicio)
    matriz = matrix_selector.get('matriz')
    excluded_dimension_ids = _aca_excluded_dimension_ids(servicio.estrategia, matriz) if servicio.estrategia_id else set()
    dimensions = get_aca_bulk_dimensions(servicio.estrategia, excluded_dimension_ids)
    dimension_headers = []
    seen_headers = {'tag', 'ut', 'equipo / componente'}
    for estrategia_dimension in dimensions:
        name = (estrategia_dimension.dimension.nombre or '').strip()
        if not name:
            continue
        key = name.lower()
        if key in seen_headers:
            continue
        seen_headers.add(key)
        dimension_headers.append(name)

    headers = [
        'TAG',
        'UT',
        'Equipo / Componente',
        *dimension_headers,
        'Criticidad final',
    ]
    rows = [
        [
            'A001',
            'MIN-CH-MOL-PULP-BOM-A001',
            'Bomba de pulpa 1',
            *(['1'] * len(dimension_headers)),
            '',
        ],
        [
            'A002',
            'MIN-CH-MOL-PULP-BOM-A002',
            'Bomba de pulpa 2',
            *(['2'] * len(dimension_headers)),
            '',
        ],
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'ACA'
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill('solid', fgColor='EAF1FF')
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    worksheet.freeze_panes = 'A2'
    for col_idx, column_cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 14), 36)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f'plantilla_aca_{servicio.codigo_servicio}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def service_aca_list(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=False)
    progress_context = _aca_progress_context(servicio, request.GET)
    criticidades = progress_context['progress_criticidades']

    columns, rows, complete_count, incomplete_count, _progress_summary = _service_aca_table_data(
        servicio,
        include_actions=True,
        criticidades=criticidades,
        include_evaluation_columns=True,
    )
    aca_groups = _service_aca_row_groups(rows)
    aca_list_items = _service_aca_list_items(aca_groups)

    other_aca_services = []
    if not rows:
        accessible_service_ids = [item.pk for item in get_accessible_services(request.user)]
        other_aca_services = list(
            models.Criticidad.objects.filter(
                aca_carga__servicio_id__in=accessible_service_ids,
            )
            .exclude(aca_carga__servicio=servicio)
            .values(
                'aca_carga__servicio_id',
                'aca_carga__servicio__codigo_servicio',
            )
            .annotate(total=Count('id'))
            .order_by('aca_carga__servicio__codigo_servicio')
        )

    context = {
        'service': servicio,
        'permission': permission,
        'columns': columns,
        'rows': rows,
        'aca_groups': aca_groups,
        'aca_list_items': aca_list_items,
        'aca_table_colspan': len(columns) + 1,
        'aca_count': len(rows),
        'complete_count': complete_count,
        'incomplete_count': incomplete_count,
        'other_aca_services': other_aca_services,
    }
    context.update(progress_context)
    return render(request, 'core/aca/service_aca_list.html', context)


@login_required
def service_aca_progress_partial(request, pk):
    servicio, _permission = _service_or_404(request, pk, edit=False)
    context = {
        'service': servicio,
    }
    context.update(_aca_progress_context(servicio, request.GET))
    return render(request, 'core/aca/partials/aca_progress_summary.html', context)


def _node_path_for_panel(node, node_by_id):
    path = []
    seen = set()
    while node and node.pk not in seen:
        seen.add(node.pk)
        path.append(node)
        node = node_by_id.get(node.parent_id)
    return list(reversed(path))


def _record_node_at_level(record, level_id, node_by_id):
    equipo = getattr(record, 'equipo', None)
    node = node_by_id.get(getattr(equipo, 'nodo_id', None))
    for path_node in _node_path_for_panel(node, node_by_id):
        if str(path_node.nivel_id) == str(level_id):
            return path_node
    return None


def _aca_panel_progress_context(servicio, params):
    hierarchy_filters = get_hierarchy_filter_options(servicio)
    selected_avance_min, selected_avance_max = _progress_range_from_params(params)
    available_node_ids = {int(node['id']) for node in hierarchy_filters.get('nodes', []) if node.get('id')}
    selected_node_ids = {
        int(node_id)
        for node_id in params.getlist('panel_nodes')
        if str(node_id).isdigit() and int(node_id) in available_node_ids
    }
    levels = hierarchy_filters.get('levels', [])
    level_order = {int(level['id']): index for index, level in enumerate(levels)}
    available_nodes = list(
        models.NodoJerarquia.objects.filter(pk__in=available_node_ids, activo=True)
        .select_related('nivel')
        .order_by('nivel__orden', 'orden', 'codigo', 'nombre')
    )
    node_by_id = {node.pk: node for node in available_nodes}
    selected_nodes = [node_by_id[node_id] for node_id in selected_node_ids if node_id in node_by_id]
    deepest_selected = None
    if selected_nodes:
        deepest_selected = max(
            selected_nodes,
            key=lambda node: level_order.get(node.nivel_id, -1),
        )
    deepest_level_id = deepest_selected.nivel_id if deepest_selected else None

    filter_node_ids = []
    if deepest_selected:
        filter_node_ids = []
        deepest_selected_ids = [
            node.pk for node in selected_nodes if node.nivel_id == deepest_level_id
        ]
        for node_id in deepest_selected_ids:
            filter_node_ids.extend(get_descendant_node_ids(node_id))

    criticidades_queryset = get_aca_criticidad_queryset(servicio).order_by(
        '-aca_carga__fecha_analisis',
        'equipo__tag_equipo',
        'id',
    )
    if filter_node_ids:
        criticidades_queryset = criticidades_queryset.filter(equipo__nodo_id__in=filter_node_ids)
    criticidades = list(criticidades_queryset)
    initial_progress_summary = build_aca_service_progress_summary(servicio, criticidades=criticidades)
    criticidades = _filter_criticidades_by_progress_range(
        criticidades,
        initial_progress_summary.get('progress_by_criticidad_id', {}),
        selected_avance_min,
        selected_avance_max,
    )

    all_company_nodes = list(
        models.NodoJerarquia.objects.filter(empresa=servicio.empresa, activo=True)
        .select_related('nivel')
    )
    all_node_by_id = {node.pk: node for node in all_company_nodes}

    selected_by_level = {}
    for node in selected_nodes:
        selected_by_level.setdefault(node.nivel_id, set()).add(node.pk)

    allowed_parent_ids = None
    panel_levels = []
    for level in levels:
        level_id = int(level['id'])
        level_nodes = [node for node in available_nodes if node.nivel_id == level_id]
        if allowed_parent_ids:
            level_nodes = [
                node for node in level_nodes
                if any(path_node.pk in allowed_parent_ids for path_node in _node_path_for_panel(node, all_node_by_id))
            ]
        level_nodes.sort(key=lambda node: (node.codigo or '', node.nombre or ''))
        for node in level_nodes:
            code = '-'.join(path_node.codigo for path_node in _node_path_for_panel(node, all_node_by_id) if path_node.codigo)
            name = node.nombre or node.codigo or 'Sin nombre'
            node.panel_label = f'{name} - {code}' if code else name
        selected_in_level = selected_by_level.get(level_id, set())
        panel_levels.append({
            'id': level_id,
            'name': level['name'],
            'nodes': level_nodes,
            'selected_ids': selected_in_level,
        })
        if selected_in_level:
            allowed_parent_ids = selected_in_level

    progress_summary = build_aca_service_progress_summary(servicio, criticidades=criticidades)
    progress_by_criticidad_id = progress_summary.get('progress_by_criticidad_id', {})
    progress_dimensions = get_aca_progress_dimensions(servicio.estrategia)
    _detail_columns, panel_detail_rows, _detail_complete, _detail_incomplete, _detail_progress = _service_aca_table_data(
        servicio,
        criticidades=criticidades,
    )
    chart_level_id = deepest_level_id or (levels[0]['id'] if levels else '')
    hierarchy_chart_summary = group_progress_by_hierarchy_level(
        criticidades,
        chart_level_id,
        progress_dimensions,
    )
    table_level_name = next((level['name'] for level in levels if str(level['id']) == str(chart_level_id)), 'Nivel')

    table_rows = []
    for record in criticidades:
        level_node = _record_node_at_level(record, chart_level_id, all_node_by_id) if chart_level_id else None
        carga = getattr(record, 'aca_carga', None)
        progress = progress_by_criticidad_id.get(record.pk, {})
        progress_percent = progress.get('progress_percent')
        node_path_ids = []
        if record.equipo and record.equipo.nodo_id:
            node_path_ids = [str(node.pk) for node in _node_path_for_panel(record.equipo.nodo, all_node_by_id)]
        table_rows.append({
            'id': record.pk,
            'level_value': (level_node.nombre or level_node.codigo) if level_node else '-',
            'ut': record.equipo.ut_display if record.equipo else '',
            'descripcion_ut': record.equipo.descripcion_ut if record.equipo else '',
            'created': timezone.localtime(carga.creado_en).strftime('%d/%m/%Y %H:%M') if carga and carga.creado_en else '',
            'updated': timezone.localtime(carga.actualizado).strftime('%d/%m/%Y %H:%M') if carga and carga.actualizado else '',
            'date_order': timezone.localtime(carga.actualizado or carga.creado_en).strftime('%Y%m%d%H%M%S') if carga and (carga.actualizado or carga.creado_en) else '',
            'tag': record.equipo.tag_display if record.equipo else '',
            'progress_percent': progress_percent if progress_percent is not None else '',
            'progress_label': progress.get('progress_label', 'N/A'),
            'progress_order': progress_percent if progress_percent is not None else -1,
            'progress_width': _progress_width_css(progress_percent),
            'progress_color': _progress_color(progress_percent),
            'progress_missing_label': ', '.join(
                item.dimension.nombre for item in progress.get('missing_dimensions', [])
            ) or 'Sin faltantes',
            'node_path_ids': ','.join(node_path_ids),
        })

    return {
        'panel_progress_summary': progress_summary,
        'panel_hierarchy_levels': panel_levels,
        'panel_selected_node_ids': selected_node_ids,
        'panel_chart_summary': hierarchy_chart_summary,
        'panel_table_rows': table_rows,
        'panel_detail_rows': panel_detail_rows,
        'panel_table_level_name': table_level_name,
        'panel_total_filtered': len(criticidades),
        'panel_is_filtered': bool(selected_node_ids) or selected_avance_min != 0 or selected_avance_max != 100,
        'panel_avance_min': selected_avance_min,
        'panel_avance_max': selected_avance_max,
    }


def _aca_global_services_context(request, selected_service_id=None):
    services = list(get_accessible_services(request.user))
    selected_service = None
    selected_summary = None
    service_rows = []

    for service in services:
        summary = build_aca_service_progress_summary(service)
        service_rows.append({
            'service': service,
            'summary': summary,
            'permission': get_service_permission(request.user, service),
        })
        if selected_service_id and str(service.pk) == str(selected_service_id):
            selected_service = service
            selected_summary = summary

    if selected_service_id and not selected_service:
        raise PermissionDenied('No tienes acceso al servicio seleccionado.')

    context = {
        'services': services,
        'service_rows': service_rows,
        'selected_service': selected_service,
        'selected_summary': selected_summary,
        'selected_service_id': str(selected_service_id or ''),
    }
    if selected_service:
        context.update(_aca_panel_progress_context(selected_service, request.GET))
    return context


@login_required
def aca_panel(request):
    context = _aca_global_services_context(request, request.GET.get('service'))
    return render(request, 'core/aca/aca_panel.html', context)


@login_required
def aca_development(request):
    context = _aca_global_services_context(request)
    for row in context.get('service_rows', []):
        summary = row.get('summary') or {}
        progress_percent = summary.get('average_progress_percent')
        row['progress_width'] = _progress_width_css(progress_percent)
        row['progress_color'] = _progress_color(progress_percent)
        row['progress_label'] = summary.get('average_progress_label', 'N/A')
        row['progress_order'] = float(progress_percent) if progress_percent is not None else -1
    return render(request, 'core/aca/aca_development.html', context)


def _aca_group_key_and_label(crit):
    carga = crit.aca_carga
    origen = (getattr(carga, 'origen', '') or '').strip()
    creado = getattr(carga, 'creado_en', None)
    created_local = timezone.localtime(creado) if creado else None
    stamp = created_local.strftime('%d/%m/%Y %H:%M') if created_local else ''
    stamp_key = created_local.strftime('%Y%m%d%H%M%S%f') if created_local else str(getattr(carga, 'pk', '') or '')
    if origen.startswith('Manual familia:'):
        return (
            f'grupo:{stamp_key}',
            'Carga grupal ACA',
            stamp,
            'grupo',
            getattr(carga, 'pk', None),
        )
    if origen == 'Manual masivo':
        return (
            f'grupo:{stamp_key}',
            'Carga grupal ACA',
            stamp,
            'grupo',
            getattr(carga, 'pk', None),
        )
    return (
        'individuales',
        'Registros individuales',
        '',
        'individual',
        None,
    )


def _service_aca_row_groups(rows):
    groups = []
    group_map = {}
    for row in rows:
        key = row.get('_group_key') or 'individuales'
        if key not in group_map:
            group = {
                'key': key,
                'label': row.get('_group_label') or 'Registros ACA',
                'meta': row.get('_group_meta') or '',
                'kind': row.get('_group_kind') or 'individual',
                'anchor': row.get('_group_anchor'),
                'is_draft': False,
                'rows': [],
            }
            group_map[key] = group
            groups.append(group)
        if row.get('carga_status') == models.Carga.STATUS_INCOMPLETO:
            group_map[key]['is_draft'] = True
        group_map[key]['rows'].append(row)
    for group in groups:
        group['count'] = len(group['rows'])
        progress_values = []
        for row in group['rows']:
            try:
                value = row.get('avance_aca_order')
                if value is not None and value >= 0:
                    progress_values.append(Decimal(str(value)))
            except Exception:
                continue
        average_progress = (sum(progress_values) / Decimal(len(progress_values))) if progress_values else None
        group['avance_aca'] = f"{_progress_width_css(average_progress)}%" if average_progress is not None else 'N/A'
        group['avance_aca_order'] = average_progress if average_progress is not None else -1
        group['avance_aca_width'] = _progress_width_css(average_progress)
        group['avance_aca_color'] = _progress_color(average_progress)
    return groups


def _service_aca_list_items(groups):
    items = []
    for group in groups:
        if group.get('kind') == 'grupo':
            items.append({'type': 'group', 'group': group})
            continue
        for row in group.get('rows', []):
            items.append({'type': 'row', 'row': row})
    return items


def _service_aca_table_data(servicio, include_actions=False, criticidades=None, include_evaluation_columns=False):
    estrategia_dims = _strategy_dimensions(
        servicio.estrategia,
        proceso=models.EstrategiaDimension.PROCESO_ACA,
    )

    estrategia_dims = [
        ed for ed in estrategia_dims
        if not _is_generated_matrix_axis_dimension(ed)
    ]
    listado_estrategia_dims = [
        ed for ed in estrategia_dims
        if getattr(ed, 'visible_en_listado_aca', True)
    ]

    if criticidades is None:
        criticidades = list(
            models.Criticidad.objects.filter(aca_carga__servicio=servicio)
            .select_related('equipo', 'equipo__nodo', 'equipo__nodo__nivel', 'aca_carga')
            .prefetch_related(
                Prefetch(
                    'dimensiones',
                    queryset=models.CriticidadDimension.objects.select_related(
                        'dimension', 'estrategia_dimension', 'catalogo_fila__catalogo', 'escala_valor', 'escala_unificada'
                    ).prefetch_related('catalogo_fila__celdas__columna')
                )
            )
            .order_by('-aca_carga__fecha_analisis', 'equipo__tag_equipo', 'id')
        )

    matrix_info = _service_matrix_selector(servicio)
    matriz = matrix_info.get('matriz')
    cells_by_classification = {}
    if matriz:
        for cell in models.MatrizRiesgoCelda.objects.filter(matriz=matriz).order_by('id'):
            classification = _calc_slug(cell.clasificacion or '')
            if classification and classification not in cells_by_classification:
                cells_by_classification[classification] = cell
    progress_summary = build_aca_service_progress_summary(servicio, criticidades=criticidades)
    progress_by_criticidad_id = progress_summary.get('progress_by_criticidad_id', {})

    columns = [
        ('avance_aca', 'Avance'),
        ('fechas', 'Fechas'),
        ('ubicacion_tecnica', 'Ubicación Técnica'),
        ('descripcion_ut', 'Descripción U.Técnica'),
        ('equipo', 'Equipo'),
        ('tag', 'TAG'),
        ('observacion', 'Observación'),
    ]
    if include_evaluation_columns:
        for ed in listado_estrategia_dims:
            columns.append((f'dim_{ed.dimension_id}', ed.dimension.nombre))
        columns.extend([
            # ('frecuencia_original', 'Frecuencia Original'),
            # ('frecuencia_normalizada', 'Frecuencia Normalizada'),
            # ('valor_cons_total', 'Valor Consecuencia Total'),
            # ('indicador_criticidad', 'Indicador Criticidad'),
        ])
    columns.append(('criticidad_final', 'Criticidad Final'))
    if include_actions:
        columns.append(('acciones', 'Acciones'))

    rows = []
    complete_count = 0
    incomplete_count = 0
    for crit in criticidades:
        crit_dims = list(crit.dimensiones.all())
        criticality_trace = _json_loads_safe(getattr(crit, 'trazabilidad_criticidad_json', ''), {})
        matrix_cell = _matrix_cell_for_criticidad_row(
            matriz,
            crit,
            crit_dims,
            cells_by_classification,
        )
        if criticality_trace.get('reglas_aplicadas') and crit.criticidad_final:
            matrix_cell = cells_by_classification.get(_calc_slug(crit.criticidad_final), matrix_cell)
        criticidad_color = criticality_trace.get('color_final') or getattr(matrix_cell, 'color', '') or ''
        progress = progress_by_criticidad_id.get(crit.pk, {})
        progress_percent = progress.get('progress_percent')
        status = models.Carga.STATUS_COMPLETO if progress_percent == Decimal('100.0') else ''
        if status == models.Carga.STATUS_COMPLETO:
            complete_count += 1
        else:
            incomplete_count += 1
        missing_names = [item.dimension.nombre for item in progress.get('missing_dimensions', [])]
        show_matrix_values = not matriz or bool(matrix_cell)
        carga = crit.aca_carga
        fecha_creado = ''
        ultimo_avance = ''
        if carga:
            if carga.creado_en:
                fecha_creado = timezone.localtime(carga.creado_en).strftime('%d/%m/%Y %H:%M')
            elif carga.fecha_analisis:
                fecha_creado = carga.fecha_analisis.strftime('%d/%m/%Y')
            if carga.actualizado:
                ultimo_avance = timezone.localtime(carga.actualizado).strftime('%d/%m/%Y %H:%M')
        row = {
            'id': crit.id,
            'cliente': servicio.codigo_servicio,
            'status': status,
            'carga_status': carga.status if carga else '',
            'avance_aca': progress.get('progress_label', 'N/A'),
            'avance_aca_order': progress_percent if progress_percent is not None else -1,
            'avance_aca_width': _progress_width_css(progress_percent),
            'avance_aca_color': _progress_color(progress_percent),
            'avance_aca_missing_label': ', '.join(missing_names) if missing_names else 'Sin faltantes',
            'fecha_creado': fecha_creado,
            'ultimo_avance': ultimo_avance,
            'fechas': f'{fecha_creado} {ultimo_avance}'.strip(),
            'ubicacion_tecnica': crit.equipo.ut_display if crit.equipo else '',
            'descripcion_ut': crit.equipo.descripcion_ut if crit.equipo else '',
            'equipo': crit.equipo.nombre_equipo if crit.equipo else '',
            'tag': crit.equipo.tag_display if crit.equipo else '',
            'observacion': crit.observacion,
            'frecuencia_original': crit.frecuencia_original,
            'frecuencia_normalizada': crit.frecuencia_normalizada,
            'valor_cons_total': crit.valor_cons_total if show_matrix_values else '',
            'indicador_criticidad': crit.indicador_criticidad if show_matrix_values else '',
            'valor_criticidad_equipo': crit.valor_criticidad_equipo if show_matrix_values else '',
            'criticidad_final': crit.criticidad_final if show_matrix_values else '',
            'criticidad_final_color': criticidad_color,
            'criticidad_final_text_color': _contrast_text_color(criticidad_color),
            'criticidad_base': criticality_trace.get('criticidad_base', ''),
            'valor_criticidad_base': criticality_trace.get('valor_base'),
            'regla_criticidad': criticality_trace.get('regla_aplicada', ''),
            'reglas_criticidad': criticality_trace.get('reglas_aplicadas', []),
        }
        group_key, group_label, group_meta, group_kind, group_anchor = _aca_group_key_and_label(crit)
        row['_group_key'] = group_key
        row['_group_label'] = group_label
        row['_group_meta'] = group_meta
        row['_group_kind'] = group_kind
        row['_group_anchor'] = group_anchor
        dims_map = {item.dimension_id: item for item in crit_dims}
        row['_dimension_details'] = []
        for ed in estrategia_dims:
            item = dims_map.get(ed.dimension_id)
            display_value = _dimension_display_value(item) if item else ''
            comment_value = (item.comentario or '').strip() if item and not display_value else ''
            row[f'dim_{ed.dimension_id}'] = display_value
            row[f'comment_dim_{ed.dimension_id}'] = comment_value
            row['_dimension_details'].append({
                'name': ed.dimension.nombre,
                'value': display_value,
                'comment': comment_value,
            })
        rows.append(row)

    return columns, rows, complete_count, incomplete_count, progress_summary


@login_required
def service_aca_export(request, pk, formato):
    servicio, _permission = _service_or_404(request, pk, edit=False)
    columns, rows, _complete_count, _incomplete_count, _progress_summary = _service_aca_table_data(
        servicio,
        include_evaluation_columns=True,
    )
    formato = (formato or '').lower()
    if formato == 'excel':
        return _export_xlsx_response(
            _export_filename('ACA', servicio, 'xlsx'),
            'Registros ACA',
            columns,
            rows,
        )
    if formato == 'pdf':
        return _export_pdf_response(
            _export_filename('ACA', servicio, 'pdf'),
            f'Registros ACA - {servicio.codigo_servicio}',
            columns,
            rows,
        )
    raise Http404('Formato de exportación no soportado.')


def _bulk_row_is_empty(row):
    if not isinstance(row, dict):
        return True
    if row.get('equipo_id') or row.get('family_id'):
        return False
    dimensions = row.get('dimensions') if isinstance(row.get('dimensions'), dict) else {}
    for value in dimensions.values():
        if not isinstance(value, dict):
            continue
        if any(value.get(key) not in (None, '') for key in [
            'catalogo_fila_id',
            'escala_valor_id',
            'valor_numerico',
            'valor_texto',
            'valor_booleano',
        ]):
            return False
    return True


def _bulk_initial_equipment_payload(servicio, limit=40):
    return _equipment_items_payload(list(get_service_equipment(servicio).order_by('tag_equipo', 'nombre_equipo', 'ut')[:limit]))


def _bulk_payload_with_equipment(servicio, raw_payload):
    if not raw_payload:
        return raw_payload
    try:
        payload = json.loads(raw_payload)
    except (TypeError, json.JSONDecodeError):
        return raw_payload
    rows = payload.get('rows') if isinstance(payload, dict) else None
    if not isinstance(rows, list):
        return raw_payload
    ids = [row.get('equipo_id') for row in rows if isinstance(row, dict) and row.get('equipo_id') and not row.get('equipo')]
    if not ids:
        return raw_payload
    equipment_map = {
        str(item['id']): item
        for item in _equipment_items_payload(get_service_equipment(servicio).filter(pk__in=ids))
    }
    for row in rows:
        if not isinstance(row, dict) or row.get('equipo'):
            continue
        equipo = equipment_map.get(str(row.get('equipo_id')))
        if equipo:
            row['equipo'] = equipo
    return json.dumps(payload, ensure_ascii=False)


def _bulk_dimension_display_value(item):
    def display_part(value):
        if value in (None, ''):
            return ''
        try:
            if value == int(value):
                return str(int(value))
        except (TypeError, ValueError):
            pass
        return str(value)

    if item is None:
        return ''
    if item.catalogo_fila_id:
        try:
            ordered_values = []
            for celda in item.catalogo_fila.celdas.all().order_by('columna__orden', 'columna_id'):
                value = display_part(celda.python_value)
                if value:
                    ordered_values.append(value)
            if not ordered_values:
                values = item.catalogo_fila.values_map()
                for key in ['valor_numerico', 'valor', 'puntaje', 'nivel', 'valor_texto', 'etiqueta', 'nombre', 'descripcion']:
                    value = display_part(values.get(key))
                    if value:
                        ordered_values.append(value)
            if ordered_values:
                return ' | '.join(dict.fromkeys(ordered_values))
        except Exception:
            pass
    if item.escala_valor_id:
        values = []
        if item.escala_valor.valor_numerico is not None:
            numeric = item.escala_valor.valor_numerico
            values.append(str(int(numeric)) if numeric == int(numeric) else str(numeric))
        label = item.escala_valor.codigo or item.escala_valor.descripcion or ''
        if label:
            values.append(label)
        if values:
            return ' | '.join(dict.fromkeys(values))
    return _dimension_display_value(item)


def _bulk_payload_from_criticidades(servicio, criticidades, matriz=None):
    criticidades = list(criticidades)
    equipment_map = {
        item['id']: item
        for item in _equipment_items_payload([crit.equipo for crit in criticidades if crit.equipo_id])
    }

    def row_payload_from_crit(crit, *, target_type='equipo', family=None, family_critics=None):
        dimensions = {}
        for item in crit.dimensiones.all():
            key = str(item.estrategia_dimension_id or item.dimension_id)
            dimensions[key] = {
                'catalogo_fila_id': item.catalogo_fila_id or '',
                'escala_valor_id': item.escala_valor_id or '',
                'valor_numerico': '' if item.valor_numerico is None else str(item.valor_numerico),
                'valor_booleano': '' if item.valor_booleano is None else item.valor_booleano,
                'valor_texto': item.valor_texto or '',
                'comentario': item.comentario or '',
                'display': _bulk_dimension_display_value(item),
            }

        matrix_cell_id = ''
        if matriz:
            cell = _resolve_complete_matrix_cell_from_dimension_records(
                matriz,
                crit.dimensiones.select_related('dimension', 'estrategia_dimension').all(),
            )
            matrix_cell_id = getattr(cell, 'pk', '') or ''

        row = {
            'criticidad_id': crit.pk if target_type == 'equipo' else '',
            'target_type': target_type,
            'target_locked': True,
            'equipo_id': crit.equipo_id if target_type == 'equipo' else '',
            'equipo': equipment_map.get(crit.equipo_id) if target_type == 'equipo' else None,
            'family_id': family.pk if family else '',
            'family': None,
            'observacion': crit.observacion or '',
            'dimensions': dimensions,
            'matrix_cell_id': matrix_cell_id,
        }
        if family:
            equipos = [item.equipo for item in family.items.all() if item.equipo_id and item.equipo]
            row['family'] = {
                'id': family.pk,
                'nombre': family.nombre,
                'equipos': _equipment_items_payload(equipos),
            }
            row['criticidad_ids'] = [item.pk for item in (family_critics or [])]
        return row

    rows = []
    consumed_ids = set()
    family_groups = {}
    for crit in criticidades:
        origen = (crit.aca_carga.origen or '').strip() if crit.aca_carga else ''
        if not origen.startswith('Manual familia:'):
            continue
        family_name = origen.replace('Manual familia:', '', 1).strip()
        if not family_name:
            continue
        family_groups.setdefault(family_name, []).append(crit)

    for family_name, family_criticidades in family_groups.items():
        group_equipment_ids = {
            crit.equipo_id
            for crit in family_criticidades
            if crit.equipo_id
        }
        family = (
            models.FamiliaEquipo.objects.filter(servicio=servicio, nombre__iexact=family_name)
            .prefetch_related('items__equipo')
            .first()
        )
        if family:
            family_equipment_ids = {
                item.equipo_id
                for item in family.items.all()
                if item.equipo_id
            }
            if not family_equipment_ids or not group_equipment_ids.issubset(family_equipment_ids):
                family = None
        if not family:
            family_candidates = models.FamiliaEquipo.objects.filter(servicio=servicio).prefetch_related('items__equipo')
            for candidate in family_candidates:
                candidate_equipment_ids = {
                    item.equipo_id
                    for item in candidate.items.all()
                    if item.equipo_id
                }
                if candidate_equipment_ids and group_equipment_ids.issubset(candidate_equipment_ids):
                    family = candidate
                    break
        if family:
            rows.append(row_payload_from_crit(
                family_criticidades[0],
                target_type='familia',
                family=family,
                family_critics=family_criticidades,
            ))
            consumed_ids.update(crit.pk for crit in family_criticidades)

    for crit in criticidades:
        if crit.pk in consumed_ids:
            continue
        rows.append(row_payload_from_crit(crit))
    return json.dumps({'rows': rows}, ensure_ascii=False)


def _bulk_group_criticidades(servicio, carga_pk):
    anchor = get_object_or_404(
        models.Carga.objects.filter(servicio=servicio),
        pk=carga_pk,
    )
    origen = (anchor.origen or '').strip()
    if origen not in {'Manual masivo'} and not origen.startswith('Manual familia:'):
        raise Http404('La carga indicada no corresponde a un grupo ACA editable.')

    return list(
        models.Criticidad.objects.filter(
            Q(aca_carga__origen='Manual masivo') | Q(aca_carga__origen__startswith='Manual familia:'),
            aca_carga__servicio=servicio,
            aca_carga__creado_en=anchor.creado_en,
        )
        .select_related('equipo', 'aca_carga')
        .prefetch_related(
            Prefetch(
                'dimensiones',
                queryset=models.CriticidadDimension.objects.select_related(
                    'dimension', 'catalogo_fila__catalogo', 'escala_valor', 'escala_unificada'
                ).prefetch_related('catalogo_fila__celdas__columna')
            )
        )
        .order_by('id')
    )


def _prepare_bulk_rows_for_save(servicio, strategy, matriz, submitted_rows, excluded_dimension_ids, is_draft, existing_by_id=None):
    equipment_qs = get_service_equipment(servicio)
    existing_by_id = existing_by_id or {}
    existing_by_equipment_id = {
        item.equipo_id: item
        for item in existing_by_id.values()
        if item.equipo_id
    }
    existing_allowed_ids = {item.pk for item in existing_by_id.values()}
    existing_equipment_ids = _existing_aca_equipment_ids(servicio, exclude_ids=existing_allowed_ids)
    family_qs = models.FamiliaEquipo.objects.filter(servicio=servicio, activa=True).prefetch_related('items__equipo')
    submitted_equipment_ids = set()
    prepared_rows = []
    errors = []

    for index, row in enumerate(submitted_rows, start=1):
        if _bulk_row_is_empty(row):
            continue

        row_errors = []
        target_type = row.get('target_type') if row.get('target_type') in {'equipo', 'familia'} else 'equipo'
        target_equipos = []
        row_origin = 'Manual masivo'
        if target_type == 'familia':
            family_id = row.get('family_id')
            familia = family_qs.filter(pk=family_id).first() if family_id else None
            if familia:
                target_equipos = sorted(
                    [item.equipo for item in familia.items.all() if item.equipo_id and item.equipo],
                    key=lambda equipo: (
                        (getattr(equipo, 'tag_display', '') or equipo.tag_equipo or '').casefold(),
                        (equipo.nombre_equipo or '').casefold(),
                        (equipo.ut or '').casefold(),
                    ),
                )
                row_origin = f'Manual familia: {familia.nombre}'
                if not target_equipos:
                    row_errors.append('la familia seleccionada no tiene equipos.')
            else:
                row_errors.append('familia requerida o inválida.')
        else:
            equipo_id = row.get('equipo_id')
            if equipo_id:
                equipo = equipment_qs.filter(pk=equipo_id).first()
                if equipo:
                    target_equipos = [equipo]
                else:
                    row_errors.append('equipo inválido o no asociado al servicio.')
            else:
                row_errors.append('equipo requerido.')

        for equipo in target_equipos:
            if equipo.pk in existing_equipment_ids:
                row_errors.append(_duplicate_aca_equipment_message(equipo))
            elif equipo.pk in submitted_equipment_ids:
                label = getattr(equipo, 'tag_display', '') or getattr(equipo, 'tag_equipo', '') or str(equipo)
                row_errors.append(f'el equipo {label} ya fue incluido en otra fila de esta carga.')

        prepared, _source_values, dimension_errors = prepare_bulk_dimension_items(
            strategy,
            row.get('dimensions') if isinstance(row.get('dimensions'), dict) else {},
            excluded_dimension_ids=excluded_dimension_ids,
            allow_incomplete=is_draft,
        )
        row_errors.extend(dimension_errors)

        selected_cell = None
        matrix_cell_id = row.get('matrix_cell_id')
        if matriz and matrix_cell_id:
            selected_cell = models.MatrizRiesgoCelda.objects.filter(
                pk=matrix_cell_id,
                matriz=matriz,
            ).select_related('probabilidad', 'impacto_nivel').first()
            if selected_cell and not _matrix_axis_values_complete(matriz, prepared):
                selected_cell = None
        if matriz and not selected_cell:
            selected_cell = _resolve_complete_matrix_cell_from_dimension_records(matriz, prepared)
        if matriz and not selected_cell and not is_draft:
            row_errors.append('no se pudo resolver la matriz de criticidad.')

        if row_errors:
            errors.append(f"Fila {index}: " + ' '.join(row_errors))
            continue

        for equipo in target_equipos:
            submitted_equipment_ids.add(equipo.pk)
            criticidad_id = row.get('criticidad_id')
            existing = existing_by_equipment_id.get(equipo.pk)
            if not existing and str(criticidad_id or '').isdigit():
                existing = existing_by_id.get(int(criticidad_id))
            prepared_rows.append({
                'existing': existing,
                'equipo': equipo,
                'observacion': str(row.get('observacion') or '').strip(),
                'prepared': prepared,
                'selected_cell': selected_cell,
                'origin': row_origin,
            })

    if not prepared_rows and not errors:
        errors.append('No hay filas válidas para guardar.')
    return prepared_rows, errors


def _delete_aca_records_by_ids(criticidad_ids):
    criticidad_ids = [pk for pk in criticidad_ids if pk]
    if not criticidad_ids:
        return
    carga_ids = list(
        models.Criticidad.objects.filter(pk__in=criticidad_ids).values_list('aca_carga_id', flat=True)
    )
    models.CriticidadDimension.objects.filter(criticidad_id__in=criticidad_ids).delete()
    models.CriticidadAdjunto.objects.filter(criticidad_id__in=criticidad_ids).delete()
    models.Criticidad.objects.filter(pk__in=criticidad_ids).delete()
    models.Carga.objects.filter(pk__in=[pk for pk in carga_ids if pk]).delete()


def _save_bulk_aca_rows(servicio, strategy, profile, prepared_rows, status, origin, existing_group=None):
    now = timezone.now()
    existing_group = list(existing_group or [])
    existing_ids = {item.pk for item in existing_group}
    seen_existing_ids = set()
    group_created = existing_group[0].aca_carga.creado_en if existing_group else now

    for offset, row in enumerate(prepared_rows):
        selected_cell = row['selected_cell']
        existing = row.get('existing')
        row_origin = row.get('origin') or origin
        if existing:
            seen_existing_ids.add(existing.pk)
            carga = existing.aca_carga
            carga.fecha_analisis = timezone.localdate()
            carga.version_carga = _aca_version_for_save(carga, status)
            carga.origen = row_origin
            carga.status = status
            carga.actualizado = now
            carga.estrategia = strategy
            carga.servicio = servicio
            carga.usuario = profile
            carga.save(update_fields=[
                'fecha_analisis',
                'version_carga',
                'origen',
                'status',
                'actualizado',
                'estrategia',
                'servicio',
                'usuario',
            ])

            evaluacion = existing
            evaluacion.equipo = row['equipo']
            evaluacion.escenario_falla = ''
            evaluacion.observacion = row['observacion']
            evaluacion.frecuencia_original = None
            evaluacion.frecuencia_normalizada = selected_cell.probabilidad.valor if selected_cell else None
            evaluacion.valor_cons_total = selected_cell.impacto_nivel.valor if selected_cell else None
            evaluacion.indicador_criticidad = ''
            evaluacion.valor_criticidad_equipo = selected_cell.resultado_num if selected_cell else None
            evaluacion.criticidad_final = selected_cell.clasificacion if selected_cell else ''
            evaluacion.save(update_fields=[
                'equipo',
                'escenario_falla',
                'observacion',
                'frecuencia_original',
                'frecuencia_normalizada',
                'valor_cons_total',
                'indicador_criticidad',
                'valor_criticidad_equipo',
                'criticidad_final',
            ])
            evaluacion.dimensiones.all().delete()
        else:
            carga = models.Carga.objects.create(
                fecha_analisis=timezone.localdate(),
                version_carga=ACA_INITIAL_VERSION,
                origen=row_origin,
                status=status,
                creado_en=group_created,
                actualizado=now,
                estrategia=strategy,
                servicio=servicio,
                usuario=profile,
            )
            evaluacion = models.Criticidad.objects.create(
                creado_en=group_created,
                aca_carga=carga,
                equipo=row['equipo'],
                escenario_falla='',
                observacion=row['observacion'],
                frecuencia_original=None,
                frecuencia_normalizada=selected_cell.probabilidad.valor if selected_cell else None,
                valor_cons_total=selected_cell.impacto_nivel.valor if selected_cell else None,
                indicador_criticidad='',
                valor_criticidad_equipo=selected_cell.resultado_num if selected_cell else None,
                criticidad_final=selected_cell.clasificacion if selected_cell else '',
            )

        _create_dimension_items(evaluacion, row['prepared'])
        if selected_cell:
            _save_matrix_dimensions(evaluacion, strategy, selected_cell)
        _sync_criticidad_resumen(evaluacion, strategy)
        _sync_aca_carga_status(evaluacion, strategy, status)

    stale_ids = existing_ids - seen_existing_ids
    _delete_aca_records_by_ids(stale_ids)


def _bulk_render_context(
    servicio,
    permission,
    strategy,
    matrix_selector,
    dimension_payload,
    errors=None,
    bulk_payload='',
    title='Nueva carga grupal ACA',
    primary_label='Guardar carga grupal',
    is_group_edit=False,
    exclude_existing_crit_ids=None,
    store_separate=False,
):
    return {
        'service': servicio,
        'permission': permission,
        'selected_strategy': strategy,
        'matrix_selector': matrix_selector,
        'dimension_payload': dimension_payload,
        'service_equipment_payload': _service_equipment_browser_payload(servicio),
        'service_equipment_endpoints': _service_equipment_endpoints(servicio),
        'initial_equipment_payload': _bulk_initial_equipment_payload(servicio),
        'service_family_payload': _service_family_payload(servicio),
        'existing_aca_equipment_ids_json': list(
            _existing_aca_equipment_ids(servicio, exclude_ids=exclude_existing_crit_ids)
        ),
        'auto_version': ACA_INITIAL_VERSION,
        'auto_fecha_analisis': timezone.localdate(),
        'bulk_errors': errors or [],
        'bulk_payload': _bulk_payload_with_equipment(servicio, bulk_payload),
        'bulk_title': title,
        'bulk_primary_label': primary_label,
        'is_group_edit': is_group_edit,
        'store_separate': store_separate,
    }


@login_required
def service_aca_bulk_new(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=True)
    if not servicio.estrategia_id:
        messages.warning(request, 'El servicio no tiene estrategia asociada. Asigna una antes de añadir ACA.')
        return redirect('service_detail', pk=servicio.pk)

    strategy = servicio.estrategia
    profile = get_profile_for_user(request.user)
    matrix_selector = _service_matrix_selector(servicio)
    matriz = matrix_selector.get('matriz') if matrix_selector else None
    excluded_dimension_ids = _aca_excluded_dimension_ids(strategy, matriz)
    dimension_payload = get_aca_bulk_dimension_payload(strategy, excluded_dimension_ids)
    raw_payload = request.POST.get('bulk_payload', '') if request.method == 'POST' else ''

    if request.method == 'POST':
        errors = []
        try:
            payload = json.loads(raw_payload or '{}')
        except json.JSONDecodeError:
            payload = {}
            errors.append('No se pudo leer la carga masiva. Revisa el formato enviado.')

        submitted_rows = payload.get('rows') if isinstance(payload, dict) else []
        if not isinstance(submitted_rows, list):
            submitted_rows = []

        is_draft = request.POST.get('save_as') == 'draft'
        store_separate = request.POST.get('store_separate') == '1'
        status = models.Carga.STATUS_INCOMPLETO if is_draft else models.Carga.STATUS_COMPLETO
        equipment_qs = get_service_equipment(servicio)
        family_qs = models.FamiliaEquipo.objects.filter(servicio=servicio, activa=True).prefetch_related('items__equipo')
        existing_equipment_ids = _existing_aca_equipment_ids(servicio)
        submitted_equipment_ids = set()
        prepared_rows = []

        for index, row in enumerate(submitted_rows, start=1):
            if _bulk_row_is_empty(row):
                continue

            row_errors = []
            target_type = row.get('target_type') if row.get('target_type') in {'equipo', 'familia'} else 'equipo'
            target_equipos = []
            row_origin = 'Manual masivo'

            if target_type == 'familia':
                family_id = row.get('family_id')
                familia = family_qs.filter(pk=family_id).first() if family_id else None
                if familia:
                    target_equipos = sorted(
                        [item.equipo for item in familia.items.all() if item.equipo_id and item.equipo],
                        key=lambda equipo: (
                            (getattr(equipo, 'tag_display', '') or equipo.tag_equipo or '').casefold(),
                            (equipo.nombre_equipo or '').casefold(),
                            (equipo.ut or '').casefold(),
                        ),
                    )
                    row_origin = f'Manual familia: {familia.nombre}'
                    if not target_equipos:
                        row_errors.append('la familia seleccionada no tiene equipos.')
                else:
                    row_errors.append('familia requerida o invalida.')
            else:
                equipo_id = row.get('equipo_id')
                if equipo_id:
                    equipo = equipment_qs.filter(pk=equipo_id).first()
                    if equipo:
                        target_equipos = [equipo]
                    else:
                        row_errors.append('equipo invalido o no asociado al servicio.')
                else:
                    row_errors.append('equipo requerido.')

            for equipo in target_equipos:
                if equipo.pk in existing_equipment_ids:
                    row_errors.append(_duplicate_aca_equipment_message(equipo))
                elif equipo.pk in submitted_equipment_ids:
                    label = getattr(equipo, 'tag_display', '') or getattr(equipo, 'tag_equipo', '') or str(equipo)
                    row_errors.append(f'el equipo {label} ya fue incluido en otra fila de esta carga.')

            prepared, _source_values, dimension_errors = prepare_bulk_dimension_items(
                strategy,
                row.get('dimensions') if isinstance(row.get('dimensions'), dict) else {},
                excluded_dimension_ids=excluded_dimension_ids,
                allow_incomplete=is_draft,
            )
            row_errors.extend(dimension_errors)

            selected_cell = None
            matrix_cell_id = row.get('matrix_cell_id')
            if matriz and matrix_cell_id:
                selected_cell = models.MatrizRiesgoCelda.objects.filter(
                    pk=matrix_cell_id,
                    matriz=matriz,
                ).select_related('probabilidad', 'impacto_nivel').first()
                if selected_cell and not _matrix_axis_values_complete(matriz, prepared):
                    selected_cell = None
            if matriz and not selected_cell:
                selected_cell = _resolve_complete_matrix_cell_from_dimension_records(matriz, prepared)
            if matriz and not selected_cell and not is_draft:
                row_errors.append('no se pudo resolver la matriz de criticidad.')

            if row_errors:
                errors.append(f"Fila {index}: " + ' '.join(row_errors))
                continue

            for equipo in target_equipos:
                submitted_equipment_ids.add(equipo.pk)
                prepared_rows.append({
                    'equipo': equipo,
                    'observacion': str(row.get('observacion') or '').strip(),
                    'prepared': prepared,
                    'selected_cell': selected_cell,
                    'origin': row_origin,
                })
        if not prepared_rows and not errors:
            errors.append('No hay filas válidas para guardar.')

        if errors:
            return render(
                request,
                'core/aca/aca_bulk_form.html',
                _bulk_render_context(
                    servicio,
                    permission,
                    strategy,
                    matrix_selector,
                    dimension_payload,
                    errors=errors,
                    bulk_payload=raw_payload,
                    store_separate=store_separate,
                ),
            )

        now = timezone.now()
        with transaction.atomic():
            for row in prepared_rows:
                row_origin = 'Manual' if store_separate else (row.get('origin') or 'Manual masivo')
                carga = models.Carga.objects.create(
                    fecha_analisis=timezone.localdate(),
                    version_carga=ACA_INITIAL_VERSION,
                    origen=row_origin,
                    status=status,
                    creado_en=now,
                    actualizado=now,
                    estrategia=strategy,
                    servicio=servicio,
                    usuario=profile,
                )
                selected_cell = row['selected_cell']
                evaluacion = models.Criticidad.objects.create(
                    creado_en=now,
                    aca_carga=carga,
                    equipo=row['equipo'],
                    escenario_falla='',
                    observacion=row['observacion'],
                    frecuencia_original=None,
                    frecuencia_normalizada=selected_cell.probabilidad.valor if selected_cell else None,
                    valor_cons_total=selected_cell.impacto_nivel.valor if selected_cell else None,
                    indicador_criticidad='',
                    valor_criticidad_equipo=selected_cell.resultado_num if selected_cell else None,
                    criticidad_final=selected_cell.clasificacion if selected_cell else '',
                )
                _create_dimension_items(evaluacion, row['prepared'])
                if selected_cell:
                    _save_matrix_dimensions(evaluacion, strategy, selected_cell)
                _sync_criticidad_resumen(evaluacion, strategy)
                _sync_aca_carga_status(evaluacion, strategy, status)

        messages.success(request, f'Se crearon {len(prepared_rows)} registros ACA correctamente.')
        return redirect('service_aca_list', pk=servicio.pk)

    return render(
        request,
        'core/aca/aca_bulk_form.html',
        _bulk_render_context(servicio, permission, strategy, matrix_selector, dimension_payload),
    )


@login_required
def service_aca_bulk_group_edit(request, pk, carga_pk):
    servicio, permission = _service_or_404(request, pk, edit=True)
    if not servicio.estrategia_id:
        messages.warning(request, 'El servicio no tiene estrategia asociada. Asigna una antes de editar ACA.')
        return redirect('service_detail', pk=servicio.pk)

    strategy = servicio.estrategia
    profile = get_profile_for_user(request.user)
    matrix_selector = _service_matrix_selector(servicio)
    matriz = matrix_selector.get('matriz') if matrix_selector else None
    excluded_dimension_ids = _aca_excluded_dimension_ids(strategy, matriz)
    dimension_payload = get_aca_bulk_dimension_payload(strategy, excluded_dimension_ids)
    existing_group = _bulk_group_criticidades(servicio, carga_pk)
    if not existing_group:
        raise Http404('No se encontraron registros para este grupo ACA.')

    origin = (existing_group[0].aca_carga.origen or 'Manual masivo').strip() or 'Manual masivo'
    initial_payload = _bulk_payload_from_criticidades(servicio, existing_group, matriz=matriz)
    raw_payload = request.POST.get('bulk_payload', '') if request.method == 'POST' else initial_payload

    if request.method == 'POST':
        errors = []
        try:
            payload = json.loads(raw_payload or '{}')
        except json.JSONDecodeError:
            payload = {}
            errors.append('No se pudo leer la carga grupal. Revisa el formato enviado.')

        submitted_rows = payload.get('rows') if isinstance(payload, dict) else []
        if not isinstance(submitted_rows, list):
            submitted_rows = []

        is_draft = request.POST.get('save_as') == 'draft'
        status = models.Carga.STATUS_INCOMPLETO if is_draft else models.Carga.STATUS_COMPLETO
        existing_by_id = {item.pk: item for item in existing_group}
        prepared_rows, row_errors = _prepare_bulk_rows_for_save(
            servicio,
            strategy,
            matriz,
            submitted_rows,
            excluded_dimension_ids,
            is_draft,
            existing_by_id=existing_by_id,
        )
        errors.extend(row_errors)

        if errors:
            return render(
                request,
                'core/aca/aca_bulk_form.html',
                _bulk_render_context(
                    servicio,
                    permission,
                    strategy,
                    matrix_selector,
                    dimension_payload,
                    errors=errors,
                    bulk_payload=raw_payload,
                    title='Editar carga grupal ACA',
                    primary_label='Guardar cambios del grupo',
                    is_group_edit=True,
                    exclude_existing_crit_ids=[item.pk for item in existing_group],
                ),
            )

        with transaction.atomic():
            _save_bulk_aca_rows(
                servicio,
                strategy,
                profile,
                prepared_rows,
                status,
                origin=origin,
                existing_group=existing_group,
            )

        messages.success(request, f'Se actualizaron {len(prepared_rows)} registros ACA del grupo.')
        return redirect('service_aca_list', pk=servicio.pk)

    return render(
        request,
        'core/aca/aca_bulk_form.html',
        _bulk_render_context(
            servicio,
            permission,
            strategy,
            matrix_selector,
            dimension_payload,
            bulk_payload=initial_payload,
            title='Editar carga grupal ACA',
            primary_label='Guardar cambios del grupo',
            is_group_edit=True,
            exclude_existing_crit_ids=[item.pk for item in existing_group],
        ),
    )


@login_required
@transaction.atomic
def service_aca_bulk_group_delete(request, pk, carga_pk):
    servicio, _permission = _service_or_404(request, pk, edit=True)
    if request.method != 'POST':
        return redirect('service_aca_list', pk=servicio.pk)

    group = _bulk_group_criticidades(servicio, carga_pk)
    if not group:
        messages.warning(request, 'No se encontraron registros para eliminar en esta carpeta ACA.')
        return redirect('service_aca_list', pk=servicio.pk)

    count = len(group)
    _delete_aca_records_by_ids([item.pk for item in group])
    messages.success(request, f'Carpeta ACA eliminada correctamente. Se eliminaron {count} registros.')
    return redirect('service_aca_list', pk=servicio.pk)


@login_required
@transaction.atomic
def service_aca_new(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=True)
    if not servicio.estrategia_id:
        messages.warning(request, 'El servicio no tiene estrategia asociada. Asigna una antes de añadir ACA.')
        return redirect('service_detail', pk=servicio.pk)

    strategy = servicio.estrategia
    profile = get_profile_for_user(request.user)
    edit_crit_id = request.POST.get('crit_id') if request.method == 'POST' else request.GET.get('edit')
    edit_crit = None
    if edit_crit_id:
        edit_crit = get_object_or_404(
            models.Criticidad.objects.select_related('aca_carga', 'equipo'),
            pk=edit_crit_id,
            aca_carga__servicio=servicio,
        )

    is_draft = request.method == 'POST' and request.POST.get('save_as') == 'draft'
    existing_is_draft = bool(
        edit_crit
        and getattr(edit_crit.aca_carga, 'status', '') == models.Carga.STATUS_INCOMPLETO
    )
    allow_incomplete = is_draft or existing_is_draft
    matrix_selector = _service_matrix_selector(servicio)
    initial_base = {}
    if edit_crit:
        initial_base = {
            'fecha_analisis': edit_crit.aca_carga.fecha_analisis if edit_crit.aca_carga else timezone.localdate(),
            'version_carga': edit_crit.aca_carga.version_carga if edit_crit.aca_carga else Decimal('1.0'),
            'origen': edit_crit.aca_carga.origen if edit_crit.aca_carga else 'Manual',
            'equipo': edit_crit.equipo,
            'observacion': edit_crit.observacion,
            'frecuencia_normalizada': edit_crit.frecuencia_normalizada,
        }
        matriz = matrix_selector.get('matriz') if matrix_selector else None
        if matriz:
            selected_cell = _resolve_complete_matrix_cell_from_dimension_records(
                matriz,
                edit_crit.dimensiones.select_related('dimension', 'estrategia_dimension').all(),
            )
            if selected_cell:
                initial_base['matrix_celda'] = selected_cell.pk

    base_form = ServicioACARegistroForm(
        request.POST or None,
        initial=initial_base,
        service=servicio,
        allow_incomplete=allow_incomplete,
    )
    excluded_dimension_ids = _aca_excluded_dimension_ids(
        strategy,
        matrix_selector.get('matriz') if matrix_selector else None,
    )
    dimension_initial = _dimension_formset_initial_from_criticidad(
        edit_crit,
        strategy,
        exclude_dimension_ids=excluded_dimension_ids,
    ) if edit_crit else None
    dimension_formset = _dimension_formset(
        request,
        strategy,
        initial=dimension_initial,
        exclude_dimension_ids=excluded_dimension_ids,
    )

    if request.method == 'POST' and base_form.is_valid() and dimension_formset.is_valid():
        attachment_payload, invalid_attachments = _record_attachment_payload(request)
        if invalid_attachments:
            base_form.add_error(
                None,
                'Formato de archivo no permitido: '
                + ', '.join(invalid_attachments)
                + '. Usa PDF, Word, Excel, PowerPoint, CSV, TXT, imagen o ZIP.',
            )
        else:
            status = models.Carga.STATUS_INCOMPLETO if is_draft else models.Carga.STATUS_COMPLETO
            now = timezone.now()
            familia = base_form.cleaned_data.get('familia_equipo') if not edit_crit else None
            observacion = base_form.cleaned_data.get('observacion') or ''

            if edit_crit:
                carga = edit_crit.aca_carga
                carga.version_carga = _aca_version_for_save(carga, status)
                carga.status = status
                carga.actualizado = now
                carga.estrategia = strategy
                carga.servicio = servicio
                carga.usuario = profile
                carga.save(update_fields=['version_carga', 'status', 'actualizado', 'estrategia', 'servicio', 'usuario'])
            elif not familia:
                carga = models.Carga.objects.create(
                    fecha_analisis=timezone.localdate(),
                    version_carga=ACA_INITIAL_VERSION,
                    origen='Manual',
                    status=status,
                    creado_en=now,
                    actualizado=now,
                    estrategia=strategy,
                    servicio=servicio,
                    usuario=profile,
                )
            else:
                carga = None

            selected_cell = base_form.cleaned_data.get('matrix_celda')
            matriz = matrix_selector.get('matriz') if matrix_selector else None
            if not selected_cell and matriz:
                prepared_items, _source_values = _prepare_dimension_items(strategy, dimension_formset)
                selected_cell = _resolve_complete_matrix_cell_from_dimension_records(matriz, prepared_items)
            elif selected_cell and matriz:
                prepared_items, _source_values = _prepare_dimension_items(strategy, dimension_formset)
                if not _matrix_axis_values_complete(matriz, prepared_items):
                    selected_cell = None
            frecuencia_normalizada = selected_cell.probabilidad.valor if selected_cell else None
            valor_cons_total = selected_cell.impacto_nivel.valor if selected_cell else None

            equipos = [base_form.cleaned_data.get('equipo')]
            if familia:
                equipos = [item.equipo for item in familia.items.select_related('equipo').order_by('orden', 'id')]

            existing_equipment_ids = _existing_aca_equipment_ids(
                servicio,
                exclude_ids=[edit_crit.pk] if edit_crit else None,
            )
            duplicate_equipos = [equipo for equipo in equipos if equipo and equipo.pk in existing_equipment_ids]
            if duplicate_equipos:
                for equipo in duplicate_equipos:
                    base_form.add_error('equipo', _duplicate_aca_equipment_message(equipo))
                return render(request, 'core/aca/aca_registro_form.html', {
                    'service': servicio,
                    'permission': permission,
                    'base_form': base_form,
                    'dimension_formset': dimension_formset,
                    'selected_strategy': strategy,
                    'matrix_selector': matrix_selector,
                    'service_equipment_payload': _service_equipment_browser_payload(servicio),
                    'service_equipment_endpoints': _service_equipment_endpoints(servicio),
                    'service_family_payload': _service_family_payload(servicio),
                    'auto_version': edit_crit.aca_carga.version_carga if edit_crit and edit_crit.aca_carga else ACA_INITIAL_VERSION,
                    'auto_fecha_analisis': edit_crit.aca_carga.fecha_analisis if edit_crit and edit_crit.aca_carga else timezone.localdate(),
                    'editing_crit': edit_crit,
                    'existing_attachments': edit_crit.adjuntos.all() if edit_crit else [],
                    'existing_aca_equipment_ids_json': list(_existing_aca_equipment_ids(servicio, exclude_ids=[edit_crit.pk] if edit_crit else None)),
                })

            version_carga = ACA_INITIAL_VERSION if familia else None
            created_count = 0
            for equipo in equipos:
                carga_actual = carga
                if familia:
                    carga_actual = models.Carga.objects.create(
                        fecha_analisis=timezone.localdate(),
                        version_carga=version_carga + (Decimal('0.1') * Decimal(created_count)),
                        origen=f'Manual familia: {familia.nombre}',
                        status=status,
                        creado_en=now,
                        actualizado=now,
                        estrategia=strategy,
                        servicio=servicio,
                        usuario=profile,
                    )
                evaluacion = edit_crit or models.Criticidad(
                    creado_en=now,
                    aca_carga=carga_actual,
                )
                evaluacion.escenario_falla = ''
                evaluacion.observacion = observacion
                evaluacion.frecuencia_original = None
                evaluacion.frecuencia_normalizada = frecuencia_normalizada
                evaluacion.valor_cons_total = valor_cons_total
                evaluacion.indicador_criticidad = ''
                evaluacion.valor_criticidad_equipo = selected_cell.resultado_num if selected_cell else None
                evaluacion.criticidad_final = selected_cell.clasificacion if selected_cell else ''
                evaluacion.aca_carga = carga_actual
                evaluacion.equipo = equipo
                evaluacion.save()

                _save_dimension_formset(evaluacion, strategy, dimension_formset)

                if selected_cell:
                    _save_matrix_dimensions(evaluacion, strategy, selected_cell)

                _sync_criticidad_resumen(evaluacion, strategy)
                _sync_aca_carga_status(evaluacion, strategy, status)
                _save_aca_attachments(evaluacion, attachment_payload, profile)
                created_count += 1

            if is_draft:
                messages.success(request, 'Borrador ACA guardado correctamente.')
            elif edit_crit:
                messages.success(request, 'Registro ACA actualizado correctamente.')
            elif familia:
                messages.success(request, f'Se crearon {created_count} registros ACA para la familia {familia.nombre}.')
            else:
                messages.success(request, 'Registro ACA creado correctamente.')
            return redirect('service_aca_list', pk=servicio.pk)

    return render(request, 'core/aca/aca_registro_form.html', {
        'service': servicio,
        'permission': permission,
        'base_form': base_form,
        'dimension_formset': dimension_formset,
        'selected_strategy': strategy,
        'matrix_selector': matrix_selector,
        'service_equipment_payload': _service_equipment_browser_payload(servicio),
        'service_equipment_endpoints': _service_equipment_endpoints(servicio),
        'service_family_payload': _service_family_payload(servicio),
        'auto_version': edit_crit.aca_carga.version_carga if edit_crit and edit_crit.aca_carga else ACA_INITIAL_VERSION,
        'auto_fecha_analisis': edit_crit.aca_carga.fecha_analisis if edit_crit and edit_crit.aca_carga else timezone.localdate(),
        'editing_crit': edit_crit,
        'existing_attachments': edit_crit.adjuntos.all() if edit_crit else [],
        'existing_aca_equipment_ids_json': list(_existing_aca_equipment_ids(servicio, exclude_ids=[edit_crit.pk] if edit_crit else None)),
    })


@login_required
def service_aca_edit(request, service_pk, crit_pk):
    servicio, _permission = _service_or_404(request, service_pk, edit=True)
    crit = get_object_or_404(
        models.Criticidad.objects.select_related('aca_carga', 'equipo'),
        pk=crit_pk,
        aca_carga__servicio=servicio,
    )

    url = f"{reverse('service_aca_new', kwargs={'pk': servicio.pk})}?edit={crit.pk}"
    return redirect(url)


def _serialize_aca_for_undo(crit):
    carga = crit.aca_carga
    return {
        'service_id': carga.servicio_id if carga else None,
        'restore_url': reverse('service_aca_restore_deleted', kwargs={'service_pk': carga.servicio_id}) if carga else '',
        'carga': {
            'id': carga.pk if carga else None,
            'fecha_analisis': carga.fecha_analisis.isoformat() if carga and carga.fecha_analisis else None,
            'version_carga': str(carga.version_carga) if carga and carga.version_carga is not None else str(ACA_INITIAL_VERSION),
            'origen': carga.origen if carga else 'Manual',
            'status': carga.status if carga else models.Carga.STATUS_INCOMPLETO,
            'creado_en': carga.creado_en.isoformat() if carga and carga.creado_en else timezone.now().isoformat(),
            'actualizado': carga.actualizado.isoformat() if carga and carga.actualizado else timezone.now().isoformat(),
            'estrategia_id': carga.estrategia_id if carga else None,
            'servicio_id': carga.servicio_id if carga else None,
            'usuario_id': carga.usuario_id if carga else None,
        },
        'criticidad': {
            'id': crit.pk,
            'equipo_id': crit.equipo_id,
            'escenario_falla': crit.escenario_falla or '',
            'observacion': crit.observacion or '',
            'frecuencia_original': str(crit.frecuencia_original) if crit.frecuencia_original is not None else None,
            'frecuencia_normalizada': str(crit.frecuencia_normalizada) if crit.frecuencia_normalizada is not None else None,
            'valor_cons_total': str(crit.valor_cons_total) if crit.valor_cons_total is not None else None,
            'indicador_criticidad': crit.indicador_criticidad or '',
            'valor_criticidad_equipo': str(crit.valor_criticidad_equipo) if crit.valor_criticidad_equipo is not None else None,
            'criticidad_final': crit.criticidad_final or '',
            'creado_en': crit.creado_en.isoformat() if crit.creado_en else timezone.now().isoformat(),
        },
        'dimensiones': [
            {
                'dimension_id': item.dimension_id,
                'estrategia_dimension_id': item.estrategia_dimension_id,
                'catalogo_fila_id': item.catalogo_fila_id,
                'escala_valor_id': item.escala_valor_id,
                'escala_unificada_id': item.escala_unificada_id,
                'valor_numerico': str(item.valor_numerico) if item.valor_numerico is not None else None,
                'valor_booleano': item.valor_booleano,
                'valor_texto': item.valor_texto or '',
                'comentario': item.comentario or '',
                'valor_secundario': str(item.valor_secundario) if item.valor_secundario is not None else None,
            }
            for item in crit.dimensiones.all()
        ],
    }


def _restore_deleted_aca_snapshot(snapshot):
    carga_data = snapshot.get('carga') or {}
    crit_data = snapshot.get('criticidad') or {}
    fecha_analisis = parse_date(carga_data.get('fecha_analisis') or '') or timezone.localdate()
    carga_creado = parse_datetime(carga_data.get('creado_en') or '') or timezone.now()
    criticidad_creado = parse_datetime(crit_data.get('creado_en') or '') or timezone.now()
    carga_id = carga_data.get('id')
    criticidad_id = crit_data.get('id')
    if criticidad_id and models.Criticidad.objects.filter(pk=criticidad_id).exists():
        raise ValueError(f'El ID ACA original {criticidad_id} ya esta ocupado.')

    existing_carga = models.Carga.objects.filter(pk=carga_id).first() if carga_id else None
    if existing_carga:
        carga = existing_carga
    else:
        carga_create_kwargs = {}
        if carga_id:
            carga_create_kwargs['id'] = carga_id
        carga = models.Carga.objects.create(
            **carga_create_kwargs,
            fecha_analisis=fecha_analisis,
            version_carga=Decimal(str(carga_data.get('version_carga') or ACA_INITIAL_VERSION)),
            origen=carga_data.get('origen') or 'Manual',
            status=carga_data.get('status') or models.Carga.STATUS_INCOMPLETO,
            creado_en=carga_creado,
            actualizado=timezone.now(),
            estrategia_id=carga_data.get('estrategia_id'),
            servicio_id=carga_data.get('servicio_id'),
            usuario_id=carga_data.get('usuario_id'),
        )

    criticidad_create_kwargs = {}
    if criticidad_id:
        criticidad_create_kwargs['id'] = criticidad_id
    criticidad = models.Criticidad.objects.create(
        **criticidad_create_kwargs,
        aca_carga=carga,
        equipo_id=crit_data.get('equipo_id'),
        escenario_falla=crit_data.get('escenario_falla') or '',
        observacion=crit_data.get('observacion') or '',
        frecuencia_original=crit_data.get('frecuencia_original'),
        frecuencia_normalizada=crit_data.get('frecuencia_normalizada'),
        valor_cons_total=crit_data.get('valor_cons_total'),
        indicador_criticidad=crit_data.get('indicador_criticidad') or '',
        valor_criticidad_equipo=crit_data.get('valor_criticidad_equipo'),
        criticidad_final=crit_data.get('criticidad_final') or '',
        creado_en=criticidad_creado,
    )
    for item in snapshot.get('dimensiones') or []:
        models.CriticidadDimension.objects.create(
            criticidad=criticidad,
            dimension_id=item.get('dimension_id'),
            estrategia_dimension_id=item.get('estrategia_dimension_id'),
            catalogo_fila_id=item.get('catalogo_fila_id'),
            escala_valor_id=item.get('escala_valor_id'),
            escala_unificada_id=item.get('escala_unificada_id'),
            valor_numerico=item.get('valor_numerico'),
            valor_booleano=item.get('valor_booleano'),
            valor_texto=item.get('valor_texto') or '',
            comentario=item.get('comentario') or '',
            valor_secundario=item.get('valor_secundario'),
        )
    return criticidad


@login_required
@transaction.atomic
def service_aca_delete(request, service_pk, crit_pk):
    servicio, _permission = _service_or_404(request, service_pk, edit=True)
    crit = get_object_or_404(
        models.Criticidad.objects.select_related('aca_carga'),
        pk=crit_pk,
        aca_carga__servicio=servicio,
    )

    if request.method == 'POST':
        carga = crit.aca_carga
        request.session['aca_deleted_record'] = _serialize_aca_for_undo(crit)
        request.session['aca_deleted_restore_url'] = reverse('service_aca_restore_deleted', kwargs={'service_pk': servicio.pk})
        crit.dimensiones.all().delete()
        crit.delete()

        if carga and not carga.criticidades.exists():
            carga.delete()

        messages.success(request, 'Registro ACA eliminado correctamente.', extra_tags='aca-undo-delete')
        return redirect('service_aca_list', pk=servicio.pk)

    return redirect('service_aca_list', pk=servicio.pk)


@login_required
@transaction.atomic
def service_aca_restore_deleted(request, service_pk):
    servicio, _permission = _service_or_404(request, service_pk, edit=True)
    snapshot = request.session.get('aca_deleted_record')
    if request.method != 'POST' or not snapshot or snapshot.get('service_id') != servicio.pk:
        messages.warning(request, 'No hay una eliminación ACA reciente para deshacer.')
        return redirect('service_aca_list', pk=servicio.pk)

    try:
        criticidad = _restore_deleted_aca_snapshot(snapshot)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect('service_aca_list', pk=servicio.pk)
    request.session.pop('aca_deleted_record', None)
    request.session.pop('aca_deleted_restore_url', None)
    messages.success(request, f'Registro ACA restaurado correctamente. ID: {criticidad.pk}.')
    return redirect('service_aca_list', pk=servicio.pk)


def aca_registro_new(request):
    servicios_editables = [
        service
        for service in get_accessible_services(request.user)
        if get_service_permission(request.user, service)['can_edit']
    ] if request.user.is_authenticated else []
    if len(servicios_editables) == 1:
        return redirect('service_aca_new', pk=servicios_editables[0].pk)
    messages.info(request, 'Selecciona primero un servicio para añadir un ACA.')
    return redirect('service_list')
