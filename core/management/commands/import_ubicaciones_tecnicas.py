from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Max

from core import models

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - se informa al ejecutar si falta.
    load_workbook = None


HEADER_ALIASES = {
    'nivel': {'nivel', 'level', 'orden nivel', 'nivel jerarquia'},
    'codigo': {
        'codigo',
        'cod',
        'codigo ut',
        'ut del nivel',
        'ut nivel',
        'codigo nivel',
        'segmento ut',
    },
    'nombre': {
        'nombre',
        'descripcion',
        'descripcion nivel',
        'nombre nivel',
        'valor',
        'denominacion',
        'denominacion de la ubicacion tecnica',
        'denominacion ubicacion tecnica',
    },
    'ut_padre': {
        'ut padre',
        'padre',
        'parent ut',
        'parent ut id',
        'parent id',
        'ut superior',
        'ubicacion padre',
    },
    'ut_completa': {'ut', 'ut completa', 'ubicacion tecnica', 'ubicacion tecnica completa', 'ruta ut'},
    'tag': {'tag', 'tag equipo', 'tag_equipo'},
    'equipo': {'equipo', 'nombre equipo', 'nombre_equipo'},
}

WIDE_LEVEL_RE = re.compile(r'^n(\d+)$')


@dataclass
class ImportRow:
    number: int
    level_name: str
    level_order: int | None
    code: str
    name: str
    path_codes: list[str]
    tag: str
    equipment_name: str


def clean_text(value):
    if value is None:
        return ''
    return str(value).strip()


def normalize_text(value):
    text = clean_text(value).lower()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
    text = re.sub(r'[_\-.]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def header_key(value):
    normalized = normalize_text(value)
    for key, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return key
    return normalized.replace(' ', '_')


def technical_code(value):
    return models._technical_segment(value)


def split_ut(value):
    raw = clean_text(value)
    if not raw:
        return []
    return [technical_code(part) for part in raw.split('-') if clean_text(part)]


def parse_level(value):
    raw = clean_text(value)
    if not raw:
        return None, ''
    if isinstance(value, (int, float)) and int(value) == value:
        order = int(value)
        return order, f'Nivel {order}'
    match = re.match(r'^\s*(\d+)\s*[-.)]?\s*(.+)?$', raw)
    if match:
        order = int(match.group(1))
        name = clean_text(match.group(2)) or f'Nivel {order}'
        return order, name
    return None, raw


class Command(BaseCommand):
    help = 'Importa estructura y valores de ubicaciones tecnicas desde CSV/XLSX.'

    def add_arguments(self, parser):
        parser.add_argument('--archivo', help='Ruta al archivo .xlsx o .csv a importar.')
        parser.add_argument('--empresa', help='ID, nombre o sigla de la empresa destino.')
        parser.add_argument('--sheet', help='Hoja del XLSX. Si se omite, usa la primera hoja.')
        parser.add_argument(
            '--estructura',
            help='Estructura base separada por comas. Ej: "Empresa,Area de negocio,Planta,Area,Sistema,Ubicacion tecnica,Equipo"',
        )
        parser.add_argument('--dry-run', action='store_true', help='Valida y muestra resumen sin guardar cambios.')
        parser.add_argument('--deactivate-missing', action='store_true', help='Desactiva nodos activos no presentes en el archivo.')
        parser.add_argument('--skip-equipos', action='store_true', help='No crea ni actualiza equipos aunque el archivo tenga TAG.')
        parser.add_argument(
            '--equipo-min-level',
            type=int,
            default=5,
            help='Nivel minimo 1-based para crear equipos automaticamente desde UT cuando el archivo no trae TAG. Default: 5.',
        )
        parser.add_argument('--crear-plantilla', help='Crea una plantilla CSV en la ruta indicada y termina.')

    def handle(self, *args, **options):
        if options.get('crear_plantilla'):
            self.create_template(Path(options['crear_plantilla']))
            return

        archivo = Path(options.get('archivo') or '')
        if not archivo.exists():
            raise CommandError(f'No existe el archivo: {archivo}')

        empresa = self.resolve_empresa(options.get('empresa'))
        raw_rows = self.read_rows(archivo, options.get('sheet'))
        rows = self.prepare_rows(
            raw_rows,
            options.get('estructura'),
            equipment_min_level=options.get('equipo_min_level') or 5,
        )
        if not rows:
            raise CommandError('El archivo no contiene filas validas para importar.')

        with transaction.atomic():
            levels_by_order = self.ensure_levels(empresa, rows, options.get('estructura'))
            imported_node_ids, equipment_count = self.import_nodes(
                empresa,
                levels_by_order,
                rows,
                create_equipment=not options.get('skip_equipos'),
            )
            if options.get('deactivate_missing'):
                missing = models.NodoJerarquia.objects.filter(
                    empresa=empresa,
                    activo=True,
                ).exclude(
                    pk__in=imported_node_ids,
                ).update(activo=False)
            else:
                missing = 0

            self.stdout.write(self.style.NOTICE(f'Empresa: {empresa}'))
            self.stdout.write(self.style.NOTICE(f'Filas procesadas: {len(rows)}'))
            self.stdout.write(self.style.NOTICE(f'Nodos activos importados/actualizados: {len(imported_node_ids)}'))
            self.stdout.write(self.style.NOTICE(f'Equipos creados/actualizados: {equipment_count}'))
            if missing:
                self.stdout.write(self.style.WARNING(f'Nodos desactivados por no venir en el archivo: {missing}'))

            if options.get('dry_run'):
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING('Dry-run: no se guardaron cambios.'))
            else:
                self.stdout.write(self.style.SUCCESS('Importacion de ubicaciones tecnicas completada.'))

    def create_template(self, path):
        path.parent.mkdir(parents=True, exist_ok=True)
        headers = ['Nivel', 'UT del nivel', 'Nombre', 'UT padre', 'UT completa', 'TAG', 'Equipo']
        rows = [
            ['Empresa', 'E', 'Empresa ejemplo', '', 'E', '', ''],
            ['Area de negocio', 'DS', 'Downstream', 'E', 'E-DS', '', ''],
            ['Planta', 'ERB', 'Refineria Bio Bio', 'E-DS', 'E-DS-ERB', '', ''],
            ['Area', 'FCCU', 'Unidad FCCU', 'E-DS-ERB', 'E-DS-ERB-FCCU', '', ''],
            ['Sistema', 'INST', 'Instrumentacion', 'E-DS-ERB-FCCU', 'E-DS-ERB-FCCU-INST', '', ''],
            ['Ubicacion tecnica', '0000001FC86', 'Lazo de control', 'E-DS-ERB-FCCU-INST', 'E-DS-ERB-FCCU-INST-0000001FC86', '', ''],
            ['Equipo', 'EQ001', 'Bomba principal', 'E-DS-ERB-FCCU-INST-0000001FC86', 'E-DS-ERB-FCCU-INST-0000001FC86-EQ001', 'TAG-001', 'Bomba principal'],
        ]
        with path.open('w', newline='', encoding='utf-8-sig') as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            writer.writerows(rows)
        self.stdout.write(self.style.SUCCESS(f'Plantilla creada: {path}'))

    def resolve_empresa(self, value):
        if not value:
            raise CommandError('Debes indicar --empresa con ID, nombre o sigla.')
        qs = models.Empresa.objects.all()
        if str(value).isdigit():
            empresa = qs.filter(pk=int(value)).first()
        else:
            empresa = qs.filter(nombre__iexact=value).first() or qs.filter(sigla__iexact=value).first()
        if not empresa:
            raise CommandError(f'No existe la empresa: {value}')
        return empresa

    def read_rows(self, path, sheet_name=None):
        suffix = path.suffix.lower()
        if suffix == '.csv':
            with path.open('r', newline='', encoding='utf-8-sig') as fh:
                return list(csv.DictReader(fh))
        if suffix in {'.xlsx', '.xlsm'}:
            if load_workbook is None:
                raise CommandError('Para importar XLSX instala openpyxl en el entorno de Django.')
            workbook = load_workbook(path, read_only=True, data_only=True)
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise CommandError(f'La hoja "{sheet_name}" no existe. Hojas: {", ".join(workbook.sheetnames)}')
                sheet = workbook[sheet_name]
            else:
                sheet = workbook[workbook.sheetnames[0]]
            return self.read_excel_sheet(sheet)
        raise CommandError('Formato no soportado. Usa .xlsx, .xlsm o .csv.')

    def read_excel_sheet(self, sheet):
        header = None
        header_row = 0
        for index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            non_empty = [clean_text(value) for value in values if clean_text(value)]
            if len(non_empty) >= 2:
                header = [header_key(value) for value in values]
                header_row = index
                break
        if not header:
            return []

        rows = []
        for index, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(clean_text(value) for value in values):
                continue
            row = {
                header[pos]: values[pos] if pos < len(values) else ''
                for pos in range(len(header))
                if header[pos]
            }
            row['_row_number'] = index
            rows.append(row)
        return rows

    def prepare_rows(self, raw_rows, structure, equipment_min_level=5):
        normalized_rows = [
            {header_key(k): clean_text(v) for k, v in raw.items()}
            for raw in raw_rows
        ]
        if self.has_wide_hierarchy_columns(normalized_rows):
            return self.prepare_wide_rows(normalized_rows, structure, equipment_min_level)
        return self.prepare_linear_rows(normalized_rows, structure)

    def has_wide_hierarchy_columns(self, rows):
        return any(
            any(WIDE_LEVEL_RE.match(key or '') for key in row.keys())
            for row in rows
        )

    def prepare_linear_rows(self, normalized_rows, structure):
        structure_names = [clean_text(item) for item in (structure or '').split(',') if clean_text(item)]
        structure_order = {normalize_text(name): pos for pos, name in enumerate(structure_names, start=1)}

        level_order_by_name = {}
        latest_path_by_order = {}
        prepared = []

        for index, row in enumerate(normalized_rows, start=2):
            number = int(row.get('_row_number') or row.get('row_number') or index)
            level_order, level_name = parse_level(row.get('nivel'))
            if level_name and normalize_text(level_name) in structure_order:
                level_order = structure_order[normalize_text(level_name)]
            elif not level_order and level_name:
                level_order = level_order_by_name.setdefault(normalize_text(level_name), len(level_order_by_name) + 1)
            elif not level_order:
                level_order = len(level_order_by_name) + 1
                level_name = structure_names[level_order - 1] if level_order <= len(structure_names) else f'Nivel {level_order}'

            if structure_names and level_order <= len(structure_names):
                level_name = structure_names[level_order - 1]

            code = row.get('codigo') or row.get('ut_completa')
            path_codes = self.path_for_row(row, level_order, latest_path_by_order)
            if not path_codes:
                raise CommandError(f'Fila {number}: falta codigo o UT.')
            code = path_codes[-1]

            for order in list(latest_path_by_order):
                if order >= level_order:
                    latest_path_by_order.pop(order, None)
            latest_path_by_order[level_order] = path_codes

            prepared.append(ImportRow(
                number=number,
                level_name=level_name,
                level_order=level_order,
                code=code,
                name=row.get('nombre') or code,
                path_codes=path_codes,
                tag=row.get('tag', ''),
                equipment_name=row.get('equipo', ''),
            ))

        return prepared

    def prepare_wide_rows(self, normalized_rows, structure, equipment_min_level=5):
        structure_names = [clean_text(item) for item in (structure or '').split(',') if clean_text(item)]
        parent_paths = {
            tuple(split_ut(row.get('ut_padre')))
            for row in normalized_rows
            if split_ut(row.get('ut_padre'))
        }
        rows_by_path = {}
        path_order = []

        def level_name_for(order):
            if structure_names and order <= len(structure_names):
                return structure_names[order - 1]
            return f'N{order - 1}'

        def remember_row(number, path_codes, name='', tag='', equipment_name=''):
            path_codes = [code for code in path_codes if code]
            if not path_codes:
                return
            level_order = len(path_codes)
            key = tuple(path_codes)
            existing = rows_by_path.get(key)
            if existing:
                if name and (not existing.name or existing.name == existing.code):
                    existing.name = name
                if tag and not existing.tag:
                    existing.tag = tag
                if equipment_name and not existing.equipment_name:
                    existing.equipment_name = equipment_name
                return

            rows_by_path[key] = ImportRow(
                number=number,
                level_name=level_name_for(level_order),
                level_order=level_order,
                code=path_codes[-1],
                name=name or path_codes[-1],
                path_codes=path_codes,
                tag=tag,
                equipment_name=equipment_name,
            )
            path_order.append(key)

        for index, row in enumerate(normalized_rows, start=2):
            number = int(row.get('_row_number') or row.get('row_number') or index)
            wide_indexes = sorted(
                int(match.group(1))
                for key in row.keys()
                for match in [WIDE_LEVEL_RE.match(key or '')]
                if match and row.get(key)
            )
            for wide_index in wide_indexes:
                path_codes = split_ut(row.get(f'n{wide_index}'))
                remember_row(
                    number,
                    path_codes,
                    name=row.get(f'n{wide_index}_nombre') or '',
                )

            full_path = split_ut(row.get('ut_completa'))
            if not full_path:
                continue

            full_ut = '-'.join(full_path)
            leaf_name = row.get('equipo') or row.get('nombre') or row.get(f'n{len(full_path) - 1}_nombre') or ''
            explicit_tag = row.get('tag')
            is_leaf = tuple(full_path) not in parent_paths
            should_create_equipment = bool(explicit_tag) or (is_leaf and len(full_path) >= equipment_min_level)
            remember_row(
                number,
                full_path,
                name=leaf_name,
                tag=explicit_tag or (full_ut if should_create_equipment else ''),
                equipment_name=row.get('equipo') or leaf_name,
            )

        return [rows_by_path[key] for key in path_order]

    def path_for_row(self, row, level_order, latest_path_by_order):
        full_path = split_ut(row.get('ut_completa'))
        if full_path:
            return full_path

        code_parts = split_ut(row.get('codigo'))
        if len(code_parts) > 1 and len(code_parts) >= level_order:
            return code_parts[:level_order]
        code = code_parts[0] if code_parts else ''
        if not code:
            return []

        parent_path = split_ut(row.get('ut_padre'))
        if parent_path:
            return parent_path + [code]
        if level_order > 1 and latest_path_by_order.get(level_order - 1):
            return latest_path_by_order[level_order - 1] + [code]
        return [code]

    def ensure_levels(self, empresa, rows, structure):
        names_by_order = {}
        structure_names = [clean_text(item) for item in (structure or '').split(',') if clean_text(item)]
        for pos, name in enumerate(structure_names, start=1):
            names_by_order[pos] = name
        for row in rows:
            names_by_order.setdefault(row.level_order or len(row.path_codes), row.level_name)
            for order in range(1, len(row.path_codes) + 1):
                names_by_order.setdefault(order, f'Nivel {order}')

        existing = list(models.NivelJerarquia.objects.filter(empresa=empresa).order_by('pk'))
        original_names = {level.pk: level.nombre for level in existing}
        if not structure_names:
            existing_name_by_order = {level.orden: level.nombre for level in existing if level.nombre}
            for order, name in list(names_by_order.items()):
                generic_name = re.match(r'^(n\d+|nivel\s+\d+)$', normalize_text(name))
                if generic_name and existing_name_by_order.get(order):
                    names_by_order[order] = existing_name_by_order[order]

        normalized_level_names = [normalize_text(name) for name in names_by_order.values()]
        if len(normalized_level_names) != len(set(normalized_level_names)):
            raise CommandError('La estructura contiene nombres de nivel duplicados.')

        max_order = max([item.orden for item in existing] or [0])
        temp_base = max_order + len(existing) + 1000
        for offset, level in enumerate(existing, start=1):
            level.orden = temp_base + offset
            level.nombre = f'__tmp_import_ut_{empresa.pk}_{level.pk}'
            level.save(update_fields=['orden', 'nombre'])

        levels_by_order = {}
        reusable_by_name = {
            normalize_text(name): level
            for level in existing
            for name in [original_names.get(level.pk, '')]
            if name
        }
        used_ids = set()
        for order in sorted(names_by_order):
            name = names_by_order[order]
            level = reusable_by_name.get(normalize_text(name))
            if level and level.pk in used_ids:
                level = None
            if level:
                level.nombre = name
                level.orden = order
                level.activo = True
                level.save(update_fields=['nombre', 'orden', 'activo'])
            else:
                level = models.NivelJerarquia.objects.create(
                    empresa=empresa,
                    nombre=name,
                    orden=order,
                    activo=True,
                )
            levels_by_order[order] = level
            used_ids.add(level.pk)

        models.NivelJerarquia.objects.filter(empresa=empresa).exclude(pk__in=used_ids).update(activo=False)
        return levels_by_order

    def import_nodes(self, empresa, levels_by_order, rows, create_equipment=True):
        path_cache = {}
        imported_ids = set()
        equipment_count = 0

        for row in rows:
            parent = None
            path_so_far = []
            leaf = None
            for order, code in enumerate(row.path_codes, start=1):
                path_so_far.append(code)
                path_key = tuple(path_so_far)
                if path_key in path_cache:
                    node = path_cache[path_key]
                    if order == len(row.path_codes):
                        node.nivel = levels_by_order[order]
                        node.nombre = row.name
                        node.activo = True
                        node.save(update_fields=['nivel', 'nombre', 'activo'])
                    parent = node
                    leaf = node
                    continue

                level = levels_by_order.get(order)
                if not level:
                    raise CommandError(f'Fila {row.number}: no existe nivel para orden {order}.')
                name = row.name if order == len(row.path_codes) else code
                node = models.NodoJerarquia.objects.filter(
                    empresa=empresa,
                    parent=parent,
                    codigo=code,
                ).first()
                if not node:
                    sibling_order = (models.NodoJerarquia.objects.filter(
                        empresa=empresa,
                        parent=parent,
                    ).aggregate(max_order=Max('orden'))['max_order'] or 0) + 1
                    node = models.NodoJerarquia.objects.create(
                        empresa=empresa,
                        nivel=level,
                        parent=parent,
                        codigo=code,
                        nombre=name,
                        orden=sibling_order,
                        activo=True,
                    )
                else:
                    node.nivel = level
                    node.nombre = name
                    node.activo = True
                    node.save(update_fields=['nivel', 'nombre', 'activo'])
                path_cache[path_key] = node
                parent = node
                leaf = node
            if leaf:
                imported_ids.add(leaf.pk)
                imported_ids.update(node.pk for node in leaf.path_nodes())
                if create_equipment and row.tag:
                    self.upsert_equipment(row, leaf)
                    equipment_count += 1

        return imported_ids, equipment_count

    def upsert_equipment(self, row, node):
        equipment = (
            models.Equipo.objects.filter(tag_equipo=row.tag).order_by('pk').first()
            or models.Equipo.objects.filter(ut=node.ut).order_by('pk').first()
        )
        if not equipment:
            equipment = models.Equipo(tag_equipo=row.tag)
        equipment.tag_equipo = row.tag
        equipment.nombre_equipo = row.equipment_name or row.name
        equipment.nodo = node
        equipment.ut = node.ut
        equipment.descripcion_ut = node.ruta_nombre[:255]
        equipment.save()
