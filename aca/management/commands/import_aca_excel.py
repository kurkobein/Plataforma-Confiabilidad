from __future__ import annotations

import unicodedata
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from aca.views import (
    ACA_INITIAL_VERSION,
    _aca_excluded_dimension_ids,
    _create_dimension_items,
    _match_catalog_dependency_row,
    _resolve_complete_matrix_cell_from_dimension_records,
    _save_matrix_dimensions,
    _service_matrix_selector,
    _sync_aca_carga_status,
    _sync_criticidad_resumen,
    prepare_bulk_dimension_items,
)
from core import models
from core.access import get_service_equipment


HEADER_ALIASES = {
    "equipo_componente": {"equipo componente", "equipo / componente", "equipo", "nombre equipo"},
    "ut": {"ut", "ubicacion tecnica", "ubicacion tecnica completa"},
    "tag": {"tag", "tag equipo"},
    "frecuencia": {"frecuencia", "frecuencia falla"},
    "impacto_operacional": {"impacto operacional", "impacto operacion", "operacional"},
    "seguridad": {"seguridad", "impacto seguridad"},
    "medio_ambiente": {"medio ambiente", "impacto medio ambiente", "impacto ambiental"},
    "costo_mantencion": {"costo mantencion", "costo mantenimiento", "impacto costo"},
    "consecuencia": {"consecuencia", "valor consecuencia total"},
    "criticidad": {"criticidad", "valor criticidad equipo"},
    "criticidad_nivel": {"criticidad nivel", "criticidad (nivel)", "criticidad final", "rango criticidad"},
}

DIMENSION_ALIASES = {
    "frecuencia": {"frecuencia", "frecuencia falla", "frecuencia normalizada"},
    "impacto_operacional": {"impacto operacional", "impacto operacion", "operacional"},
    "seguridad": {"seguridad", "impacto seguridad"},
    "medio_ambiente": {"medio ambiente", "impacto medio ambiente", "impacto ambiental"},
    "costo_mantencion": {"costo mantencion", "costo mantenimiento", "impacto costo"},
    "consecuencia": {"consecuencia", "valor consecuencia total"},
    "criticidad": {"criticidad", "valor criticidad equipo"},
    "criticidad_nivel": {"criticidad nivel", "criticidad final", "rango criticidad", "criticidad (nivel)"},
}


def normalize(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    for char in "\n\r\t()[]/":
        text = text.replace(char, " ")
    text = " ".join(text.replace("-", " ").replace("_", " ").split())
    return text


def decimal_or_none(value):
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return None


def text_or_blank(value):
    return "" if value is None else str(value).strip()


class Command(BaseCommand):
    help = "Importa registros ACA desde Excel usando el flujo normal de Carga, Criticidad, dimensiones y matriz."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Ruta del archivo Excel.")
        parser.add_argument("--sheet", default="ACA", help="Hoja del Excel. Default: ACA.")
        parser.add_argument("--service", required=True, help="ID, codigo o descripcion del servicio.")
        parser.add_argument("--dry-run", action="store_true", help="Valida y revierte todos los cambios.")
        parser.add_argument("--confirm", action="store_true", help="Ejecuta la carga real.")
        parser.add_argument("--replace", action="store_true", help="Borra cargas previas del mismo archivo/origen para el servicio.")
        parser.add_argument("--create-missing-equipment", action="store_true", help="Crea equipos que no existan por TAG/UT.")
        parser.add_argument("--limit", type=int, help="Limita la cantidad de filas procesadas.")

    def handle(self, *args, **options):
        if not options["dry_run"] and not options["confirm"]:
            raise CommandError("Debes usar --dry-run para revisar o --confirm para guardar.")

        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"No existe el archivo: {file_path}")

        service = self.resolve_service(options["service"])
        if not service.estrategia_id:
            raise CommandError(f"El servicio {service.codigo_servicio} no tiene estrategia asociada.")

        wb = load_workbook(file_path, read_only=True, data_only=True)
        if options["sheet"] not in wb.sheetnames:
            raise CommandError(f"El archivo no tiene hoja '{options['sheet']}'. Hojas: {', '.join(wb.sheetnames)}")

        rows = self.read_rows(wb[options["sheet"]], limit=options.get("limit"))
        if not rows:
            raise CommandError("No se encontraron filas ACA para importar.")

        origin = f"ACA Excel: {file_path.name}"
        now = timezone.now()

        with transaction.atomic():
            if options["replace"]:
                self.delete_previous(service, origin)

            result = self.import_rows(
                service=service,
                rows=rows,
                origin=origin,
                create_missing_equipment=options["create_missing_equipment"],
                now=now,
            )

            self.print_summary(service, origin, rows, result, dry_run=options["dry_run"])

            if options["dry_run"]:
                transaction.set_rollback(True)
                return

        self.stdout.write(self.style.SUCCESS("Carga ACA completada correctamente."))

    def resolve_service(self, value):
        query = models.Servicio.objects.select_related("empresa", "estrategia")
        if str(value).isdigit():
            service = query.filter(pk=int(value)).first()
            if service:
                return service
        service = query.filter(codigo_servicio=value).first()
        if service:
            return service
        service = query.filter(descripcion=value).first()
        if service:
            return service
        service = query.filter(descripcion__icontains=value).first()
        if service:
            return service
        raise CommandError(f"No se encontro servicio para: {value}")

    def resolve_header_key(self, header):
        normalized = normalize(header)
        for key, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                return key
        return normalized.replace(" ", "_")

    def read_rows(self, ws, limit=None):
        iterator = ws.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            return []
        keys = [self.resolve_header_key(header) for header in headers]
        rows = []
        for excel_row, values in enumerate(iterator, start=2):
            payload = {"_excel_row": excel_row}
            empty = True
            for key, value in zip(keys, values):
                if value not in (None, ""):
                    empty = False
                payload[key] = value
            if empty:
                continue
            rows.append(payload)
            if limit and len(rows) >= limit:
                break
        return rows

    def dimension_map(self, strategy, excluded_dimension_ids):
        qs = (
            models.EstrategiaDimension.objects.filter(
                estrategia=strategy,
                activo=True,
                proceso_uso__in=[
                    models.EstrategiaDimension.PROCESO_ACA,
                    models.EstrategiaDimension.PROCESO_AMBOS,
                ],
            )
            .exclude(dimension_id__in=excluded_dimension_ids)
            .select_related("dimension")
            .order_by("orden", "id")
        )
        by_alias = {}
        for estrategia_dimension in qs:
            normalized_name = normalize(estrategia_dimension.dimension.nombre)
            for source_key, aliases in DIMENSION_ALIASES.items():
                if normalized_name in aliases:
                    by_alias[source_key] = estrategia_dimension
                    break
        return by_alias

    def row_dimensions(self, row, dimension_by_alias):
        payload = {}
        for source_key, estrategia_dimension in dimension_by_alias.items():
            raw = row.get(source_key)
            if raw in (None, ""):
                continue
            numeric = decimal_or_none(raw)
            catalog_row = self.match_catalog_row(estrategia_dimension, raw)
            payload[str(estrategia_dimension.pk)] = {
                "catalogo_fila_id": catalog_row.pk if catalog_row else "",
                "valor_numerico": "" if numeric is None else str(numeric),
                "valor_texto": "" if numeric is not None else text_or_blank(raw),
                "display": text_or_blank(raw),
            }
        return payload

    def match_catalog_row(self, estrategia_dimension, raw):
        try:
            catalogo = estrategia_dimension.catalogo
        except models.DimensionCatalogo.DoesNotExist:
            return None
        if not catalogo or not catalogo.activa:
            return None

        numeric = decimal_or_none(raw)
        if numeric is not None:
            matched = _match_catalog_dependency_row(catalogo, numeric)
            if matched:
                return matched

        raw_text = normalize(raw)
        if not raw_text:
            return None
        for fila in catalogo.filas.prefetch_related("celdas__columna").order_by("orden", "id"):
            if normalize(fila.etiqueta) == raw_text:
                return fila
            values = fila.values_map()
            for value in values.values():
                if normalize(value) == raw_text:
                    return fila
        return None

    def resolve_equipment(self, service, row, create_missing):
        tag = text_or_blank(row.get("tag"))
        ut = text_or_blank(row.get("ut"))
        name = text_or_blank(row.get("equipo_componente")) or tag or ut

        equipment_qs = get_service_equipment(service)
        equipo = None
        if tag:
            equipo = equipment_qs.filter(tag_equipo=tag).first()
        if not equipo and ut:
            equipo = equipment_qs.filter(ut=ut).first()
        if not equipo and tag:
            equipo = models.Equipo.objects.filter(tag_equipo=tag).first()
        if not equipo and ut:
            equipo = models.Equipo.objects.filter(ut=ut).first()

        if equipo:
            models.ServicioEquipo.objects.get_or_create(servicio=service, equipo=equipo)
            return equipo, False

        if not create_missing:
            return None, False

        equipo = models.Equipo.objects.create(
            tag_equipo=(tag or ut or name)[:100],
            nombre_equipo=name[:200],
            ut=ut[:200],
            descripcion_ut=name[:255],
        )
        models.ServicioEquipo.objects.get_or_create(servicio=service, equipo=equipo)
        return equipo, True

    def delete_previous(self, service, origin):
        cargas = models.Carga.objects.filter(servicio=service, origen=origin)
        carga_ids = list(cargas.values_list("id", flat=True))
        criticidad_ids = list(models.Criticidad.objects.filter(aca_carga_id__in=carga_ids).values_list("id", flat=True))
        dim_count = models.CriticidadDimension.objects.filter(criticidad_id__in=criticidad_ids).delete()[0]
        crit_count = models.Criticidad.objects.filter(id__in=criticidad_ids).delete()[0]
        carga_count = models.Carga.objects.filter(id__in=carga_ids).delete()[0]
        self.stdout.write(self.style.WARNING(
            f"Registros previos eliminados: cargas={carga_count}, ACA={crit_count}, dimensiones={dim_count}"
        ))

    def import_rows(self, service, rows, origin, create_missing_equipment, now):
        strategy = service.estrategia
        matrix_selector = _service_matrix_selector(service)
        matriz = matrix_selector.get("matriz")
        excluded_dimension_ids = _aca_excluded_dimension_ids(strategy, matriz)
        dimension_by_alias = self.dimension_map(strategy, excluded_dimension_ids)
        profile = models.Usuario.objects.filter(correo_corporativo__icontains="importador").first()

        result = Counter()
        warnings = []
        existing_equipment_ids = set(
            models.Criticidad.objects.filter(
                aca_carga__servicio=service,
                equipo_id__isnull=False,
            ).values_list("equipo_id", flat=True)
        )
        submitted_equipment_ids = set()

        for row in rows:
            excel_row = row["_excel_row"]
            equipo, created_equipment = self.resolve_equipment(service, row, create_missing_equipment)
            if not equipo:
                warnings.append(f"Fila {excel_row}: equipo no encontrado por TAG/UT.")
                result["omitidas"] += 1
                continue
            if created_equipment:
                result["equipos_creados"] += 1
            if equipo.pk in existing_equipment_ids:
                warnings.append(f"Fila {excel_row}: el equipo {equipo.tag_display} ya tiene un ACA en el servicio.")
                result["omitidas"] += 1
                continue
            if equipo.pk in submitted_equipment_ids:
                warnings.append(f"Fila {excel_row}: el equipo {equipo.tag_display} esta repetido en el archivo.")
                result["omitidas"] += 1
                continue

            row_dimension_payload = self.row_dimensions(row, dimension_by_alias)
            prepared, _source_values, errors = prepare_bulk_dimension_items(
                strategy,
                row_dimension_payload,
                excluded_dimension_ids=excluded_dimension_ids,
                allow_incomplete=False,
            )
            if errors:
                warnings.append(f"Fila {excel_row}: {' '.join(errors)}")
                result["omitidas"] += 1
                continue

            selected_cell = _resolve_complete_matrix_cell_from_dimension_records(matriz, prepared) if matriz else None
            if matriz and not selected_cell:
                warnings.append(f"Fila {excel_row}: no se pudo resolver la matriz de criticidad.")
                result["omitidas"] += 1
                continue

            carga = models.Carga.objects.create(
                fecha_analisis=timezone.localdate(),
                version_carga=ACA_INITIAL_VERSION,
                origen=origin,
                status=models.Carga.STATUS_COMPLETO,
                creado_en=now,
                actualizado=now,
                estrategia=strategy,
                servicio=service,
                usuario=profile,
            )
            criticidad = models.Criticidad.objects.create(
                creado_en=now,
                aca_carga=carga,
                equipo=equipo,
                escenario_falla="",
                observacion="",
                frecuencia_original=None,
                frecuencia_normalizada=selected_cell.probabilidad.valor if selected_cell else decimal_or_none(row.get("frecuencia")),
                valor_cons_total=selected_cell.impacto_nivel.valor if selected_cell else decimal_or_none(row.get("consecuencia")),
                indicador_criticidad="",
                valor_criticidad_equipo=selected_cell.resultado_num if selected_cell else decimal_or_none(row.get("criticidad")),
                criticidad_final=selected_cell.clasificacion if selected_cell else text_or_blank(row.get("criticidad_nivel")),
            )
            created_dimensions = _create_dimension_items(criticidad, prepared)
            if selected_cell:
                _save_matrix_dimensions(criticidad, strategy, selected_cell)
            _sync_criticidad_resumen(criticidad, strategy)
            _sync_aca_carga_status(criticidad, strategy, models.Carga.STATUS_COMPLETO)

            result["cargas"] += 1
            result["aca"] += 1
            result["dimensiones"] += len(created_dimensions)
            result["catalogo_fila_asignadas"] += sum(1 for item in created_dimensions if item.catalogo_fila_id)
            submitted_equipment_ids.add(equipo.pk)

        result["warnings"] = warnings
        return result

    def print_summary(self, service, origin, rows, result, dry_run):
        self.stdout.write(f"Servicio: {service.codigo_servicio}")
        self.stdout.write(f"Origen: {origin}")
        self.stdout.write(f"Filas leidas: {len(rows)}")
        self.stdout.write(f"Cargas creadas: {result['cargas']}")
        self.stdout.write(f"ACA creados: {result['aca']}")
        self.stdout.write(f"Dimensiones creadas: {result['dimensiones']}")
        self.stdout.write(f"Dimensiones con catalogo_fila_id: {result['catalogo_fila_asignadas']}")
        self.stdout.write(f"Equipos creados: {result['equipos_creados']}")
        self.stdout.write(f"Filas omitidas: {result['omitidas']}")
        warnings = result.get("warnings") or []
        if warnings:
            self.stdout.write(self.style.WARNING(f"Warnings: {len(warnings)}"))
            for warning in warnings[:30]:
                self.stdout.write(f"- {warning}")
            if len(warnings) > 30:
                self.stdout.write(f"... {len(warnings) - 30} warnings adicionales")
        if dry_run:
            self.stdout.write("No se guardo nada porque es dry-run.")
