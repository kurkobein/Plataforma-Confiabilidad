import json
import re
import html
import uuid
import unicodedata
from datetime import timedelta
from io import BytesIO
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db import connection, transaction
from django.db.models import Count, Q, TextField, Prefetch
from django.http import HttpResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from .access import get_accessible_services, get_profile_for_user, get_service_equipment, get_service_permission
from .user_sync import archive_profile, sync_profile_from_auth_user
from .forms import (
    EmailLoginForm,
    EquipoBulkUploadForm,
    get_form_for_key,
)
from .registry import MODEL_REGISTRY, get_registered_model
from . import models
from core.services.equipment_import import execute_equipment_import, preview_equipment_import

MAX_LIST_COLUMNS = 6
_RANGE_RE = re.compile(r'^\s*(-?\d+(?:[.,]\d+)?)\s*-\s*(-?\d+(?:[.,]\d+)?)\s*$')
_MATRIX_DECIMAL_STEP = Decimal('0.01')


def _export_value(value):
    if value is None:
        return ''
    if isinstance(value, Decimal):
        return format(value, 'f')
    if hasattr(value, 'strftime'):
        return value.strftime('%d/%m/%Y')
    return str(value)


def _export_filename(prefix, service, extension):
    code = re.sub(r'[^A-Za-z0-9_-]+', '_', str(service.codigo_servicio or service.pk)).strip('_')
    date_label = timezone.localdate().strftime('%Y%m%d')
    return f'{prefix}_{code}_{date_label}.{extension}'


def _export_xlsx_response(filename, sheet_name, columns, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31] or 'Registros'
    headers = [label for _key, label in columns]
    worksheet.append(headers)

    for row in rows:
        worksheet.append([_export_value(row.get(key)) for key, _label in columns])

    header_fill = PatternFill('solid', fgColor='E8EEF7')
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    for row_cells in worksheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    for index, _header in enumerate(headers, start=1):
        letter = get_column_letter(index)
        max_length = max(
            len(str(worksheet.cell(row=row_idx, column=index).value or ''))
            for row_idx in range(1, worksheet.max_row + 1)
        )
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 42)

    worksheet.freeze_panes = 'A2'
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _export_pdf_response(filename, title, columns, rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    body_style = ParagraphStyle('ExportBody', parent=styles['BodyText'], fontSize=8, leading=10)
    label_style = ParagraphStyle(
        'ExportLabel',
        parent=styles['BodyText'],
        fontSize=8,
        leading=10,
        fontName='Helvetica-Bold',
    )

    story = [
        Paragraph(html.escape(title), styles['Title']),
        Paragraph(f'Total registros: {len(rows)}', styles['Normal']),
        Spacer(1, 8),
    ]

    for row_index, row in enumerate(rows, start=1):
        if row_index > 1:
            story.append(Spacer(1, 8))
        story.append(Paragraph(f'Registro {row_index}', styles['Heading3']))
        table_data = []
        for key, label in columns:
            value = html.escape(_export_value(row.get(key))).replace('\n', '<br/>')
            table_data.append([
                Paragraph(html.escape(label), label_style),
                Paragraph(value or '-', body_style),
            ])
        table = Table(table_data, colWidths=[52 * mm, 220 * mm], repeatRows=0)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D5DBE5')),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F6FA')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(table)
        if row_index < len(rows) and row_index % 2 == 0:
            story.append(PageBreak())

    if not rows:
        story.append(Paragraph('No hay registros para exportar.', styles['Normal']))

    document.build(story)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response




def _can_delete(request, config, obj=None):
    if config.get('delete_superuser_only', False):
        return request.user.is_superuser
    return True


def _ensure_delete_allowed(request, config, obj=None):
    if not _can_delete(request, config, obj=obj):
        raise PermissionDenied('Solo un administrador puede eliminar este registro.')


def _delete_by_materialized_ids(model, ids):
    ids = list(ids)
    if not ids:
        return 0
    deleted_count, _deleted_by_model = model.objects.filter(id__in=ids).delete()
    return deleted_count


def _delete_service_related_data(service):
    counts = {}

    carga_ids = list(models.Carga.objects.filter(servicio=service).values_list('id', flat=True))
    criticidad_ids = list(models.Criticidad.objects.filter(aca_carga_id__in=carga_ids).values_list('id', flat=True))
    rcm_ids = list(models.RCM.objects.filter(carga_id__in=carga_ids).values_list('id', flat=True))
    fmea_ids = list(models.FMEA_FMECA.objects.filter(rcm_id__in=rcm_ids).values_list('id', flat=True))
    tarea_ids = list(models.TareaRCM.objects.filter(fmea_id__in=fmea_ids).values_list('id', flat=True))

    pauta_ids = list(models.Pauta.objects.filter(servicio=service).values_list('id', flat=True))
    plantilla_ids = list(models.PlantillaPauta.objects.filter(servicio=service).values_list('id', flat=True))
    familia_ids = list(models.FamiliaEquipo.objects.filter(servicio=service).values_list('id', flat=True))

    counts['valores_tarea_rcm'] = _delete_by_materialized_ids(
        models.ValorCampoTareaRCM,
        models.ValorCampoTareaRCM.objects.filter(tarea_id__in=tarea_ids).values_list('id', flat=True),
    )
    counts['tareas_rcm'] = _delete_by_materialized_ids(models.TareaRCM, tarea_ids)
    counts['evaluaciones_fmea'] = _delete_by_materialized_ids(
        models.EvaluacionFMEA,
        models.EvaluacionFMEA.objects.filter(fmea_id__in=fmea_ids).values_list('id', flat=True),
    )
    counts['fmea'] = _delete_by_materialized_ids(models.FMEA_FMECA, fmea_ids)
    counts['adjuntos_rcm'] = _delete_by_materialized_ids(
        models.RCMAdjunto,
        models.RCMAdjunto.objects.filter(rcm_id__in=rcm_ids).values_list('id', flat=True),
    )
    counts['rcm'] = _delete_by_materialized_ids(models.RCM, rcm_ids)

    counts['adjuntos_aca'] = _delete_by_materialized_ids(
        models.CriticidadAdjunto,
        models.CriticidadAdjunto.objects.filter(criticidad_id__in=criticidad_ids).values_list('id', flat=True),
    )
    counts['dimensiones_aca'] = _delete_by_materialized_ids(
        models.CriticidadDimension,
        models.CriticidadDimension.objects.filter(criticidad_id__in=criticidad_ids).values_list('id', flat=True),
    )
    counts['criticidades'] = _delete_by_materialized_ids(models.Criticidad, criticidad_ids)
    counts['cargas'] = _delete_by_materialized_ids(models.Carga, carga_ids)

    counts['tareas_pauta'] = _delete_by_materialized_ids(
        models.PautaTarea,
        models.PautaTarea.objects.filter(pauta_id__in=pauta_ids).values_list('id', flat=True),
    )
    counts['pautas'] = _delete_by_materialized_ids(models.Pauta, pauta_ids)
    counts['mapeos_plantilla'] = _delete_by_materialized_ids(
        models.MapeoPlantillaPauta,
        models.MapeoPlantillaPauta.objects.filter(plantilla_id__in=plantilla_ids).values_list('id', flat=True),
    )
    counts['plantillas_pauta'] = _delete_by_materialized_ids(models.PlantillaPauta, plantilla_ids)
    counts['reglas_pauta'] = _delete_by_materialized_ids(
        models.ReglaGeneracionPauta,
        models.ReglaGeneracionPauta.objects.filter(servicio=service).values_list('id', flat=True),
    )

    counts['items_familia'] = _delete_by_materialized_ids(
        models.FamiliaEquipoItem,
        models.FamiliaEquipoItem.objects.filter(familia_id__in=familia_ids).values_list('id', flat=True),
    )
    counts['familias_equipo'] = _delete_by_materialized_ids(models.FamiliaEquipo, familia_ids)
    counts['escenarios_falla'] = _delete_by_materialized_ids(
        models.EscenarioFalla,
        models.EscenarioFalla.objects.filter(servicio=service).values_list('id', flat=True),
    )
    counts['accesos'] = _delete_by_materialized_ids(
        models.AccesoUsuario,
        models.AccesoUsuario.objects.filter(servicio=service).values_list('id', flat=True),
    )
    counts['equipos_servicio'] = _delete_by_materialized_ids(
        models.ServicioEquipo,
        models.ServicioEquipo.objects.filter(servicio=service).values_list('id', flat=True),
    )
    return counts


def _merge_delete_counts(target, source, prefix=''):
    for key, value in source.items():
        target[f'{prefix}{key}'] = target.get(f'{prefix}{key}', 0) + value
    return target


def _delete_nodes_by_depth(node_ids):
    nodes = list(models.NodoJerarquia.objects.filter(id__in=list(node_ids)).select_related('parent'))
    if not nodes:
        return 0
    node_by_id = {node.pk: node for node in nodes}

    def depth(node):
        current = node
        seen = set()
        value = 0
        while current and current.pk not in seen:
            seen.add(current.pk)
            current = node_by_id.get(current.parent_id)
            if current:
                value += 1
        return value

    deleted = 0
    for node in sorted(nodes, key=depth, reverse=True):
        deleted += _delete_by_materialized_ids(models.NodoJerarquia, [node.pk])
    return deleted


def _delete_equipment_related_data(equipment_ids):
    counts = {}
    equipment_ids = list(dict.fromkeys(equipment_ids))
    if not equipment_ids:
        return counts

    criticidad_ids = list(
        models.Criticidad.objects.filter(equipo_id__in=equipment_ids).values_list('id', flat=True)
    )
    rcm_ids = list(
        models.RCM.objects.filter(equipo_id__in=equipment_ids).values_list('id', flat=True)
    )
    fmea_ids = list(
        models.FMEA_FMECA.objects.filter(rcm_id__in=rcm_ids).values_list('id', flat=True)
    )
    tarea_ids = list(
        models.TareaRCM.objects.filter(fmea_id__in=fmea_ids).values_list('id', flat=True)
    )
    pauta_ids = list(
        models.Pauta.objects.filter(equipo_id__in=equipment_ids).values_list('id', flat=True)
    )

    counts['valores_tarea_rcm_equipo'] = _delete_by_materialized_ids(
        models.ValorCampoTareaRCM,
        models.ValorCampoTareaRCM.objects.filter(tarea_id__in=tarea_ids).values_list('id', flat=True),
    )
    counts['tareas_rcm_equipo'] = _delete_by_materialized_ids(models.TareaRCM, tarea_ids)
    counts['evaluaciones_fmea_equipo'] = _delete_by_materialized_ids(
        models.EvaluacionFMEA,
        models.EvaluacionFMEA.objects.filter(fmea_id__in=fmea_ids).values_list('id', flat=True),
    )
    counts['fmea_equipo'] = _delete_by_materialized_ids(models.FMEA_FMECA, fmea_ids)
    counts['adjuntos_rcm_equipo'] = _delete_by_materialized_ids(
        models.RCMAdjunto,
        models.RCMAdjunto.objects.filter(rcm_id__in=rcm_ids).values_list('id', flat=True),
    )
    counts['rcm_equipo'] = _delete_by_materialized_ids(models.RCM, rcm_ids)
    counts['adjuntos_aca_equipo'] = _delete_by_materialized_ids(
        models.CriticidadAdjunto,
        models.CriticidadAdjunto.objects.filter(criticidad_id__in=criticidad_ids).values_list('id', flat=True),
    )
    counts['dimensiones_aca_equipo'] = _delete_by_materialized_ids(
        models.CriticidadDimension,
        models.CriticidadDimension.objects.filter(criticidad_id__in=criticidad_ids).values_list('id', flat=True),
    )
    counts['criticidades_equipo'] = _delete_by_materialized_ids(models.Criticidad, criticidad_ids)
    counts['tareas_pauta_equipo'] = _delete_by_materialized_ids(
        models.PautaTarea,
        models.PautaTarea.objects.filter(pauta_id__in=pauta_ids).values_list('id', flat=True),
    )
    counts['pautas_equipo'] = _delete_by_materialized_ids(models.Pauta, pauta_ids)
    counts['items_familia_equipo'] = _delete_by_materialized_ids(
        models.FamiliaEquipoItem,
        models.FamiliaEquipoItem.objects.filter(equipo_id__in=equipment_ids).values_list('id', flat=True),
    )
    counts['servicios_equipo'] = _delete_by_materialized_ids(
        models.ServicioEquipo,
        models.ServicioEquipo.objects.filter(equipo_id__in=equipment_ids).values_list('id', flat=True),
    )
    counts['componentes_equipo'] = _delete_by_materialized_ids(
        models.ComponenteEquipo,
        models.ComponenteEquipo.objects.filter(equipo_id__in=equipment_ids).values_list('id', flat=True),
    )
    counts['equipos'] = _delete_by_materialized_ids(models.Equipo, equipment_ids)
    return counts


def _strategy_delete_blockers(strategy):
    blockers = []
    estrategia_dimension_ids = list(
        models.EstrategiaDimension.objects.filter(estrategia=strategy).values_list('id', flat=True)
    )
    tipo_tarea_ids = list(
        models.TipoTareaEstrategia.objects.filter(estrategia=strategy).values_list('id', flat=True)
    )
    campo_tarea_ids = list(
        models.CampoTareaEstrategia.objects.filter(
            tipo_tarea_estrategia_id__in=tipo_tarea_ids
        ).values_list('id', flat=True)
    )
    catalogo_ids = list(
        models.DimensionCatalogo.objects.filter(
            estrategia_dimension_id__in=estrategia_dimension_ids
        ).values_list('id', flat=True)
    )
    catalogo_fila_ids = list(
        models.DimensionCatalogoFila.objects.filter(catalogo_id__in=catalogo_ids).values_list('id', flat=True)
    )
    escala_ids = list(
        models.EscalaValor.objects.filter(
            estrategia_dimension_id__in=estrategia_dimension_ids
        ).values_list('id', flat=True)
    )

    checks = [
        ('cargas ACA/RCM/FMECA', models.Carga.objects.filter(estrategia=strategy).count()),
        (
            'evaluaciones ACA asociadas a dimensiones de la estrategia',
            models.CriticidadDimension.objects.filter(
                Q(estrategia_dimension_id__in=estrategia_dimension_ids)
                | Q(catalogo_fila_id__in=catalogo_fila_ids)
                | Q(escala_valor_id__in=escala_ids)
            ).count(),
        ),
        (
            'evaluaciones FMECA asociadas a dimensiones de la estrategia',
            models.EvaluacionFMEA.objects.filter(
                Q(estrategia_dimension_id__in=estrategia_dimension_ids)
                | Q(catalogo_fila_id__in=catalogo_fila_ids)
                | Q(escala_valor_id__in=escala_ids)
            ).count(),
        ),
        (
            'tareas RCM/FMECA asociadas a tipos de tarea de la estrategia',
            models.TareaRCM.objects.filter(tipo_tarea_estrategia_id__in=tipo_tarea_ids).count(),
        ),
        (
            'valores dinámicos de tareas RCM/FMECA asociados a campos de la estrategia',
            models.ValorCampoTareaRCM.objects.filter(campo_id__in=campo_tarea_ids).count(),
        ),
    ]
    for label, count in checks:
        if count:
            blockers.append(f'{count} {label}')
    return blockers


def _delete_strategy_related_data(strategy):
    counts = {}
    estrategia_dimension_ids = list(
        models.EstrategiaDimension.objects.filter(estrategia=strategy).values_list('id', flat=True)
    )
    catalogo_ids = list(
        models.DimensionCatalogo.objects.filter(
            estrategia_dimension_id__in=estrategia_dimension_ids
        ).values_list('id', flat=True)
    )
    catalogo_fila_ids = list(
        models.DimensionCatalogoFila.objects.filter(catalogo_id__in=catalogo_ids).values_list('id', flat=True)
    )
    catalogo_columna_ids = list(
        models.DimensionCatalogoColumna.objects.filter(catalogo_id__in=catalogo_ids).values_list('id', flat=True)
    )
    tipo_tarea_ids = list(
        models.TipoTareaEstrategia.objects.filter(estrategia=strategy).values_list('id', flat=True)
    )
    campo_tarea_ids = list(
        models.CampoTareaEstrategia.objects.filter(
            tipo_tarea_estrategia_id__in=tipo_tarea_ids
        ).values_list('id', flat=True)
    )
    matriz_ids = list(
        models.MatrizRiesgo.objects.filter(estrategia=strategy).values_list('id', flat=True)
    )
    probabilidad_ids = list(
        models.NivelProbabilidad.objects.filter(matriz_id__in=matriz_ids).values_list('id', flat=True)
    )
    impacto_ids = list(
        models.NivelImpacto.objects.filter(matriz_id__in=matriz_ids).values_list('id', flat=True)
    )

    nullable_relations = [
        ('servicios_desvinculados', models.Servicio.objects.filter(estrategia=strategy)),
        ('accesos_desvinculados', models.AccesoUsuario.objects.filter(estrategia=strategy)),
        ('plantillas_desvinculadas', models.PlantillaPauta.objects.filter(estrategia=strategy)),
        ('pautas_desvinculadas', models.Pauta.objects.filter(estrategia=strategy)),
        ('reglas_pauta_desvinculadas', models.ReglaGeneracionPauta.objects.filter(estrategia=strategy)),
    ]
    for key, queryset in nullable_relations:
        counts[key] = queryset.update(estrategia=None)

    counts['celdas_matriz'] = _delete_by_materialized_ids(
        models.MatrizRiesgoCelda,
        models.MatrizRiesgoCelda.objects.filter(matriz_id__in=matriz_ids).values_list('id', flat=True),
    )
    counts['niveles_probabilidad'] = _delete_by_materialized_ids(models.NivelProbabilidad, probabilidad_ids)
    counts['niveles_consecuencia'] = _delete_by_materialized_ids(models.NivelImpacto, impacto_ids)
    counts['matrices'] = _delete_by_materialized_ids(models.MatrizRiesgo, matriz_ids)

    counts['celdas_catalogo'] = _delete_by_materialized_ids(
        models.DimensionCatalogoCelda,
        models.DimensionCatalogoCelda.objects.filter(
            Q(fila_id__in=catalogo_fila_ids) | Q(columna_id__in=catalogo_columna_ids)
        ).values_list('id', flat=True),
    )
    counts['columnas_catalogo'] = _delete_by_materialized_ids(models.DimensionCatalogoColumna, catalogo_columna_ids)
    counts['filas_catalogo'] = _delete_by_materialized_ids(models.DimensionCatalogoFila, catalogo_fila_ids)
    counts['catalogos'] = _delete_by_materialized_ids(models.DimensionCatalogo, catalogo_ids)
    counts['escalas_valor'] = _delete_by_materialized_ids(
        models.EscalaValor,
        models.EscalaValor.objects.filter(
            estrategia_dimension_id__in=estrategia_dimension_ids
        ).values_list('id', flat=True),
    )
    counts['campos_tarea'] = _delete_by_materialized_ids(models.CampoTareaEstrategia, campo_tarea_ids)
    counts['tipos_tarea'] = _delete_by_materialized_ids(models.TipoTareaEstrategia, tipo_tarea_ids)
    counts['dimensiones_estrategia'] = _delete_by_materialized_ids(
        models.EstrategiaDimension,
        estrategia_dimension_ids,
    )
    return counts


def _delete_company_related_data(company):
    counts = {}
    service_ids = list(models.Servicio.objects.filter(empresa=company).values_list('id', flat=True))
    strategy_ids = list(models.Estrategia.objects.filter(empresa=company).values_list('id', flat=True))
    node_ids = list(models.NodoJerarquia.objects.filter(empresa=company).values_list('id', flat=True))
    equipment_ids = set(
        models.Equipo.objects.filter(nodo_id__in=node_ids).values_list('id', flat=True)
    )
    equipment_ids.update(
        models.ServicioEquipo.objects.filter(servicio_id__in=service_ids).values_list('equipo_id', flat=True)
    )

    for service in models.Servicio.objects.filter(id__in=service_ids):
        _merge_delete_counts(counts, _delete_service_related_data(service), prefix='servicio_')
        counts['servicios'] = counts.get('servicios', 0) + _delete_by_materialized_ids(models.Servicio, [service.pk])

    _merge_delete_counts(counts, _delete_equipment_related_data(equipment_ids), prefix='empresa_')

    plantilla_ids = list(
        models.PlantillaPauta.objects.filter(empresa=company).values_list('id', flat=True)
    )
    counts['mapeos_plantilla_empresa'] = _delete_by_materialized_ids(
        models.MapeoPlantillaPauta,
        models.MapeoPlantillaPauta.objects.filter(plantilla_id__in=plantilla_ids).values_list('id', flat=True),
    )
    counts['plantillas_empresa'] = _delete_by_materialized_ids(models.PlantillaPauta, plantilla_ids)

    for strategy in models.Estrategia.objects.filter(id__in=strategy_ids):
        _merge_delete_counts(counts, _delete_strategy_related_data(strategy), prefix='estrategia_')
        counts['estrategias'] = counts.get('estrategias', 0) + _delete_by_materialized_ids(models.Estrategia, [strategy.pk])

    counts['accesos_empresa'] = _delete_by_materialized_ids(
        models.AccesoUsuario,
        models.AccesoUsuario.objects.filter(empresa=company).values_list('id', flat=True),
    )
    counts['usuarios_desvinculados'] = models.Usuario.all_objects.filter(empresa=company).update(empresa=None)
    counts['usuarios_eliminados_desvinculados'] = models.UsuarioEliminado.objects.filter(empresa=company).update(empresa=None)
    counts['valores_nivel_jerarquia'] = _delete_by_materialized_ids(
        models.ValorNivelJerarquia,
        models.ValorNivelJerarquia.objects.filter(empresa=company).values_list('id', flat=True),
    )
    counts['nodos_jerarquia'] = _delete_nodes_by_depth(node_ids)
    counts['niveles_jerarquia'] = _delete_by_materialized_ids(
        models.NivelJerarquia,
        models.NivelJerarquia.objects.filter(empresa=company).values_list('id', flat=True),
    )
    return counts


def empresa_logo(request, empresa_id):
    empresa = models.Empresa.objects.filter(id=empresa_id).first()

    if not empresa or not empresa.logo or empresa.nombre == "Mindco":
        return HttpResponse(status=204)

    return HttpResponse(empresa.logo, content_type="image/png")

def _get_config(model_key: str):
    config = get_registered_model(model_key)
    if not config:
        raise Http404('Módulo no encontrado')
    config = dict(config)
    config['db_table'] = config['model']._meta.db_table
    return config


def _ensure_admin_access(request):
    if not request.user.is_authenticated:
        raise PermissionDenied('Debes iniciar sesión.')
    if not request.user.is_superuser:
        raise PermissionDenied('Solo el superusuario puede acceder a este módulo.')


def _ensure_direct_crud_allowed(config):
    if not config.get('allow_direct_crud', True):
        raise Http404('Este módulo se administra desde un editor especí­fico.')


FORM_TEMPLATE_BY_MODEL = {
    'empresa': 'core/forms/empresa_form.html',
    'cargo': 'core/forms/cargo_form.html',
    'usuario': 'core/forms/usuario_form.html',
    'componente': 'core/forms/componente_form.html',
    'equipo': 'core/forms/equipo_form.html',
    'dimension': 'core/forms/dimension_form.html',
    'escalaunificada': 'core/forms/escalaunificada_form.html',
    'estrategia': 'core/forms/estrategia_form.html',
    'servicio': 'core/forms/servicio_form.html',
    'accesousuario': 'core/forms/accesousuario_form.html',
    'carga': 'core/forms/acacarga_form.html',
    'criticidad': 'core/forms/criticidad_form.html',
    'matrizriesgo': 'core/forms/matrizriesgo_form.html',
}


def _form_template_for(model_key):
    return FORM_TEMPLATE_BY_MODEL.get(model_key, 'core/model_form.html')


def _visible_fields(model_class):
    return [f for f in model_class._meta.fields if f.name != 'id']


def _list_fields(model_class):
    base = []
    exclude = [id]
    for field in _visible_fields(model_class):
        if isinstance(field, TextField) and not field.is_relation:
            continue
        base.append(field)
    return base[:MAX_LIST_COLUMNS] or _visible_fields(model_class)[:MAX_LIST_COLUMNS]


def _detail_fields(model_class):
    return _visible_fields(model_class)


def _equipment_location_context(form=None, obj=None):
    selected_node_id = None
    selected_empresa_id = None
    if form is not None and getattr(form, 'is_bound', False):
        selected_node_id = form.data.get(form.add_prefix('nodo'))
        selected_empresa_id = form.data.get(form.add_prefix('empresa'))
    if not selected_node_id and obj is not None:
        selected_node_id = getattr(obj, 'nodo_id', None)
    if not selected_empresa_id and obj is not None and getattr(obj, 'nodo_id', None):
        selected_empresa_id = obj.nodo.empresa_id
    if not selected_node_id and form is not None:
        selected_node_id = form.initial.get('nodo')
        selected_node_id = getattr(selected_node_id, 'pk', selected_node_id)
    if not selected_empresa_id and form is not None:
        selected_empresa_id = form.initial.get('empresa')
        selected_empresa_id = getattr(selected_empresa_id, 'pk', selected_empresa_id)

    levels = [
        {
            'id': level.pk,
            'empresa_id': level.empresa_id,
            'order': level.orden,
            'name': level.nombre,
        }
        for level in models.NivelJerarquia.objects.filter(
            activo=True,
        ).select_related(
            'empresa',
        ).order_by(
            'empresa__nombre',
            'orden',
        )
    ]
    nodes = []
    if selected_node_id:
        try:
            selected_node_id_int = int(selected_node_id)
        except (TypeError, ValueError):
            selected_node_id_int = None
        if selected_node_id_int:
            rows_by_id = _equipment_path_rows({selected_node_id_int})
            for node_id in _path_ids_for_node(selected_node_id_int, rows_by_id):
                row = rows_by_id.get(node_id)
                if not row:
                    continue
                path_ids = _path_ids_for_node(node_id, rows_by_id)
                nodes.append({
                    'id': row['id'],
                    'empresa_id': row['empresa_id'],
                    'parent_id': row['parent_id'],
                    'level_id': row['nivel_id'],
                    'level_order': row['nivel__orden'],
                    'code': row['codigo'],
                    'name': row['nombre'],
                    'label': f"{row['codigo']} - {row['nombre']}",
                    'ut': _node_ut_from_path(path_ids, rows_by_id),
                    'route': _node_route_from_path(path_ids, rows_by_id),
                })
    return {
        'tech_location_levels_json': json.dumps(levels, ensure_ascii=False),
        'tech_location_nodes_json': json.dumps(nodes, ensure_ascii=False),
        'tech_location_selected_id': str(selected_node_id or ''),
        'tech_location_empresa_id': str(selected_empresa_id or ''),
        'equipment_services_json': json.dumps([
            {
                'id': service.pk,
                'empresa_id': service.empresa_id,
                'label': f'{service.codigo_servicio} - {service.descripcion}'.strip(' -'),
            }
            for service in models.Servicio.objects.select_related('empresa').order_by('empresa__nombre', 'codigo_servicio')
        ], ensure_ascii=False),
        'tech_location_nodes_url_template': reverse(
            'hierarchy_values_nodes',
            kwargs={'empresa_id': 0},
        ).replace('/0/', '/__empresa__/'),
    }


def _path_ids_for_node(node_id, rows_by_id):
    path = []
    seen = set()
    current_id = node_id
    while current_id and current_id not in seen:
        row = rows_by_id.get(current_id)
        if not row:
            break
        seen.add(current_id)
        path.append(current_id)
        current_id = row.get('parent_id')
    return list(reversed(path))


def _node_ut_from_path(path_ids, rows_by_id):
    return '-'.join(
        models._technical_segment(rows_by_id[node_id]['codigo'])
        for node_id in path_ids
        if node_id in rows_by_id
    )


def _node_route_from_path(path_ids, rows_by_id):
    parts = []
    for node_id in path_ids:
        row = rows_by_id.get(node_id)
        if not row:
            continue
        parts.append(f"{row['nivel__nombre']}: {row['codigo']} - {row['nombre']}")
    return ' > '.join(parts)


def _service_equipment_browser_payload(service):
    if not service:
        return {'levels': [], 'nodes': [], 'equipment': []}

    levels = [
        {
            'id': level.pk,
            'order': level.orden,
            'name': level.nombre,
        }
        for level in models.NivelJerarquia.objects.filter(
            empresa=service.empresa,
            activo=True,
        ).order_by('orden')
    ]

    return {
        'levels': levels,
        'nodes': [],
        'equipment': [],
    }


def _service_equipment_count(service):
    if not service:
        return 0
    return get_service_equipment(service).order_by().count()


def _service_equipment_endpoints(service):
    if not service:
        return {'nodes': '', 'search': '', 'detail_template': ''}
    return {
        'nodes': reverse('service_equipment_nodes', kwargs={'pk': service.pk}),
        'search': reverse('service_equipment_search', kwargs={'pk': service.pk}),
        'detail_template': reverse(
            'service_equipment_detail',
            kwargs={'pk': service.pk, 'equipment_pk': 0},
        ).replace('/0/', '/__id__/'),
    }


def _equipment_path_rows(node_ids):
    rows_by_id = {}
    pending = {node_id for node_id in node_ids if node_id}
    while pending:
        rows = list(
            models.NodoJerarquia.objects.filter(
                pk__in=pending,
            ).values(
                'id',
                'empresa_id',
                'parent_id',
                'nivel_id',
                'nivel__orden',
                'nivel__nombre',
                'codigo',
                'nombre',
                'activo',
            )
        )
        pending = set()
        for row in rows:
            rows_by_id[row['id']] = row
            parent_id = row.get('parent_id')
            if parent_id and parent_id not in rows_by_id:
                pending.add(parent_id)
    return rows_by_id


def _equipment_items_payload(equipment_items):
    equipment_items = list(equipment_items)
    rows_by_id = _equipment_path_rows({equipo.nodo_id for equipo in equipment_items if equipo.nodo_id})
    payload = []
    for equipo in equipment_items:
        path_ids = _path_ids_for_node(equipo.nodo_id, rows_by_id) if equipo.nodo_id else []
        payload.append({
            'id': equipo.pk,
            'ut': equipo.ut_display or '',
            'descripcion_ut': equipo.descripcion_ut or '',
            'equipo': equipo.nombre_equipo or '',
            'tag': equipo.tag_display or '',
            'node_id': equipo.nodo_id,
            'path_node_ids': path_ids,
            'path_text': _node_route_from_path(path_ids, rows_by_id),
        })
    return payload


def _service_equipment_search_queryset(service, node=None, query=''):
    qs = get_service_equipment(service)
    if node:
        subtree_ids = _active_subtree_ids(node)
        node_ut = node.ut
        qs = qs.filter(
            Q(nodo_id__in=subtree_ids)
            | Q(ut__iexact=node_ut)
            | Q(ut__istartswith=f'{node_ut}-')
        )

    query = (query or '').strip()
    if query:
        qs = qs.filter(
            Q(ut__icontains=query)
            | Q(descripcion_ut__icontains=query)
            | Q(nombre_equipo__icontains=query)
            | Q(tag_equipo__icontains=query)
        )
    return qs.distinct().order_by('tag_equipo', 'nombre_equipo', 'ut')


def _active_subtree_ids(node):
    ids = []
    frontier = [node.pk]
    while frontier:
        current_ids = list(
            models.NodoJerarquia.objects.filter(
                empresa=node.empresa,
                pk__in=frontier,
                activo=True,
            ).values_list(
                'pk',
                flat=True,
            )
        )
        if not current_ids:
            break
        ids.extend(current_ids)
        frontier = list(
            models.NodoJerarquia.objects.filter(
                empresa=node.empresa,
                parent_id__in=current_ids,
                activo=True,
            ).values_list(
                'pk',
                flat=True,
            )
        )
    return ids


@login_required
def dashboard(request):
    accessible_services_qs = get_accessible_services(request.user).select_related('empresa', 'estrategia')
    accessible_service_ids = list(accessible_services_qs.values_list('id', flat=True))
    servicios = list(accessible_services_qs[:8])
    service_ids = [service.pk for service in servicios]

    aca_evaluated_by_service = {}
    fmeca_evaluated_by_service = {}
    if service_ids:
        aca_evaluated_by_service = {
            row['aca_carga__servicio_id']: row['total']
            for row in models.Criticidad.objects.filter(
                aca_carga__servicio_id__in=service_ids,
                equipo_id__isnull=False,
            ).values('aca_carga__servicio_id').annotate(total=Count('equipo_id', distinct=True))
        }
        fmeca_evaluated_by_service = {
            row['carga__servicio_id']: row['total']
            for row in models.RCM.objects.filter(
                carga__servicio_id__in=service_ids,
                equipo_id__isnull=False,
                fmea_fmeca__isnull=False,
            ).values('carga__servicio_id').annotate(total=Count('equipo_id', distinct=True))
        }
        strategy_ids = {service.estrategia_id for service in servicios if service.estrategia_id}
        strategies_with_matrix = set(
            models.MatrizRiesgo.objects
            .filter(estrategia_id__in=strategy_ids)
            .values_list('estrategia_id', flat=True)
        )
        strategies_with_dimensions = set(
            models.EstrategiaDimension.objects
            .filter(estrategia_id__in=strategy_ids, activo=True)
            .values_list('estrategia_id', flat=True)
        )
        services_with_equipment = set(
            models.ServicioEquipo.objects
            .filter(servicio_id__in=service_ids)
            .values_list('servicio_id', flat=True)
        )
    else:
        strategies_with_matrix = set()
        strategies_with_dimensions = set()
        services_with_equipment = set()
    for service in servicios:
        service.dashboard_aca_evaluated = aca_evaluated_by_service.get(service.pk, 0)
        service.dashboard_fmeca_evaluated = fmeca_evaluated_by_service.get(service.pk, 0)
        alerts = []
        if not service.estrategia_id:
            alerts.append('Servicio sin estrategia')
        else:
            if service.estrategia_id not in strategies_with_dimensions:
                alerts.append('Servicio sin tablas de evaluación')
            if service.estrategia_id not in strategies_with_matrix:
                alerts.append('Servicio sin matriz de riesgo')
        if service.pk not in services_with_equipment:
            alerts.append('Servicio sin equipos vinculados')
        service.dashboard_alerts = alerts

    activos_evaluados_ACA = (
        models.Equipo.objects
        .filter(criticidades__aca_carga__servicio_id__in=accessible_service_ids)
        .distinct()
        .count()
        if accessible_service_ids else 0
    )
    activos_evaluados_FMECA = (
        models.Equipo.objects
        .filter(
            registros_rcm__fmea_fmeca__isnull=False,
            registros_rcm__carga__servicio_id__in=accessible_service_ids,
        )
        .distinct()
        .count()
        if accessible_service_ids else 0
    )
    activos_evaluados_total = (
        models.Equipo.objects
        .filter(
            Q(criticidades__aca_carga__servicio_id__in=accessible_service_ids)
            | Q(
                registros_rcm__fmea_fmeca__isnull=False,
                registros_rcm__carga__servicio_id__in=accessible_service_ids,
            )
        )
        .distinct()
        .count()
        if accessible_service_ids else 0
    )
    editable_services = [s for s in accessible_services_qs if get_service_permission(request.user, s)['can_edit']]
    return render(request, 'core/dashboard.html', {
        'service_cards': servicios,
        'quick_stats': [
            {'label': 'Servicios accesibles', 'count': len(accessible_service_ids)},
            {'label': 'Servicios editables', 'count': len(editable_services)},
            {'label': 'Activos evaluados', 'count': activos_evaluados_total},
            {'label': 'Activos evaluados en ACA', 'count': activos_evaluados_ACA},
            {'label': 'Activos evaluados en FMECA', 'count': activos_evaluados_FMECA},
        ]
    })


@login_required
def model_list(request, model_key):
    config = _get_config(model_key)
    _ensure_direct_crud_allowed(config)
    model = config['model']
    if model_key == 'equipo':
        try:
            requested_page_size = int(request.GET.get('page_size') or 50)
        except (TypeError, ValueError):
            requested_page_size = 50
        page_size = requested_page_size if requested_page_size in {10, 25, 50, 100} else 50
    else:
        page_size = 100
    equipment_filter_context = {}

    editable_service_ids = set()
    model_list_stats = []
    if model_key == 'servicio':
        qs = get_accessible_services(request.user)
        if request.user.is_superuser:
            editable_service_ids = set(qs.values_list('id', flat=True))
        else:
            editable_service_ids = {
                service.id
                for service in qs
                if get_service_permission(request.user, service)['can_edit']
            }
        today = timezone.localdate()
        next_30_days = today + timedelta(days=30)
        scheduled_qs = qs.exclude(status='cerrado')
        upcoming_starts = scheduled_qs.filter(fecha_inicio__range=(today, next_30_days))
        upcoming_finishes = scheduled_qs.filter(fecha_fin__range=(today, next_30_days))
        nearest_finish = (
            scheduled_qs
            .filter(fecha_fin__gte=today)
            .order_by('fecha_fin', 'codigo_servicio')
            .values('codigo_servicio', 'fecha_fin')
            .first()
        )
        service_field_order = [
            'status',
            'codigo_servicio',
            'fecha_inicio',
            'fecha_fin',
            'creado_en',
        ]
        model_list_stats = [
            {
                'label': 'Servicios accesibles',
                'value': qs.count(),
                'detail': 'Disponibles para tu usuario',
            },
            {
                'label': 'Próximos a iniciar',
                'value': upcoming_starts.count(),
                'detail': 'Dentro de los próximos 30 días',
            },
            {
                'label': 'Próximos a finalizar',
                'value': upcoming_finishes.count(),
                'detail': 'Dentro de los próximos 30 días',
            },
            {
                'label': 'Más próximo a finalizar',
                'value': nearest_finish['codigo_servicio'] if nearest_finish else '-',
                'detail': nearest_finish['fecha_fin'].strftime('%d/%m/%Y') if nearest_finish else 'Sin fechas futuras registradas',
            },
        ]
    else:
        qs = model.objects.all()
        if model_key == 'equipo':
            qs = qs.only(
                'id',
                'tag_equipo',
                'nombre_equipo',
                'ut',
                'descripcion_ut',
                'nodo_id',
            )
            equipment_company_ids = (
                models.Equipo.objects
                .filter(nodo__empresa_id__isnull=False)
                .values_list('nodo__empresa_id', flat=True)
                .distinct()
            )
            service_company_ids = (
                models.ServicioEquipo.objects
                .filter(servicio__empresa_id__isnull=False)
                .values_list('servicio__empresa_id', flat=True)
                .distinct()
            )
            equipment_companies = models.Empresa.objects.filter(
                Q(id__in=equipment_company_ids) | Q(id__in=service_company_ids)
            ).distinct().order_by('nombre', 'sigla')
            selected_empresa_id = (request.GET.get('empresa') or '').strip()
            selected_service_id = (request.GET.get('servicio') or '').strip()
            selected_node_id = (request.GET.get('node_id') or '').strip()
            selected_service = None
            if selected_service_id:
                selected_service = models.Servicio.objects.filter(pk=selected_service_id).select_related('empresa').first()
                if selected_service and not selected_empresa_id:
                    selected_empresa_id = str(selected_service.empresa_id)
            selected_node = None
            if selected_node_id:
                selected_node = models.NodoJerarquia.objects.filter(pk=selected_node_id, activo=True).select_related('empresa').first()
                if selected_node:
                    selected_empresa_id = str(selected_node.empresa_id)
            selected_empresa = None
            if selected_empresa_id:
                selected_empresa = models.Empresa.objects.filter(pk=selected_empresa_id).first()
            if selected_empresa:
                qs = qs.filter(
                    Q(nodo__empresa=selected_empresa)
                    | Q(servicios_equipo__servicio__empresa=selected_empresa)
                )
            if selected_service:
                qs = qs.filter(servicios_equipo__servicio=selected_service)
            if selected_node:
                subtree_ids = _active_subtree_ids(selected_node)
                node_ut = selected_node.ut
                qs = qs.filter(
                    Q(nodo_id__in=subtree_ids)
                    | Q(ut__iexact=node_ut)
                    | Q(ut__istartswith=f'{node_ut}-')
                )

            selected_levels = []
            if selected_empresa:
                selected_levels = list(
                    models.NivelJerarquia.objects
                    .filter(empresa=selected_empresa, activo=True)
                    .order_by('orden')
                    .values('id', 'nombre', 'orden')
                )
            selected_nodes = []
            if selected_node:
                rows_by_id = _equipment_path_rows({selected_node.pk})
                for node_id in _path_ids_for_node(selected_node.pk, rows_by_id):
                    row = rows_by_id.get(node_id)
                    if not row:
                        continue
                    path_ids = _path_ids_for_node(node_id, rows_by_id)
                    selected_nodes.append({
                        'id': row['id'],
                        'empresa_id': row['empresa_id'],
                        'parent_id': row['parent_id'],
                        'level_id': row['nivel_id'],
                        'level_order': row['nivel__orden'],
                        'code': row['codigo'],
                        'name': row['nombre'],
                        'label': f"{row['codigo']} - {row['nombre']}",
                        'ut': _node_ut_from_path(path_ids, rows_by_id),
                        'route': _node_route_from_path(path_ids, rows_by_id),
                    })
            equipment_filter_context = {
                'equipment_filter_enabled': True,
                'equipment_companies': equipment_companies,
                'equipment_services': models.Servicio.objects.select_related('empresa').order_by('codigo_servicio'),
                'equipment_selected_empresa_id': str(selected_empresa_id or ''),
                'equipment_selected_service_id': str(selected_service_id or ''),
                'equipment_selected_node_id': str(selected_node_id or ''),
                'equipment_levels': selected_levels,
                'equipment_levels_json': json.dumps([
                    {'id': item['id'], 'name': item['nombre'], 'order': item['orden']}
                    for item in selected_levels
                ], ensure_ascii=False),
                'equipment_nodes_json': json.dumps(selected_nodes, ensure_ascii=False),
                'equipment_nodes_url_template': reverse(
                    'hierarchy_values_nodes',
                    kwargs={'empresa_id': 0},
                ).replace('/0/', '/__empresa__/'),
            }

    search = request.GET.get('q', '').strip()
    if search and model_key == 'equipo':
        normalized_search = search.upper()
        if '-' in normalized_search:
            qs = qs.filter(
                Q(ut__iexact=normalized_search)
                | Q(ut__istartswith=f'{normalized_search}-')
                | Q(ut__icontains=normalized_search)
                | Q(tag_equipo__icontains=search)
                | Q(nombre_equipo__icontains=search)
            )
        else:
            qs = qs.filter(
                Q(tag_equipo__icontains=search)
                | Q(nombre_equipo__icontains=search)
                | Q(ut__icontains=search)
                | Q(descripcion_ut__icontains=search)
            )
    elif search and config['search_fields']:
        clause = Q()
        for field_name in config['search_fields']:
            clause |= Q(**{f'{field_name}__icontains': search})
        qs = qs.filter(clause)

    if model_key == 'equipo':
        qs = qs.distinct()

    total_count = qs.count()
    paginator = Paginator(qs, page_size)
    page_obj = paginator.get_page(request.GET.get('page'))
    rows = page_obj.object_list
    equipment_hierarchy_values = {}
    if model_key == 'equipo' and equipment_filter_context.get('equipment_levels'):
        rows_by_id = _equipment_path_rows({row.nodo_id for row in rows if row.nodo_id})
        for row in rows:
            path_ids = _path_ids_for_node(row.nodo_id, rows_by_id) if row.nodo_id else []
            parts = {}
            for node_id in path_ids:
                node_row = rows_by_id.get(node_id)
                if not node_row:
                    continue
                parts[node_row['nivel_id']] = f"{node_row['codigo']} - {node_row['nombre']}"
            equipment_hierarchy_values[row.pk] = parts

    list_fields = [
        field for field in _list_fields(model)
        if field.name != 'id'
        and not (model_key == 'empresa' and field.name == 'logo')
        and not (model_key == 'equipo' and field.name == 'nodo')
    ]

    if model_key == 'servicio':
        service_field_order = [
            'codigo_servicio',
            'status',
            'fecha_inicio',
            'fecha_fin',
            'creado_en',
        ]
        service_fields_by_name = {field.name: field for field in list_fields}
        list_fields = [
            service_fields_by_name[field_name]
            for field_name in service_field_order
            if field_name in service_fields_by_name
        ]

    list_query = request.GET.copy()
    list_query.pop('page', None)
    context = {
        'config': config,
        'model_key': model_key,
        'rows': rows,
        'search': search,
        'list_fields': list_fields,
        'editable_service_ids': editable_service_ids,
        'total_count': total_count,
        'page_obj': page_obj,
        'paginator': paginator,
        'page_size': page_size,
        'list_query_string': list_query.urlencode(),
        'equipment_hierarchy_values': equipment_hierarchy_values,
        'model_list_stats': model_list_stats,
    }
    context.update(equipment_filter_context)
    return render(request, 'core/model_list.html', context)


@login_required
def equipment_bulk_example(request, formato):
    _ensure_admin_access(request)
    formato = (formato or '').strip().lower()
    examples = {
        'mindco_simple': {
            'filename': 'ejemplo_equipos_mindco_simple.xlsx',
            'sheet': 'Mindco simple',
            'headers': [
                'Empresa',
                'Planta / Sitio',
                'Area',
                'Sistema',
                'SubSistema',
                'Ubicacion Tecnica',
                'TAG',
                'Equipo / Componente',
                'Descripcion U.T.',
            ],
            'rows': [
                ['Minera Inventada S.A.', 'Planta Norte', 'Molienda', 'Transporte de pulpa', 'Bombas', 'MIN-PLN-MOL-PUL-BOM', 'A001', 'Bomba de pulpa 1', 'Linea de bombeo de pulpa planta norte'],
                ['Minera Inventada S.A.', 'Planta Norte', 'Chancado', 'Correas transportadoras', 'Correa principal', 'MIN-PLN-CHA-COR-PRI', 'A002', 'Correa transportadora 2', 'Sistema de transporte de mineral chancado'],
            ],
        },
        'sap_uts': {
            'filename': 'ejemplo_equipos_sap_uts.xlsx',
            'sheet': 'SAP UTS',
            'headers': [
                'Ubicacion tecnica',
                'Denominacion',
                'N0',
                'N0_nombre',
                'N1',
                'N1_nombre',
                'N2',
                'N2_nombre',
                'N3',
                'N3_nombre',
                'N4',
                'N4_nombre',
            ],
            'rows': [
                ['MIN-PLN-MOL-PUL-BOM-A101', 'Bomba centrifuga de pulpa A101', 'MIN', 'Minera Inventada S.A.', 'MIN-PLN', 'Planta Norte', 'MIN-PLN-MOL', 'Molienda', 'MIN-PLN-MOL-PUL', 'Transporte de pulpa', 'MIN-PLN-MOL-PUL-BOM-A101', 'Bomba centrifuga de pulpa A101'],
                ['MIN-PLN-CHA-COR-PRI-A102', 'Correa transportadora principal A102', 'MIN', 'Minera Inventada S.A.', 'MIN-PLN', 'Planta Norte', 'MIN-PLN-CHA', 'Chancado', 'MIN-PLN-CHA-COR', 'Correas transportadoras', 'MIN-PLN-CHA-COR-PRI-A102', 'Correa transportadora principal A102'],
            ],
        },
    }
    example = examples.get(formato)
    if not example:
        raise Http404('Formato de ejemplo no soportado.')

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = example['sheet']
    worksheet.append(example['headers'])
    for row in example['rows']:
        worksheet.append(row)

    header_fill = PatternFill('solid', fgColor='E8EEF7')
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for index, header in enumerate(example['headers'], start=1):
        letter = get_column_letter(index)
        max_length = max(
            len(str(worksheet.cell(row=row_idx, column=index).value or ''))
            for row_idx in range(1, worksheet.max_row + 1)
        )
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, len(header) + 2, 14), 42)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{example["filename"]}"'
    return response


def _delete_session_upload_file(path):
    if not path or not default_storage.exists(path):
        return
    try:
        default_storage.delete(path)
    except OSError:
        pass


def store_session_upload(request, session_key, uploaded_file, folder):
    previous = request.session.get(session_key)
    if previous:
        _delete_session_upload_file(previous.get('path'))
    safe_name = re.sub(r'[^A-Za-z0-9._-]+', '_', uploaded_file.name or 'archivo.xlsx').strip('_') or 'archivo.xlsx'
    path = default_storage.save(f'tmp_uploads/{folder}/{uuid.uuid4().hex}_{safe_name}', uploaded_file)
    ref = {'path': path, 'name': uploaded_file.name or safe_name}
    request.session[session_key] = ref
    request.session.modified = True
    return ref


def open_session_upload(request, session_key):
    ref = request.session.get(session_key) or {}
    path = ref.get('path')
    if not path or not default_storage.exists(path):
        return None, None
    file_obj = default_storage.open(path, 'rb')
    file_obj.name = ref.get('name') or path.rsplit('/', 1)[-1]
    return file_obj, ref


def clear_session_upload(request, session_key):
    ref = request.session.pop(session_key, None)
    request.session.modified = True
    if ref:
        _delete_session_upload_file(ref.get('path'))


@login_required
def equipment_bulk_upload(request):
    _ensure_admin_access(request)
    report = None
    preview_only = True
    upload_session_key = 'equipment_bulk_upload_file'
    if request.method == 'POST':
        form = EquipoBulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data.get('archivo')
            if uploaded_file:
                store_session_upload(request, upload_session_key, uploaded_file, 'equipment')
            stored_file, _stored_ref = open_session_upload(request, upload_session_key)
            if not stored_file:
                form.add_error('archivo', 'Selecciona un archivo Excel para continuar.')
            else:
                form.cleaned_data['archivo'] = stored_file
        if form.is_valid():
            try:
                preview = preview_equipment_import(
                    form.cleaned_data['archivo'],
                    form.cleaned_data['empresa'],
                    servicio=form.cleaned_data.get('servicio'),
                    sheet_name=form.cleaned_data.get('hoja') or None,
                    format=form.cleaned_data.get('formato') or 'auto',
                    last_segment_is_equipment=None,
                    last_level_is_equipment=True,
                )
                if request.POST.get('action') == 'confirm':
                    preview_only = False
                    report = execute_equipment_import(preview, user=request.user)
                else:
                    report = preview

                if not preview_only:
                    clear_session_upload(request, upload_session_key)

                if preview_only:
                    if report.errors:
                        messages.warning(request, 'La previsualizacion encontro errores. Corrige el archivo antes de confirmar.')
                    else:
                        messages.info(request, 'Previsualizacion lista. Puedes confirmar la carga sin volver a seleccionar el archivo.')
                elif report.equipment_created or report.equipment_updated:
                    messages.success(
                        request,
                        f'Carga masiva completada: {report.equipment_created} equipos creados y {report.equipment_updated} actualizados.',
                    )
                else:
                    messages.warning(request, 'No se crearon ni actualizaron equipos. Revisa las observaciones de la carga.')
            except ValueError as exc:
                messages.error(request, str(exc))
            finally:
                try:
                    form.cleaned_data['archivo'].close()
                except Exception:
                    pass
    else:
        clear_session_upload(request, upload_session_key)
        form = EquipoBulkUploadForm()
    stored_upload = request.session.get(upload_session_key)

    return render(request, 'core/equipment_bulk_upload.html', {
        'form': form,
        'report': report,
        'preview_only': preview_only,
        'stored_upload': stored_upload,
        'equipment_upload_services': list(
            models.Servicio.objects
            .select_related('empresa')
            .order_by('empresa__nombre', 'codigo_servicio')
            .values('id', 'codigo_servicio', 'descripcion', 'empresa_id', 'empresa__sigla')
        ),
    })


@login_required
def model_detail(request, model_key, pk):
    config = _get_config(model_key)
    _ensure_direct_crud_allowed(config)

    if model_key == 'servicio':
        _service_or_404(request, pk, edit=False)
        return redirect('service_detail', pk=pk)

    _ensure_admin_access(request)
    return redirect('model_list', model_key=model_key)


@login_required
def model_create(request, model_key):
    _ensure_admin_access(request)
    if model_key == 'matrizriesgo':
        return redirect('matriz_builder_new')
    config = _get_config(model_key)
    _ensure_direct_crud_allowed(config)
    FormClass = get_form_for_key(model_key)
    form_kwargs = {}
    if model_key == 'servicio':
        creador_usuario = get_profile_for_user(request.user)
        if not creador_usuario:
            creador_usuario = sync_profile_from_auth_user(request.user)
        form_kwargs['creador_usuario'] = creador_usuario

    if request.method == 'POST':
        form = FormClass(request.POST, request.FILES, **form_kwargs)
        if form.is_valid():
            obj = form.save()
            if model_key == 'servicio':
                return redirect('service_detail', pk=obj.pk)
            return redirect('model_list', model_key=model_key)
    else:
        form = FormClass(**form_kwargs)

    context = {
        'config': config,
        'model_key': model_key,
        'form': form,
        'mode': 'create',
    }
    if model_key == 'equipo':
        context.update(_equipment_location_context(form=form))
    return render(request, _form_template_for(model_key), context)


@login_required
def model_update(request, model_key, pk):
    config = _get_config(model_key)
    _ensure_direct_crud_allowed(config)

    if model_key == 'servicio':
        obj, permission = _service_or_404(request, pk, edit=True)
    else:
        _ensure_admin_access(request)
        obj = get_object_or_404(config['model'], pk=pk)
        permission = None

    FormClass = get_form_for_key(model_key)

    if request.method == 'POST':
        form = FormClass(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            obj = form.save()
            if model_key == 'servicio':
                return redirect('service_detail', pk=obj.pk)
            return redirect('model_list', model_key=model_key)
    else:
        form = FormClass(instance=obj)

    context = {
        'config': config,
        'model_key': model_key,
        'form': form,
        'object': obj,
        'mode': 'update',
        'permission': permission,
    }
    if model_key == 'equipo':
        context.update(_equipment_location_context(form=form, obj=obj))
    return render(request, _form_template_for(model_key), context)


@login_required
@transaction.atomic
def model_delete(request, model_key, pk):
    _ensure_admin_access(request)
    config = _get_config(model_key)
    _ensure_direct_crud_allowed(config)

    lookup_qs = config['model']
    if model_key == 'usuario' and hasattr(config['model'], 'all_objects'):
        lookup_qs = config['model'].all_objects

    obj = get_object_or_404(lookup_qs, pk=pk)

    _ensure_delete_allowed(request, config, obj)

    if request.method == 'POST':
        if model_key == 'usuario':
            if getattr(obj, 'is_deleted', False):
                messages.info(request, 'El usuario ya estaba archivado.')
            else:
                archive_profile(
                    obj,
                    deleted_by=request.user,
                    reason='Eliminado desde la plataforma',
                )
                messages.success(
                    request,
                    'Usuario archivado correctamente. Ya no aparecerá en listados ni opciones.'
                )
            return redirect('model_list', model_key=model_key)
        if model_key == 'matrizriesgo':
            models.MatrizRiesgoCelda.objects.filter(matriz=obj).delete()
            models.NivelImpacto.objects.filter(matriz=obj).delete()
            models.NivelProbabilidad.objects.filter(matriz=obj).delete()

            obj.delete()

            messages.success(request, 'Matriz eliminada correctamente.')
            return redirect('model_list', model_key=model_key)
        if model_key == 'servicio':
            counts = _delete_service_related_data(obj)
            obj.delete()
            removed_children = sum(counts.values())
            detail = f' Se eliminaron {removed_children} registros asociados.' if removed_children else ''
            messages.success(request, f'Servicio eliminado correctamente.{detail}')
            return redirect('model_list', model_key=model_key)
        if model_key == 'empresa':
            counts = _delete_company_related_data(obj)
            obj.delete()
            removed_children = sum(counts.values())
            detail = f' Se eliminaron, desvincularon o limpiaron {removed_children} registros asociados.' if removed_children else ''
            messages.success(request, f'Empresa eliminada correctamente.{detail}')
            return redirect('model_list', model_key=model_key)
        if model_key == 'equipo':
            counts = _delete_equipment_related_data([obj.pk])
            removed_children = sum(counts.values()) - counts.get('equipos', 0)
            detail = f' Se eliminaron {removed_children} registros asociados.' if removed_children else ''
            messages.success(request, f'Equipo eliminado correctamente.{detail}')
            return redirect('model_list', model_key=model_key)
        if model_key == 'estrategia':
            blockers = _strategy_delete_blockers(obj)
            if blockers:
                messages.error(
                    request,
                    'No se puede eliminar la estrategia porque tiene datos históricos asociados: '
                    + '; '.join(blockers)
                    + '. Elimina primero esos registros o usa una estrategia sin cargas/evaluaciones.'
                )
                return redirect('model_list', model_key=model_key)
            counts = _delete_strategy_related_data(obj)
            obj.delete()
            removed_children = sum(counts.values())
            detail = f' Se eliminaron o desvincularon {removed_children} registros asociados.' if removed_children else ''
            messages.success(request, f'Estrategia eliminada correctamente.{detail}')
            return redirect('model_list', model_key=model_key)
        
        obj.delete()
        messages.success(request, 'Registro eliminado correctamente.')
        return redirect('model_list', model_key=model_key)
    equipment_delete_summary = None
    if model_key == 'equipo':
        equipment_delete_summary = {
            'servicios': models.ServicioEquipo.objects.filter(equipo=obj).count(),
            'familias': models.FamiliaEquipoItem.objects.filter(equipo=obj).count(),
            'componentes': models.ComponenteEquipo.objects.filter(equipo=obj).count(),
            'aca': models.Criticidad.objects.filter(equipo=obj).count(),
            'rcm': models.RCM.objects.filter(equipo=obj).count(),
            'pautas': models.Pauta.objects.filter(equipo=obj).count(),
        }

    return render(request, 'core/model_delete.html', {
        'config': config,
        'model_key': model_key,
        'object': obj,
        'can_delete': _can_delete(request, config, obj),
        'equipment_delete_summary': equipment_delete_summary,
    })

# ---------------------------------------------------------------------------
# Helpers generales
# ---------------------------------------------------------------------------
def _json_payload(request, key, default=None):
    raw = request.POST.get(key, '')
    if not raw:
        return default
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return default


def _decimal_or_none(value):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value).strip().replace(',', '.'))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _quantize_matrix_decimal(value):
    parsed = _decimal_or_none(value)
    if parsed is None:
        return None
    return parsed.quantize(_MATRIX_DECIMAL_STEP)


def _format_matrix_decimal(value):
    parsed = _quantize_matrix_decimal(value)
    if parsed is None:
        return ''
    text = format(parsed, 'f').rstrip('0').rstrip('.')
    return text or '0'


def _parse_matrix_range(range_text):
    match = _RANGE_RE.match(str(range_text or '').strip())
    if not match:
        return None
    start = _quantize_matrix_decimal(match.group(1))
    end = _quantize_matrix_decimal(match.group(2))
    if start is None or end is None or end < start:
        return None
    return start, end


def _int_or_default(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _json_loads_safe(value, default=None):
    if default is None:
        default = {}
    if not value:
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return default


def _json_safe(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, dict):
        return {k: _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(v) for v in value]
    return value


def _login_coordinate(value, *, minimum, maximum):
    try:
        coordinate = Decimal(str(value).strip().replace(',', '.'))
    except (InvalidOperation, AttributeError):
        return Decimal('0.000000')
    if coordinate < minimum or coordinate > maximum:
        return Decimal('0.000000')
    return coordinate.quantize(Decimal('0.000001'))


def _login_location_is_valid(request):
    latitud = request.POST.get('latitud')
    longitud = request.POST.get('longitud')
    if not str(latitud or '').strip() or not str(longitud or '').strip():
        return False
    try:
        latitud = Decimal(str(latitud).strip().replace(',', '.'))
        longitud = Decimal(str(longitud).strip().replace(',', '.'))
    except (InvalidOperation, AttributeError):
        return False
    return Decimal('-90') <= latitud <= Decimal('90') and Decimal('-180') <= longitud <= Decimal('180')


def _record_login_location(request, profile):
    if not profile:
        return
    latitud = _login_coordinate(
        request.POST.get('latitud'),
        minimum=Decimal('-90'),
        maximum=Decimal('90'),
    )
    longitud = _login_coordinate(
        request.POST.get('longitud'),
        minimum=Decimal('-180'),
        maximum=Decimal('180'),
    )
    login_record = models.InicioSesion.objects.create(
        usuario=profile,
        hora=timezone.now(),
        latitud=latitud,
        longitud=longitud,
    )
    hora_chile = timezone.localtime(login_record.hora).replace(tzinfo=None)
    with connection.cursor() as cursor:
        cursor.execute(
            'UPDATE reliability_iniciosesion SET hora = %s WHERE id = %s',
            [hora_chile, login_record.pk],
        )


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = EmailLoginForm(request.POST or None)
    if request.method == 'POST':
        form_is_valid = form.is_valid()
        location_is_valid = _login_location_is_valid(request)
        if form_is_valid and not location_is_valid:
            form.add_error(None, 'Debes autorizar y registrar tu ubicación para iniciar sesión.')
        if form_is_valid and location_is_valid:
            login(request, form.cleaned_data['auth_user'])
            _record_login_location(request, form.cleaned_data.get('perfil_usuario'))
            messages.success(request, 'Sesión iniciada correctamente.')
            return redirect(request.GET.get('next') or 'dashboard')
    return render(request, 'core/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


def _service_or_404(request, pk, edit=False):
    servicio = get_object_or_404(
        models.Servicio.objects.select_related('empresa', 'estrategia', 'responsable_usuario', 'creado_por_usuario'),
        pk=pk,
    )
    permission = get_service_permission(request.user, servicio)
    if edit and not permission['can_edit']:
        raise PermissionDenied('No tienes permiso para editar este servicio.')
    if not edit and not permission['can_view']:
        raise PermissionDenied('No tienes permiso para ver este servicio.')
    return servicio, permission


def _strategy_dimensions(estrategia, proceso=None):
    if not estrategia:
        return []
    qs = (
        models.EstrategiaDimension.objects.filter(estrategia=estrategia, activo=True)
        .select_related('dimension')
        .prefetch_related(
            'escalas_valor__escala_unificada',
            'catalogo__columnas',
            'catalogo__filas__celdas__columna',
        )
        .order_by('orden', 'id')
    )
    if proceso:
        process_values = [proceso, models.EstrategiaDimension.PROCESO_AMBOS]
        if proceso == getattr(models.EstrategiaDimension, 'PROCESO_FMECA', 'fmeca'):
            process_values.extend(getattr(models.EstrategiaDimension, 'PROCESO_FMECA_ALIASES', ()))
        qs = qs.filter(proceso_uso__in=list(dict.fromkeys(process_values)))
    return list(qs)


def _dimension_display_value(item):
    if item is None:
        return ''

    if item.valor_booleano is not None:
        return 'Sí­' if item.valor_booleano else 'No'

    if item.valor_numerico is not None:
        if item.valor_numerico == int(item.valor_numerico):
            return str(int(item.valor_numerico))
        return str(item.valor_numerico)

    if item.escala_valor_id:
        if item.escala_valor.valor_numerico is not None:
            if item.escala_valor.valor_numerico == int(item.escala_valor.valor_numerico):
                return str(int(item.escala_valor.valor_numerico))
            return str(item.escala_valor.valor_numerico)
        return item.escala_valor.codigo or item.escala_valor.descripcion or ''

    if item.valor_texto:
        return item.valor_texto

    if item.catalogo_fila_id:
        return item.catalogo_fila.etiqueta or ''

    return ''


def _catalog_preview(estrategia_dimension):
    catalogo = getattr(estrategia_dimension, 'catalogo', None)
    if not catalogo:
        return None
    columnas = list(catalogo.columnas.all().order_by('orden'))
    filas = []
    for fila in catalogo.filas.all().order_by('orden'):
        value_map = fila.values_map()
        cells = []
        for col in columnas:
            cells.append(value_map.get(col.clave_interna, fila.etiqueta if col.clave_interna == 'etiqueta' else ''))
        filas.append({'pk': fila.pk, 'cells': cells, 'label': fila.etiqueta or f'Fila {fila.orden}'})
    return {
        'nombre': catalogo.nombre,
        'columnas': columnas,
        'filas': filas,
    }


def _quantize_decimal(value):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _matrix_axis_level_info(levels, estrategia_dimension):
    levels = list(levels)
    if not levels:
        return {'mode': 'sin_niveles', 'levels': []}

    scale_levels = []
    if estrategia_dimension is not None:
        try:
            scale_levels = list(estrategia_dimension.escalas_valor.select_related('escala_unificada').order_by('nivel_ordinal', 'id'))
        except Exception:
            scale_levels = []

    use_unified = len(scale_levels) >= len(levels) and all(
        getattr(scale_levels[idx], 'escala_unificada_id', None) and 1 <= int(scale_levels[idx].escala_unificada.nivel) <= 5
        for idx in range(len(levels))
    )

    info = []
    if use_unified:
        mode = 'escala_unificada'
        for idx, level in enumerate(levels, start=1):
            unified = int(scale_levels[idx - 1].escala_unificada.nivel)
            position = 0 if unified <= 1 else (unified - 1) / 4
            info.append({
                'idx': idx,
                'obj': level,
                'nombre': level.nombre,
                'valor': _quantize_decimal(level.valor),
                'descripcion': level.descripcion,
                'position': position,
                'canonical_slot': unified,
                'source': 'Escala unificada',
                'unified_level': unified,
            })
        return {'mode': mode, 'levels': info}

    numeric_values = [_quantize_decimal(level.valor) for level in levels]
    unique_values = {value for value in numeric_values if value is not None}
    if len(levels) > 1 and len(unique_values) > 1:
        mode = 'valor_relativo'
        min_value = min(unique_values)
        max_value = max(unique_values)
        span = max_value - min_value
        for idx, level in enumerate(levels, start=1):
            current = _quantize_decimal(level.valor)
            position = float((current - min_value) / span) if current is not None and span not in (None, 0) else (idx - 1) / (len(levels) - 1)
            canonical = max(1, min(5, int(round(1 + (position * 4)))))
            info.append({
                'idx': idx,
                'obj': level,
                'nombre': level.nombre,
                'valor': current,
                'descripcion': level.descripcion,
                'position': position,
                'canonical_slot': canonical,
                'source': 'Valor relativo',
                'unified_level': None,
            })
        return {'mode': mode, 'levels': info}

    mode = 'orden_relativo'
    denominator = max(len(levels) - 1, 1)
    for idx, level in enumerate(levels, start=1):
        position = (idx - 1) / denominator
        canonical = max(1, min(5, int(round(1 + (position * 4)))))
        info.append({
            'idx': idx,
            'obj': level,
            'nombre': level.nombre,
            'valor': _quantize_decimal(level.valor),
            'descripcion': level.descripcion,
            'position': position,
            'canonical_slot': canonical,
            'source': 'Orden relativo',
            'unified_level': None,
        })
    return {'mode': mode, 'levels': info}


def _closest_original_level(level_info, canonical_slot):
    target_position = 0 if canonical_slot <= 1 else (canonical_slot - 1) / 4
    levels = level_info.get('levels', [])
    if not levels:
        return None
    return min(
        levels,
        key=lambda item: (abs(item['position'] - target_position), abs(item['idx'] - canonical_slot), item['idx'])
    )


def _homologation_method_label(level_info):
    mode = level_info.get('mode')
    return {
        'escala_unificada': 'Exacta por escala unificada',
        'valor_relativo': 'Proyección por valor relativo',
        'orden_relativo': 'Proyección por orden relativo',
        'sin_niveles': 'Sin homologación',
    }.get(mode, 'Proyección')


def _build_homologated_matrix_preview(matriz):
    prob_levels = list(matriz.niveles_probabilidad.order_by('orden_visual', 'id'))
    impact_levels = list(matriz.niveles_impacto.order_by('orden_visual', 'id'))
    cell_map = {
        (cell.probabilidad.orden_visual, cell.impacto_nivel.orden_visual): cell
        for cell in matriz.celdas.select_related('probabilidad', 'impacto_nivel').all()
    }

    prob_info = _matrix_axis_level_info(prob_levels, matriz.dimension_probabilidad)
    impact_info = _matrix_axis_level_info(impact_levels, matriz.dimension_impacto)

    prob_defs = []
    impact_defs = []
    cell_payload = {}

    for slot in range(1, 6):
        source_prob = _closest_original_level(prob_info, slot)
        source_impact = _closest_original_level(impact_info, slot)
        prob_defs.append({
            'idx': slot,
            'nombre': f'P{slot}',
            'valor': slot,
            'descripcion': f"Ref. {source_prob['nombre']}" if source_prob else 'Sin referencia',
        })
        impact_defs.append({
            'idx': slot,
            'nombre': f'C{slot}',
            'valor': slot,
            'descripcion': f"Ref. {source_impact['nombre']}" if source_impact else 'Sin referencia',
        })

    for prob_slot in range(1, 6):
        src_prob = _closest_original_level(prob_info, prob_slot)
        if not src_prob:
            continue
        for impact_slot in range(1, 6):
            src_impact = _closest_original_level(impact_info, impact_slot)
            if not src_impact:
                continue
            original_cell = cell_map.get((src_prob['idx'], src_impact['idx']))
            if original_cell:
                result_num = original_cell.resultado_num
                clasificacion = original_cell.clasificacion
                color = original_cell.color
            else:
                prob_value = _quantize_decimal(src_prob.get('valor')) or Decimal(prob_slot)
                impact_value = _quantize_decimal(src_impact.get('valor')) or Decimal(impact_slot)
                result_num = prob_value * impact_value
                matched = _match_legend(result_num, _legend_from_matrix(matriz))
                clasificacion, color = matched
            cell_payload[(prob_slot, impact_slot)] = {
                'prob_idx': prob_slot,
                'impact_idx': impact_slot,
                'resultado_num': result_num,
                'clasificacion': clasificacion,
                'color': color,
                'calcular': True,
            }

    preview = _matrix_preview_from_defs(
        'probabilidad',
        prob_defs,
        impact_defs,
        cell_payload,
        matriz.estrategia,
        matriz.dimension_probabilidad,
        matriz.dimension_impacto,
        _legend_from_matrix(matriz),
    )

    return {
        'preview': preview,
        'prob_mapping': prob_info['levels'],
        'impact_mapping': impact_info['levels'],
        'prob_method': _homologation_method_label(prob_info),
        'impact_method': _homologation_method_label(impact_info),
        'source_shape': f"{len(prob_levels)}x{len(impact_levels)}",
        'target_shape': '5x5',
        'is_exact': prob_info.get('mode') == 'escala_unificada' and impact_info.get('mode') == 'escala_unificada',
    }


def _build_matrix_original_preview(matriz):
    prob_levels = list(matriz.niveles_probabilidad.order_by('orden_visual', 'id'))
    impact_levels = list(matriz.niveles_impacto.order_by('orden_visual', 'id'))
    prob_defs = _matrix_level_dicts(prob_levels, len(prob_levels) or 5, 'p')
    impact_defs = _matrix_level_dicts(impact_levels, len(impact_levels) or 5, 'c')
    existing_cells = {
        (cell.probabilidad.orden_visual, cell.impacto_nivel.orden_visual): cell
        for cell in matriz.celdas.select_related('probabilidad', 'impacto_nivel').all()
    }
    return _matrix_preview_from_defs(
        'probabilidad',
        prob_defs,
        impact_defs,
        existing_cells,
        matriz.estrategia,
        matriz.dimension_probabilidad,
        matriz.dimension_impacto,
        _legend_from_matrix(matriz),
    )


def _matrix_stats(preview):
    rows = preview.get('rows', [])
    cols = preview.get('x_defs', [])
    values = []
    for row in rows:
        for cell in row.get('cells', []):
            value = _quantize_decimal(cell.get('resultado_num'))
            if value is not None:
                values.append(value)
    return {
        'rows': len(rows),
        'cols': len(cols),
        'min_value': min(values) if values else None,
        'max_value': max(values) if values else None,
    }


# ---------------------------------------------------------------------------
# Helpers compartidos de dimensiones y matrices
# ---------------------------------------------------------------------------
def _catalog_row_primary_numeric(row):
    if row is None:
        return None
    values = row.values_map()
    for key in ['valor_numerico', 'valor_principal', 'valor', 'nivel', 'puntaje']:
        if key in values and values.get(key) not in (None, ''):
            return _decimal_or_none(values.get(key))
    for key, value in values.items():
        if key in {'limite_inferior', 'limite_superior', 'desde', 'hasta', 'min', 'max', 'minimo', 'mí­nimo', 'maximo', 'máximo'}:
            continue
        decimal_value = _decimal_or_none(value)
        if decimal_value is not None:
            return decimal_value
    return None


def _catalog_row_text(row):
    if row is None:
        return ''
    values = row.values_map()
    for key in ['etiqueta', 'nombre', 'descripcion', 'texto', 'codigo']:
        value = values.get(key)
        if value not in (None, ''):
            return str(value)
    return str(row.etiqueta or '')


def _catalog_row_boolean(row):
    if row is None:
        return None
    values = row.values_map()
    for key in ['valor_booleano', 'booleano', 'flag']:
        value = values.get(key)
        if value in (True, False):
            return value
        if value not in (None, ''):
            return str(value).strip().lower() in {'true', '1', 'si', 'sí­', 'yes'}
    return None


def _calc_slug(value):
    text = str(value or '').strip()
    if 'Ã' in text or 'Â' in text:
        try:
            text = text.encode('latin1').decode('utf-8')
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    text = unicodedata.normalize('NFKD', text.lower())
    text = ''.join(char for char in text if not unicodedata.combining(char))
    return re.sub(r'[^a-z0-9]+', '_', text).strip('_')


def _matrix_legend_config(matriz):
    payload = _json_loads_safe(getattr(matriz, 'leyenda_json', ''), [])
    if isinstance(payload, dict):
        return payload
    if isinstance(payload, list):
        return {'items': payload}
    return {}


def _matrix_resolution_mode(matriz):
    mode = (_matrix_legend_config(matriz).get('modo_resolucion') or '').strip()
    valid_modes = {choice[0] for choice in models.MatrizRiesgo.RESOLUCION_CHOICES}
    return mode if mode in valid_modes else models.MatrizRiesgo.RESOLUCION_EXACTA


def _matrix_generation_config(matriz):
    config = _matrix_legend_config(matriz).get('generacion') if matriz else {}
    return config if isinstance(config, dict) else {}


def _matrix_generation_mode(matriz):
    return str(_matrix_generation_config(matriz).get('modo') or models.MatrizRiesgo.MODO_MANUAL).strip()


def _matrix_result_from_axis_values(prob_val, impact_val):
    prob_val = _decimal_or_none(prob_val)
    impact_val = _decimal_or_none(impact_val)
    if prob_val is not None and impact_val is not None:
        return (prob_val * impact_val).quantize(_MATRIX_DECIMAL_STEP)
    if impact_val is not None:
        return impact_val.quantize(_MATRIX_DECIMAL_STEP)
    if prob_val is not None:
        return prob_val.quantize(_MATRIX_DECIMAL_STEP)
    return None


def _matrix_threshold_axis_values(matriz, prob_val, impact_val):
    prob_generated = _is_generated_matrix_axis_dimension(
        getattr(matriz, 'dimension_probabilidad', None),
        'probabilidad',
    )
    impact_generated = _is_generated_matrix_axis_dimension(
        getattr(matriz, 'dimension_impacto', None),
        'impacto',
    )
    if prob_generated and not impact_generated:
        prob_val = None
    if impact_generated and not prob_generated:
        impact_val = None
    return prob_val, impact_val


def _matrix_cell_for_result_floor(matriz, result_value):
    result_value = _quantize_matrix_decimal(result_value)
    if not matriz or result_value is None:
        return None
    return models.MatrizRiesgoCelda.objects.filter(
        matriz=matriz,
        resultado_num__lte=result_value,
    ).select_related(
        'probabilidad',
        'impacto_nivel',
    ).order_by('-resultado_num', '-id').first()


def _matrix_range_level_for_value(levels, ranges, value):
    value = _decimal_or_none(value)
    if value is None:
        return None
    ordered_levels = list(levels)
    if not ordered_levels:
        return None
    normalized_ranges = ranges if isinstance(ranges, list) else []
    for idx, level in enumerate(ordered_levels):
        if idx >= len(normalized_ranges):
            break
        range_item = normalized_ranges[idx] if isinstance(normalized_ranges[idx], dict) else {}
        start = _decimal_or_none(range_item.get('desde'))
        end = _decimal_or_none(range_item.get('hasta'))
        if end is None:
            continue
        if start is None:
            start = Decimal('0')
        if range_item.get('desde_exclusivo') is True:
            lower_match = value > start
        else:
            lower_match = value >= start
        if lower_match and value <= end:
            return level
    last_range = normalized_ranges[-1] if normalized_ranges and isinstance(normalized_ranges[-1], dict) else {}
    first_range = normalized_ranges[0] if normalized_ranges and isinstance(normalized_ranges[0], dict) else {}
    last_end = _decimal_or_none(last_range.get('hasta')) if last_range else None
    if last_end is not None and value > last_end:
        return ordered_levels[-1]
    first_start = _decimal_or_none(first_range.get('desde')) if first_range else None
    if first_start is not None and value < first_start:
        return ordered_levels[0]
    return None


def _matrix_cell_for_auto_axis_ranges(matriz, prob_val, impact_val):
    config = _matrix_generation_config(matriz)
    if config.get('modo') != models.MatrizRiesgo.MODO_AUTOMATICA_MAXIMO_TEORICO:
        return None
    prob_levels = matriz.niveles_probabilidad.order_by('orden_visual', 'id')
    impact_levels = matriz.niveles_impacto.order_by('orden_visual', 'id')
    prob_level = _matrix_range_level_for_value(prob_levels, config.get('rangos_eje_x'), prob_val)
    impact_level = _matrix_range_level_for_value(impact_levels, config.get('rangos_eje_y'), impact_val)
    if not prob_level or not impact_level:
        return None
    return models.MatrizRiesgoCelda.objects.filter(
        matriz=matriz,
        probabilidad=prob_level,
        impacto_nivel=impact_level,
    ).select_related(
        'probabilidad',
        'impacto_nivel',
    ).order_by('id').first()


def _matrix_cell_for_axis_values(matriz, prob_val, impact_val):
    if not matriz:
        return None
    if _matrix_generation_mode(matriz) == models.MatrizRiesgo.MODO_AUTOMATICA_MAXIMO_TEORICO:
        ranged_cell = _matrix_cell_for_auto_axis_ranges(matriz, prob_val, impact_val)
        if ranged_cell:
            return ranged_cell

    if prob_val is not None and impact_val is not None:
        exact_cell = models.MatrizRiesgoCelda.objects.filter(
            matriz=matriz,
            probabilidad__valor=prob_val,
            impacto_nivel__valor=impact_val,
        ).select_related(
            'probabilidad',
            'impacto_nivel',
        ).order_by('id').first()
        if exact_cell:
            return exact_cell

    if _matrix_resolution_mode(matriz) == models.MatrizRiesgo.RESOLUCION_UMBRAL_RESULTADO:
        prob_val, impact_val = _matrix_threshold_axis_values(matriz, prob_val, impact_val)
        return _matrix_cell_for_result_floor(
            matriz,
            _matrix_result_from_axis_values(prob_val, impact_val),
        )

    return None

# ---------------------------------------------------------------------------
# Constructor visual de matrices
# ---------------------------------------------------------------------------
def _matrix_level_dicts(levels, count, prefix):
    data = []
    levels = list(levels)
    for idx in range(1, count + 1):
        obj = levels[idx - 1] if idx <= len(levels) else None
        value = getattr(obj, 'valor', None) if obj else None
        name = getattr(obj, 'nombre', None) if obj else None
        if prefix.lower().startswith('c') and name == f'I{idx}':
            name = f'C{idx}'
        data.append({
            'idx': idx,
            'nombre': name or f'{prefix.upper()}{idx}',
            'valor': _json_safe(value) if value is not None else idx,
            'descripcion': getattr(obj, 'descripcion', None) or '',
        })
    return data


def _level_defs_from_strategy_dimension(estrategia_dimension, count, prefix):
    if not estrategia_dimension:
        return _matrix_level_dicts([], count, prefix)

    escalas = list(estrategia_dimension.escalas_valor.order_by('nivel_ordinal', 'id'))
    if not escalas:
        try:
            catalogo = estrategia_dimension.catalogo
            filas_catalogo = list(
                catalogo.filas.prefetch_related('celdas__columna').order_by('orden', 'id')
            )
        except models.DimensionCatalogo.DoesNotExist:
            filas_catalogo = []
        if filas_catalogo:
            data = []
            for idx in range(1, count + 1):
                fila = filas_catalogo[idx - 1] if idx <= len(filas_catalogo) else None
                numeric_value = _catalog_row_primary_numeric(fila) if fila else None
                data.append({
                    'idx': idx,
                    'nombre': _catalog_row_text(fila) if fila else f'{prefix.upper()}{idx}',
                    'valor': _json_safe(numeric_value) if numeric_value is not None else idx,
                    'descripcion': fila.etiqueta if fila else '',
                })
            return data

    data = []
    for idx in range(1, count + 1):
        escala = escalas[idx - 1] if idx <= len(escalas) else None
        data.append({
            'idx': idx,
            'nombre': (escala.descripcion or escala.codigo) if escala else f'{prefix.upper()}{idx}',
            'valor': _json_safe(escala.valor_numerico) if escala else idx,
            'descripcion': escala.codigo if escala and escala.codigo else '',
        })
    return data


def _normalize_matrix_level_defs(defs, count, prefix):
    defs = defs if isinstance(defs, list) else []
    normalized = []
    for idx in range(1, count + 1):
        raw = defs[idx - 1] if idx <= len(defs) and isinstance(defs[idx - 1], dict) else {}
        value = _quantize_matrix_decimal(raw.get('valor'))
        normalized.append({
            'idx': idx,
            'nombre': str(raw.get('nombre') or f'{prefix.upper()}{idx}'),
            'valor': value if value is not None else Decimal(idx),
            'descripcion': str(raw.get('descripcion') or ''),
        })
    return normalized


def _definitions_from_request(request, prob_count, impact_count, fallback_prob, fallback_impact):
    prob_defs = _json_payload(request, 'prob_levels_json', fallback_prob) or fallback_prob
    impact_defs = _json_payload(request, 'impact_levels_json', fallback_impact) or fallback_impact
    prob_defs = _normalize_matrix_level_defs(prob_defs if isinstance(prob_defs, list) else fallback_prob, prob_count, 'p')
    impact_defs = _normalize_matrix_level_defs(impact_defs if isinstance(impact_defs, list) else fallback_impact, impact_count, 'c')

    return prob_defs, impact_defs


def _cell_data_from_request(request):
    payload = _json_payload(request, 'matrix_cells_json', []) or []
    result = {}
    rows = payload if isinstance(payload, list) else []
    for row in rows:
        if not isinstance(row, list):
            continue
        for cell in row:
            if not isinstance(cell, dict):
                continue
            prob_idx = _int_or_default(cell.get('prob_idx'), None)
            impact_idx = _int_or_default(cell.get('impact_idx'), None)
            if not prob_idx or not impact_idx:
                continue
            result[(prob_idx, impact_idx)] = {
                'prob_idx': prob_idx,
                'impact_idx': impact_idx,
                'clasificacion': str(cell.get('clasificacion') or ''),
                'color': str(cell.get('color') or '#2a2a3a'),
                'resultado_num': _quantize_matrix_decimal(cell.get('resultado_num')),
                'calcular': bool(cell.get('calcular', True)),
            }
    return result


def _matrix_result_values(prob_defs, impact_defs):
    values = []
    for prob in prob_defs:
        for impact in impact_defs:
            prob_value = _decimal_or_none(prob.get('valor'))
            impact_value = _decimal_or_none(impact.get('valor'))
            if prob_value is None or impact_value is None:
                continue
            values.append((prob_value * impact_value).quantize(_MATRIX_DECIMAL_STEP))
    return values


def _matrix_value_bounds(prob_defs, impact_defs):
    values = _matrix_result_values(prob_defs, impact_defs)
    return (min(values), max(values)) if values else (1, 1)


def _safe_legend_items(raw_items):
    if isinstance(raw_items, dict):
        raw_items = raw_items.get('items') or raw_items.get('legend') or raw_items.get('leyenda') or []
    if not isinstance(raw_items, list):
        return []
    cleaned = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        name = str(item.get('name') or '').strip()
        range_text = str(item.get('range') or '').strip()
        color = str(item.get('color') or '#6c63ff').strip() or '#6c63ff'
        if not name and not range_text:
            continue
        cleaned.append({'name': name, 'range': range_text, 'color': color})
    return cleaned


def _matrix_legend_payload(legend_items, modo_resolucion=None, criticality_rules=None):
    return {
        'items': legend_items or [],
        'modo_resolucion': modo_resolucion or models.MatrizRiesgo.RESOLUCION_EXACTA,
        'reglas_criticidad': criticality_rules or [],
    }


def _default_legend_for_bounds(min_value, max_value):
    min_value = _quantize_matrix_decimal(min_value) or Decimal('0.00')
    max_value = _quantize_matrix_decimal(max_value) or min_value
    if max_value < min_value:
        max_value = min_value
    if max_value == min_value or max_value - min_value < Decimal('0.04'):
        return [{'name': 'Bajo', 'range': f'{_format_matrix_decimal(min_value)}-{_format_matrix_decimal(max_value)}', 'color': '#2ecc71'}]
    span = max_value - min_value
    step = (span / Decimal('4')).quantize(_MATRIX_DECIMAL_STEP) if span else Decimal('1.00')
    if step <= 0:
        step = Decimal('1.00')
    labels = [('Bajo', '#2ecc71'), ('Medio', '#f1c40f'), ('Alto', '#e67e22'), ('Crí­tico', '#e74c3c')]
    items = []
    current = min_value
    for idx, (label, color) in enumerate(labels):
        end = min_value + (step * (idx + 1)) if idx < 3 else max_value
        end = _quantize_matrix_decimal(end) or max_value
        end = min(max_value, max(current, end))
        items.append({
            'name': label,
            'range': f'{_format_matrix_decimal(current)}-{_format_matrix_decimal(end)}',
            'color': color,
        })
        current = end + _MATRIX_DECIMAL_STEP
    items[-1]['range'] = f"{items[-1]['range'].split('-')[0]}-{_format_matrix_decimal(max_value)}"
    return items


def _validate_legend_items(raw_items, min_value, max_value, result_values=None):
    items = _safe_legend_items(raw_items)
    if not items:
        return None, 'Debes definir al menos un rango en la leyenda.'

    parsed = []
    for item in items:
        if not item['name']:
            return None, 'Cada rango de la leyenda debe tener un nombre.'
        range_values = _parse_matrix_range(item['range'])
        if not range_values:
            return None, f"El rango '{item['range']}' no tiene un formato valido. Usa por ejemplo 0,1-4,5."
        start, end = range_values
        if end < start:
            return None, f"El rango '{item['range']}' no es válido porque el final es menor que el inicio."
        parsed.append({'name': item['name'], 'range': f'{_format_matrix_decimal(start)}-{_format_matrix_decimal(end)}', 'color': item['color'], 'start': start, 'end': end})

    parsed.sort(key=lambda item: item['start'])
    min_value = _quantize_matrix_decimal(min_value) or Decimal('0.00')
    max_value = _quantize_matrix_decimal(max_value) or min_value
    values_to_cover = result_values if result_values is not None else [min_value, max_value]
    uncovered = []
    for value in values_to_cover:
        value = _quantize_matrix_decimal(value)
        if value is None:
            continue
        if not any(item['start'] <= value <= item['end'] for item in parsed):
            uncovered.append(value)
    if uncovered:
        samples = ', '.join(_format_matrix_decimal(value) for value in sorted(set(uncovered))[:5])
        return None, f'La leyenda no cubre los valores calculados: {samples}. Ajusta los rangos para incluirlos.'
    return [{'name': item['name'], 'range': item['range'], 'color': item['color']} for item in parsed], None


def _legend_from_matrix(matriz, min_value=None, max_value=None):
    try:
        legend = json.loads(matriz.leyenda_json or '[]')
    except json.JSONDecodeError:
        legend = []
    legend = _safe_legend_items(legend)
    if legend:
        return legend
    if min_value is None or max_value is None:
        prob_defs = _matrix_level_dicts(matriz.niveles_probabilidad.order_by('orden_visual', 'id'), matriz.niveles_probabilidad.count() or 5, 'prob')
        impact_defs = _matrix_level_dicts(matriz.niveles_impacto.order_by('orden_visual', 'id'), matriz.niveles_impacto.count() or 5, 'impact')
        min_value, max_value = _matrix_value_bounds(prob_defs, impact_defs)
    return _default_legend_for_bounds(min_value, max_value)


def _match_legend(result_value, legend_items):
    result_value = _quantize_matrix_decimal(result_value)
    if result_value is None:
        return '', '#2a2a3a'
    for item in legend_items:
        range_values = _parse_matrix_range(item['range'])
        if not range_values:
            continue
        start, end = range_values
        if start <= result_value <= end:
            return item['name'], item['color']
    return '', '#2a2a3a'


def _matrix_preview_from_defs(selected_axis, prob_defs, impact_defs, existing_cells=None, strategy=None, prob_dim=None, impact_dim=None, legend_items=None):
    existing_cells = existing_cells or {}
    legend_items = legend_items or _default_legend_for_bounds(*_matrix_value_bounds(prob_defs, impact_defs))
    x_defs = impact_defs if selected_axis == 'impacto' else prob_defs
    y_defs = prob_defs if selected_axis == 'impacto' else impact_defs
    rows = []
    for y_idx, y_def in enumerate(y_defs, start=1):
        row = {'header': y_def, 'cells': []}
        for x_idx, x_def in enumerate(x_defs, start=1):
            prob_idx = y_idx if selected_axis == 'impacto' else x_idx
            impact_idx = x_idx if selected_axis == 'impacto' else y_idx
            existing = existing_cells.get((prob_idx, impact_idx))
            if existing and hasattr(existing, 'resultado_num'):
                result_num = existing.resultado_num
                clasificacion = existing.clasificacion
                color = existing.color
                calcular = existing.calcular
            elif isinstance(existing, dict):
                result_num = existing.get('resultado_num')
                clasificacion = existing.get('clasificacion') or ''
                color = existing.get('color') or '#2a2a3a'
                calcular = existing.get('calcular', True)
            else:
                prob_value = _decimal_or_none(prob_defs[prob_idx - 1].get('valor'))
                impact_value = _decimal_or_none(impact_defs[impact_idx - 1].get('valor'))
                result_num = ((prob_value or Decimal('0')) * (impact_value or Decimal('0'))).quantize(_MATRIX_DECIMAL_STEP)
                clasificacion, color = _match_legend(result_num, legend_items)
                calcular = True
            row['cells'].append({
                'prob_idx': prob_idx,
                'impact_idx': impact_idx,
                'resultado_num': _json_safe(result_num),
                'clasificacion': clasificacion,
                'color': color,
                'calcular': calcular,
            })
        rows.append(row)
    return {
        'x_axis': selected_axis,
        'x_defs': x_defs,
        'rows': rows,
        'prob_defs': prob_defs,
        'impact_defs': impact_defs,
        'strategy': strategy,
        'prob_dim': prob_dim,
        'impact_dim': impact_dim,
    }


def _matrix_ui_payload(matrix_preview, stored_legend=None):
    return {
        'rows_count': len(matrix_preview['rows']),
        'cols_count': len(matrix_preview['x_defs']),
        'prob_defs': _json_safe(matrix_preview['prob_defs']),
        'impact_defs': _json_safe(matrix_preview['impact_defs']),
        'legend': stored_legend or [],
    }


@transaction.atomic
def _sync_matrix_levels(matriz, model_cls, definitions, delete_stale=True):
    existing = list(model_cls.objects.filter(matriz=matriz).order_by('orden_visual', 'id'))
    keep_ids = []
    result = []
    for idx, definition in enumerate(definitions, start=1):
        obj = existing[idx - 1] if idx <= len(existing) else model_cls(matriz=matriz)
        obj.nombre = str(definition.get('nombre') or f'N{idx}')
        value = _quantize_matrix_decimal(definition.get('valor'))
        obj.valor = value if value is not None else Decimal(idx)
        obj.descripcion = str(definition.get('descripcion') or '')
        obj.orden_visual = idx
        obj.save()
        keep_ids.append(obj.pk)
        result.append(obj)
    if delete_stale:
        model_cls.objects.filter(matriz=matriz).exclude(pk__in=keep_ids).delete()
    return result


def _next_strategy_order(estrategia):
    current = models.EstrategiaDimension.objects.filter(estrategia=estrategia, activo=True).order_by('-orden').first()
    return (current.orden if current else 0) + 1


def _matrix_axis_prefix(axis):
    return 'Eje X' if axis == 'probabilidad' else 'Eje Y'


def _is_generated_matrix_axis_dimension(estrategia_dimension, axis=None):
    if not estrategia_dimension:
        return False
    name = (estrategia_dimension.dimension.nombre or '').strip().lower()
    prefixes = []
    if axis in (None, 'probabilidad'):
        prefixes.append('eje x - ')
        prefixes.append('probabilidad - ')
    if axis in (None, 'impacto'):
        prefixes.append('eje y - ')
        prefixes.append('impacto - ')
        prefixes.append('consecuencia - ')
    return any(name.startswith(prefix) for prefix in prefixes)


def _update_generated_matrix_axis_dimension(estrategia_dimension, axis, matrix_name):
    prefix = _matrix_axis_prefix(axis)
    dimension = estrategia_dimension.dimension
    dimension.nombre = f'{prefix} - {matrix_name}'
    dimension.descripcion = f'Dimension creada automaticamente para la matriz {matrix_name}'
    dimension.tipo_funcional = axis
    dimension.tipo_dato = 'numerico'
    dimension.save()
    estrategia_dimension.obligatorio = True
    estrategia_dimension.proceso_uso = models.EstrategiaDimension.PROCESO_ACA
    estrategia_dimension.activo = True
    estrategia_dimension.save(update_fields=['obligatorio', 'proceso_uso', 'activo'])
    return estrategia_dimension


def _create_generated_matrix_axis_dimension(estrategia, axis, matrix_name, order):
    prefix = _matrix_axis_prefix(axis)
    dimension = models.Dimension.objects.create(
        nombre=f'{prefix} - {matrix_name}',
        descripcion=f'Dimension creada automaticamente para la matriz {matrix_name}',
        tipo_funcional=axis,
        tipo_dato='numerico',
    )
    return models.EstrategiaDimension.objects.create(
        estrategia=estrategia,
        dimension=dimension,
        orden=order,
        obligatorio=True,
        proceso_uso=models.EstrategiaDimension.PROCESO_ACA,
        activo=True,
    )


def _resolve_matrix_axis_dimension(estrategia, axis, matrix_name, next_order, selected=None, existing=None):
    if selected and selected.estrategia_id == estrategia.pk:
        return selected, next_order

    if existing and existing.estrategia_id == estrategia.pk and _is_generated_matrix_axis_dimension(existing, axis):
        return _update_generated_matrix_axis_dimension(existing, axis, matrix_name), next_order

    created = _create_generated_matrix_axis_dimension(estrategia, axis, matrix_name, next_order)
    return created, next_order + 1


def _ensure_matrix_strategy_dimensions(
    estrategia,
    matrix_name,
    prob_defs,
    impact_defs,
    existing_prob=None,
    existing_impact=None,
    selected_prob=None,
    selected_impact=None,
):
    next_order = _next_strategy_order(estrategia)
    prob_dim, next_order = _resolve_matrix_axis_dimension(
        estrategia,
        'probabilidad',
        matrix_name,
        next_order,
        selected=selected_prob,
        existing=existing_prob,
    )
    impact_dim, next_order = _resolve_matrix_axis_dimension(
        estrategia,
        'impacto',
        matrix_name,
        next_order,
        selected=selected_impact,
        existing=existing_impact,
    )
    return prob_dim, impact_dim


@transaction.atomic
def _persist_matrix_grid(matriz, prob_defs, impact_defs, cell_payload):
    prob_levels = _sync_matrix_levels(matriz, models.NivelProbabilidad, prob_defs, delete_stale=False)
    impact_levels = _sync_matrix_levels(matriz, models.NivelImpacto, impact_defs, delete_stale=False)
    keep_prob_level_ids = [level.pk for level in prob_levels]
    keep_impact_level_ids = [level.pk for level in impact_levels]

    prob_by_idx = {idx + 1: level for idx, level in enumerate(prob_levels)}
    impact_by_idx = {idx + 1: level for idx, level in enumerate(impact_levels)}

    keep_ids = []
    existing = {
        (cell.probabilidad.orden_visual, cell.impacto_nivel.orden_visual): cell
        for cell in models.MatrizRiesgoCelda.objects.filter(matriz=matriz).select_related('probabilidad', 'impacto_nivel')
    }

    for prob_idx, prob in prob_by_idx.items():
        for impact_idx, impact in impact_by_idx.items():
            payload = cell_payload.get((prob_idx, impact_idx), {})
            cell = existing.get((prob_idx, impact_idx)) or models.MatrizRiesgoCelda(matriz=matriz, probabilidad=prob, impacto_nivel=impact)
            result_num = payload.get('resultado_num')
            if result_num is None:
                prob_value = _decimal_or_none(prob.valor)
                impact_value = _decimal_or_none(impact.valor)
                result_num = ((prob_value or Decimal('0')) * (impact_value or Decimal('0'))).quantize(_MATRIX_DECIMAL_STEP)
            else:
                result_num = _quantize_matrix_decimal(result_num) or Decimal('0.00')
            cell.resultado_num = result_num
            cell.clasificacion = str(payload.get('clasificacion') or '')
            cell.color = str(payload.get('color') or '#2a2a3a')
            cell.calcular = bool(payload.get('calcular', True))
            cell.save()
            keep_ids.append(cell.pk)

    models.MatrizRiesgoCelda.objects.filter(matriz=matriz).exclude(pk__in=keep_ids).delete()
    models.NivelProbabilidad.objects.filter(matriz=matriz).exclude(pk__in=keep_prob_level_ids).delete()
    models.NivelImpacto.objects.filter(matriz=matriz).exclude(pk__in=keep_impact_level_ids).delete()
