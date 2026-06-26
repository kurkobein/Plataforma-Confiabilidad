from __future__ import annotations

import re
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.dateparse import parse_date

from core import models
from core.services.criticality_rules import sync_fmeca_criticality_rules

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


PROCESS_RCM_VALUES = {'fmeca', 'rcm', 'rcm_fmea', 'global', 'ambos'}

BASE_ALIASES = {
    'items': ['items', 'item'],
    'validado': ['validado'],
    'empresa': ['empresa', 'cliente'],
    'planta_sitio': ['planta sitio', 'planta / sitio', 'planta', 'sitio'],
    'area': ['area', 'área'],
    'proceso': ['proceso'],
    'descripcion_equipo': ['descripcion del equipo', 'descripción del equipo', 'equipo'],
    'tag': ['tag', 'tag equipo'],
    'ut': ['ubicac tecnica', 'ubicac.técnica', 'ubicacion tecnica', 'ubicación técnica', 'ut'],
    'sistema': ['sistema'],
    'sub_sistema': ['sub sistema', 'subsistema', 'sub-sistema'],
    'item_mantenible': ['item mantenible', 'ítem mantenible', 'item/mantenible'],
    'componente': ['componente'],
    'sap_equipo': ['n equipo sap', 'n° equipo sap', 'numero equipo sap'],
    'funcion': ['funcion', 'función'],
    'falla_funcional': ['falla funcional'],
    'modo_de_falla': ['modo de falla'],
    'efecto': ['efecto consecuencia de la falla', 'efecto/consecuencia de la falla', 'efecto'],
    'op': ['op'],
    'sld': ['sld'],
    'ma': ['ma'],
    'ir': ['i&r', 'iyr', 'i r'],
    'severidad': ['severidad'],
    'ocurrencia': ['ocurrencia'],
    'deteccion': ['deteccion', 'detección'],
    'npr': ['npr'],
}

BASE_ALIASES['descripcion_equipo'].extend([
    'nombre equipo',
    'descripcion equipo',
    'descripción equipo',
    'equipo componente',
    'equipo / componente',
    'equipo/componente',
])
BASE_ALIASES['funcion'].extend(['funcion equipo', 'función equipo'])
BASE_ALIASES['modo_de_falla'].extend(['ds modo de falla elemento problema causa raiz'])
BASE_ALIASES['efecto'].extend(['efecto de la falla', 'ds efecto de modo de falla'])
BASE_ALIASES['deteccion'].extend(['detencion', 'detención'])

DIMENSION_ALIASES = {
    'impacto operacion': ['op', 'operacion', 'impacto operacion'],
    'impacto seguridad': ['sld', 'seguridad', 'impacto seguridad'],
    'impacto medio ambiente': ['ma', 'medio ambiente', 'impacto medio ambiente'],
    'impacto reputacion': ['i&r', 'i r', 'iyr', 'ir', 'reputacion', 'impacto reputacion'],
    'severidad': ['severidad'],
    'ocurrencia': ['ocurrencia'],
    'deteccion': ['deteccion'],
    'npr': ['npr'],
    'criticidad': ['criticidad', 'criticidad rcm', 'criticidad fmea'],
    'rango npr': ['rango npr', 'rango', 'clasificacion npr', 'clasificación npr'],
}

DIMENSION_ALIASES['deteccion'].extend(['detencion', 'detención'])

STRUCTURAL_FILL_DOWN_ALIASES = {
    'Area': BASE_ALIASES['area'],
    'Proceso': BASE_ALIASES['proceso'],
    'Descripcion del Equipo': BASE_ALIASES['descripcion_equipo'],
    'TAG': BASE_ALIASES['tag'],
    'Ubicacion Tecnica': BASE_ALIASES['ut'],
    'Sistema': BASE_ALIASES['sistema'],
    'Sub Sistema': BASE_ALIASES['sub_sistema'],
    'Item Mantenible': BASE_ALIASES['item_mantenible'],
    'Componente': BASE_ALIASES['componente'],
    'N Equipo SAP': BASE_ALIASES['sap_equipo'],
    'Funcion': BASE_ALIASES['funcion'],
    'Falla Funcional': BASE_ALIASES['falla_funcional'],
    'Modo de Falla': BASE_ALIASES['modo_de_falla'],
    'Efecto': BASE_ALIASES['efecto'],
    'Tactica': ['tactica', 'táctica'],
}

NUMERIC_ZERO_ALIASES = {
    'OP': BASE_ALIASES['op'],
    'SLD': BASE_ALIASES['sld'],
    'MA': BASE_ALIASES['ma'],
    'I&R': BASE_ALIASES['ir'],
    'Ocurrencia': BASE_ALIASES['ocurrencia'],
    'Deteccion': BASE_ALIASES['deteccion'],
}

TASK_ALIASES = {
    'actual': {
        'descripcion': ['actividad de mantenimiento actual'],
        'repuesto': ['repuestos involucrados actuales', 'repuestos involucrados (actuales)'],
        'costo_total': ['costo mantenimiento actual'],
        'puesto_trabajo': ['pto trabajo sap'],
        'plan_sap': ['plan sap'],
        'descripcion_plan': ['descripcion plan', 'descripción plan'],
        'hoja_ruta': ['hoja de ruta'],
        'operacion_hoja_ruta': ['operacion', 'operación'],
        'frecuencia_texto': ['frecuencia'],
        'hh': ['hh sap'],
    },
    'primaria': {
        'tactica': ['tactica ftm cbm rtf ffm', 'táctica ftm-cbm-rtf-ffm', 'estrategia', 'estrategia cbm otf ftm'],
        'descripcion': [
            'actividad de mantenimiento primaria',
            'actividad primaria',
            'accion recomendada',
            'acción recomendada',
            'tarea de mtto',
            'ds descripcion de la tarea accion para prevenir el modo de falla',
        ],
        'componente_involucrado': ['items componente', 'items / componente'],
        'operacion_hoja_ruta': ['operacion hruta_1', 'operación hruta_1'],
        'operacion_pauta': ['operacion pauta', 'operación pauta'],
        'limite_aceptable': ['limite aceptable', 'límite aceptable'],
        'puesto_trabajo': ['pto trabajo', 'pto. trabajo'],
        'frecuencia_texto': ['frecuencia s', 'frecuencia (s)', 'frecuencia', 'no freq'],
        'frecuencia_unidad': ['cd freq unit horas'],
        'cantidad_personas': ['cant pers ejecutante', 'cant. pers ejecutante', 'no persons'],
        'especialidad': ['especialidad', 'cd esp electrico mecanico instrumentista lubricacion moncon', 'esp'],
        'duracion_min': ['dur min', 'dur (min)', 'no time min'],
        'duracion_hr': ['dur hr', 'dur (hr)'],
        'parametros': ['parametros'],
        'riesgo_material': ['riesgo material de la tarea'],
        'repuesto': ['repuesto para pm cod sap o descripcion', 'descrip repuesto'],
        'procedimiento_trabajo': ['procedimiento de trabajo'],
        'descripcion_plan': ['descripcion plan', 'descripción plan'],
        'hoja_ruta': ['hr', 'hoja de ruta'],
        'pauta': ['codigo pauta', 'código pauta', 'pauta final'],
        'titulo_pauta': ['titulo pauta', 'título pauta'],
        'hh': ['hh'],
        'estado_equipo': ['estado equipo'],
    },
    'secundaria': {
        'descripcion': ['actividad de mantenimiento secundaria', 'actividad secundaria', 'ds accion secundaria', 'ds acción secundaria'],
        'limite_aceptable': ['limite aceptable secundaria si existe', 'límite aceptable secundaria si existe'],
        'puesto_trabajo': ['pto trabajo secundaria si existe', 'pto. trabajo secundaria si existe'],
        'cantidad_personas': ['pers secundaria si existe', 'cantidad de personas'],
        'duracion_hr': ['dur secundaria si existe', 'duracion trabajo en hrs', 'duración trabajo en hrs'],
        'hh': ['hh secundaria si existe'],
        'componente_involucrado': ['componente involucrado'],
        'numero_parte': ['n parte n item', 'n° parte / n° ítem', 'numero parte item'],
        'numero_sap': ['n sap', 'n° sap'],
        'oportunidad_mejora': ['obs', 'observacion', 'observación'],
    },
}

TASK_TYPE_ALIASES = {
    'actual': ['actual', 'existente'],
    'primaria': ['primaria', 'principal', 'tarea rcm'],
    'secundaria': ['secundaria', 'complementaria'],
}


@dataclass
class ResolvedValue:
    raw: object = None
    number: Decimal | None = None
    text: str = ''
    catalog_row: object = None
    scale_value: object = None
    source: str = 'directo'

    @property
    def has_value(self):
        return self.number is not None or self.text not in ('', None) or self.catalog_row is not None or self.scale_value is not None


@dataclass
class ImportStats:
    filas_leidas: int = 0
    filas_importadas: int = 0
    filas_omitidas: int = 0
    equipos_creados: int = 0
    equipos_reutilizados: int = 0
    servicio_equipo_creados: int = 0
    cargas_creadas: int = 0
    rcm_creados: int = 0
    fmea_creados: int = 0
    evaluaciones_creadas: int = 0
    evaluaciones_calculadas: int = 0
    dependientes_resueltas: int = 0
    dependientes_minimo_cero: int = 0
    tareas_creadas: int = 0
    filas_omitidas_npr_cero: int = 0
    header_row: int | None = None
    start_row: int | None = None
    sheet_name: str = ''
    warnings: list[str] = field(default_factory=list)
    unresolved_dimensions: Counter = field(default_factory=Counter)
    fill_down_stats: dict[str, Counter] = field(default_factory=dict)

    def warn(self, message):
        self.warnings.append(message)

    def warn_unresolved_dimension(self, dimension_name):
        self.unresolved_dimensions[dimension_name] += 1

    def count_fill_down(self, column_name, event):
        self.fill_down_stats.setdefault(column_name, Counter())[event] += 1


def clean_text(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def normalize_key(value):
    text = clean_text(value).lower()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
    text = text.replace('&', ' y ')
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def slug_key(value):
    return normalize_key(value).replace(' ', '_')


def parse_decimal(value):
    if value in (None, ''):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raw = clean_text(value)
    if not raw:
        return None
    raw = raw.replace('%', '').replace(' ', '')
    if ',' in raw and '.' in raw:
        raw = raw.replace('.', '').replace(',', '.')
    else:
        raw = raw.replace(',', '.')
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return None


def value_is_empty(value):
    return value is None or clean_text(value) == ''


def int_or_none(value):
    number = parse_decimal(value)
    if number is None:
        return None
    if number == number.to_integral_value():
        return int(number)
    return None


def get_attr_any(obj, names, default=''):
    for name in names:
        if hasattr(obj, name):
            value = getattr(obj, name)
            if value not in (None, ''):
                return value
    return default


def set_attr_if_exists(obj, field_names, value):
    for field_name in field_names:
        if hasattr(obj, field_name):
            setattr(obj, field_name, value)
            return field_name
    return None


def source_ref_candidates(source_ref):
    if source_ref in (None, ''):
        return []
    if isinstance(source_ref, dict):
        candidates = []
        ed_id = source_ref.get('estrategia_dimension_id') or source_ref.get('estrategiaDimensionId') or source_ref.get('ed_id')
        dim_id = source_ref.get('dimension_id') or source_ref.get('dimensionId')
        if ed_id not in (None, ''):
            candidates.extend([f'ed:{ed_id}', f'estrategia_dimension:{ed_id}'])
        if dim_id not in (None, ''):
            candidates.extend([f'dim:{dim_id}', str(dim_id)])
        for key in ['campo', 'nombre', 'source', 'fuente', 'dependencia', 'depende_de', 'campo_fuente', 'dimension_fuente']:
            value = source_ref.get(key)
            if value not in (None, '') and not isinstance(value, dict):
                candidates.append(str(value))
        return candidates
    return [str(source_ref)]


def json_loads_safe(value, default=None):
    import json

    if default is None:
        default = {}
    if not value:
        return default
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return default


class Command(BaseCommand):
    help = 'Importa registros RCM/FMEA desde Excel respetando dimensiones configuradas.'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True, help='Ruta del Excel RCM.xlsx.')
        parser.add_argument('--service', required=True, help='ID, codigo o descripcion del servicio.')
        parser.add_argument('--sheet', default='RCM', help='Hoja del Excel. Default: RCM. Si no existe, intenta RCM-FMECA.')
        parser.add_argument('--header-row', type=int, default=None, help='Fila de encabezados. Si se omite, se autodetecta.')
        parser.add_argument('--start-row', type=int, default=None, help='Primera fila de datos. Si se omite, usa la siguiente al encabezado.')
        parser.add_argument('--limit', type=int, help='Limita la cantidad de filas procesadas.')
        parser.add_argument('--dry-run', action='store_true', help='Ejecuta y revierte todos los cambios.')
        parser.add_argument('--replace', action='store_true', help='Reemplaza cargas previas del mismo archivo y servicio.')
        parser.add_argument('--date', help='Fecha de analisis YYYY-MM-DD. Si se omite, usa hoy.')
        parser.add_argument('--create-task-types', action='store_true', help='Crea tipos de tarea actual/primaria/secundaria si no existen.')
        parser.add_argument('--debug-mapping', action='store_true', help='Muestra como se mapearon las dimensiones RCM contra columnas del Excel.')
        parser.add_argument('--debug-fill-down', action='store_true', help='Muestra estadisticas de fill-down y ceros aplicados por columna.')

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError('openpyxl no esta instalado.')

        path = Path(options['file'])
        if not path.exists():
            raise CommandError(f'No existe el archivo: {path}')

        service = self.resolve_service(options['service'])
        strategy = service.estrategia
        if not strategy:
            raise CommandError(f'El servicio {service} no tiene estrategia asociada.')

        analysis_date = parse_date(options.get('date') or '') or timezone.localdate()
        origin = f'RCM Excel: {path.name}'
        wb = load_workbook(path, data_only=True, read_only=True)
        sheet_name = options['sheet']
        if sheet_name not in wb.sheetnames and sheet_name == 'RCM' and 'RCM-FMECA' in wb.sheetnames:
            sheet_name = 'RCM-FMECA'
        if sheet_name not in wb.sheetnames:
            raise CommandError(f'El archivo no tiene hoja "{options["sheet"]}". Hojas: {", ".join(wb.sheetnames)}')
        ws = wb[sheet_name]

        header_row = options.get('header_row') or self.detect_header_row(ws)
        if not header_row:
            raise CommandError('No se detecto una fila de encabezados compatible con RCM/FMECA.')
        start_row = options.get('start_row') or header_row + 1
        header_map = self.build_header_map(ws, header_row)
        if not header_map:
            raise CommandError('No se detectaron encabezados validos.')

        stats = ImportStats()
        stats.header_row = header_row
        stats.start_row = start_row
        stats.sheet_name = sheet_name
        dims = self.get_active_rcm_dimensions(strategy)
        if options.get('debug_mapping') or options.get('dry_run'):
            self.print_debug_mapping(dims, header_map, ws, start_row)
        task_types = self.resolve_task_types(strategy, options['create_task_types'], stats)

        with transaction.atomic():
            if options['replace']:
                self.replace_previous(service, origin)

            processed = 0
            last_values = {}
            for row_number in range(start_row, ws.max_row + 1):
                if options.get('limit') and processed >= options['limit']:
                    break
                raw_row = ws[row_number]
                if self.raw_row_is_empty(raw_row, header_map):
                    continue
                row = self.build_row_context(raw_row, header_map, last_values, stats)
                if self.row_is_empty(row, header_map):
                    continue
                stats.filas_leidas += 1
                processed += 1
                try:
                    self.import_row(
                        row=row,
                        row_number=row_number,
                        header_map=header_map,
                        service=service,
                        strategy=strategy,
                        dims=dims,
                        task_types=task_types,
                        analysis_date=analysis_date,
                        origin=origin,
                        stats=stats,
                    )
                except Exception as exc:
                    stats.filas_omitidas += 1
                    stats.warn(f'Fila {row_number}: omitida por error: {exc}')

            if options.get('debug_fill_down'):
                self.print_fill_down_debug(stats)
            self.print_summary(service, strategy, path, options, stats)

            if options['dry_run']:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING('Dry-run: no se guardaron cambios.'))
            else:
                self.stdout.write(self.style.SUCCESS('Importacion RCM completada.'))

    def resolve_service(self, value):
        qs = models.Servicio.objects.select_related('empresa', 'estrategia')
        value = clean_text(value)
        if value.isdigit():
            service = qs.filter(pk=int(value)).first()
            if service:
                return service
        service = qs.filter(codigo_servicio__iexact=value).first()
        if service:
            return service
        service = qs.filter(Q(codigo_servicio__icontains=value) | Q(descripcion__icontains=value)).first()
        if service:
            return service
        raise CommandError(f'No se encontro servicio para: {value}')

    def build_header_map(self, ws, header_row):
        header_map = {}
        self.header_labels = {}
        for cell in ws[header_row]:
            label = clean_text(cell.value)
            if not label:
                continue
            key = normalize_key(label)
            header_map.setdefault(key, cell.column)
            self.header_labels.setdefault(cell.column, label)
        return header_map

    def detect_header_row(self, ws, max_rows=30):
        best_row = None
        best_score = 0
        alias_groups = list(BASE_ALIASES.values())
        for aliases in TASK_ALIASES.values():
            alias_groups.extend(aliases.values())
        alias_groups.extend(DIMENSION_ALIASES.values())
        normalized_aliases = {normalize_key(alias) for group in alias_groups for alias in group}
        critical_aliases = {
            normalize_key(alias)
            for key in ['tag', 'ut', 'funcion', 'falla_funcional', 'modo_de_falla', 'efecto', 'descripcion_equipo']
            for alias in BASE_ALIASES[key]
        }
        hierarchy_aliases = {
            normalize_key(alias)
            for key in ['empresa', 'planta_sitio', 'area', 'sistema', 'sub_sistema', 'tag']
            for alias in BASE_ALIASES[key]
        }

        for row_number in range(1, min(ws.max_row, max_rows) + 1):
            labels = [normalize_key(cell.value) for cell in ws[row_number] if clean_text(cell.value)]
            if not labels:
                continue
            label_set = set(labels)
            score = sum(3 for label in label_set if label in normalized_aliases)
            score += sum(2 for label in label_set if label in critical_aliases)
            score += sum(1 for label in label_set if label in hierarchy_aliases)
            has_rcm_signal = bool(label_set & critical_aliases)
            has_hierarchy_signal = len(label_set & hierarchy_aliases) >= 3
            if (has_rcm_signal or has_hierarchy_signal) and score > best_score:
                best_row = row_number
                best_score = score
        return best_row

    def row_is_empty(self, row, header_map):
        important = [
            'tag',
            'ut',
            'empresa',
            'planta_sitio',
            'area',
            'sistema',
            'sub_sistema',
            'funcion',
            'falla_funcional',
            'modo_de_falla',
            'efecto',
            'descripcion_equipo',
        ]
        return all(value_is_empty(self.get_cell(row, header_map, BASE_ALIASES[name])) for name in important)

    def raw_row_is_empty(self, row, header_map):
        return all(value_is_empty(self.row_value(row, col)) for col in set(header_map.values()))

    def row_value(self, row, col, default=''):
        if isinstance(row, dict):
            return row.get(col, default)
        return row[col - 1].value

    def get_cell(self, row, header_map, aliases, default=''):
        for alias in aliases:
            col = header_map.get(normalize_key(alias))
            if col:
                return self.row_value(row, col, default)
        return default

    def column_for_aliases(self, header_map, aliases):
        for alias in aliases:
            col = header_map.get(normalize_key(alias))
            if col:
                return col
        return None

    def build_row_context(self, row, header_map, last_values, stats):
        context = {col: self.row_value(row, col) for col in set(header_map.values())}

        for label, aliases in STRUCTURAL_FILL_DOWN_ALIASES.items():
            col = self.column_for_aliases(header_map, aliases)
            if not col:
                continue
            raw = self.row_value(row, col)
            if not value_is_empty(raw):
                last_values[col] = raw
                context[col] = raw
                stats.count_fill_down(label, 'reales')
            elif col in last_values and not value_is_empty(last_values[col]):
                context[col] = last_values[col]
                stats.count_fill_down(label, 'heredados')
            else:
                context[col] = ''
                stats.count_fill_down(label, 'vacios_finales')

        for label, aliases in NUMERIC_ZERO_ALIASES.items():
            col = self.column_for_aliases(header_map, aliases)
            if not col:
                continue
            raw = self.row_value(row, col)
            if value_is_empty(raw):
                context[col] = Decimal('0')
                stats.count_fill_down(label, 'cero_por_vacio')
            else:
                context[col] = raw
                stats.count_fill_down(label, 'reales')
        return context

    def print_debug_mapping(self, dims, header_map, ws, start_row):
        self.stdout.write(self.style.NOTICE('Mapeo de dimensiones RCM contra Excel:'))
        for ed in dims:
            name = ed.dimension.nombre
            if self.is_calculated(ed):
                self.stdout.write(self.style.NOTICE(f'- {name} -> calculada'))
                continue
            if self.dependency_ref(ed):
                self.stdout.write(self.style.NOTICE(f'- {name} -> dependiente'))
                continue
            col = self.match_dimension_to_excel_column(ed, header_map)
            if not col:
                self.stdout.write(self.style.WARNING(f'- {name} -> no encontrada'))
                continue
            label = getattr(self, 'header_labels', {}).get(col, f'Columna {col}')
            sample = self.first_non_empty_column_value(ws, col, start_row)
            sample_text = f' | primer valor: {sample}' if not value_is_empty(sample) else ' | primer valor: vacio'
            self.stdout.write(self.style.NOTICE(f'- {name} -> {label}{sample_text}'))

    def first_non_empty_column_value(self, ws, col, start_row):
        return ws.cell(row=start_row, column=col).value

    def get_active_rcm_dimensions(self, strategy):
        allowed = set()
        choices = dict(models.EstrategiaDimension.PROCESO_USO_CHOICES)
        for value in PROCESS_RCM_VALUES:
            if value in choices:
                allowed.add(value)
        allowed.update(getattr(models.EstrategiaDimension, 'PROCESO_FMECA_ALIASES', ('fmeca', 'rcm')))
        allowed.add(models.EstrategiaDimension.PROCESO_AMBOS)
        return list(
            models.EstrategiaDimension.objects.filter(
                estrategia=strategy,
                activo=True,
                proceso_uso__in=allowed,
            )
            .select_related('dimension')
            .prefetch_related(
                'escalas_valor__escala_unificada',
                'catalogo__filas__celdas__columna',
                'catalogo__columnas',
            )
            .order_by('orden', 'id')
        )

    def import_row(self, row, row_number, header_map, service, strategy, dims, task_types, analysis_date, origin, stats):
        resolved, eval_stats = self.resolve_dimensions(row, header_map, dims, stats)
        npr_value = self.find_resolved_value(resolved, ['npr'])
        if npr_value and npr_value.number == Decimal('0'):
            stats.filas_omitidas_npr_cero += 1
            stats.filas_omitidas += 1
            return

        equipment, created = self.get_or_create_equipment(row, header_map, service, stats)
        if created:
            stats.equipos_creados += 1
        else:
            stats.equipos_reutilizados += 1

        _, se_created = models.ServicioEquipo.objects.get_or_create(servicio=service, equipo=equipment)
        if se_created:
            stats.servicio_equipo_creados += 1

        now = timezone.now()
        carga = models.Carga.objects.create(
            fecha_analisis=analysis_date,
            version_carga=Decimal('1.0'),
            origen=origin,
            status=models.Carga.STATUS_COMPLETO,
            creado_en=now,
            actualizado=now,
            estrategia=strategy,
            servicio=service,
            usuario=None,
        )
        stats.cargas_creadas += 1

        rcm = models.RCM.objects.create(
            carga=carga,
            equipo=equipment,
            criticidad=None,
            fecha_analisis=analysis_date,
            estado=models.Carga.STATUS_COMPLETO,
            componente=clean_text(self.get_cell(row, header_map, BASE_ALIASES['componente'])),
            funcion=clean_text(self.get_cell(row, header_map, BASE_ALIASES['funcion'])),
            falla_funcional=clean_text(self.get_cell(row, header_map, BASE_ALIASES['falla_funcional'])),
            modo_de_falla=clean_text(self.get_cell(row, header_map, BASE_ALIASES['modo_de_falla'])),
            causa='',
            efecto=clean_text(self.get_cell(row, header_map, BASE_ALIASES['efecto'])),
        )
        stats.rcm_creados += 1

        fmea = models.FMEA_FMECA.objects.create(rcm=rcm)
        stats.fmea_creados += 1

        for ed, value in resolved.items():
            if self.save_evaluation(fmea, ed, value):
                stats.evaluaciones_creadas += 1
        stats.evaluaciones_calculadas += eval_stats['calculadas']
        stats.dependientes_resueltas += eval_stats['dependientes']
        stats.dependientes_minimo_cero += eval_stats['dependientes_minimo_cero']

        criticidad_value = self.find_resolved_value(resolved, ['criticidad', 'criticidad rcm', 'criticidad fmea'])
        if criticidad_value and criticidad_value.number is not None and criticidad_value.number == criticidad_value.number.to_integral_value():
            rcm.criticidad = int(criticidad_value.number)
            rcm.save(update_fields=['criticidad'])

        sync_fmeca_criticality_rules(rcm, fmea, service)

        self.create_tasks(fmea, row, header_map, task_types, stats)
        stats.filas_importadas += 1

    def get_or_create_equipment(self, row, header_map, service=None, stats=None):
        tag = clean_text(self.get_cell(row, header_map, BASE_ALIASES['tag']))
        ut = clean_text(self.get_cell(row, header_map, BASE_ALIASES['ut'])).upper()
        if not ut and service is not None:
            ut = self.build_ut_from_hierarchy(row, header_map, service, stats)
        name = clean_text(self.get_cell(row, header_map, BASE_ALIASES['descripcion_equipo'])) or tag or ut
        nodo = self.find_node_by_ut(ut, service) if service is not None else None
        equipo = None
        if ut:
            equipo = models.Equipo.objects.filter(ut__iexact=ut).first()
        if not equipo and tag:
            equipo = models.Equipo.objects.filter(tag_equipo__iexact=tag).first()
        if equipo:
            changed = []
            if name and equipo.nombre_equipo != name:
                equipo.nombre_equipo = name
                changed.append('nombre_equipo')
            if ut and equipo.ut != ut:
                equipo.ut = ut
                changed.append('ut')
            if nodo and equipo.nodo_id != nodo.pk:
                equipo.nodo = nodo
                changed.append('nodo')
            if changed:
                equipo.save(update_fields=changed)
            return equipo, False
        return models.Equipo.objects.create(
            tag_equipo=tag or self.tag_from_ut(ut) or 'SIN-TAG',
            nombre_equipo=name or 'Equipo importado RCM',
            ut=ut or tag or 'SIN-UT',
            descripcion_ut='',
            nodo=nodo,
        ), True

    def build_ut_from_hierarchy(self, row, header_map, service, stats=None):
        segments = []
        parent = None
        hierarchy = [
            ('empresa', 1),
            ('planta_sitio', 2),
            ('area', 3),
            ('sistema', 4),
            ('sub_sistema', 5),
        ]
        for key, level_order in hierarchy:
            raw = clean_text(self.get_cell(row, header_map, BASE_ALIASES[key]))
            if not raw:
                continue
            node = self.match_hierarchy_node(service, level_order, raw, parent)
            if node:
                segments.append(self.technical_segment(node.codigo))
                parent = node
                continue
            fallback = self.technical_segment(raw)
            if fallback:
                segments.append(fallback)
            if stats:
                stats.warn(f'No se encontro nodo jerarquico para "{raw}" en nivel {level_order}; se uso "{fallback}" como codigo.')

        tag = self.technical_segment(clean_text(self.get_cell(row, header_map, BASE_ALIASES['tag'])))
        if tag and (not segments or segments[-1] != tag):
            segments.append(tag)
        return '-'.join(segment for segment in segments if segment)

    def match_hierarchy_node(self, service, level_order, raw, parent=None):
        raw_norm = normalize_key(raw)
        if not raw_norm:
            return None
        qs = models.NodoJerarquia.objects.filter(
            empresa=service.empresa,
            activo=True,
            nivel__orden=level_order,
        ).select_related('nivel', 'parent')
        if parent is not None:
            preferred = list(qs.filter(parent=parent))
            fallback = list(qs)
            nodes = preferred or fallback
        else:
            nodes = list(qs)
        for node in nodes:
            candidates = [
                node.codigo,
                node.nombre,
                f'{node.codigo} {node.nombre}',
                f'{node.codigo} - {node.nombre}',
            ]
            if raw_norm in {normalize_key(candidate) for candidate in candidates}:
                return node
        for node in nodes:
            if normalize_key(node.nombre) in raw_norm or raw_norm in normalize_key(node.nombre):
                return node
        return None

    def find_node_by_ut(self, ut, service):
        if not ut or service is None:
            return None
        parts = [self.technical_segment(part) for part in clean_text(ut).split('-') if clean_text(part)]
        if not parts:
            return None
        candidates = models.NodoJerarquia.objects.filter(
            empresa=service.empresa,
            activo=True,
            codigo__iexact=parts[-2] if len(parts) > 1 else parts[-1],
        ).select_related('nivel', 'parent')
        for node in candidates:
            if node.ut.upper() == '-'.join(parts[:len(node.path_nodes())]).upper():
                return node
        return None

    def technical_segment(self, value):
        text = clean_text(value).upper()
        text = unicodedata.normalize('NFD', text)
        text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
        text = re.sub(r'[^A-Z0-9]+', '', text)
        return text

    def tag_from_ut(self, ut):
        last = clean_text(ut).split('-')[-1] if ut else ''
        return last.lstrip('0') or last

    def resolve_task_types(self, strategy, create_missing, stats):
        result = {}
        existing = list(models.TipoTareaEstrategia.objects.filter(estrategia=strategy, activo=True))
        for task_key, aliases in TASK_TYPE_ALIASES.items():
            tipo = next(
                (
                    item for item in existing
                    if normalize_key(item.codigo) in {normalize_key(alias) for alias in aliases}
                    or any(normalize_key(alias) in normalize_key(item.nombre) for alias in aliases)
                ),
                None,
            )
            if not tipo and create_missing:
                tipo = models.TipoTareaEstrategia.objects.create(
                    estrategia=strategy,
                    nombre=task_key.capitalize(),
                    codigo=task_key,
                    orden=len(existing) + len(result) + 1,
                    activo=True,
                )
                stats.warn(f'Se creo TipoTareaEstrategia "{task_key}" para la estrategia {strategy}.')
            if not tipo:
                stats.warn(f'No existe TipoTareaEstrategia para "{task_key}". Se omitiran esas tareas.')
            result[task_key] = tipo
        return result

    def create_tasks(self, fmea, row, header_map, task_types, stats):
        order = 0
        for task_key in ['actual', 'primaria', 'secundaria']:
            tipo = task_types.get(task_key)
            if not tipo:
                continue
            task_data = self.task_payload(task_key, row, header_map)
            if not task_data.get('descripcion'):
                continue
            order += 1
            task = models.TareaRCM.objects.create(
                fmea=fmea,
                tipo_tarea_estrategia=tipo,
                orden=order,
                estado=models.TareaRCM.ESTADO_ACTIVO,
                **task_data,
            )
            self.save_dynamic_task_values(task, row, header_map)
            stats.tareas_creadas += 1

    def task_payload(self, task_key, row, header_map):
        aliases = TASK_ALIASES[task_key]
        text_fields = {
            'descripcion',
            'tactica',
            'limite_aceptable',
            'parametros',
            'riesgo_material',
            'especialidad',
            'puesto_trabajo',
            'estado_equipo',
            'frecuencia_unidad',
            'frecuencia_texto',
            'plan_sap',
            'descripcion_plan',
            'hoja_ruta',
            'texto_hoja_ruta',
            'operacion_hoja_ruta',
            'texto_operacion',
            'operacion_pauta',
            'pauta',
            'titulo_pauta',
            'repuesto',
            'componente_involucrado',
            'numero_parte',
            'numero_sap',
            'procedimiento_trabajo',
            'oportunidad_mejora',
        }
        decimal_fields = {
            'frecuencia_valor',
            'duracion_min',
            'duracion_hr',
            'cantidad_personas',
            'hh',
            'costo_hh',
            'costo_repuestos',
            'tarifa_servicios',
            'costo_total',
        }
        payload = {}
        for field_name in text_fields:
            if field_name in aliases:
                payload[field_name] = clean_text(self.get_cell(row, header_map, aliases[field_name]))
        for field_name in decimal_fields:
            if field_name in aliases:
                payload[field_name] = parse_decimal(self.get_cell(row, header_map, aliases[field_name]))
        payload.setdefault('descripcion', '')
        return payload

    def save_dynamic_task_values(self, task, row, header_map):
        fields = task.tipo_tarea_estrategia.campos.filter(activo=True)
        for field in fields:
            raw = self.get_cell(row, header_map, [field.nombre, field.clave], default='')
            if value_is_empty(raw):
                continue
            defaults = {'valor_texto': '', 'valor_numero': None, 'valor_booleano': None, 'valor_fecha': None}
            if field.tipo_dato in {models.CampoTareaEstrategia.TIPO_NUMERO, models.CampoTareaEstrategia.TIPO_DECIMAL}:
                defaults['valor_numero'] = parse_decimal(raw)
                if defaults['valor_numero'] is None:
                    defaults['valor_texto'] = clean_text(raw)
            elif field.tipo_dato == models.CampoTareaEstrategia.TIPO_BOOLEANO:
                defaults['valor_booleano'] = normalize_key(raw) in {'si', 'sí', 'true', '1', 'x'}
            elif field.tipo_dato == models.CampoTareaEstrategia.TIPO_FECHA:
                defaults['valor_fecha'] = parse_date(clean_text(raw))
                if defaults['valor_fecha'] is None:
                    defaults['valor_texto'] = clean_text(raw)
            else:
                defaults['valor_texto'] = clean_text(raw)
            models.ValorCampoTareaRCM.objects.update_or_create(tarea=task, campo=field, defaults=defaults)

    def resolve_dimensions(self, row, header_map, dims, stats):
        resolved = {}
        value_context = {}
        eval_stats = {'calculadas': 0, 'dependientes': 0, 'dependientes_minimo_cero': 0}

        for ed in dims:
            if self.is_calculated(ed) or self.dependency_ref(ed):
                continue
            raw = self.dimension_raw_value(ed, row, header_map)
            if value_is_empty(raw):
                continue
            value = self.match_catalog_or_scale(ed, raw) or self.direct_value(ed, raw)
            if value.has_value:
                resolved[ed] = value
                self.remember_value(value_context, ed, value)

        for _pass in range(10):
            changed = False
            for ed in dims:
                if ed in resolved:
                    continue
                value = None
                if self.is_calculated(ed):
                    value = self.evaluate_calculated_dimension(ed, value_context)
                    if value and value.has_value:
                        eval_stats['calculadas'] += 1
                elif self.dependency_ref(ed):
                    value = self.resolve_dependent_dimension(ed, value_context)
                    if value and value.has_value:
                        if value.source == 'dependiente_minimo_cero':
                            eval_stats['dependientes_minimo_cero'] += 1
                        else:
                            eval_stats['dependientes'] += 1
                if value and value.has_value:
                    resolved[ed] = value
                    self.remember_value(value_context, ed, value)
                    changed = True
            if not changed:
                break

        for ed in dims:
            if ed.obligatorio and ed not in resolved:
                if self.strict_dimension_key(ed.dimension.nombre) == 'criticidad':
                    continue
                stats.warn_unresolved_dimension(ed.dimension.nombre)
        return resolved, eval_stats

    def dimension_raw_value(self, ed, row, header_map):
        col = self.match_dimension_to_excel_column(ed, header_map)
        if col:
            return self.row_value(row, col)
        return ''

    def match_dimension_to_excel_column(self, ed, header_map):
        strict_aliases = self.strict_dimension_aliases(ed.dimension.nombre)
        if strict_aliases is not None:
            for alias in strict_aliases:
                col = header_map.get(normalize_key(alias))
                if col:
                    return col
            return None

        candidates = [ed.dimension.nombre, slug_key(ed.dimension.nombre)]
        try:
            catalogo = ed.catalogo
            if catalogo and catalogo.campo:
                candidates.append(catalogo.campo)
        except models.DimensionCatalogo.DoesNotExist:
            pass
        candidates.extend(self.dimension_aliases(ed.dimension.nombre))

        for candidate in candidates:
            col = header_map.get(normalize_key(candidate))
            if col:
                return col

        normalized_candidates = [
            normalize_key(candidate)
            for candidate in candidates
            if len(normalize_key(candidate)) >= 4
        ]
        for header_key, col in header_map.items():
            if len(header_key) < 4:
                continue
            if self.strict_dimension_key(header_key) in {'npr', 'criticidad'}:
                continue
            if any(candidate in header_key or header_key in candidate for candidate in normalized_candidates):
                return col
        return None

    def strict_dimension_key(self, name):
        normalized = normalize_key(name)
        words = set(normalized.split())
        if normalized == 'npr':
            return 'npr'
        if normalized in {'criticidad', 'criticidad rcm', 'criticidad fmea'}:
            return 'criticidad'
        if normalized == 'rango npr' or ('rango' in words and 'npr' in words) or normalized == 'clasificacion npr':
            return 'rango npr'
        return ''

    def strict_dimension_aliases(self, name):
        key = self.strict_dimension_key(name)
        if not key:
            return None
        return DIMENSION_ALIASES[key]

    def dimension_aliases(self, name):
        normalized = normalize_key(name)
        aliases = []
        for alias_key, alias_group in DIMENSION_ALIASES.items():
            alias_key_norm = normalize_key(alias_key)
            if self.strict_dimension_key(alias_key_norm):
                if normalized == alias_key_norm:
                    aliases.extend(alias_group)
                continue
            if normalized == alias_key_norm or alias_key_norm in normalized or normalized in alias_key_norm:
                aliases.extend(alias_group)
        for alias_group in BASE_ALIASES.values():
            if normalized in {normalize_key(item) for item in alias_group}:
                aliases.extend(alias_group)
        if normalized in {'i y r', 'i r', 'ir', 'iyr'}:
            aliases.extend(['I&R', 'IYR', 'I R', 'IR'])
        if normalized == 'deteccion':
            aliases.append('Detección')
        deduped = []
        seen = set()
        for alias in aliases:
            alias_norm = normalize_key(alias)
            if alias_norm not in seen:
                deduped.append(alias)
                seen.add(alias_norm)
        return deduped

    def is_calculated(self, ed):
        return bool(clean_text(getattr(ed.dimension, 'tipo_calculo', '')))

    def dependency_ref(self, ed):
        config = json_loads_safe(ed.dimension.config_calculo, {})
        if not isinstance(config, dict):
            return ''
        for key in ['dependencia', 'depende_de', 'source', 'fuente', 'campo_fuente', 'dimension_fuente']:
            value = config.get(key)
            if value not in (None, ''):
                return value
        if config.get('estrategia_dimension_id') or config.get('dimension_id'):
            return config
        return ''

    def direct_value(self, ed, raw):
        number = parse_decimal(raw)
        if number is not None and ed.dimension.tipo_dato in {'numerico', 'ordinal', 'tabla'}:
            return ResolvedValue(raw=raw, number=number, text=clean_text(raw), source='directo')
        return ResolvedValue(raw=raw, text=clean_text(raw), source='directo')

    def match_catalog_or_scale(self, ed, raw):
        raw_text = clean_text(raw)
        raw_norm = normalize_key(raw_text)
        raw_number = parse_decimal(raw)

        for scale in ed.escalas_valor.all():
            if raw_number is not None and scale.valor_numerico == raw_number:
                return ResolvedValue(raw=raw, number=scale.valor_numerico, text=scale.codigo or scale.descripcion or raw_text, scale_value=scale, source='escala')
            candidates = [scale.codigo, scale.descripcion]
            if scale.escala_unificada_id:
                candidates.extend([scale.escala_unificada.significado, scale.escala_unificada.interpretacion])
            if raw_norm and raw_norm in {normalize_key(item) for item in candidates if item}:
                return ResolvedValue(raw=raw, number=scale.valor_numerico, text=scale.codigo or scale.descripcion or raw_text, scale_value=scale, source='escala')

        try:
            catalogo = ed.catalogo
        except models.DimensionCatalogo.DoesNotExist:
            catalogo = None
        if not catalogo:
            return None

        for catalog_row in catalogo.filas.all().order_by('orden', 'id'):
            values = catalog_row.values_map()
            candidates = [catalog_row.etiqueta] + [clean_text(value) for value in values.values()]
            if raw_norm and raw_norm in {normalize_key(item) for item in candidates if item}:
                return self.value_from_catalog_row(catalog_row, raw=raw, source='catalogo')
            for value in values.values():
                if raw_number is not None and parse_decimal(value) == raw_number:
                    return self.value_from_catalog_row(catalog_row, raw=raw, source='catalogo')
        return None

    def value_from_catalog_row(self, catalog_row, raw=None, source='dependiente'):
        values = catalog_row.values_map()
        number = self.catalog_primary_numeric(values)
        text = self.catalog_primary_text(catalog_row, values)
        return ResolvedValue(raw=raw, number=number, text=text, catalog_row=catalog_row, source=source)

    def catalog_primary_numeric(self, values):
        for key in ['valor_numerico', 'valor_principal', 'valor', 'nivel', 'puntaje', 'indicador']:
            if key in values:
                number = parse_decimal(values.get(key))
                if number is not None:
                    return number
        for key, value in values.items():
            if key in {'limite_inferior', 'limite_superior', 'desde', 'hasta', 'min', 'max', 'minimo', 'maximo'}:
                continue
            number = parse_decimal(value)
            if number is not None:
                return number
        return None

    def catalog_primary_text(self, catalog_row, values):
        for key in ['etiqueta', 'nombre', 'descripcion', 'texto', 'codigo', 'indicador', 'valor']:
            value = catalog_row.etiqueta if key == 'etiqueta' else values.get(key)
            if value not in (None, ''):
                return clean_text(value)
        return clean_text(catalog_row.etiqueta)

    def remember_value(self, value_context, ed, value):
        keys = [
            f'ed:{ed.pk}',
            f'estrategia_dimension:{ed.pk}',
            f'dim:{ed.dimension_id}',
            str(ed.dimension_id),
            ed.dimension.nombre,
            slug_key(ed.dimension.nombre),
            normalize_key(ed.dimension.nombre),
        ]
        try:
            catalogo = ed.catalogo
            if catalogo and catalogo.campo:
                keys.extend([catalogo.campo, slug_key(catalogo.campo)])
        except models.DimensionCatalogo.DoesNotExist:
            pass
        for alias in self.dimension_aliases(ed.dimension.nombre):
            keys.extend([alias, slug_key(alias), normalize_key(alias)])
        for key in keys:
            value_context[key] = value

    def source_value(self, value_context, ref):
        for candidate in source_ref_candidates(ref):
            for key in [candidate, slug_key(candidate), normalize_key(candidate)]:
                value = value_context.get(key)
                if value and value.number is not None:
                    return value
        return None

    def evaluate_calculated_dimension(self, ed, value_context):
        config = json_loads_safe(ed.dimension.config_calculo, {})
        tipo = normalize_key(ed.dimension.tipo_calculo)
        steps = self.calculation_steps(tipo, config)
        result = None
        for step_type, operands, mode in steps:
            values = []
            for operand in operands:
                if isinstance(operand, dict) and (operand.get('resultado') is True or operand.get('tipo') == 'resultado'):
                    if result is not None:
                        values.append(result)
                    continue
                source = self.source_value(value_context, operand)
                if source and source.number is not None:
                    values.append(source.number)
            if mode == 'ponderado' and step_type == 'suma':
                values = self.weighted_values(operands, values)
                if values is None:
                    return None
            result = self.calculate(step_type, values)
            if result is None:
                return None
        return ResolvedValue(number=result, text=clean_text(result), source='calculado') if result is not None else None

    def calculation_steps(self, tipo, config):
        if not isinstance(config, dict):
            config = {}
        raw_steps = config.get('pasos') or config.get('steps') or []
        if isinstance(raw_steps, list) and raw_steps:
            steps = []
            for step in raw_steps:
                if not isinstance(step, dict):
                    continue
                op = normalize_key(step.get('operacion') or step.get('tipo_calculo') or step.get('operation') or tipo)
                operands = step.get('operandos') or step.get('campos') or step.get('sources') or []
                mode = step.get('modo') or ('ponderado' if step.get('ponderado') is True else '')
                steps.append((op, operands if isinstance(operands, list) else [], mode))
            return steps
        operands = config.get('operandos') or config.get('campos') or config.get('sources') or []
        return [(tipo, operands if isinstance(operands, list) else [], '')]

    def weighted_values(self, operands, values):
        if not operands or len(operands) != len(values):
            return None
        weights = []
        for operand in operands:
            raw = operand.get('peso', operand.get('ponderador', operand.get('weight'))) if isinstance(operand, dict) else None
            try:
                weights.append(Decimal(str(raw).replace(',', '.')) if raw not in (None, '') else None)
            except (InvalidOperation, TypeError, ValueError):
                return None
        if all(weight is None for weight in weights):
            equal_weight = Decimal('1') / Decimal(len(weights))
            weights = [equal_weight] * len(weights)
        elif any(weight is None for weight in weights):
            assigned = sum((weight for weight in weights if weight is not None), Decimal('0'))
            missing_count = sum(1 for weight in weights if weight is None)
            remaining = Decimal('1') - assigned
            if remaining < 0:
                return None
            missing_weight = remaining / Decimal(missing_count)
            weights = [missing_weight if weight is None else weight for weight in weights]
        if any(weight < 0 or weight > 1 for weight in weights):
            return None
        if abs(sum(weights, Decimal('0')) - Decimal('1')) > Decimal('0.0001'):
            return None
        return [value * weight for value, weight in zip(values, weights)]

    def calculate(self, tipo, values):
        if not values:
            return None
        if tipo == 'suma':
            return sum(values, Decimal('0'))
        if tipo == 'resta':
            result = values[0]
            for value in values[1:]:
                result -= value
            return result
        if tipo in {'multiplicacion', 'multiplicacion'}:
            result = Decimal('1')
            for value in values:
                result *= value
            return result
        if tipo == 'division':
            result = values[0]
            for value in values[1:]:
                if value == 0:
                    return None
                result /= value
            return result
        if tipo in {'maximo', 'maximo'}:
            return max(values)
        if tipo in {'minimo', 'minimo'}:
            return min(values)
        return None

    def resolve_dependent_dimension(self, ed, value_context):
        ref = self.dependency_ref(ed)
        source = self.source_value(value_context, ref)
        if not source or source.number is None:
            return None
        try:
            catalogo = ed.catalogo
        except models.DimensionCatalogo.DoesNotExist:
            return None
        row = self.match_dependency_row(catalogo, source.number)
        source_label = 'dependiente'
        if not row and source.number == Decimal('0'):
            row = self.match_minimum_dependency_row(catalogo)
            source_label = 'dependiente_minimo_cero'
        if not row:
            return None
        return self.value_from_catalog_row(row, raw=source.number, source=source_label)

    def match_dependency_row(self, catalogo, source_number):
        if catalogo.tipo == 'opciones':
            for row in catalogo.filas.all().order_by('orden', 'id'):
                values = row.values_map()
                if self.catalog_primary_numeric(values) == source_number:
                    return row
            return None

        rows = list(catalogo.filas.all().order_by('orden', 'id'))
        for idx, row in enumerate(rows):
            values = row.values_map()
            lower = self.catalog_bound(values, ['limite_inferior', 'desde', 'min', 'minimo', 'mínimo'])
            upper = self.catalog_bound(values, ['limite_superior', 'hasta', 'max', 'maximo', 'máximo'])
            if lower is not None and source_number < lower:
                continue
            if upper is not None and source_number > upper:
                continue
            # El superior se considera inclusivo, especialmente el ultimo rango.
            if upper is not None and source_number == upper:
                return row
            return row
        return None

    def match_minimum_dependency_row(self, catalogo):
        if catalogo.tipo == 'opciones':
            return None
        rows = list(catalogo.filas.all().order_by('orden', 'id'))
        if not rows:
            return None

        lowest_row = None
        lowest_lower = None
        for row in rows:
            values = row.values_map()
            lower = self.catalog_bound(values, ['limite_inferior', 'desde', 'min', 'minimo', 'mí­nimo'])
            if lower is None:
                continue
            if lowest_lower is None or lower < lowest_lower:
                lowest_lower = lower
                lowest_row = row
        return lowest_row or rows[0]

    def catalog_bound(self, values, keys):
        normalized_values = {normalize_key(key): value for key, value in values.items()}
        for key in keys:
            value = normalized_values.get(normalize_key(key))
            if value not in (None, ''):
                return parse_decimal(value)
        return None

    def save_evaluation(self, fmea, ed, value):
        defaults = {
            'valor_numerico': None,
            'valor_texto': value.text or '',
            'catalogo_fila': value.catalog_row,
            'escala_valor': value.scale_value,
        }
        if value.number is not None and value.number == value.number.to_integral_value():
            defaults['valor_numerico'] = int(value.number)
        elif value.number is not None and not defaults['valor_texto']:
            defaults['valor_texto'] = clean_text(value.number)
        models.EvaluacionFMEA.objects.update_or_create(
            fmea=fmea,
            estrategia_dimension=ed,
            defaults=defaults,
        )
        return True

    def find_resolved_value(self, resolved, names):
        normalized_names = {normalize_key(name) for name in names}
        for ed, value in resolved.items():
            if normalize_key(ed.dimension.nombre) in normalized_names:
                return value
        return None

    def replace_previous(self, service, origin):
        carga_ids = list(models.Carga.objects.filter(servicio=service, origen=origin).values_list('id', flat=True))
        if not carga_ids:
            return
        try:
            rcm_ids = list(models.RCM.objects.filter(carga_id__in=carga_ids).values_list('id', flat=True))
            fmea_ids = list(models.FMEA_FMECA.objects.filter(rcm_id__in=rcm_ids).values_list('id', flat=True))
            tarea_ids = list(models.TareaRCM.objects.filter(fmea_id__in=fmea_ids).values_list('id', flat=True))
            if tarea_ids:
                models.ValorCampoTareaRCM.objects.filter(tarea_id__in=tarea_ids).delete()
                models.TareaRCM.objects.filter(id__in=tarea_ids).delete()
            if fmea_ids:
                models.EvaluacionFMEA.objects.filter(fmea_id__in=fmea_ids).delete()
                models.FMEA_FMECA.objects.filter(id__in=fmea_ids).delete()
            if rcm_ids:
                models.RCM.objects.filter(id__in=rcm_ids).delete()
            models.Carga.objects.filter(id__in=carga_ids).delete()
        except IntegrityError as exc:
            raise CommandError(f'No se pudo reemplazar la carga previa por restricciones FK: {exc}') from exc

    def print_fill_down_debug(self, stats):
        self.stdout.write(self.style.NOTICE('Debug fill-down:'))
        self.stdout.write(self.style.NOTICE(f'- Total filas procesadas: {stats.filas_leidas}'))
        self.stdout.write(self.style.NOTICE(f'- Filas omitidas por NPR=0: {stats.filas_omitidas_npr_cero}'))
        self.stdout.write(self.style.NOTICE(f'- Filas importables: {stats.filas_importadas}'))
        debug_order = list(NUMERIC_ZERO_ALIASES.keys()) + list(STRUCTURAL_FILL_DOWN_ALIASES.keys())
        for label in debug_order:
            counters = stats.fill_down_stats.get(label)
            if not counters:
                continue
            parts = []
            for key in ['reales', 'heredados', 'cero_por_vacio', 'vacios_finales']:
                if counters.get(key):
                    parts.append(f'{key}={counters[key]}')
            self.stdout.write(self.style.NOTICE(f'- {label}: {", ".join(parts)}'))

    def print_summary(self, service, strategy, path, options, stats):
        self.stdout.write(self.style.NOTICE(f'Servicio detectado: {service}'))
        self.stdout.write(self.style.NOTICE(f'Estrategia usada: {strategy}'))
        self.stdout.write(self.style.NOTICE(f'Archivo: {path}'))
        self.stdout.write(self.style.NOTICE(f'Hoja: {stats.sheet_name or options["sheet"]}'))
        if stats.header_row:
            self.stdout.write(self.style.NOTICE(f'Fila encabezados: {stats.header_row}'))
        if stats.start_row:
            self.stdout.write(self.style.NOTICE(f'Primera fila datos: {stats.start_row}'))
        self.stdout.write(self.style.NOTICE(f'Filas leidas: {stats.filas_leidas}'))
        self.stdout.write(self.style.NOTICE(f'Filas importadas: {stats.filas_importadas}'))
        self.stdout.write(self.style.NOTICE(f'Filas omitidas: {stats.filas_omitidas}'))
        self.stdout.write(self.style.NOTICE(f'Filas omitidas por NPR=0: {stats.filas_omitidas_npr_cero}'))
        self.stdout.write(self.style.NOTICE(f'Equipos creados: {stats.equipos_creados}'))
        self.stdout.write(self.style.NOTICE(f'Equipos reutilizados: {stats.equipos_reutilizados}'))
        self.stdout.write(self.style.NOTICE(f'ServicioEquipo creados: {stats.servicio_equipo_creados}'))
        self.stdout.write(self.style.NOTICE(f'Cargas creadas: {stats.cargas_creadas}'))
        self.stdout.write(self.style.NOTICE(f'RCM creados: {stats.rcm_creados}'))
        self.stdout.write(self.style.NOTICE(f'FMEA creados: {stats.fmea_creados}'))
        self.stdout.write(self.style.NOTICE(f'Evaluaciones creadas: {stats.evaluaciones_creadas}'))
        self.stdout.write(self.style.NOTICE(f'Evaluaciones calculadas: {stats.evaluaciones_calculadas}'))
        self.stdout.write(self.style.NOTICE(f'Dependientes resueltas: {stats.dependientes_resueltas}'))
        self.stdout.write(self.style.NOTICE(f'Dependientes resueltas con minimo por fuente 0: {stats.dependientes_minimo_cero}'))
        self.stdout.write(self.style.NOTICE(f'Tareas creadas: {stats.tareas_creadas}'))
        warning_total = len(stats.warnings) + sum(stats.unresolved_dimensions.values())
        if warning_total:
            self.stdout.write(self.style.WARNING(f'Warnings: {warning_total}'))
            for dimension_name, count in stats.unresolved_dimensions.most_common():
                self.stdout.write(self.style.WARNING(f'- {dimension_name}: no resuelta en {count} filas'))

            detail_limit = None if int(options.get('verbosity', 1)) >= 2 else 50
            detail_warnings = stats.warnings if detail_limit is None else stats.warnings[:detail_limit]
            for warning in detail_warnings:
                self.stdout.write(self.style.WARNING(f'- {warning}'))
            if detail_limit is not None and len(stats.warnings) > detail_limit:
                self.stdout.write(self.style.WARNING(f'- ... {len(stats.warnings) - detail_limit} warnings adicionales. Usa --verbosity 2 para verlos.'))
