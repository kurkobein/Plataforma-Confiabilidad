import re
import unicodedata
from dataclasses import dataclass, field

from django.db.models import Max

from core import models

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - validated at runtime
    load_workbook = None


HEADER_ALIASES = {
    'tag': ['tag', 'tag equipo', 'tag de equipo', 'codigo equipo'],
    'nombre': ['nombre', 'nombre equipo', 'equipo', 'descripcion equipo'],
    'ut': ['ut', 'ubicacion tecnica', 'ubicac tecnica', 'u tecnica', 'ubicacion'],
    'descripcion': ['descripcion', 'descripcion ut', 'descripcion u tecnica', 'descripcion ubicacion tecnica'],
}


@dataclass
class EquipmentImportReport:
    rows_read: int = 0
    created: int = 0
    skipped: int = 0
    missing_required: int = 0
    invalid_ut: int = 0
    duplicate_ut_file: int = 0
    duplicate_ut_db: int = 0
    duplicate_tag_db: int = 0
    materialized_nodes: int = 0
    warnings: list = field(default_factory=list)


def _normalize_key(value):
    value = unicodedata.normalize('NFKD', str(value or ''))
    value = ''.join(char for char in value if not unicodedata.combining(char))
    value = re.sub(r'[^a-zA-Z0-9]+', ' ', value).strip().lower()
    return re.sub(r'\s+', ' ', value)


def _normalize_ut(value):
    return '-'.join(
        models._technical_segment(part)
        for part in str(value or '').strip().split('-')
        if part.strip()
    )


def _build_header_map(row):
    headers = {}
    for index, cell in enumerate(row):
        key = _normalize_key(cell.value)
        if key:
            headers[key] = index
    return headers


def _find_column(header_map, field_key):
    for alias in HEADER_ALIASES[field_key]:
        normalized = _normalize_key(alias)
        if normalized in header_map:
            return header_map[normalized]
    return None


def _find_header(ws):
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=20), start=1):
        header_map = _build_header_map(row)
        if all(_find_column(header_map, key) is not None for key in ('tag', 'nombre', 'ut')):
            return row_number, header_map
    raise ValueError('No se encontraron encabezados validos. Usa columnas TAG, Nombre y UT.')


def _row_value(row, columns, field_key):
    column = columns.get(field_key)
    if column is None or column >= len(row):
        return ''
    value = row[column].value
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _find_simple_value(empresa, level, code):
    value = models.ValorNivelJerarquia.objects.filter(
        empresa=empresa,
        nivel=level,
        codigo__iexact=code,
        activo=True,
    ).first()
    if value:
        return value
    return next(
        (
            item for item in models.ValorNivelJerarquia.objects.filter(
                empresa=empresa,
                nivel=level,
                activo=True,
            )
            if models._technical_segment(item.codigo) == code
        ),
        None,
    )


def _materialize_simple_value(empresa, level, parent, simple_value):
    sibling_order = (
        models.NodoJerarquia.objects.filter(
            empresa=empresa,
            parent=parent,
            nivel=level,
        ).aggregate(max_order=Max('orden')).get('max_order')
        or 0
    ) + 1
    node, created = models.NodoJerarquia.objects.get_or_create(
        empresa=empresa,
        parent=parent,
        codigo=simple_value.codigo,
        defaults={
            'nivel': level,
            'nombre': simple_value.nombre,
            'orden': sibling_order,
            'activo': True,
        },
    )
    if not created:
        changed_fields = []
        if node.nivel_id != level.pk:
            node.nivel = level
            changed_fields.append('nivel')
        if node.nombre != simple_value.nombre:
            node.nombre = simple_value.nombre
            changed_fields.append('nombre')
        if not node.activo:
            node.activo = True
            changed_fields.append('activo')
        if changed_fields:
            node.save(update_fields=changed_fields)
    return node, created


def _resolve_node_from_ut(empresa, ut):
    codes = [code for code in _normalize_ut(ut).split('-') if code]
    levels = list(
        models.NivelJerarquia.objects.filter(
            empresa=empresa,
            activo=True,
        ).order_by('orden')
    )
    if not levels:
        return None, 'La empresa no tiene estructura de ubicacion tecnica configurada.'
    if len(codes) != len(levels):
        return None, f'La UT tiene {len(codes)} componentes y la estructura requiere {len(levels)} niveles.'

    parent = None
    resolved = None
    materialized_count = 0
    for index, (level, code) in enumerate(zip(levels, codes), start=1):
        candidates = models.NodoJerarquia.objects.filter(
            empresa=empresa,
            nivel=level,
            parent=parent,
            activo=True,
        )
        node = next(
            (
                candidate for candidate in candidates
                if models._technical_segment(candidate.codigo) == code
            ),
            None,
        )
        if not node:
            simple_value = _find_simple_value(empresa, level, code)
            if simple_value:
                node, created = _materialize_simple_value(empresa, level, parent, simple_value)
                if created:
                    materialized_count += 1
            else:
                parent_label = f' bajo {parent.codigo} - {parent.nombre}' if parent else ''
                return None, f'No existe "{code}" para el nivel {index} ({level.nombre}){parent_label}.', materialized_count
        if not node:
            parent_label = f' bajo {parent.codigo} - {parent.nombre}' if parent else ''
            return None, f'No se pudo relacionar "{code}" para el nivel {index} ({level.nombre}){parent_label}.', materialized_count
        parent = node
        resolved = node
    return resolved, '', materialized_count


def _append_warning(report, message):
    report.warnings.append(message)


def import_equipment_excel(archivo, empresa):
    if load_workbook is None:
        raise ValueError('openpyxl no esta instalado en el entorno de Django.')
    archivo.seek(0)
    wb = load_workbook(archivo, data_only=True, read_only=True)
    ws = wb.active
    header_row, header_map = _find_header(ws)
    columns = {key: _find_column(header_map, key) for key in HEADER_ALIASES}
    report = EquipmentImportReport()
    seen_ut = set()

    for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        tag = _row_value(row, columns, 'tag')
        nombre = _row_value(row, columns, 'nombre')
        ut_raw = _row_value(row, columns, 'ut')
        descripcion = _row_value(row, columns, 'descripcion')
        if not any((tag, nombre, ut_raw, descripcion)):
            continue
        report.rows_read += 1
        if not tag or not nombre or not ut_raw:
            report.skipped += 1
            report.missing_required += 1
            _append_warning(report, f'Fila {row_number}: faltan TAG, Nombre o UT.')
            continue

        normalized_ut = _normalize_ut(ut_raw)
        if normalized_ut in seen_ut:
            report.skipped += 1
            report.duplicate_ut_file += 1
            _append_warning(report, f'Fila {row_number}: UT repetida dentro del archivo ({normalized_ut}).')
            continue
        seen_ut.add(normalized_ut)

        if models.Equipo.objects.filter(ut__iexact=normalized_ut).exists():
            report.skipped += 1
            report.duplicate_ut_db += 1
            _append_warning(report, f'Fila {row_number}: ya existe un equipo con la UT {normalized_ut}.')
            continue
        if models.Equipo.objects.filter(tag_equipo__iexact=tag).exists():
            report.skipped += 1
            report.duplicate_tag_db += 1
            _append_warning(report, f'Fila {row_number}: ya existe un equipo con TAG {tag}.')
            continue

        node, error, materialized_count = _resolve_node_from_ut(empresa, normalized_ut)
        report.materialized_nodes += materialized_count
        if not node:
            report.skipped += 1
            report.invalid_ut += 1
            _append_warning(report, f'Fila {row_number}: {error}')
            continue

        models.Equipo.objects.create(
            tag_equipo=tag[:100],
            nombre_equipo=nombre[:200],
            ut=node.ut[:200],
            descripcion_ut=descripcion[:255],
            nodo=node,
        )
        report.created += 1

    return report
