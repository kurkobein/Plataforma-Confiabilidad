import json
import re
from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date

from core import models
from core.access import get_accessible_services, get_service_equipment
from rcm.forms import RCMExcelBulkUploadForm, RCMRegistroForm, RCMTaskFormSet
from rcm.field_options import (
    RCM_FIELD_OPTION_LABELS,
    rcm_field_option_key,
)
from rcm.import_excel import Command as RCMExcelImportCommand
from rcm.import_excel import BASE_ALIASES, ImportStats, TASK_ALIASES, clean_text
from rcm.services.progress import (
    build_fmeca_service_progress_summary,
    compatible_hierarchy_node_ids,
    filter_fmeca_by_hierarchy,
    get_fmeca_progress_dimensions,
    get_fmeca_queryset,
    get_hierarchy_filter_options,
    group_fmeca_progress_by_hierarchy_level,
)
from core.services.criticality_rules import (
    sync_fmeca_criticality_rules,
)
from core.views import (
    _decimal_or_none,
    _equipment_items_payload,
    _export_filename,
    _export_pdf_response,
    _export_value,
    _export_xlsx_response,
    _json_loads_safe,
    _json_payload,
    _json_safe,
    _service_equipment_browser_payload,
    _service_equipment_endpoints,
    _service_or_404,
    clear_session_upload,
    open_session_upload,
    store_session_upload,
)


RCM_EXTRA_FIELD_CONFIG = RCM_FIELD_OPTION_LABELS


def _rcm_field_upload_session_key(service_pk, field_name):
    return f'rcm_excel_field_{service_pk}_{field_name}'


def _clear_rcm_field_uploads(request, service_pk):
    for field_name in RCM_EXTRA_FIELD_CONFIG:
        clear_session_upload(
            request,
            _rcm_field_upload_session_key(service_pk, field_name),
        )


def _fallback_excel_header_row(ws, max_rows=30):
    for row_number in range(1, min(ws.max_row, max_rows) + 1):
        if any(clean_text(cell.value) for cell in ws[row_number]):
            return row_number
    return 1


def _rcm_field_preview_payload(uploaded_file, requested_sheet=''):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    uploaded_file.seek(0)
    workbook = load_workbook(BytesIO(uploaded_file.read()), data_only=True, read_only=True)
    if not workbook.sheetnames:
        workbook.close()
        raise ValueError('El archivo no contiene hojas.')
    selected_sheet = requested_sheet if requested_sheet in workbook.sheetnames else workbook.sheetnames[0]
    ws = workbook[selected_sheet]
    importer = RCMExcelImportCommand()
    header_row = importer.detect_header_row(ws) or _fallback_excel_header_row(ws)
    preview_end = min(ws.max_row, header_row + 10)
    columns = []
    for column in range(1, min(ws.max_column, 80) + 1):
        label = clean_text(ws.cell(header_row, column).value)
        has_data = bool(label) or any(
            clean_text(ws.cell(row_number, column).value)
            for row_number in range(header_row + 1, preview_end + 1)
        )
        if has_data:
            columns.append({
                'index': column,
                'letter': get_column_letter(column),
                'label': label or f'Columna {get_column_letter(column)}',
            })
    rows = []
    for row_number in range(header_row + 1, preview_end + 1):
        values = {
            str(column['index']): clean_text(ws.cell(row_number, column['index']).value)
            for column in columns
        }
        if any(values.values()):
            rows.append({'index': row_number, 'values': values})
    result = {
        'sheets': workbook.sheetnames,
        'selected_sheet': selected_sheet,
        'header_row': header_row,
        'columns': columns,
        'rows': rows,
    }
    workbook.close()
    return result


def _rcm_import_field_options_from_stored_excel(
    request,
    service,
    field_name,
    sheet_name,
    header_row,
    column,
):
    from openpyxl import load_workbook

    stored_file, _stored_ref = open_session_upload(
        request,
        _rcm_field_upload_session_key(service.pk, field_name),
    )
    if not stored_file:
        raise ValueError('Vuelve a seleccionar el archivo Excel.')
    try:
        stored_file.seek(0)
        workbook = load_workbook(BytesIO(stored_file.read()), data_only=True, read_only=True)
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError('La hoja seleccionada no existe.')
        ws = workbook[sheet_name]
        header_row = int(header_row or 1)
        column = int(column or 0)
        if column < 1 or column > ws.max_column:
            workbook.close()
            raise ValueError('La columna seleccionada no existe.')

        source_rows = []
        for row_number in range(header_row + 1, ws.max_row + 1):
            value = clean_text(ws.cell(row_number, column).value)
            if not value:
                continue
            source_rows.append({
                'row': row_number,
                'value': value,
                'key': rcm_field_option_key(value),
            })
        workbook.close()
    finally:
        stored_file.close()

    existing_options = {
        option.clave_normalizada: option
        for option in models.RCMCampoOpcion.objects.filter(
            servicio=service,
            campo=field_name,
        )
    }
    seen_in_file = {}
    duplicate_file = []
    duplicate_database = []
    new_options = []
    for source in source_rows:
        key = source['key']
        if not key:
            continue
        if key in seen_in_file:
            duplicate_file.append({
                'row': source['row'],
                'first_row': seen_in_file[key]['row'],
                'value': source['value'],
            })
            continue
        seen_in_file[key] = source
        if key in existing_options:
            duplicate_database.append({
                'row': source['row'],
                'value': source['value'],
                'existing_value': existing_options[key].valor,
            })
            continue
        new_options.append(models.RCMCampoOpcion(
            servicio=service,
            campo=field_name,
            valor=source['value'],
            clave_normalizada=key,
            activo=True,
        ))

    if new_options:
        models.RCMCampoOpcion.objects.bulk_create(
            new_options,
            batch_size=500,
            ignore_conflicts=True,
        )

    duplicate_examples = []
    for item in duplicate_file:
        duplicate_examples.append(
            f'fila {item["row"]}: "{item["value"]}" repite la fila {item["first_row"]}'
        )
    for item in duplicate_database:
        duplicate_examples.append(
            f'fila {item["row"]}: "{item["value"]}" ya existe como "{item["existing_value"]}"'
        )
    duplicate_total = len(duplicate_file) + len(duplicate_database)
    return {
        'created': len(new_options),
        'source_rows': len(source_rows),
        'duplicates': duplicate_total,
        'duplicates_file': len(duplicate_file),
        'duplicates_database': len(duplicate_database),
        'duplicate_examples': duplicate_examples[:8],
    }


@login_required
def rcm_index(request):
    servicios = get_accessible_services(request.user)
    if servicios.count() == 1:
        return redirect('service_fmeca_list', pk=servicios.first().pk)
    messages.info(request, 'Selecciona un servicio para ver su desarrollo FMECA.')
    return redirect('service_list')


def _fmeca_service_rows(request):
    services = list(get_accessible_services(request.user))
    rows = []
    for service in services:
        summary = build_fmeca_service_progress_summary(service)
        progress_percent = summary.get('average_progress_percent')
        rows.append({
            'service': service,
            'summary': summary,
            'progress_width': _progress_width_css(progress_percent),
            'progress_color': _progress_color(progress_percent),
            'progress_label': summary.get('average_progress_label', 'N/A'),
            'progress_order': float(progress_percent) if progress_percent is not None else -1,
        })
    return services, rows


@login_required
def fmeca_development(request):
    _services, rows = _fmeca_service_rows(request)
    return render(request, 'fmeca_development.html', {
        'service_rows': rows,
    })


@login_required
def fmeca_panel(request):
    services, rows = _fmeca_service_rows(request)
    selected_service_id = request.GET.get('service') or ''
    selected_service = None
    context = {
        'services': services,
        'service_rows': rows,
        'selected_service_id': selected_service_id,
    }
    if selected_service_id:
        selected_service = next((service for service in services if str(service.pk) == str(selected_service_id)), None)
        if not selected_service:
            raise PermissionDenied('No tienes acceso al servicio seleccionado.')
        fmeas = get_fmeca_queryset(selected_service)
        hierarchy_filters = get_hierarchy_filter_options(selected_service)
        available_node_ids = {
            int(node['id']) for node in hierarchy_filters.get('nodes', []) if node.get('id')
        }
        selected_panel_node_ids = {
            int(node_id)
            for node_id in request.GET.getlist('panel_nodes')
            if str(node_id).isdigit() and int(node_id) in available_node_ids
        }
        legacy_progress_nodo = request.GET.get('progress_nodo') or ''
        if not selected_panel_node_ids and legacy_progress_nodo:
            selected_panel_node_ids = {
                int(value.strip())
                for value in legacy_progress_nodo.replace(';', ',').split(',')
                if value.strip().isdigit() and int(value.strip()) in available_node_ids
            }
        selected_panel_node_ids = compatible_hierarchy_node_ids(
            hierarchy_filters,
            selected_panel_node_ids,
        )
        levels = hierarchy_filters.get('levels', [])
        level_order = {int(level['id']): index for index, level in enumerate(levels)}
        node_by_id = {
            int(node['id']): node for node in hierarchy_filters.get('nodes', []) if node.get('id')
        }
        selected_nodes = [node_by_id[node_id] for node_id in selected_panel_node_ids if node_id in node_by_id]
        deepest_level_id = None
        if selected_nodes:
            deepest_level_id = max(
                selected_nodes,
                key=lambda node: level_order.get(int(node.get('level_id') or 0), -1),
            ).get('level_id')
        deepest_node_ids = [
            node['id'] for node in selected_nodes if str(node.get('level_id')) == str(deepest_level_id)
        ]
        selected_by_level = {}
        for node in selected_nodes:
            selected_by_level.setdefault(int(node.get('level_id') or 0), set()).add(int(node['id']))
        panel_hierarchy_levels = []
        for level in levels:
            level_id = int(level['id'])
            level_nodes = [
                node for node in hierarchy_filters.get('nodes', [])
                if str(node.get('level_id')) == str(level_id)
            ]
            selected_in_level = selected_by_level.get(level_id, set())
            panel_hierarchy_levels.append({
                **level,
                'nodes': level_nodes,
                'selected_ids': selected_in_level,
            })
        selected_progress_nodo = ','.join(str(node_id) for node_id in deepest_node_ids)
        selected_progress_nivel = str(deepest_level_id or '')
        selected_avance_min, selected_avance_max = _progress_range_from_params(request.GET)
        if selected_progress_nodo:
            fmeas = filter_fmeca_by_hierarchy(fmeas, selected_progress_nodo)
        fmeas = list(fmeas)
        progress_summary = build_fmeca_service_progress_summary(selected_service, fmeas=fmeas)
        progress_by_fmea_id = progress_summary.get('progress_by_fmea_id', {})
        fmeas = _filter_fmeas_by_progress_range(
            fmeas,
            progress_by_fmea_id,
            selected_avance_min,
            selected_avance_max,
        )
        if selected_avance_min != 0 or selected_avance_max != 100:
            progress_summary = build_fmeca_service_progress_summary(selected_service, fmeas=fmeas)
            progress_by_fmea_id = progress_summary.get('progress_by_fmea_id', {})
        area_level = next(
            (
                level for level in levels
                if str(level.get('name') or '').strip().casefold() in {'area', 'área'}
            ),
            None,
        )
        chart_level_id = deepest_level_id or (
            area_level.get('id') if area_level else (levels[0].get('id') if levels else '')
        )
        progress_dimensions = get_fmeca_progress_dimensions(selected_service.estrategia)
        panel_chart_summary = group_fmeca_progress_by_hierarchy_level(
            fmeas,
            chart_level_id,
            progress_dimensions,
        )
        panel_table_level_name = next(
            (
                level.get('name')
                for level in levels
                if str(level.get('id')) == str(chart_level_id)
            ),
            'Nivel',
        )
        table_rows = []
        for fmea in fmeas:
            rcm = fmea.rcm
            equipo = rcm.equipo
            carga = rcm.carga
            progress = progress_by_fmea_id.get(fmea.pk, {})
            progress_percent = progress.get('progress_percent')
            table_rows.append({
                'id': rcm.pk,
                'ut': equipo.ut_display if equipo else '',
                'descripcion_ut': equipo.descripcion_ut if equipo else '',
                'equipo': equipo.nombre_equipo if equipo else '',
                'tag': equipo.tag_display if equipo else '',
                'created': timezone.localtime(carga.creado_en).strftime('%d/%m/%Y %H:%M') if carga and carga.creado_en else '',
                'updated': timezone.localtime(carga.actualizado).strftime('%d/%m/%Y %H:%M') if carga and carga.actualizado else '',
                'date_order': timezone.localtime(carga.actualizado or carga.creado_en).strftime('%Y%m%d%H%M%S') if carga and (carga.actualizado or carga.creado_en) else '',
                'modo_de_falla': rcm.modo_de_falla,
                'progress_percent': progress_percent if progress_percent is not None else '',
                'progress_label': progress.get('progress_label', 'N/A'),
                'progress_order': float(progress_percent) if progress_percent is not None else -1,
                'progress_width': _progress_width_css(progress_percent),
                'progress_color': _progress_color(progress_percent),
            })
        context.update({
            'selected_service': selected_service,
            'progress_summary': progress_summary,
            'hierarchy_filters': hierarchy_filters,
            'panel_hierarchy_levels': panel_hierarchy_levels,
            'selected_progress_nivel': selected_progress_nivel,
            'selected_progress_nodo': selected_progress_nodo,
            'selected_progress_node_ids': [str(node_id) for node_id in selected_panel_node_ids],
            'selected_avance_min': selected_avance_min,
            'selected_avance_max': selected_avance_max,
            'panel_chart_summary': panel_chart_summary,
            'panel_table_level_name': panel_table_level_name,
            'panel_total_filtered': len(fmeas),
            'table_rows': table_rows,
        })
    return render(request, 'fmeca_panel.html', context)


def _safe_task_slug(value, fallback='campo'):
    text = str(value or '').strip().lower()
    text = re.sub(r'[^a-z0-9áéíóúñ]+', '_', text, flags=re.IGNORECASE)
    return text.strip('_')[:100] or fallback


def _task_field_options_json(values):
    if isinstance(values, list):
        cleaned = [str(item).strip() for item in values if str(item).strip()]
        return json.dumps(cleaned, ensure_ascii=False) if cleaned else ''
    text = str(values or '').strip()
    if not text:
        return ''
    cleaned = [item.strip() for item in re.split(r'[\n,;]+', text) if item.strip()]
    return json.dumps(cleaned, ensure_ascii=False) if cleaned else ''


def _task_field_options_list(raw):
    parsed = _json_loads_safe(raw, [])
    if isinstance(parsed, list):
        return [str(item) for item in parsed]
    return []


def _service_family_payload(servicio):
    payload = []
    familias = (
        models.FamiliaEquipo.objects.filter(servicio=servicio, activa=True)
        .prefetch_related('items__equipo')
        .order_by('nombre')
    )
    for familia in familias:
        equipos = [item.equipo for item in familia.items.all()]
        payload.append({
            'id': familia.pk,
            'nombre': familia.nombre,
            'equipos': _equipment_items_payload(equipos),
        })
    return payload


def _progress_color(progress_percent):
    if progress_percent is None:
        return '#94a3b8'
    try:
        percent = max(0, min(100, float(progress_percent)))
    except (TypeError, ValueError):
        percent = 0
    hue = int(percent * 1.2)
    return f'hsl({hue}, 72%, 43%)'


def _progress_width_css(progress_percent):
    if progress_percent is None:
        return '0'
    try:
        percent = max(0, min(100, float(progress_percent)))
    except (TypeError, ValueError):
        percent = 0
    return f'{percent:.1f}'.rstrip('0').rstrip('.')


def _task_config_payload(estrategia):
    tipos = (
        models.TipoTareaEstrategia.objects.filter(estrategia=estrategia, activo=True)
        .prefetch_related('campos')
        .order_by('orden', 'nombre')
    )
    payload = []
    for tipo in tipos:
        payload.append({
            'id': tipo.pk,
            'nombre': tipo.nombre,
            'codigo': tipo.codigo,
            'orden': tipo.orden,
            'activo': tipo.activo,
            'campos': [
                {
                    'id': campo.pk,
                    'nombre': campo.nombre,
                    'clave': campo.clave,
                    'tipo_dato': campo.tipo_dato,
                    'opciones': _task_field_options_list(campo.opciones_json),
                    'obligatorio': campo.obligatorio,
                    'orden': campo.orden,
                    'activo': campo.activo,
                }
                for campo in tipo.campos.all().order_by('orden', 'nombre')
                if campo.activo
            ],
        })
    return payload


def _record_attachment_payload(request):
    files = request.FILES.getlist('adjuntos')
    if not files:
        return [], []
    allowed = set(models.RECORD_ATTACHMENT_EXTENSIONS)
    invalid = []
    payload = []
    for uploaded in files:
        extension = (uploaded.name.rsplit('.', 1)[-1] if '.' in uploaded.name else '').lower()
        if extension not in allowed:
            invalid.append(uploaded.name)
            continue
        payload.append({
            'name': uploaded.name,
            'content': uploaded.read(),
        })
    return payload, invalid


def _save_rcm_attachments(rcm, attachment_payload, usuario):
    now = timezone.now()
    for item in attachment_payload:
        adjunto = models.RCMAdjunto(
            rcm=rcm,
            nombre_original=item['name'],
            creado_en=now,
            usuario=usuario,
        )
        adjunto.archivo.save(item['name'], ContentFile(item['content']), save=False)
        adjunto.save()


def _save_task_config(estrategia, payload):
    payload = payload if isinstance(payload, list) else []
    existing_types = {
        str(item.pk): item
        for item in models.TipoTareaEstrategia.objects.filter(estrategia=estrategia).prefetch_related('campos')
    }
    existing_types_by_code = {
        item.codigo: item
        for item in existing_types.values()
    }
    keep_type_ids = set()
    valid_field_types = dict(models.CampoTareaEstrategia.TIPO_DATO_CHOICES)

    def unique_code(base_code):
        base_code = (base_code or 'tarea')[:80]
        candidate = base_code
        stem = base_code[:74]
        suffix = 2
        while candidate in existing_types_by_code:
            candidate = f'{stem}_{suffix}'[:80]
            suffix += 1
        return candidate

    for type_idx, item in enumerate(payload, start=1):
        if not isinstance(item, dict):
            continue
        nombre = str(item.get('nombre') or '').strip() or f'Tarea {type_idx}'
        requested_code = _safe_task_slug(item.get('codigo') or nombre, f'tarea_{type_idx}')[:80]
        active = item.get('activo', True) is not False
        raw_id = str(item.get('id') or '').strip()

        tipo = existing_types.get(raw_id)
        if not tipo:
            tipo = models.TipoTareaEstrategia(estrategia=estrategia)
            tipo.codigo = unique_code(requested_code)
        tipo.nombre = nombre
        tipo.orden = type_idx
        tipo.activo = active
        tipo.save()
        existing_types_by_code[tipo.codigo] = tipo
        keep_type_ids.add(tipo.pk)

        existing_fields = {
            str(field.pk): field
            for field in tipo.campos.all()
        }
        existing_fields_by_key = {
            field.clave: field
            for field in existing_fields.values()
        }
        keep_field_ids = set()
        fields = item.get('campos') if isinstance(item.get('campos'), list) else []
        for field_idx, field_data in enumerate(fields, start=1):
            if not isinstance(field_data, dict):
                continue
            field_name = str(field_data.get('nombre') or '').strip() or f'Campo {field_idx}'
            field_key = _safe_task_slug(field_data.get('clave') or field_name, f'campo_{field_idx}')
            tipo_dato = str(field_data.get('tipo_dato') or models.CampoTareaEstrategia.TIPO_TEXTO).strip()
            if tipo_dato not in valid_field_types:
                tipo_dato = models.CampoTareaEstrategia.TIPO_TEXTO
            raw_field_id = str(field_data.get('id') or '').strip()
            campo = existing_fields.get(raw_field_id) or existing_fields_by_key.get(field_key)
            if not campo:
                campo = models.CampoTareaEstrategia(tipo_tarea_estrategia=tipo)
            campo.nombre = field_name
            campo.clave = field_key
            campo.tipo_dato = tipo_dato
            campo.opciones_json = _task_field_options_json(field_data.get('opciones'))
            campo.obligatorio = field_data.get('obligatorio', False) is True
            campo.orden = field_idx
            campo.activo = field_data.get('activo', True) is not False
            campo.save()
            keep_field_ids.add(campo.pk)

        tipo.campos.exclude(pk__in=keep_field_ids).update(activo=False)

    models.TipoTareaEstrategia.objects.filter(estrategia=estrategia).exclude(pk__in=keep_type_ids).update(activo=False)


@login_required
def service_rcm_task_config(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=True)
    if not servicio.estrategia_id:
        messages.warning(request, 'El servicio debe tener una estrategia antes de configurar tareas RCM/FMECA.')
        return redirect('service_detail', pk=servicio.pk)

    if request.method == 'POST':
        payload = _json_payload(request, 'payload_json', []) or []
        with transaction.atomic():
            _save_task_config(servicio.estrategia, payload)
        messages.success(request, 'La configuración de tareas RCM/FMECA se guardó correctamente.')
        return redirect('service_fmeca_task_config', pk=servicio.pk)

    return render(request, 'service_rcm_task_config.html', {
        'service': servicio,
        'permission': permission,
        'task_payload_json': json.dumps(_json_safe(_task_config_payload(servicio.estrategia)), ensure_ascii=False),
        'field_types': models.CampoTareaEstrategia.TIPO_DATO_CHOICES,
    })


def _rcm_sheet_score(importer, ws):
    header_row = importer.detect_header_row(ws)
    if not header_row:
        return 0, None
    header_map = importer.build_header_map(ws, header_row)
    score = 0
    critical = [
        ('funcion', 5),
        ('falla_funcional', 6),
        ('modo_de_falla', 6),
        ('efecto', 5),
        ('tag', 3),
        ('ut', 3),
        ('descripcion_equipo', 3),
        ('npr', 3),
    ]
    for key, weight in critical:
        if importer.column_for_aliases(header_map, BASE_ALIASES[key]):
            score += weight
    for task_aliases in TASK_ALIASES.values():
        for aliases in task_aliases.values():
            if importer.column_for_aliases(header_map, aliases):
                score += 1
    sheet_name = clean_text(ws.title).lower()
    if any(term in sheet_name for term in ('rcm', 'fmea', 'fmeca')):
        score += 8
    return score, header_row


def _select_rcm_excel_sheet(importer, workbook, requested_sheet=''):
    requested_sheet = clean_text(requested_sheet)
    if requested_sheet:
        if requested_sheet in workbook.sheetnames:
            ws = workbook[requested_sheet]
            return requested_sheet, ws, importer.detect_header_row(ws)
        lower_map = {name.lower(): name for name in workbook.sheetnames}
        matched = lower_map.get(requested_sheet.lower())
        if matched:
            ws = workbook[matched]
            return matched, ws, importer.detect_header_row(ws)
        raise ValueError(f'El archivo no tiene hoja "{requested_sheet}". Hojas: {", ".join(workbook.sheetnames)}')

    preferred = [name for name in workbook.sheetnames if any(term in name.lower() for term in ('rcm', 'fmea', 'fmeca'))]
    candidates = preferred + [name for name in workbook.sheetnames if name not in preferred]
    best = (0, None, None, None)
    for name in candidates:
        score, header_row = _rcm_sheet_score(importer, workbook[name])
        if score > best[0]:
            best = (score, name, workbook[name], header_row)
    if not best[1]:
        raise ValueError(f'No se detectó una hoja RCM/FMECA compatible. Hojas: {", ".join(workbook.sheetnames)}')
    return best[1], best[2], best[3]


def _rcm_excel_sample_rows(importer, ws, header_map, start_row, limit=10):
    sample = []
    stats = ImportStats()
    last_values = {}
    dims = []
    try:
        dims = importer.get_active_rcm_dimensions(getattr(importer, '_sample_strategy', None)) if getattr(importer, '_sample_strategy', None) else []
    except Exception:
        dims = []
    for row_number in range(start_row, ws.max_row + 1):
        if len(sample) >= limit:
            break
        raw_row = ws[row_number]
        if importer.raw_row_is_empty(raw_row, header_map):
            continue
        row = importer.build_row_context(raw_row, header_map, last_values, stats)
        if importer.row_is_empty(row, header_map):
            continue
        status = 'Listo'
        npr = ''
        if dims:
            try:
                resolved, _eval_stats = importer.resolve_dimensions(row, header_map, dims, stats)
                npr_value = importer.find_resolved_value(resolved, ['npr'])
                if npr_value and npr_value.number is not None:
                    npr = str(npr_value.number)
                    if npr_value.number == Decimal('0'):
                        status = 'Omitida por NPR=0'
            except Exception:
                pass
        tasks_found = 0
        for task_key in ['actual', 'primaria', 'secundaria']:
            if importer.task_payload(task_key, row, header_map).get('descripcion'):
                tasks_found += 1
        sample.append({
            'excel_row': row_number,
            'tag': clean_text(importer.get_cell(row, header_map, BASE_ALIASES['tag'])),
            'ut': clean_text(importer.get_cell(row, header_map, BASE_ALIASES['ut'])),
            'equipo': clean_text(importer.get_cell(row, header_map, BASE_ALIASES['descripcion_equipo'])),
            'componente': clean_text(importer.get_cell(row, header_map, BASE_ALIASES['componente'])),
            'funcion': clean_text(importer.get_cell(row, header_map, BASE_ALIASES['funcion'])),
            'falla_funcional': clean_text(importer.get_cell(row, header_map, BASE_ALIASES['falla_funcional'])),
            'modo': clean_text(importer.get_cell(row, header_map, BASE_ALIASES['modo_de_falla'])),
            'efecto': clean_text(importer.get_cell(row, header_map, BASE_ALIASES['efecto'])),
            'npr': npr or clean_text(importer.get_cell(row, header_map, BASE_ALIASES['npr'])),
            'tasks_found': tasks_found,
            'status': status,
        })
    return sample


def _run_rcm_excel_upload(
    servicio,
    uploaded_file,
    sheet_name='',
    replace=False,
    create_task_types=False,
    preview=True,
):
    from openpyxl import load_workbook

    if not servicio.estrategia_id:
        raise ValueError(f'El servicio {servicio.codigo_servicio} no tiene estrategia asociada.')
    uploaded_file.seek(0)
    workbook = load_workbook(BytesIO(uploaded_file.read()), data_only=True, read_only=True)
    importer = RCMExcelImportCommand()
    resolved_sheet, ws, detected_header_row = _select_rcm_excel_sheet(importer, workbook, sheet_name)
    if not detected_header_row:
        raise ValueError('No se detectó una fila de encabezados compatible con RCM/FMECA.')
    header_map = importer.build_header_map(ws, detected_header_row)
    if not header_map:
        raise ValueError('No se detectaron encabezados válidos.')

    start_row = detected_header_row + 1
    origin = f'RCM Excel: {uploaded_file.name}'
    stats = ImportStats(header_row=detected_header_row, start_row=start_row, sheet_name=resolved_sheet)
    dims = importer.get_active_rcm_dimensions(servicio.estrategia)
    importer._sample_strategy = servicio.estrategia
    sample_rows = _rcm_excel_sample_rows(importer, ws, header_map, start_row)
    task_types = importer.resolve_task_types(servicio.estrategia, create_task_types, stats)
    analysis_date = timezone.localdate()

    with transaction.atomic():
        if replace:
            importer.replace_previous(servicio, origin)

        processed = 0
        last_values = {}
        for row_number in range(start_row, ws.max_row + 1):
            raw_row = ws[row_number]
            if importer.raw_row_is_empty(raw_row, header_map):
                continue
            row = importer.build_row_context(raw_row, header_map, last_values, stats)
            if importer.row_is_empty(row, header_map):
                continue
            stats.filas_leidas += 1
            processed += 1
            try:
                importer.import_row(
                    row=row,
                    row_number=row_number,
                    header_map=header_map,
                    service=servicio,
                    strategy=servicio.estrategia,
                    dims=dims,
                    task_types=task_types,
                    analysis_date=analysis_date,
                    origin=origin,
                    stats=stats,
                )
            except Exception as exc:
                stats.filas_omitidas += 1
                stats.warn(f'Fila {row_number}: omitida por error: {exc}')

        if preview:
            transaction.set_rollback(True)

    warning_total = len(stats.warnings) + sum(stats.unresolved_dimensions.values())
    warnings = list(stats.warnings)
    warnings.extend(
        f'Dimensión "{dimension_name}" no resuelta en {count} fila(s).'
        for dimension_name, count in stats.unresolved_dimensions.most_common()
    )
    workbook.close()
    return {
        'sheet_name': resolved_sheet,
        'origin': origin,
        'header_row': stats.header_row,
        'start_row': stats.start_row,
        'rows_read': stats.filas_leidas,
        'rows_imported': stats.filas_importadas,
        'rows_skipped': stats.filas_omitidas,
        'skipped_npr_zero': stats.filas_omitidas_npr_cero,
        'equipos_creados': stats.equipos_creados,
        'equipos_reutilizados': stats.equipos_reutilizados,
        'servicio_equipo_creados': stats.servicio_equipo_creados,
        'cargas': stats.cargas_creadas,
        'rcm': stats.rcm_creados,
        'fmea': stats.fmea_creados,
        'evaluaciones': stats.evaluaciones_creadas,
        'evaluaciones_calculadas': stats.evaluaciones_calculadas,
        'dependientes': stats.dependientes_resueltas,
        'tareas': stats.tareas_creadas,
        'warning_total': warning_total,
        'warnings': warnings,
        'sample_rows': sample_rows,
    }


@login_required
def service_rcm_excel_field_preview(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=True)
    if not permission.get('can_edit'):
        raise PermissionDenied('No tienes permisos para configurar esta carga.')
    if request.method != 'POST':
        return JsonResponse({'error': 'Metodo no permitido.'}, status=405)

    field_name = clean_text(request.POST.get('field'))
    if field_name not in RCM_EXTRA_FIELD_CONFIG:
        return JsonResponse({'error': 'Campo de destino no valido.'}, status=400)

    uploaded_file = request.FILES.get('archivo')
    session_key = _rcm_field_upload_session_key(servicio.pk, field_name)
    if uploaded_file:
        extension = uploaded_file.name.rsplit('.', 1)[-1].lower() if '.' in uploaded_file.name else ''
        if extension not in {'xlsx', 'xlsm'}:
            return JsonResponse({'error': 'Selecciona un archivo Excel .xlsx o .xlsm.'}, status=400)
        store_session_upload(request, session_key, uploaded_file, 'rcm_fields')

    if request.POST.get('action') == 'confirm':
        try:
            with transaction.atomic():
                result = _rcm_import_field_options_from_stored_excel(
                    request,
                    servicio,
                    field_name,
                    sheet_name=clean_text(request.POST.get('sheet')),
                    header_row=request.POST.get('header_row'),
                    column=request.POST.get('column'),
                )
            clear_session_upload(request, session_key)
            return JsonResponse({
                'ok': True,
                'field': field_name,
                'field_label': RCM_EXTRA_FIELD_CONFIG[field_name],
                'message': (
                    f'{RCM_EXTRA_FIELD_CONFIG[field_name]}: '
                    f'{result["created"]} opciones nuevas cargadas.'
                ),
                **result,
            })
        except Exception as exc:
            return JsonResponse({'error': str(exc)}, status=400)

    stored_file, stored_ref = open_session_upload(request, session_key)
    if not stored_file:
        return JsonResponse({'error': 'Selecciona un archivo Excel para continuar.'}, status=400)
    try:
        payload = _rcm_field_preview_payload(
            stored_file,
            requested_sheet=clean_text(request.POST.get('sheet')),
        )
        payload['file_name'] = (stored_ref or {}).get('name') or ''
        payload['field'] = field_name
        payload['field_label'] = RCM_EXTRA_FIELD_CONFIG[field_name]
        return JsonResponse(payload)
    except Exception as exc:
        return JsonResponse({'error': f'No se pudo leer el Excel: {exc}'}, status=400)
    finally:
        stored_file.close()


@login_required
def service_rcm_excel_upload(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=True)
    if not permission.get('can_edit'):
        raise PermissionDenied('No tienes permisos para cargar registros RCM/FMECA en este servicio.')

    report = None
    preview_only = True
    upload_session_key = f'rcm_excel_upload_file_{servicio.pk}'
    if request.method == 'POST':
        form = RCMExcelBulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.cleaned_data.get('archivo')
            if uploaded_file:
                store_session_upload(request, upload_session_key, uploaded_file, 'rcm')
            stored_file, _stored_ref = open_session_upload(request, upload_session_key)
            if not stored_file:
                form.add_error('archivo', 'Selecciona un archivo Excel para continuar.')
            else:
                form.cleaned_data['archivo'] = stored_file
        if form.is_valid():
            preview_only = request.POST.get('action') != 'confirm'
            try:
                report = _run_rcm_excel_upload(
                    servicio,
                    form.cleaned_data['archivo'],
                    sheet_name=form.cleaned_data.get('hoja') or '',
                    replace=form.cleaned_data.get('replace'),
                    create_task_types=form.cleaned_data.get('create_task_types'),
                    preview=preview_only,
                )
                if preview_only:
                    messages.info(request, 'Previsualización RCM/FMECA lista. Puedes confirmar sin volver a seleccionar el archivo.')
                else:
                    try:
                        form.cleaned_data['archivo'].close()
                    except Exception:
                        pass
                    clear_session_upload(request, upload_session_key)
                    messages.success(request, f'Carga RCM/FMECA completada: {report["rcm"]} registros creados.')
            except Exception as exc:
                messages.error(request, str(exc))
            finally:
                try:
                    form.cleaned_data['archivo'].close()
                except Exception:
                    pass
    else:
        clear_session_upload(request, upload_session_key)
        _clear_rcm_field_uploads(request, servicio.pk)
        form = RCMExcelBulkUploadForm()
    stored_upload = request.session.get(upload_session_key)

    return render(request, 'service_rcm_excel_upload.html', {
        'service': servicio,
        'permission': permission,
        'form': form,
        'report': report,
        'preview_only': preview_only,
        'stored_upload': stored_upload,
    })


@login_required
def service_rcm_excel_template(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=False)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    headers = [
        'Equipo / Componente',
        'TAG',
        'Ubicación Técnica',
        'Componente',
        'Función',
        'Falla Funcional',
        'Modo de Falla',
        'Efecto/Consecuencia de la Falla',
        'OP',
        'SLD',
        'MA',
        'I&R',
        'Severidad',
        'Ocurrencia',
        'Detección',
        'NPR',
        'Criticidad',
        'Rango NPR',
        'Actividad de Mantenimiento Primaria',
        'Táctica FTM-CBM-RTF-FFM',
        'ITEMS / COMPONENTE',
        'Limite Aceptable',
        'Pto. Trabajo',
        'Frecuencia (s)',
        'Cant. Pers ejecutante',
        'Dur (min)',
        'Dur (hr)',
        'HH',
        'Estado Equipo',
        'Actividad de Mantenimiento Secundaria',
        'Componente Involucrado',
        'N° Parte / N° Ítem',
        'N° SAP',
        'Obs',
    ]
    rows = [
        [
            'Bomba de pulpa 1',
            'A001',
            'MIN-CH-MOL-PULP-BOM-A001',
            'Impulsor',
            'Transportar pulpa hacia clasificación',
            'Pérdida de caudal',
            'Desgaste abrasivo del impulsor',
            'Pérdida de eficiencia y riesgo de detención del circuito',
            4, 3, 2, 1, '', 3, 4, '', 'Alta', '',
            'Inspeccionar desgaste de impulsor y holguras',
            'CBM',
            'Impulsor bomba',
            'Holgura dentro del rango definido',
            'Mecánico',
            'Mensual',
            1,
            60,
            1,
            1,
            'Operando',
            'Revisar disponibilidad de repuesto crítico',
            'Impulsor',
            'REP-001',
            '30001234',
            'Ejemplo de tarea secundaria',
        ],
        [
            'Motor ventilador',
            'A002',
            'MIN-CH-SER-VEN-EXT-A002',
            'Rodamientos',
            'Entregar ventilación al área',
            'No entrega flujo de aire requerido',
            'Falla de rodamiento por lubricación deficiente',
            'Aumento de temperatura y detención del ventilador',
            3, 2, 2, 1, '', 2, 5, '', 'Media', '',
            'Medir vibración y temperatura de rodamientos',
            'CBM',
            'Rodamiento motor',
            'Vibración bajo límite definido',
            'Mecánico',
            'Quincenal',
            1,
            45,
            0.75,
            0.75,
            'Operando',
            '',
            '',
            '',
            '',
            '',
        ],
    ]

    workbook = Workbook()
    ws = workbook.active
    ws.title = 'RCM-FMEA'
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill('solid', fgColor='EAF1FF')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[letter].width = min(max(max_length + 2, 14), 42)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f'plantilla_rcm_fmeca_{servicio.codigo_servicio}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def service_rcm_list(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=False)
    fmeas = get_fmeca_queryset(servicio)
    selected_progress_nivel = request.GET.get('progress_nivel') or ''
    selected_progress_nodo = request.GET.get('progress_nodo') or ''
    selected_avance_min, selected_avance_max = _progress_range_from_params(request.GET)
    selected_progress_node_ids = [
        value.strip()
        for value in selected_progress_nodo.replace(';', ',').split(',')
        if value.strip().isdigit()
    ]
    if selected_progress_nodo:
        fmeas = filter_fmeca_by_hierarchy(fmeas, selected_progress_nodo)
    fmeas = list(fmeas)
    progress_summary = build_fmeca_service_progress_summary(servicio, fmeas=fmeas)
    progress_by_fmea_id = progress_summary.get('progress_by_fmea_id', {})
    fmeas = _filter_fmeas_by_progress_range(
        fmeas,
        progress_by_fmea_id,
        selected_avance_min,
        selected_avance_max,
    )
    if selected_avance_min != 0 or selected_avance_max != 100:
        progress_summary = build_fmeca_service_progress_summary(servicio, fmeas=fmeas)
        progress_by_fmea_id = progress_summary.get('progress_by_fmea_id', {})
    rows, rcm_count, fmea_count, fmeca_count = _service_rcm_rows(servicio, fmeas=fmeas)
    for row in rows:
        progress = progress_by_fmea_id.get(row.get('fmea').pk if row.get('fmea') else None, {})
        progress_percent = progress.get('progress_percent')
        row['avance_fmeca'] = progress.get('progress_label', 'N/A')
        row['avance_fmeca_order'] = float(progress_percent) if progress_percent is not None else -1
        row['avance_fmeca_width'] = _progress_width_css(progress_percent)
        row['avance_fmeca_color'] = _progress_color(progress_percent)
        row['avance_fmeca_missing_label'] = ', '.join(
            item.dimension.nombre for item in progress.get('missing_dimensions', [])
        ) or 'Sin faltantes'

    hierarchy_filters = get_hierarchy_filter_options(servicio)
    evaluation_columns = []
    if servicio.estrategia_id:
        evaluation_columns = list(
            models.EstrategiaDimension.objects.filter(
                estrategia=servicio.estrategia,
                activo=True,
                visible_en_listado_fmeca=True,
                proceso_uso__in=[
                    models.EstrategiaDimension.PROCESO_FMECA,
                    models.EstrategiaDimension.PROCESO_AMBOS,
                ],
            )
            .select_related('dimension')
            .order_by('orden', 'id')
        )

    return render(request, 'service_rcm_list.html', {
        'service': servicio,
        'permission': permission,
        'rows': rows,
        'rcm_count': rcm_count,
        'fmea_count': fmea_count,
        'fmeca_count': fmeca_count,
        'progress_summary': progress_summary,
        'hierarchy_filters': hierarchy_filters,
        'selected_progress_nivel': selected_progress_nivel,
        'selected_progress_nodo': selected_progress_nodo,
        'selected_progress_node_ids': selected_progress_node_ids,
        'selected_avance_min': selected_avance_min,
        'selected_avance_max': selected_avance_max,
        'evaluation_columns': evaluation_columns,
    })


def _progress_range_from_params(params):
    def parse(value, fallback):
        try:
            number = int(str(value).strip())
        except (TypeError, ValueError):
            return fallback
        return max(0, min(100, number))

    min_value = parse(params.get('avance_min'), 0)
    max_value = parse(params.get('avance_max'), 100)
    if min_value > max_value:
        min_value, max_value = max_value, min_value
    return min_value, max_value


def _filter_fmeas_by_progress_range(fmeas, progress_by_fmea_id, min_value, max_value):
    if min_value <= 0 and max_value >= 100:
        return list(fmeas)
    filtered = []
    for fmea in fmeas:
        progress = progress_by_fmea_id.get(fmea.pk, {})
        progress_percent = progress.get('progress_percent')
        if progress_percent is None:
            continue
        try:
            value = float(progress_percent)
        except (TypeError, ValueError):
            continue
        if min_value <= value <= max_value:
            filtered.append(fmea)
    return filtered


def _service_rcm_rows(servicio, fmeas=None):
    if fmeas is None:
        fmeas = list(get_fmeca_queryset(servicio).order_by('-rcm__fecha_analisis', '-rcm__id'))
    rows = []
    for fmea in fmeas:
        registro = fmea.rcm
        criticality_trace = _json_loads_safe(getattr(registro, 'trazabilidad_criticidad_json', ''), {})
        evaluations = []
        tasks = []
        if fmea:
            evaluations = list(
                fmea.evaluaciones.select_related('estrategia_dimension__dimension').order_by(
                    'estrategia_dimension__orden',
                    'estrategia_dimension_id',
                )
            )
            tasks = list(
                fmea.tareas_rcm.select_related(
                    'tipo_tarea_estrategia',
                )
                .prefetch_related('valores_campos')
                .order_by('orden', 'id')
            )
            tasks = [task for task in tasks if _task_has_meaningful_content(task)]
        rows.append({
            'registro': registro,
            'fmea': fmea,
            'evaluations': evaluations,
            'evaluations_by_ed': {evaluation.estrategia_dimension_id: evaluation for evaluation in evaluations},
            'tasks': tasks,
            'criticality_trace': criticality_trace,
        })

    rcm_count = len(rows)
    fmea_count = sum(1 for row in rows if not row['registro'].criticidad)
    fmeca_count = sum(1 for row in rows if row['registro'].criticidad)
    return (
        rows,
        rcm_count,
        fmea_count,
        fmeca_count,
    )


def _service_rcm_export_data(servicio):
    rows, _rcm_count, _fmea_count, _fmeca_count = _service_rcm_rows(servicio)
    columns = [
        ('id', 'ID'),
        ('fecha_creacion', 'Fecha creación'),
        ('fecha_modificacion', 'Fecha modificación'),
        ('tipo', 'Tipo'),
        ('estado', 'Estado'),
        ('ubicacion_tecnica', 'U.T.'),
        ('equipo', 'Equipo'),
        ('tag', 'TAG'),
        ('componente', 'Componente'),
        ('funcion', 'Función'),
        ('falla_funcional', 'Falla funcional'),
        ('modo_de_falla', 'Modo de falla'),
        ('efecto', 'Efecto'),
        ('observacion', 'Observación'),
        ('criticidad', 'Criticidad'),
        ('evaluacion', 'Evaluación'),
        ('tareas', 'Tareas'),
    ]
    records = []
    for item in rows:
        registro = item['registro']
        evaluations = [
            f'{evaluation.estrategia_dimension.dimension.nombre}: {_export_value(evaluation.valor_display) or "-"}'
            for evaluation in item['evaluations']
        ]
        tasks = [
            f'{task.tipo_tarea_estrategia.nombre}: {task.descripcion}'
            for task in item['tasks']
        ]
        records.append({
            'id': registro.pk,
            'fecha_creacion': registro.carga.creado_en if registro.carga_id else '',
            'fecha_modificacion': registro.carga.actualizado if registro.carga_id else '',
            'tipo': registro.tipo_analisis,
            'estado': registro.get_estado_display(),
            'ubicacion_tecnica': registro.equipo.ut_display if registro.equipo else '',
            'equipo': registro.equipo.nombre_equipo if registro.equipo else '',
            'tag': registro.equipo.tag_display if registro.equipo else '',
            'componente': registro.componente or '',
            'funcion': registro.funcion or '',
            'falla_funcional': registro.falla_funcional,
            'modo_de_falla': registro.modo_de_falla,
            'efecto': registro.efecto,
            'observacion': registro.observacion,
            'criticidad': registro.criticidad,
            'evaluacion': '\n'.join(evaluations),
            'tareas': '\n'.join(tasks),
        })
    return columns, records


@login_required
def service_rcm_export(request, pk, formato):
    servicio, _permission = _service_or_404(request, pk, edit=False)
    columns, rows = _service_rcm_export_data(servicio)
    formato = (formato or '').lower()
    if formato == 'excel':
        return _export_xlsx_response(
            _export_filename('RCM', servicio, 'xlsx'),
            'Desarrollo FMECA',
            columns,
            rows,
        )
    if formato == 'pdf':
        return _export_pdf_response(
            _export_filename('RCM', servicio, 'pdf'),
            f'Desarrollo FMECA - {servicio.codigo_servicio}',
            columns,
            rows,
        )
    raise Http404('Formato de exportación no soportado.')


def _format_rcm_timestamp(value):
    if not value:
        return ''
    if hasattr(value, 'hour'):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime('%d/%m/%Y %H:%M')
    if hasattr(value, 'strftime'):
        return value.strftime('%d/%m/%Y')
    return str(value)


def _rcm_date_meta(rcm=None):
    now = timezone.localtime(timezone.now())
    if rcm and rcm.carga_id:
        return {
            'creado_en': _format_rcm_timestamp(rcm.carga.creado_en),
            'actualizado': _format_rcm_timestamp(rcm.carga.actualizado),
        }
    return {
        'creado_en': _format_rcm_timestamp(now),
        'actualizado': _format_rcm_timestamp(now),
    }


def _rcm_form_initial_from_instance(rcm):
    return {
        'equipo': rcm.equipo_id,
        'fecha_analisis': rcm.fecha_analisis,
        'estado': rcm.estado,
        'criticidad': rcm.criticidad,
        'componente': rcm.componente or '',
        'funcion': rcm.funcion or '',
        'falla_funcional': rcm.falla_funcional,
        'modo_de_falla': rcm.modo_de_falla,
        'efecto': rcm.efecto,
        'observacion': rcm.observacion,
    }


def _rcm_task_initials(rcm):
    try:
        fmea = rcm.fmea_fmeca
    except models.FMEA_FMECA.DoesNotExist:
        return [{}]
    tasks = list(
        fmea.tareas_rcm.select_related(
            'tipo_tarea_estrategia',
        )
        .prefetch_related('valores_campos')
        .order_by('orden', 'id')
    )
    tasks = [task for task in tasks if _task_has_meaningful_content(task)]
    initials = []
    for task in tasks:
        dynamic_values = {
            value.campo.clave: str(_json_safe(value.valor_display))
            for value in task.valores_campos.select_related('campo').all()
            if value.valor_display not in (None, '')
        }
        initials.append({
            'id': task.pk,
            'valores_json': json.dumps(dynamic_values, ensure_ascii=False),
            'tipo_tarea_estrategia': task.tipo_tarea_estrategia_id,
            'descripcion': task.descripcion,
            'tactica': task.tactica,
            'limite_aceptable': task.limite_aceptable,
            'parametros': task.parametros,
            'riesgo_material': task.riesgo_material,
            'especialidad': task.especialidad,
            'puesto_trabajo': task.puesto_trabajo,
            'estado_equipo': task.estado_equipo,
            'frecuencia_valor': task.frecuencia_valor,
            'frecuencia_unidad': task.frecuencia_unidad,
            'frecuencia_texto': task.frecuencia_texto,
            'duracion_min': task.duracion_min,
            'duracion_hr': task.duracion_hr,
            'cantidad_personas': task.cantidad_personas,
            'hh': task.hh,
            'plan_sap': task.plan_sap,
            'descripcion_plan': task.descripcion_plan,
            'hoja_ruta': task.hoja_ruta,
            'texto_hoja_ruta': task.texto_hoja_ruta,
            'operacion_hoja_ruta': task.operacion_hoja_ruta,
            'texto_operacion': task.texto_operacion,
            'operacion_pauta': task.operacion_pauta,
            'pauta': task.pauta,
            'titulo_pauta': task.titulo_pauta,
            'repuesto': task.repuesto,
            'componente_involucrado': task.componente_involucrado,
            'numero_parte': task.numero_parte,
            'numero_sap': task.numero_sap,
            'procedimiento_trabajo': task.procedimiento_trabajo,
            'costo_hh': task.costo_hh,
            'costo_repuestos': task.costo_repuestos,
            'tarifa_servicios': task.tarifa_servicios,
            'costo_total': task.costo_total,
            'oportunidad_mejora': task.oportunidad_mejora,
            'estado': task.estado,
        })
    return initials


def _rcm_task_formset(request, servicio, rcm=None):
    initial = _rcm_task_initials(rcm) if rcm else []
    return RCMTaskFormSet(
        request.POST or None,
        prefix='tasks',
        initial=initial,
        form_kwargs={'estrategia': servicio.estrategia},
    )


_TASK_TEXT_FIELDS = {
    'descripcion',
    'tactica',
    'limite_aceptable',
    'parametros',
    'riesgo_material',
    'especialidad',
    'puesto_trabajo',
    'estado_equipo',
    'frecuencia_unidad',
    'frecuencia_texto',
    'plan_sap',
    'descripcion_plan',
    'hoja_ruta',
    'texto_hoja_ruta',
    'operacion_hoja_ruta',
    'texto_operacion',
    'operacion_pauta',
    'pauta',
    'titulo_pauta',
    'repuesto',
    'componente_involucrado',
    'numero_parte',
    'numero_sap',
    'procedimiento_trabajo',
    'oportunidad_mejora',
}
_TASK_DECIMAL_FIELDS = {
    'frecuencia_valor',
    'duracion_min',
    'duracion_hr',
    'cantidad_personas',
    'hh',
    'costo_hh',
    'costo_repuestos',
    'tarifa_servicios',
    'costo_total',
}


def _task_has_meaningful_content(task):
    dynamic_values = list(task.valores_campos.all()) if hasattr(task, '_prefetched_objects_cache') else list(task.valores_campos.all())
    for value in dynamic_values:
        if value.valor_display not in (None, ''):
            return True

    type_name = (task.tipo_tarea_estrategia.nombre if task.tipo_tarea_estrategia_id else '') or ''
    description = (task.descripcion or '').strip()
    if description and description != type_name:
        return True

    for name in _TASK_TEXT_FIELDS - {'descripcion'}:
        if str(getattr(task, name, '') or '').strip():
            return True
    for name in _TASK_DECIMAL_FIELDS:
        if getattr(task, name, None) is not None:
            return True
    return False


def _task_config_for_type(tipo_tarea):
    if not tipo_tarea:
        return []
    return list(tipo_tarea.campos.filter(activo=True).order_by('orden', 'nombre'))


def _task_display_from_dynamic(fields, values):
    for field in fields:
        value = values.get(field.clave)
        if value not in (None, '', []):
            return str(value)
    return ''


def _apply_task_dynamic_values_to_fixed_fields(task, fields, values, cleaned):
    for name in _TASK_TEXT_FIELDS:
        setattr(task, name, cleaned.get(name) or '')
    for name in _TASK_DECIMAL_FIELDS:
        setattr(task, name, cleaned.get(name))

    for field in fields:
        target = field.clave
        if target not in _TASK_TEXT_FIELDS and target not in _TASK_DECIMAL_FIELDS:
            continue
        raw = values.get(field.clave)
        if raw in (None, '', []):
            continue
        if target in _TASK_DECIMAL_FIELDS:
            parsed = _decimal_or_none(raw)
            if parsed is not None:
                setattr(task, target, parsed)
        else:
            setattr(task, target, str(raw))

    if not task.descripcion:
        task.descripcion = _task_display_from_dynamic(fields, values) or cleaned['tipo_tarea_estrategia'].nombre


def _save_task_dynamic_values(task, fields, values):
    keep_ids = []
    for field in fields:
        raw = values.get(field.clave)
        if raw in (None, '', []):
            continue
        defaults = {
            'valor_texto': '',
            'valor_numero': None,
            'valor_booleano': None,
            'valor_fecha': None,
        }
        if field.tipo_dato in {
            models.CampoTareaEstrategia.TIPO_NUMERO,
            models.CampoTareaEstrategia.TIPO_DECIMAL,
        }:
            defaults['valor_numero'] = _decimal_or_none(raw)
            if defaults['valor_numero'] is None:
                defaults['valor_texto'] = str(raw)
        elif field.tipo_dato == models.CampoTareaEstrategia.TIPO_BOOLEANO:
            defaults['valor_booleano'] = str(raw).lower() in {'1', 'true', 'si', 'sí', 'on', 'yes'}
        elif field.tipo_dato == models.CampoTareaEstrategia.TIPO_FECHA:
            defaults['valor_fecha'] = parse_date(str(raw))
            if defaults['valor_fecha'] is None:
                defaults['valor_texto'] = str(raw)
        else:
            defaults['valor_texto'] = str(raw)

        value_obj, _ = models.ValorCampoTareaRCM.objects.update_or_create(
            tarea=task,
            campo=field,
            defaults=defaults,
        )
        keep_ids.append(value_obj.pk)

    task.valores_campos.exclude(pk__in=keep_ids).delete()


def _save_rcm_tasks(fmea, task_formset):
    saved_ids = []
    existing_tasks = {
        item.pk: item
        for item in fmea.tareas_rcm.all()
    }

    active_forms = []
    for form_idx, task_form in enumerate(task_formset.forms, start=1):
        if not hasattr(task_form, 'cleaned_data'):
            continue
        cleaned = task_form.cleaned_data
        task_id = cleaned.get('id')
        if cleaned.get('DELETE'):
            if task_id and task_id in existing_tasks:
                models.ValorCampoTareaRCM.objects.filter(tarea=existing_tasks[task_id]).delete()
                existing_tasks[task_id].delete()
            continue
        if getattr(task_form, 'empty_task', False):
            continue
        sort_order = cleaned['tipo_tarea_estrategia'].orden or form_idx
        active_forms.append((sort_order, form_idx, task_form))

    active_forms.sort(key=lambda item: (item[0], item[1]))
    for order, (_sort_order, _form_idx, task_form) in enumerate(active_forms, start=1):
        cleaned = task_form.cleaned_data
        task_id = cleaned.get('id')
        configured_fields = _task_config_for_type(cleaned['tipo_tarea_estrategia'])
        dynamic_values = cleaned.get('dynamic_values') or {}
        task = existing_tasks.get(task_id) if task_id else models.TareaRCM(fmea=fmea)
        task.fmea = fmea
        task.tipo_tarea_estrategia = cleaned['tipo_tarea_estrategia']
        _apply_task_dynamic_values_to_fixed_fields(task, configured_fields, dynamic_values, cleaned)
        task.orden = order
        task.estado = cleaned.get('estado') or models.TareaRCM.ESTADO_ACTIVO
        task.save()
        _save_task_dynamic_values(task, configured_fields, dynamic_values)
        saved_ids.append(task.pk)

    stale_tasks = fmea.tareas_rcm.exclude(pk__in=saved_ids)
    models.ValorCampoTareaRCM.objects.filter(tarea__in=stale_tasks).delete()
    stale_tasks.delete()


def _save_rcm_form(
    servicio,
    permission,
    form,
    task_formset=None,
    rcm=None,
    equipo=None,
    attachment_payload=None,
    origin='RCM Manual',
    created_at=None,
):
    cleaned = form.cleaned_data
    now = timezone.now()
    created_at = created_at or now
    selected_equipment = equipo or cleaned['equipo']
    attachment_payload = attachment_payload or []

    if rcm:
        carga = rcm.carga
        carga.fecha_analisis = cleaned['fecha_analisis']
        carga.status = cleaned['estado']
        carga.actualizado = now
        carga.estrategia = servicio.estrategia
        carga.servicio = servicio
        if permission.get('profile'):
            carga.usuario = permission.get('profile')
        carga.save(update_fields=['fecha_analisis', 'status', 'actualizado', 'estrategia', 'servicio', 'usuario'])

        rcm.equipo = selected_equipment
        rcm.criticidad = cleaned.get('criticidad')
        rcm.fecha_analisis = cleaned['fecha_analisis']
        rcm.estado = cleaned['estado']
        rcm.componente = cleaned.get('componente') or ''
        rcm.funcion = cleaned.get('funcion') or ''
        rcm.falla_funcional = cleaned['falla_funcional']
        rcm.modo_de_falla = cleaned['modo_de_falla']
        rcm.causa = ''
        rcm.efecto = cleaned['efecto']
        rcm.observacion = cleaned.get('observacion') or ''
        rcm.save(update_fields=[
            'equipo',
            'criticidad',
            'fecha_analisis',
            'estado',
            'componente',
            'funcion',
            'falla_funcional',
            'modo_de_falla',
            'causa',
            'efecto',
            'observacion',
        ])
    else:
        carga = models.Carga.objects.create(
            fecha_analisis=cleaned['fecha_analisis'],
            version_carga=Decimal('1.0'),
            usuario=permission.get('profile'),
            servicio=servicio,
            estrategia=servicio.estrategia,
            origen=origin,
            status=cleaned['estado'],
            creado_en=created_at,
            actualizado=now,
        )
        rcm = models.RCM.objects.create(
            carga=carga,
            equipo=selected_equipment,
            criticidad=cleaned.get('criticidad'),
            fecha_analisis=cleaned['fecha_analisis'],
            estado=cleaned['estado'],
            componente=cleaned.get('componente') or '',
            funcion=cleaned.get('funcion') or '',
            falla_funcional=cleaned['falla_funcional'],
            modo_de_falla=cleaned['modo_de_falla'],
            causa='',
            efecto=cleaned['efecto'],
            observacion=cleaned.get('observacion') or '',
        )

    fmea, _created = models.FMEA_FMECA.objects.get_or_create(rcm=rcm)

    saved_dimension_ids = []
    for evaluation in getattr(form, 'cleaned_dimension_evaluations', []):
        estrategia_dimension = evaluation['estrategia_dimension']
        saved_dimension_ids.append(estrategia_dimension.pk)
        models.EvaluacionFMEA.objects.update_or_create(
            fmea=fmea,
            estrategia_dimension=estrategia_dimension,
            defaults={
                'valor_numerico': evaluation.get('valor_numerico'),
                'valor_texto': evaluation.get('valor_texto') or '',
                'catalogo_fila': evaluation.get('catalogo_fila'),
                'escala_valor': evaluation.get('escala_valor'),
            },
        )
    models.EvaluacionFMEA.objects.filter(fmea=fmea).exclude(
        estrategia_dimension_id__in=saved_dimension_ids,
    ).delete()
    sync_fmeca_criticality_rules(rcm, fmea, servicio)
    if task_formset is not None:
        _save_rcm_tasks(fmea, task_formset)
    _save_rcm_attachments(rcm, attachment_payload, permission.get('profile'))
    return rcm


def _fmeca_bulk_dimension_payload(service):
    metadata_form = RCMRegistroForm(service=service)
    runtime_by_field = {
        item['field_name']: item
        for item in metadata_form.dimension_runtime_payload
    }
    dimensions = []
    for item in metadata_form.impact_dimensions:
        runtime = dict(runtime_by_field.get(item['field_name'], {}))
        runtime.update({
            'field_name': item['field_name'],
            'nombre': item['dimension'].nombre,
            'required': item['estrategia_dimension'].obligatorio,
            'is_calculated': item['is_calculated'],
            'options': [
                {
                    'value': row['pk'],
                    'label': ' | '.join(
                        str(value) for value in row.get('cells', [])
                        if value not in (None, '')
                    ) or f'Opcion {index}',
                    'numeric': row.get('value_numeric', ''),
                }
                for index, row in enumerate(item.get('option_rows') or [], start=1)
            ],
        })
        dimensions.append(runtime)

    field_options = {}
    for field_name in ('falla_funcional', 'modo_de_falla', 'efecto'):
        field_options[field_name] = [
            {'value': value, 'label': label}
            for value, label in metadata_form.fields[field_name].choices
            if value
        ]
    return dimensions, field_options


def _fmeca_bulk_row_is_empty(row):
    if not isinstance(row, dict):
        return True
    if row.get('equipo_id') or row.get('family_id'):
        return False
    for field_name in (
        'componente',
        'funcion',
        'falla_funcional',
        'modo_de_falla',
        'efecto',
        'criticidad',
        'observacion',
    ):
        if str(row.get(field_name) or '').strip():
            return False
    dimensions = row.get('dimensions') if isinstance(row.get('dimensions'), dict) else {}
    return not any(value not in (None, '') for value in dimensions.values())


def _fmeca_bulk_attachment_payload(request, row_id):
    files = request.FILES.getlist(f'row_files_{row_id}')
    if not files:
        return [], []
    allowed = set(models.RECORD_ATTACHMENT_EXTENSIONS)
    payload = []
    invalid = []
    for uploaded in files:
        extension = (uploaded.name.rsplit('.', 1)[-1] if '.' in uploaded.name else '').lower()
        if extension not in allowed:
            invalid.append(uploaded.name)
            continue
        payload.append({
            'name': uploaded.name,
            'content': uploaded.read(),
        })
    return payload, invalid


def _fmeca_bulk_context(
    service,
    permission,
    *,
    dimension_payload,
    field_options,
    errors=None,
    raw_payload='',
    selected_date=None,
):
    initial_equipment = list(
        get_service_equipment(service)
        .order_by('tag_equipo', 'nombre_equipo', 'ut')[:50]
    )
    return {
        'service': service,
        'permission': permission,
        'dimension_payload': dimension_payload,
        'field_options': field_options,
        'service_equipment_payload': _service_equipment_browser_payload(service),
        'service_equipment_endpoints': _service_equipment_endpoints(service),
        'initial_equipment_payload': _equipment_items_payload(initial_equipment),
        'service_family_payload': _service_family_payload(service),
        'bulk_errors': errors or [],
        'bulk_payload': raw_payload or json.dumps({'rows': []}),
        'selected_date': selected_date or timezone.localdate(),
    }


@login_required
def service_fmeca_bulk_new(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=True)
    if not servicio.estrategia_id:
        messages.warning(
            request,
            'El servicio debe tener una estrategia antes de registrar FMECA.',
        )
        return redirect('service_detail', pk=servicio.pk)

    dimension_payload, field_options = _fmeca_bulk_dimension_payload(servicio)
    raw_payload = request.POST.get('bulk_payload', '') if request.method == 'POST' else ''
    selected_date = parse_date(request.POST.get('fecha_analisis', '')) or timezone.localdate()

    if request.method == 'POST':
        errors = []
        try:
            payload = json.loads(raw_payload or '{}')
        except (TypeError, json.JSONDecodeError):
            payload = {}
            errors.append('No se pudo leer la carga grupal.')
        submitted_rows = payload.get('rows') if isinstance(payload, dict) else []
        if not isinstance(submitted_rows, list):
            submitted_rows = []

        is_draft = request.POST.get('save_as') == 'draft'
        status = models.Carga.STATUS_INCOMPLETO if is_draft else models.Carga.STATUS_COMPLETO
        equipment_qs = get_service_equipment(servicio)
        families = (
            models.FamiliaEquipo.objects.filter(servicio=servicio, activa=True)
            .prefetch_related('items__equipo')
        )
        prepared_rows = []

        for index, row in enumerate(submitted_rows, start=1):
            if _fmeca_bulk_row_is_empty(row):
                continue
            row_errors = []
            target_type = row.get('target_type') if row.get('target_type') in {'equipo', 'familia'} else 'equipo'
            target_equipment = []
            if target_type == 'familia':
                family = families.filter(pk=row.get('family_id')).first() if row.get('family_id') else None
                if family:
                    target_equipment = [
                        item.equipo
                        for item in family.items.all()
                        if item.equipo_id and item.equipo
                    ]
                    if not target_equipment:
                        row_errors.append('La familia seleccionada no tiene equipos.')
                else:
                    row_errors.append('Selecciona una familia valida.')
            else:
                equipment = equipment_qs.filter(pk=row.get('equipo_id')).first() if row.get('equipo_id') else None
                if equipment:
                    target_equipment = [equipment]
                else:
                    row_errors.append('Selecciona un equipo valido.')

            row_id = row.get('row_id') if str(row.get('row_id') or '').isdigit() else index
            attachments, invalid_files = _fmeca_bulk_attachment_payload(request, row_id)
            if invalid_files:
                row_errors.append('Archivos no permitidos: ' + ', '.join(invalid_files) + '.')

            form_data = {
                'equipo': target_equipment[0].pk if target_equipment else '',
                'familia_equipo': '',
                'fecha_analisis': selected_date.isoformat(),
                'estado': status,
                'criticidad': row.get('criticidad', ''),
                'componente': row.get('componente', ''),
                'funcion': row.get('funcion', ''),
                'falla_funcional': row.get('falla_funcional', ''),
                'modo_de_falla': row.get('modo_de_falla', ''),
                'efecto': row.get('efecto', ''),
                'observacion': row.get('observacion', ''),
            }
            row_dimensions = row.get('dimensions') if isinstance(row.get('dimensions'), dict) else {}
            for dimension in dimension_payload:
                form_data[dimension['field_name']] = row_dimensions.get(dimension['field_name'], '')

            row_form = RCMRegistroForm(
                form_data,
                service=servicio,
                allow_incomplete=is_draft,
            )
            if not row_form.is_valid():
                for field_name, field_errors in row_form.errors.items():
                    field = row_form.fields.get(field_name)
                    label = field.label if field else 'Fila'
                    for error in field_errors:
                        row_errors.append(f'{label}: {error}')
            if row_errors:
                errors.append(f'Fila {index}: ' + ' '.join(row_errors))
                continue
            prepared_rows.append({
                'form': row_form,
                'equipment': target_equipment,
                'attachments': attachments,
            })

        if not prepared_rows and not errors:
            errors.append('No hay filas válidas para guardar.')
        if errors:
            return render(
                request,
                'fmeca_bulk_form.html',
                _fmeca_bulk_context(
                    servicio,
                    permission,
                    dimension_payload=dimension_payload,
                    field_options=field_options,
                    errors=errors,
                    raw_payload=raw_payload,
                    selected_date=selected_date,
                ),
            )

        created_count = 0
        group_created_at = timezone.now()
        with transaction.atomic():
            for prepared in prepared_rows:
                for equipment in prepared['equipment']:
                    _save_rcm_form(
                        servicio,
                        permission,
                        prepared['form'],
                        equipo=equipment,
                        attachment_payload=prepared['attachments'],
                        origin='FMECA grupal',
                        created_at=group_created_at,
                    )
                    created_count += 1
        messages.success(request, f'Se crearon {created_count} registros FMECA correctamente.')
        return redirect('service_fmeca_list', pk=servicio.pk)

    return render(
        request,
        'fmeca_bulk_form.html',
        _fmeca_bulk_context(
            servicio,
            permission,
            dimension_payload=dimension_payload,
            field_options=field_options,
            raw_payload=raw_payload,
            selected_date=selected_date,
        ),
    )


def _service_aca_criticality_payload(servicio):
    def key_segment(value):
        return (str(value or '').strip().split('-')[-1] or '').strip().upper()

    items = []
    records = (
        models.Criticidad.objects.filter(
            aca_carga__servicio=servicio,
            equipo_id__isnull=False,
        )
        .select_related('equipo', 'aca_carga')
        .order_by('equipo__tag_equipo', '-aca_carga__fecha_analisis', '-id')
    )
    for record in records:
        numeric_value = record.valor_criticidad_equipo
        if numeric_value is None:
            continue
        try:
            fmeca_value = int(Decimal(numeric_value).quantize(Decimal('1'), rounding=ROUND_HALF_UP))
        except (TypeError, ValueError):
            continue
        equipo = record.equipo
        fecha = record.aca_carga.fecha_analisis if record.aca_carga_id else None
        label_parts = [
            f'ACA #{record.pk}',
            f'Valor {numeric_value}',
        ]
        if record.criticidad_final:
            label_parts.append(record.criticidad_final)
        if fecha:
            label_parts.append(fecha.strftime('%d/%m/%Y'))
        items.append({
            'id': record.pk,
            'equipo_id': record.equipo_id,
            'tag': getattr(equipo, 'tag_equipo', '') if equipo else '',
            'ut': getattr(equipo, 'ut', '') if equipo else '',
            'ut_segment': key_segment(getattr(equipo, 'ut', '') if equipo else ''),
            'value': fmeca_value,
            'numeric_value': str(numeric_value),
            'classification': record.criticidad_final or '',
            'label': ' | '.join(label_parts),
            'equipment_label': ' | '.join(
                item for item in [
                    getattr(equipo, 'ut_display', '') if equipo else '',
                    getattr(equipo, 'tag_display', '') if equipo else '',
                    getattr(equipo, 'nombre_equipo', '') if equipo else '',
                ]
                if item
            ),
        })
    return items


@login_required
def service_rcm_new(request, pk):
    servicio, permission = _service_or_404(request, pk, edit=True)
    form = RCMRegistroForm(
        request.POST or None,
        service=servicio,
        initial={
            'fecha_analisis': timezone.localdate(),
            'estado': models.Carga.STATUS_COMPLETO,
        },
    )
    task_formset = _rcm_task_formset(request, servicio)

    if request.method == 'POST' and form.is_valid() and task_formset.is_valid():
        attachment_payload, invalid_attachments = _record_attachment_payload(request)
        if invalid_attachments:
            form.add_error(
                None,
                'Formato de archivo no permitido: '
                + ', '.join(invalid_attachments)
                + '. Usa PDF, Word, Excel, PowerPoint, CSV, TXT, imagen o ZIP.',
            )
        else:
            with transaction.atomic():
                familia = form.cleaned_data.get('familia_equipo')
                if familia:
                    equipos = [item.equipo for item in familia.items.select_related('equipo').order_by('orden', 'id')]
                    for equipo in equipos:
                        _save_rcm_form(
                            servicio,
                            permission,
                            form,
                            task_formset=task_formset,
                            equipo=equipo,
                            attachment_payload=attachment_payload,
                        )
                    messages.success(request, f'Se crearon {len(equipos)} registros RCM/FMECA para la familia {familia.nombre}.')
                else:
                    _save_rcm_form(servicio, permission, form, task_formset=task_formset, attachment_payload=attachment_payload)
                    messages.success(request, 'Registro RCM/FMECA creado correctamente.')
            return redirect('service_fmeca_list', pk=servicio.pk)

    return render(request, 'service_rcm_form.html', {
        'service': servicio,
        'permission': permission,
        'form': form,
        'task_formset': task_formset,
        'task_types_available': models.TipoTareaEstrategia.objects.filter(estrategia=servicio.estrategia, activo=True).exists(),
        'task_field_config_payload': json.dumps(_json_safe(_task_config_payload(servicio.estrategia)), ensure_ascii=False) if servicio.estrategia_id else '[]',
        'editing_rcm': None,
        'rcm_date_meta': _rcm_date_meta(),
        'form_title': 'Nuevo registro FMECA',
        'submit_label': 'Guardar registro FMECA',
        'service_equipment_payload': _service_equipment_browser_payload(servicio),
        'service_equipment_endpoints': _service_equipment_endpoints(servicio),
        'service_family_payload': _service_family_payload(servicio),
        'service_aca_criticality_payload': _service_aca_criticality_payload(servicio),
        'existing_attachments': [],
    })


@login_required
def service_rcm_edit(request, service_pk, rcm_pk):
    servicio, permission = _service_or_404(request, service_pk, edit=True)
    rcm = get_object_or_404(
        models.RCM.objects.select_related('carga', 'equipo', 'fmea_fmeca').prefetch_related('fmea_fmeca__evaluaciones'),
        pk=rcm_pk,
        carga__servicio=servicio,
    )
    form = RCMRegistroForm(
        request.POST or None,
        service=servicio,
        rcm=rcm,
        initial=_rcm_form_initial_from_instance(rcm),
    )
    task_formset = _rcm_task_formset(request, servicio, rcm=rcm)

    if request.method == 'POST' and form.is_valid() and task_formset.is_valid():
        attachment_payload, invalid_attachments = _record_attachment_payload(request)
        if invalid_attachments:
            form.add_error(
                None,
                'Formato de archivo no permitido: '
                + ', '.join(invalid_attachments)
                + '. Usa PDF, Word, Excel, PowerPoint, CSV, TXT, imagen o ZIP.',
            )
        else:
            with transaction.atomic():
                _save_rcm_form(
                    servicio,
                    permission,
                    form,
                    task_formset=task_formset,
                    rcm=rcm,
                    attachment_payload=attachment_payload,
                )
            messages.success(request, 'Registro RCM/FMECA actualizado correctamente.')
            return redirect('service_fmeca_list', pk=servicio.pk)

    return render(request, 'service_rcm_form.html', {
        'service': servicio,
        'permission': permission,
        'form': form,
        'task_formset': task_formset,
        'task_types_available': models.TipoTareaEstrategia.objects.filter(estrategia=servicio.estrategia, activo=True).exists(),
        'task_field_config_payload': json.dumps(_json_safe(_task_config_payload(servicio.estrategia)), ensure_ascii=False) if servicio.estrategia_id else '[]',
        'editing_rcm': rcm,
        'rcm_date_meta': _rcm_date_meta(rcm),
        'form_title': 'Editar registro RCM',
        'submit_label': 'Actualizar registro RCM',
        'service_equipment_payload': _service_equipment_browser_payload(servicio),
        'service_equipment_endpoints': _service_equipment_endpoints(servicio),
        'service_family_payload': _service_family_payload(servicio),
        'service_aca_criticality_payload': _service_aca_criticality_payload(servicio),
        'existing_attachments': rcm.adjuntos.all(),
    })


@login_required
@transaction.atomic
def service_rcm_delete(request, service_pk, rcm_pk):
    servicio, permission = _service_or_404(request, service_pk, edit=True)
    rcm = get_object_or_404(
        models.RCM.objects.select_related('carga'),
        pk=rcm_pk,
        carga__servicio=servicio,
    )

    if request.method == 'POST':
        carga = rcm.carga
        models.ValorCampoTareaRCM.objects.filter(tarea__fmea__rcm=rcm).delete()
        models.TareaRCM.objects.filter(fmea__rcm=rcm).delete()
        models.EvaluacionFMEA.objects.filter(fmea__rcm=rcm).delete()
        models.FMEA_FMECA.objects.filter(rcm=rcm).delete()
        rcm.delete()
        if carga and not carga.criticidades.exists() and not models.RCM.objects.filter(carga=carga).exists():
            carga.delete()
        messages.success(request, 'Registro RCM/FMECA eliminado correctamente.')
        return redirect('service_fmeca_list', pk=servicio.pk)

    return redirect('service_fmeca_list', pk=servicio.pk)
