from __future__ import annotations

from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from core.models import (
    AcaCarga,
    Cargo,
    Componente,
    ComponenteEquipo,
    Criticidad,
    CriticidadDimension,
    Dimension,
    Empresa,
    Equipo,
    Estrategia,
    EstrategiaDimension,
    Metodologia,
    Servicio,
    ServicioEquipo,
    Sistema,
    Usuario,
)

HEADER_MAP = {
    "CLIENTE": "servicio_codigo",
    "Ubicación Técnica": "ubicacion_tecnica",
    "Sistema": "sistema",
    "Descripción  U.Técnica": "descripcion_ut",
    "Equipo": "equipo_tipo",
    "TAG": "tag",
    "División": "division",
    "MTTR Equipo (Hr)": "mttr_equipo_hr",
    "MTTR Equipo en producto final (hr)": "mttr_equipo_prod_hr",
    "Pérdida Total (unidad)": "perdida_total_unidad",
    "Pérdida Total \n(unidad)": "perdida_total_unidad",
    "Impacta Produccion": "impacta_produccion",
    "Producción Fuera de Norma": "produccion_fuera_norma",
    "Impacta Calidad": "impacto_calidad_flag",
    "Escenario de Falla": "escenario_falla",
    "Frecuencia Falla": "frecuencia_falla",
    "Frecuencia Normalizada": "frecuencia_normalizada",
    "Impacto Seguridad": "impacto_seguridad",
    "Impacto Ambiental": "impacto_ambiental",
    "Impacto Operación": "impacto_operacion",
    "Impacto Entorno": "impacto_entorno",
    "Impacto Calidad": "impacto_calidad",
    "Impacto Costo": "impacto_costo",
    "Valor Consecuencia Total": "valor_consecuencia_total",
    "Indicador de Criticidad [ C - P ]": "indicador_criticidad",
    "Indicador de Criticidad\n[ C - P ]": "indicador_criticidad",
    "Valor Criticidad Equipo": "valor_criticidad_equipo",
    "Criticidad Final": "criticidad_final",
    "Estructuras Soportantes": "estructuras_soportantes",
    "Distribucion Electrica": "distribucion_electrica",
    "Legislación": "legislacion",
    "Asset Integrity": "asset_integrity",
    "Stand by": "stand_by",
    "Asistencia Técnica Crítica": "asistencia_tecnica_critica",
    "Repuesto Crítico / Estratégico": "repuesto_critico_estrategico",
}

SERVICE_COMPANY = {
    "PUC-011": "Pucobre",
    "CDC-072": "Codelco",
    "CDC-0072": "Codelco",
    "ERB-001": "ENAP",
}

SERVICE_DESCRIPTIONS = {
    "PUC-011": "ACA Pucobre",
    "CDC-072": "ACA Codelco Chuquicamata",
    "CDC-0072": "ACA Codelco Chuquicamata",
    "ERB-001": "ACA ENAP Refinería Bío Bío",
}

DIM_SPECS = {
    "MTTR Equipo (Hr)": ("atributo", "numerico", "mttr_equipo_hr"),
    "MTTR Equipo en producto final (hr)": ("atributo", "numerico", "mttr_equipo_prod_hr"),
    "Pérdida Total (unidad)": ("atributo", "numerico", "perdida_total_unidad"),
    "Impacta Produccion": ("atributo", "booleano", "impacta_produccion"),
    "Producción Fuera de Norma": ("atributo", "numerico", "produccion_fuera_norma"),
    "Impacta Calidad": ("atributo", "booleano", "impacto_calidad_flag"),
    "Frecuencia Falla": ("probabilidad", "numerico", "frecuencia_falla"),
    "Frecuencia Normalizada": ("probabilidad", "numerico", "frecuencia_normalizada"),
    "Impacto Seguridad": ("impacto", "numerico", "impacto_seguridad"),
    "Impacto Ambiental": ("impacto", "numerico", "impacto_ambiental"),
    "Impacto Operación": ("impacto", "numerico", "impacto_operacion"),
    "Impacto Entorno": ("impacto", "numerico", "impacto_entorno"),
    "Impacto Calidad": ("impacto", "numerico", "impacto_calidad"),
    "Impacto Costo": ("impacto", "numerico", "impacto_costo"),
    "Valor Consecuencia Total": ("resultado", "numerico", "valor_consecuencia_total"),
    "Indicador de Criticidad [ C - P ]": ("resultado", "texto", "indicador_criticidad"),
    "Valor Criticidad Equipo": ("resultado", "numerico", "valor_criticidad_equipo"),
    "Criticidad Final": ("resultado", "texto", "criticidad_final"),
    "Estructuras Soportantes": ("atributo", "booleano", "estructuras_soportantes"),
    "Distribucion Electrica": ("atributo", "booleano", "distribucion_electrica"),
    "Legislación": ("impacto", "numerico", "legislacion"),
    "Asset Integrity": ("atributo", "booleano", "asset_integrity"),
    "Stand by": ("atributo", "booleano", "stand_by"),
    "Asistencia Técnica Crítica": ("atributo", "booleano", "asistencia_tecnica_critica"),
    "Repuesto Crítico / Estratégico": ("atributo", "booleano", "repuesto_critico_estrategico"),
}

BOOLEAN_STRINGS = {
    "SI": True,
    "SÍ": True,
    "TRUE": True,
    "1": True,
    "X": True,
    "NO": False,
    "FALSE": False,
    "0": False,
}


def clean_header(value):
    return str(value or "").replace("\n", " ").replace("  ", " ").strip()


def qtext(value):
    if value is None:
        return ""
    return str(value).strip()


def d(value):
    if value in (None, "", "None"):
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def parse_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    return BOOLEAN_STRINGS.get(str(value).strip().upper())


def slug(text):
    return qtext(text).lower().replace(" ", "_").replace("/", "_").replace("-", "_")[:100] or "sin_codigo"


class Command(BaseCommand):
    help = "Carga Base_Completa.xlsx al modelo core. Crea una Criticidad por cada fila del Excel."

    def add_arguments(self, parser):
        parser.add_argument("--base", default="Base_Completa.xlsx", help="Ruta al archivo Base_Completa.xlsx")
        parser.add_argument("--replace", action="store_true", help="Elimina cargas previas del mismo origen y vuelve a importar")
        parser.add_argument("--dry-run", action="store_true", help="Prueba sin guardar cambios")

    def handle(self, *args, **options):
        base_path = Path(options["base"])
        if not base_path.exists():
            raise CommandError(f"No existe el archivo base: {base_path}")

        wb = load_workbook(base_path, data_only=True, read_only=True)
        if "ACA" not in wb.sheetnames:
            raise CommandError("El archivo no tiene hoja 'ACA'")

        with transaction.atomic():
            ctx = self.bootstrap()
            rows = self.read_base_rows(wb["ACA"])
            read_counts = Counter(r["servicio_codigo"] for r in rows)
            self.stdout.write(self.style.NOTICE(f"Filas leídas: {len(rows)} | {dict(read_counts)}"))

            dimensiones = self.ensure_dimensions(ctx["metodologia"])
            estrategias, servicios = self.ensure_strategies_and_services(ctx)

            if options["replace"]:
                self.delete_previous(base_path.name)

            inserted_counts = self.import_rows(rows, ctx, servicios, estrategias, dimensiones, base_path.name)
            self.stdout.write(self.style.SUCCESS(f"Criticidades creadas: {sum(inserted_counts.values())} | {dict(inserted_counts)}"))

            if options["dry_run"]:
                raise CommandError("Dry run solicitado: rollback ejecutado")

        self.stdout.write(self.style.SUCCESS("Carga completada correctamente."))

    def bootstrap(self):
        empresa_mindco, _ = Empresa.objects.get_or_create(nombre="Mindco", defaults={"sigla": "MINDCO", "estado": "Activo"})
        empresas = {"Mindco": empresa_mindco}
        for nombre in {v for v in SERVICE_COMPANY.values()}:
            empresas[nombre], _ = Empresa.objects.get_or_create(nombre=nombre, defaults={"sigla": nombre.upper()[:20], "estado": "Activo"})

        cargo, _ = Cargo.objects.get_or_create(
            nombre_cargo="Ingeniero de Confiabilidad",
            defaults={"area": "Reliability", "jefatura": "Confiabilidad"},
        )
        usuario, _ = Usuario.all_objects.get_or_create(
            correo_corporativo="importador.aca@mindco.cl",
            defaults={"nombre_completo": "Importador ACA", "empresa": empresa_mindco, "cargo": cargo},
        )
        metodologia, _ = Metodologia.objects.get_or_create(
            abreviatura="ACA",
            defaults={"nombre": "Análisis de Criticidad de Activos", "descripcion": "Carga automática desde Base_Completa.xlsx"},
        )
        return {"empresas": empresas, "usuario": usuario, "metodologia": metodologia}

    def ensure_dimensions(self, metodologia):
        # El modelo actual Dimension no tiene FK a Metodologia, por eso se crean por nombre global.
        dimensiones = {}
        for name, (tipo_funcional, tipo_dato, _field) in DIM_SPECS.items():
            dim, _ = Dimension.objects.get_or_create(
                nombre=name,
                defaults={"descripcion": name, "tipo_funcional": tipo_funcional, "tipo_dato": tipo_dato},
            )
            changed = False
            if dim.tipo_funcional != tipo_funcional:
                dim.tipo_funcional = tipo_funcional
                changed = True
            if dim.tipo_dato != tipo_dato:
                dim.tipo_dato = tipo_dato
                changed = True
            if changed:
                dim.save(update_fields=["tipo_funcional", "tipo_dato"])
            dimensiones[name] = dim
        return dimensiones

    def ensure_strategies_and_services(self, ctx):
        estrategias = {}
        servicios = {}
        for code, company_name in SERVICE_COMPANY.items():
            empresa = ctx["empresas"][company_name]
            estrategia, _ = Estrategia.objects.get_or_create(
                empresa=empresa,
                nombre=f"Estrategia {code}",
                defaults={"version": Decimal("1.0"), "descripcion": f"Estrategia ACA importada para {code}", "activa": True},
            )
            estrategias[code] = estrategia
            servicio, _ = Servicio.objects.get_or_create(
                codigo_servicio=code,
                defaults={
                    "descripcion": SERVICE_DESCRIPTIONS.get(code, code),
                    "empresa": empresa,
                    "estrategia": estrategia,
                    "responsable_usuario": ctx["usuario"],
                    "creado_por_usuario": ctx["usuario"],
                    "fecha_inicio": timezone.localdate(),
                    "status": "activo",
                    "creado_en": timezone.now(),
                },
            )
            updates = {}
            if servicio.empresa_id != empresa.id:
                updates["empresa"] = empresa
            if servicio.estrategia_id != estrategia.id:
                updates["estrategia"] = estrategia
            if updates:
                for k, v in updates.items():
                    setattr(servicio, k, v)
                servicio.save()
            servicios[code] = servicio
        # alias por si el Excel viene con CDC-0072 pero el servicio real esperado es CDC-072
        if "CDC-072" in servicios:
            servicios["CDC-0072"] = servicios["CDC-072"]
            estrategias["CDC-0072"] = estrategias["CDC-072"]
        return estrategias, servicios

    def ensure_strategy_dimension(self, estrategia, dimension, orden):
        obj, created = EstrategiaDimension.objects.get_or_create(
            estrategia=estrategia,
            dimension=dimension,
            defaults={"orden": orden, "obligatorio": False, "activo": True},
        )
        if created:
            return obj
        updates = {}
        if not obj.orden:
            updates["orden"] = orden
        if hasattr(obj, "activo") and not obj.activo:
            updates["activo"] = True
        if updates:
            for k, v in updates.items():
                setattr(obj, k, v)
            obj.save()
        return obj

    def delete_previous(self, base_filename):
        origenes = [f"{base_filename}:{svc}" for svc in ("PUC-011", "CDC-072", "ERB-001")]
        cargas = AcaCarga.objects.filter(origen__in=origenes)
        criticidades = Criticidad.objects.filter(aca_carga__in=cargas)
        dim_deleted = CriticidadDimension.objects.filter(criticidad__in=criticidades).delete()[0]
        crit_deleted = criticidades.delete()[0]
        carga_deleted = cargas.delete()[0]
        self.stdout.write(self.style.WARNING(
            f"Eliminado previo: cargas={carga_deleted}, criticidades={crit_deleted}, dimensiones={dim_deleted}"
        ))

    def read_base_rows(self, ws):
        iterator = ws.iter_rows(values_only=True)
        raw_headers = next(iterator)
        headers = [clean_header(h) for h in raw_headers]
        header_keys = [HEADER_MAP.get(h, slug(h)) for h in headers]
        rows = []
        for raw_row in iterator:
            payload = {}
            empty = True
            for key, value in zip(header_keys, raw_row):
                if value not in (None, ""):
                    empty = False
                payload[key] = value
            if empty:
                continue
            svc = qtext(payload.get("servicio_codigo"))
            if not svc:
                continue
            if svc == "CDC-0072":
                svc = "CDC-072"
            payload["servicio_codigo"] = svc
            rows.append(payload)
        return rows

    def get_or_create_sistema(self, empresa, name):
        nombre = qtext(name) or f"{empresa.nombre} · General"
        codigo = slug(nombre)
        obj, _ = Sistema.objects.get_or_create(
            empresa=empresa,
            nombre_sistema=nombre[:200],
            defaults={"codigo_sistema": codigo},
        )
        return obj

    def get_or_create_component(self, name):
        nombre = qtext(name) or "SIN_COMPONENTE"
        obj, _ = Componente.objects.get_or_create(nombre=nombre[:200], defaults={"descripcion": nombre})
        return obj

    def get_or_create_equipo(self, empresa, row):
        sistema = self.get_or_create_sistema(empresa, row.get("sistema"))
        tag = qtext(row.get("tag")) or qtext(row.get("ubicacion_tecnica")) or "SIN_TAG"
        nombre = qtext(row.get("descripcion_ut")) or qtext(row.get("equipo_tipo")) or tag
        obj, _ = Equipo.objects.get_or_create(
            sistema=sistema,
            tag_equipo=tag[:100],
            defaults={
                "nombre_equipo": nombre[:200],
                "ut": qtext(row.get("ubicacion_tecnica"))[:200],
                "descripcion_ut": qtext(row.get("descripcion_ut"))[:255],
                "otros_posibles": f"División: {qtext(row.get('division'))}" if qtext(row.get("division")) else "",
            },
        )
        componente = self.get_or_create_component(row.get("equipo_tipo"))
        ComponenteEquipo.objects.get_or_create(equipo=obj, componente=componente)
        return obj

    def import_rows(self, rows, ctx, servicios, estrategias, dimensiones, base_filename):
        cargas = {}
        for svc_code in ("PUC-011", "CDC-072", "ERB-001"):
            servicio = servicios[svc_code]
            carga = AcaCarga.objects.create(
                servicio=servicio,
                estrategia=estrategias[svc_code],
                origen=f"{base_filename}:{svc_code}",
                fecha_analisis=timezone.localdate(),
                version_carga=Decimal("1.0"),
                usuario=ctx["usuario"],
                creado_en=timezone.now(),
                actualizado=timezone.now(),
            )
            cargas[svc_code] = carga

        estrategia_dimension_cache = {}
        inserted = Counter()

        for idx, row in enumerate(rows, start=1):
            svc_code = row["servicio_codigo"]
            servicio = servicios.get(svc_code)
            if not servicio:
                continue

            equipo = self.get_or_create_equipo(servicio.empresa, row)
            ServicioEquipo.objects.get_or_create(servicio=servicio, equipo=equipo)

            # CLAVE: create(), no get_or_create(). Cada fila Excel debe generar una Criticidad.
            criticidad = Criticidad.objects.create(
                aca_carga=cargas[svc_code],
                equipo=equipo,
                escenario_falla=qtext(row.get("escenario_falla")),
                frecuencia_original=d(row.get("frecuencia_falla")),
                frecuencia_normalizada=d(row.get("frecuencia_normalizada")),
                valor_cons_total=d(row.get("valor_consecuencia_total")),
                indicador_criticidad=qtext(row.get("indicador_criticidad"))[:100],
                valor_criticidad_equipo=d(row.get("valor_criticidad_equipo")),
                criticidad_final=qtext(row.get("criticidad_final"))[:30],
                creado_en=timezone.now(),
            )
            inserted[svc_code] += 1

            for orden, (dim_name, (_tipo_funcional, tipo_dato, source_key)) in enumerate(DIM_SPECS.items(), start=1):
                raw = row.get(source_key)
                if raw in (None, ""):
                    continue
                dim = dimensiones[dim_name]
                cache_key = (servicio.estrategia_id, dim.id)
                estrategia_dimension = estrategia_dimension_cache.get(cache_key)
                if estrategia_dimension is None:
                    estrategia_dimension = self.ensure_strategy_dimension(servicio.estrategia, dim, orden)
                    estrategia_dimension_cache[cache_key] = estrategia_dimension

                cd = CriticidadDimension(
                    criticidad=criticidad,
                    dimension=dim,
                    estrategia_dimension=estrategia_dimension,
                    valor_texto="",
                )
                booleano = parse_bool(raw)
                numero = d(raw)
                if tipo_dato == "booleano" and booleano is not None:
                    cd.valor_booleano = booleano
                    cd.valor_texto = "Sí" if booleano else "No"
                elif numero is not None:
                    cd.valor_numerico = numero
                else:
                    cd.valor_texto = qtext(raw)
                cd.save()

            if idx % 500 == 0:
                self.stdout.write(f"Procesadas {idx} filas...")

        return inserted
