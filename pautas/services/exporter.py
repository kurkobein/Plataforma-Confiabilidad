from __future__ import annotations

from copy import copy
from decimal import Decimal
from io import BytesIO

from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from core import models


PAUTA_FIELDS = [
    'codigo',
    'nombre',
    'area',
    'ubicacion_tecnica',
    'frecuencia',
    'especialidad',
    'estado_equipo',
    'estrategia_mantenimiento',
    'cantidad_personas',
    'duracion_horas',
    'hh_total',
    'estado',
    'origen',
]

TAREA_FIELDS = [
    'orden',
    'componente',
    'actividad',
    'limite_aceptable',
    'observacion',
    'tipo_tarea',
    'frecuencia',
    'pto_trabajo',
    'cantidad_personas',
    'duracion_horas',
    'hh',
    'estado_equipo',
]

SYSTEM_FIELDS = [
    'correlativo',
    'fecha_generacion',
]

RCM_FIELDS = [
    'equipo',
    'tag',
    'ubicacion_tecnica',
    'descripcion_ut',
    'fecha_analisis',
    'estado',
    'criticidad',
    'falla_funcional',
    'modo_de_falla',
    'causa',
    'efecto',
]


class PautaExportError(Exception):
    pass


def _load_workbook_from_template(plantilla):
    if not plantilla or not plantilla.archivo:
        raise PautaExportError('La pauta no tiene una plantilla Excel asociada.')
    try:
        return load_workbook(plantilla.archivo.path)
    except Exception as exc:
        raise PautaExportError(f'No se pudo abrir la plantilla Excel: {exc}') from exc


def listar_hojas_plantilla(plantilla):
    wb = _load_workbook_from_template(plantilla)
    return wb.sheetnames


def listar_celdas_usadas(plantilla, hoja):
    wb = _load_workbook_from_template(plantilla)
    if hoja not in wb.sheetnames:
        raise PautaExportError(f'La hoja "{hoja}" no existe en la plantilla.')
    ws = wb[hoja]
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ''):
                cells.append({
                    'celda': cell.coordinate,
                    'valor': str(cell.value),
                })
    return cells[:500]


def _color_to_hex(color_obj):
    color = getattr(color_obj, 'rgb', None)
    if color is not None:
        color = str(color)
    if color and len(color) == 8 and color != '00000000':
        return f'#{color[-6:]}'
    return ''


def _cell_fill_hex(cell):
    fill = cell.fill
    if not fill or fill.fill_type in (None, 'none'):
        return ''
    return _color_to_hex(fill.fgColor)


def _cell_font_hex(cell):
    return _color_to_hex(cell.font.color) if cell.font and cell.font.color else ''


def _border_style(value):
    return '1px solid #cbd5e1' if value else ''


def _cell_border_css(cell):
    border = cell.border
    if not border:
        return ''
    rules = []
    for side_name, css_name in [
        ('left', 'border-left'),
        ('right', 'border-right'),
        ('top', 'border-top'),
        ('bottom', 'border-bottom'),
    ]:
        side = getattr(border, side_name, None)
        style = getattr(side, 'style', None)
        if style:
            rules.append(f'{css_name}: {_border_style(style)};')
    return ' '.join(rules)


def _column_width_px(ws, col_index):
    letter = get_column_letter(col_index)
    width = ws.column_dimensions[letter].width
    if width is None:
        width = 8.43
    return max(int(float(width) * 7 + 12), 42)


def _row_height_px(ws, row_index):
    height = ws.row_dimensions[row_index].height
    if height is None:
        height = 18
    return max(int(float(height) * 1.35), 24)


def _merge_lookup(ws):
    lookup = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        master = f'{get_column_letter(min_col)}{min_row}'
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                lookup[(row, col)] = {
                    'range': str(merged_range),
                    'master': master,
                    'is_master': row == min_row and col == min_col,
                    'rowspan': max_row - min_row + 1,
                    'colspan': max_col - min_col + 1,
                }
    return lookup


def listar_grid_plantilla(plantilla, hoja, max_rows=80, max_cols=30):
    wb = _load_workbook_from_template(plantilla)
    if hoja not in wb.sheetnames:
        raise PautaExportError(f'La hoja "{hoja}" no existe en la plantilla.')
    ws = wb[hoja]
    row_count = min(ws.max_row or 1, max_rows)
    col_count = min(ws.max_column or 1, max_cols)
    merge_lookup = _merge_lookup(ws)
    columns = [
        {
            'letter': get_column_letter(index),
            'width': _column_width_px(ws, index),
        }
        for index in range(1, col_count + 1)
    ]
    rows = []
    for row_index in range(1, row_count + 1):
        cells = []
        for col_index in range(1, col_count + 1):
            merge_info = merge_lookup.get((row_index, col_index))
            if merge_info and not merge_info['is_master']:
                continue
            cell = ws.cell(row=row_index, column=col_index)
            master_cell = merge_info['master'] if merge_info else cell.coordinate
            alignment = cell.alignment
            font = cell.font
            style_parts = [
                f'background-color: {_cell_fill_hex(cell)};' if _cell_fill_hex(cell) else '',
                f'color: {_cell_font_hex(cell)};' if _cell_font_hex(cell) else '',
                f'font-size: {font.sz}px;' if font and font.sz else '',
                f'font-weight: {"700" if font and font.bold else "400"};',
                f'text-align: {alignment.horizontal};' if alignment and alignment.horizontal else '',
                f'vertical-align: {alignment.vertical};' if alignment and alignment.vertical else '',
                _cell_border_css(cell),
            ]
            cells.append({
                'coord': cell.coordinate,
                'master_cell': master_cell,
                'merge_range': merge_info['range'] if merge_info else '',
                'rowspan': merge_info['rowspan'] if merge_info else 1,
                'colspan': merge_info['colspan'] if merge_info else 1,
                'value': '' if cell.value is None else str(cell.value),
                'fill': _cell_fill_hex(cell),
                'bold': bool(cell.font and cell.font.bold),
                'align': getattr(cell.alignment, 'horizontal', '') or '',
                'style': ' '.join(part for part in style_parts if part),
            })
        rows.append({
            'index': row_index,
            'height': _row_height_px(ws, row_index),
            'cells': cells,
        })
    return {
        'columns': columns,
        'rows': rows,
        'truncated_rows': (ws.max_row or 1) > row_count,
        'truncated_cols': (ws.max_column or 1) > col_count,
        'max_row': ws.max_row or 1,
        'max_col': ws.max_column or 1,
    }


def _normalize_value(value):
    if value is None:
        return ''
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, 'strftime'):
        return value
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _object_value(obj, field_path):
    current = obj
    for part in str(field_path or '').split('.'):
        if not part:
            continue
        if current is None:
            return ''
        current = getattr(current, part, '')
        if callable(current) and not hasattr(current, '_meta'):
            try:
                current = current()
            except TypeError:
                return ''
    return _normalize_value(current)
    return value


def _rcm_value(rcm, campo):
    if not rcm:
        return ''
    equipo = getattr(rcm, 'equipo', None)
    if campo == 'equipo':
        return getattr(equipo, 'nombre_equipo', '') or ''
    if campo == 'tag':
        return getattr(equipo, 'tag_display', '') or ''
    if campo == 'ubicacion_tecnica':
        return getattr(equipo, 'ut', '') or ''
    if campo == 'descripcion_ut':
        return getattr(equipo, 'descripcion_ut', '') or ''
    return _normalize_value(getattr(rcm, campo, ''))


def _first_rcm_for_pauta(pauta):
    task_ref = pauta.tareas.filter(origen_modelo='TareaRCM', origen_id__isnull=False).order_by('orden', 'id').first()
    if task_ref:
        return _rcm_for_pauta_tarea(task_ref)
    return None


def _first_tarea_rcm_for_pauta(pauta):
    task_ref = pauta.tareas.filter(origen_modelo='TareaRCM', origen_id__isnull=False).order_by('orden', 'id').first()
    return _tarea_rcm_for_pauta_tarea(task_ref) if task_ref else None


def _tarea_rcm_for_pauta_tarea(tarea):
    if not tarea or tarea.origen_modelo != 'TareaRCM' or not tarea.origen_id:
        return None
    try:
        return models.TareaRCM.objects.select_related('fmea__rcm__equipo').get(pk=tarea.origen_id)
    except models.TareaRCM.DoesNotExist:
        return None


def _rcm_for_pauta_tarea(tarea):
    task = _tarea_rcm_for_pauta_tarea(tarea)
    if not task:
        return None
    return task.fmea.rcm if task.fmea_id else None


def _fmea_for_rcm(rcm):
    if not rcm:
        return None
    try:
        return rcm.fmea_fmeca
    except models.FMEA_FMECA.DoesNotExist:
        return None


def _evaluation_value(fmea, estrategia_dimension_id):
    if not fmea or not estrategia_dimension_id:
        return ''
    try:
        evaluation = (
            models.EvaluacionFMEA.objects
            .select_related('catalogo_fila', 'escala_valor')
            .get(fmea=fmea, estrategia_dimension_id=estrategia_dimension_id)
        )
    except (models.EvaluacionFMEA.DoesNotExist, ValueError, TypeError):
        return ''
    if evaluation.valor_texto not in (None, ''):
        return evaluation.valor_texto
    if evaluation.catalogo_fila_id:
        values = evaluation.catalogo_fila.values_map()
        for key in ['valor', 'indicador', 'codigo', 'nombre', 'etiqueta', 'descripcion', 'texto']:
            value = values.get(key)
            if value not in (None, ''):
                return _normalize_value(value)
        return evaluation.catalogo_fila.etiqueta or ''
    if evaluation.escala_valor_id:
        escala = evaluation.escala_valor
        return escala.codigo or escala.descripcion or _normalize_value(escala.valor_numerico)
    if evaluation.valor_numerico is not None:
        return evaluation.valor_numerico
    return ''


def _task_dynamic_value(tarea_rcm, campo_id):
    if not tarea_rcm or not campo_id:
        return ''
    try:
        value = models.ValorCampoTareaRCM.objects.get(tarea=tarea_rcm, campo_id=campo_id)
    except (models.ValorCampoTareaRCM.DoesNotExist, ValueError, TypeError):
        return ''
    return _normalize_value(value.valor_display)


def _equipment_ut_level_value(equipo, level_id, attr):
    if not equipo or not level_id:
        return ''
    try:
        level_id = int(level_id)
    except (TypeError, ValueError):
        return ''

    node = getattr(equipo, 'nodo', None)
    if node:
        for path_node in node.path_nodes():
            if path_node.nivel_id == level_id:
                return getattr(path_node, attr, '') or ''

    try:
        level = models.NivelJerarquia.objects.get(pk=level_id)
    except models.NivelJerarquia.DoesNotExist:
        return ''
    if attr != 'codigo':
        return ''
    parts = [part for part in str(getattr(equipo, 'ut', '') or '').split('-') if part]
    index = int(level.orden or 0) - 1
    if 0 <= index < len(parts):
        return parts[index]
    return ''


def _legacy_source_path(item):
    source_path = (item.get('source_path') or '').strip()
    if source_path:
        return source_path
    origen = (item.get('origen') or '').strip().lower()
    campo = (item.get('campo') or '').strip()
    if not origen or not campo:
        return ''
    return f'{origen}.{campo}'


def resolve_source_path(source_path, context):
    source_path = (source_path or '').strip()
    if not source_path:
        return ''
    if source_path.startswith('evaluacion:'):
        return _evaluation_value(context.get('fmea'), source_path.split(':', 1)[1])
    if source_path.startswith('tarea_campo:'):
        return _task_dynamic_value(context.get('tarea_rcm'), source_path.split(':', 1)[1])
    if source_path.startswith('equipo_ut:'):
        _prefix, level_id, attr = (source_path.split(':') + ['', ''])[:3]
        return _equipment_ut_level_value(context.get('equipo'), level_id, attr or 'codigo')
    if source_path == 'fijo.valor':
        return context.get('valor_fijo', '')

    if '.' not in source_path:
        return ''
    origin, field_path = source_path.split('.', 1)
    if origin == 'sistema':
        system = context.get('sistema') or {}
        return _normalize_value(system.get(field_path, ''))
    if origin == 'rcm' and field_path in {'equipo', 'tag', 'ubicacion_tecnica', 'descripcion_ut'}:
        return _rcm_value(context.get('rcm'), field_path)
    return _object_value(context.get(origin), field_path)


def _header_context(pauta, item=None):
    first_task = pauta.tareas.order_by('orden', 'id').first()
    tarea_rcm = _tarea_rcm_for_pauta_tarea(first_task)
    rcm = _first_rcm_for_pauta(pauta)
    if not rcm and tarea_rcm and tarea_rcm.fmea_id:
        rcm = tarea_rcm.fmea.rcm
    return {
        'pauta': pauta,
        'tarea': first_task,
        'tarea_rcm': tarea_rcm,
        'servicio': pauta.servicio,
        'equipo': pauta.equipo or getattr(rcm, 'equipo', None),
        'rcm': rcm,
        'fmea': _fmea_for_rcm(rcm),
        'sistema': {
            'fecha_generacion': timezone.localdate(),
            'correlativo': 1,
            'usuario': '',
        },
        'valor_fijo': (item or {}).get('valor', ''),
    }


def _task_context(pauta_tarea, indice, item=None):
    tarea_rcm = _tarea_rcm_for_pauta_tarea(pauta_tarea)
    rcm = tarea_rcm.fmea.rcm if tarea_rcm and tarea_rcm.fmea_id else _rcm_for_pauta_tarea(pauta_tarea)
    return {
        'pauta': pauta_tarea.pauta,
        'tarea': pauta_tarea,
        'tarea_rcm': tarea_rcm,
        'rcm': rcm,
        'fmea': tarea_rcm.fmea if tarea_rcm and tarea_rcm.fmea_id else _fmea_for_rcm(rcm),
        'equipo': getattr(rcm, 'equipo', None),
        'servicio': pauta_tarea.pauta.servicio,
        'sistema': {
            'correlativo': indice,
            'fecha_generacion': timezone.localdate(),
        },
        'valor_fijo': (item or {}).get('valor', ''),
    }


def resolver_valor_mapeo(item, pauta):
    source_path = _legacy_source_path(item)
    if source_path:
        return resolve_source_path(source_path, _header_context(pauta, item))
    origen = (item.get('origen') or '').strip().lower()
    campo = (item.get('campo') or '').strip()
    if origen == 'pauta':
        return _normalize_value(getattr(pauta, campo, ''))
    if origen == 'rcm':
        return _rcm_value(_first_rcm_for_pauta(pauta), campo)
    if origen == 'sistema':
        if campo == 'fecha_generacion':
            return timezone.localdate()
        if campo == 'correlativo':
            return 1
        return ''
    if origen == 'fijo':
        return item.get('valor', '')
    return ''


def resolver_valor_tarea(item, tarea, indice):
    source_path = _legacy_source_path(item)
    if source_path:
        return resolve_source_path(source_path, _task_context(tarea, indice, item))
    origen = (item.get('origen') or '').strip().lower()
    campo = (item.get('campo') or '').strip()
    if origen == 'tarea':
        return _normalize_value(getattr(tarea, campo, ''))
    if origen == 'pauta':
        return _normalize_value(getattr(tarea.pauta, campo, ''))
    if origen == 'rcm':
        return _rcm_value(_rcm_for_pauta_tarea(tarea), campo)
    if origen == 'sistema':
        if campo == 'correlativo':
            return indice
        if campo == 'fecha_generacion':
            return timezone.localdate()
        return ''
    if origen == 'fijo':
        return item.get('valor', '')
    return ''


def _sheet_or_error(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        raise PautaExportError(f'La hoja "{sheet_name}" no existe en la plantilla.')
    return wb[sheet_name]


def aplicar_mapeo_celdas(wb, pauta, config):
    for item in config.get('celdas', []):
        hoja = item.get('hoja') or config.get('hoja_principal') or wb.sheetnames[0]
        celda = (item.get('celda') or '').strip()
        if not celda:
            continue
        ws = _sheet_or_error(wb, hoja)
        try:
            ws[celda] = resolver_valor_mapeo(item, pauta)
        except Exception as exc:
            raise PautaExportError(f'No se pudo escribir la celda {hoja}!{celda}: {exc}') from exc


def copiar_formato_fila(ws, fila_origen, fila_destino):
    if fila_origen == fila_destino:
        return
    ws.row_dimensions[fila_destino].height = ws.row_dimensions[fila_origen].height
    for column in range(1, ws.max_column + 1):
        source = ws.cell(row=fila_origen, column=column)
        target = ws.cell(row=fila_destino, column=column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def aplicar_mapeo_tablas(wb, pauta, config):
    tareas = list(pauta.tareas.all().order_by('orden', 'id'))
    for table in config.get('tablas', []):
        hoja = table.get('hoja') or config.get('hoja_principal') or wb.sheetnames[0]
        ws = _sheet_or_error(wb, hoja)
        try:
            fila_inicio = int(table.get('fila_inicio') or 1)
            fila_template = int(table.get('fila_template') or fila_inicio)
        except (TypeError, ValueError):
            raise PautaExportError(f'La tabla "{table.get("nombre", "tareas")}" tiene filas inválidas.')

        for index, tarea in enumerate(tareas, start=1):
            row_number = fila_inicio + index - 1
            for column in table.get('columnas', []):
                col = (column.get('columna') or '').strip().upper()
                if not col:
                    continue
                ws[f'{col}{row_number}'] = resolver_valor_tarea(column, tarea, index)


def generar_excel_pauta(pauta):
    plantilla = pauta.plantilla
    if not plantilla:
        raise PautaExportError('La pauta no tiene plantilla asociada.')
    mapeo = getattr(plantilla, 'mapeo', None)
    if not mapeo or not mapeo.config:
        raise PautaExportError('La plantilla no tiene mapeo configurado.')

    wb = _load_workbook_from_template(plantilla)
    config = dict(mapeo.config or {})
    if mapeo.hoja_principal:
        config.setdefault('hoja_principal', mapeo.hoja_principal)
    aplicar_mapeo_celdas(wb, pauta, config)
    aplicar_mapeo_tablas(wb, pauta, config)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
