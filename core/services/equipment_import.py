import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from django.db import transaction
from django.db.models import Max, Q

from core import models

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


FORMAT_AUTO = 'auto'
FORMAT_MINDCO = 'mindco_simple'
FORMAT_SAP = 'sap_uts'
FORMAT_CHOICES = (FORMAT_AUTO, FORMAT_MINDCO, FORMAT_SAP)
HEADER_SCAN_ROWS = 50

UT_HEADERS = (
    'ut',
    'ubicacion tecnica',
    'ubicacion tecnica',
    'ubicac tecnica',
    'u tecnica',
)
TAG_HEADERS = ('tag', 'tag equipo', 'codigo equipo', 'n equipo', 'n equipo sap')
EQUIPMENT_NAME_HEADERS = (
    'equipo',
    'equipo componente',
    'equipo componente',
    'nombre',
    'nombre equipo',
    'descripcion equipo',
    'descripcion del equipo',
    'denominacion de la ubicacion tecnica',
)
DESCRIPTION_HEADERS = (
    'descripcion',
    'descripcion ut',
    'descripcion u tecnica',
    'denominacion de la ubicacion tecnica',
)
COMPANY_HEADERS = ('empresa', 'cliente')
SERVICE_HEADERS = ('servicio', 'codigo servicio', 'codigo_servicio')
MINDCO_HIERARCHY_HEADERS = (
    'empresa',
    'planta',
    'planta sitio',
    'sitio',
    'area',
    'sistema',
    'subsistema',
    'sub sistema',
    'nivel 1',
    'nivel 2',
    'nivel 3',
    'nivel 4',
    'nivel 5',
    'nivel 6',
)
NON_HIERARCHY_HEADERS = set(
    UT_HEADERS
    + TAG_HEADERS
    + EQUIPMENT_NAME_HEADERS
    + DESCRIPTION_HEADERS
    + COMPANY_HEADERS
    + SERVICE_HEADERS
    + ('level', 'parent ut id', 'parent_ut_id')
)
MINDCO_NON_HIERARCHY_HEADERS = set(
    UT_HEADERS
    + TAG_HEADERS
    + EQUIPMENT_NAME_HEADERS
    + DESCRIPTION_HEADERS
    + SERVICE_HEADERS
    + ('level', 'parent ut id', 'parent_ut_id')
)


def _normalized_aliases(aliases):
    return {normalize_header(alias) for alias in aliases}


@dataclass
class EquipmentImportReport:
    rows_read: int = 0
    valid_rows: int = 0
    skipped: int = 0
    format_detected: str = ''
    levels_detected: list[str] = field(default_factory=list)
    nodes_to_create: int = 0
    nodes_reused: int = 0
    nodes_created: int = 0
    equipment_to_create: int = 0
    equipment_to_update: int = 0
    equipment_created: int = 0
    equipment_updated: int = 0
    service_equipment_to_create: int = 0
    service_equipment_created: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    normalized_rows: list[dict[str, Any]] = field(default_factory=list)


def normalize_header(value):
    value = unicodedata.normalize('NFKD', str(value or ''))
    value = ''.join(char for char in value if not unicodedata.combining(char))
    value = re.sub(r'[^a-zA-Z0-9]+', ' ', value).strip().lower()
    return re.sub(r'\s+', ' ', value)


def normalize_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def technical_segment(value):
    return models._technical_segment(value)


def split_ut_path(ut):
    return [technical_segment(part) for part in str(ut or '').split('-') if str(part).strip()]


def build_cumulative_paths(parts):
    paths = []
    current = []
    for part in parts:
        current.append(part)
        paths.append('-'.join(current))
    return paths


def _first_value(row, aliases):
    for alias in _normalized_aliases(aliases):
        value = row['values_by_key'].get(alias, '')
        if value != '':
            return value
    return ''


def _has_header(headers, aliases):
    normalized = {item['key'] for item in headers}
    return any(alias in normalized for alias in _normalized_aliases(aliases))


def detect_excel_format(headers):
    normalized = {item['key'] for item in headers}
    if {'n0', 'n0 nombre', 'n1', 'n1 nombre'}.issubset(normalized):
        return FORMAT_SAP
    if _has_header(headers, UT_HEADERS) and (_has_header(headers, TAG_HEADERS) or _has_header(headers, EQUIPMENT_NAME_HEADERS)):
        return FORMAT_MINDCO
    hierarchy_headers = _mindco_hierarchy_headers(headers)
    if hierarchy_headers and _has_header(headers, TAG_HEADERS):
        return FORMAT_MINDCO
    raise ValueError('No se pudo detectar el formato. Usa auto, mindco_simple o sap_uts y revisa encabezados.')


def _header_match_score(normalized):
    normalized_set = set(normalized)
    sap_headers = {'n0', 'n0 nombre', 'n1', 'n1 nombre'}
    if sap_headers.issubset(normalized_set):
        return 100 + len([value for value in normalized if value])

    non_hierarchy = _normalized_aliases(MINDCO_NON_HIERARCHY_HEADERS)
    known_hierarchy = _normalized_aliases(MINDCO_HIERARCHY_HEADERS)
    hierarchy_count = sum(1 for value in normalized if value and value not in non_hierarchy)
    known_hierarchy_count = sum(1 for value in normalized if value in known_hierarchy)
    has_ut = any(alias in normalized_set for alias in _normalized_aliases(UT_HEADERS))
    has_tag = any(alias in normalized_set for alias in _normalized_aliases(TAG_HEADERS))
    has_equipment_name = any(alias in normalized_set for alias in _normalized_aliases(EQUIPMENT_NAME_HEADERS))

    if has_ut and (has_tag or has_equipment_name):
        return 70 + known_hierarchy_count * 20 + hierarchy_count + (10 if has_tag else 0) + (8 if has_equipment_name else 0)
    if has_tag and hierarchy_count >= 1:
        return 50 + known_hierarchy_count * 20 + hierarchy_count + (8 if has_equipment_name else 0)
    return 0


def _find_header_candidate(worksheet):
    best = None
    for row_number, cells in enumerate(worksheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS), start=1):
        raw_headers = [normalize_text(cell.value) for cell in cells]
        normalized = [normalize_header(value) for value in raw_headers]
        score = _header_match_score(normalized)
        if score <= 0:
            continue
        candidate = {
            'score': score,
            'row_number': row_number,
            'raw_headers': raw_headers,
            'normalized': normalized,
        }
        if best is None or score > best['score']:
            best = candidate
    return best


def _candidate_data_score(worksheet, candidate):
    headers = []
    for index, label in enumerate(candidate['raw_headers']):
        key = normalize_header(label)
        if key:
            headers.append({'index': index, 'label': label, 'key': key})
    ut_index = next((item['index'] for item in headers if item['key'] in _normalized_aliases(UT_HEADERS)), None)
    if ut_index is None:
        return 0

    non_hierarchy = _normalized_aliases(MINDCO_NON_HIERARCHY_HEADERS)
    hierarchy_indexes = [
        item['index']
        for item in headers
        if item['key'] not in non_hierarchy and item['index'] < ut_index
    ]
    max_ut_parts = 0
    max_hierarchy_values = 0
    for cells in worksheet.iter_rows(min_row=candidate['row_number'] + 1, max_row=candidate['row_number'] + 10):
        ut_value = normalize_text(cells[ut_index].value if ut_index < len(cells) else '')
        if ut_value:
            max_ut_parts = max(max_ut_parts, len(split_ut_path(ut_value)))
        hierarchy_values = 0
        for index in hierarchy_indexes:
            value = normalize_text(cells[index].value if index < len(cells) else '')
            if value:
                hierarchy_values += 1
        max_hierarchy_values = max(max_hierarchy_values, hierarchy_values)
    return max_ut_parts * 5 + max_hierarchy_values * 2


def read_xlsx_rows(file, sheet_name=None, limit=None):
    if load_workbook is None:
        raise ValueError('openpyxl no está instalado en el entorno de Django.')
    file.seek(0)
    workbook = load_workbook(BytesIO(file.read()), read_only=True, data_only=True)
    selected_sheet_name = sheet_name
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f'La hoja "{sheet_name}" no existe en el archivo.')
        worksheet = workbook[sheet_name]
        candidate = _find_header_candidate(worksheet)
    else:
        candidate = None
        worksheet = None
        for current_worksheet in workbook.worksheets:
            current_candidate = _find_header_candidate(current_worksheet)
            if not current_candidate:
                continue
            current_candidate['score'] += _candidate_data_score(current_worksheet, current_candidate)
            if candidate is None or current_candidate['score'] > candidate['score']:
                candidate = current_candidate
                worksheet = current_worksheet
                selected_sheet_name = current_worksheet.title
        if worksheet is None:
            worksheet = workbook.active

    header_row = candidate['row_number'] if candidate else None
    headers = []
    if header_row is None:
        detail = f' en la hoja "{selected_sheet_name}"' if selected_sheet_name else ''
        raise ValueError(f'No se encontraron encabezados válidos{detail} en las primeras {HEADER_SCAN_ROWS} filas.')
    for index, label in enumerate(candidate['raw_headers']):
        key = normalize_header(label)
        if key:
            headers.append({'index': index, 'label': label, 'key': key})

    rows = []
    for row_number, cells in enumerate(worksheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
        values_by_key = {}
        values_by_label = {}
        has_data = False
        for header in headers:
            value = normalize_text(cells[header['index']].value if header['index'] < len(cells) else '')
            if value:
                has_data = True
            values_by_key[header['key']] = value
            values_by_label[header['label']] = value
        if not has_data:
            continue
        rows.append({
            'row_number': row_number,
            'values_by_key': values_by_key,
            'values_by_label': values_by_label,
        })
        if limit and len(rows) >= limit:
            break
    workbook.close()
    return headers, rows


def _resolve_empresa(row, selected_empresa):
    raw_empresa = _first_value(row, COMPANY_HEADERS)
    if not raw_empresa:
        return selected_empresa
    empresa = models.Empresa.objects.filter(Q(sigla__iexact=raw_empresa) | Q(nombre__iexact=raw_empresa)).first()
    return empresa or selected_empresa


def _mindco_hierarchy_headers(headers):
    ut_index = next((item['index'] for item in headers if item['key'] in _normalized_aliases(UT_HEADERS)), None)
    non_hierarchy = _normalized_aliases(MINDCO_NON_HIERARCHY_HEADERS)
    candidates = []
    for header in headers:
        if header['key'] in non_hierarchy:
            continue
        if ut_index is not None and header['index'] > ut_index:
            continue
        candidates.append(header)
    return candidates


def _mindco_last_segment_is_equipment(row, parts, explicit_value):
    if explicit_value is not None:
        return bool(explicit_value)
    tag = technical_segment(_first_value(row, TAG_HEADERS))
    return bool(tag and parts and tag == parts[-1])


def build_hierarchy_from_mindco_row(row, headers, last_segment_is_equipment=None):
    hierarchy_headers = _mindco_hierarchy_headers(headers)
    ut = _first_value(row, UT_HEADERS)
    parts = split_ut_path(ut)
    tag_value = _first_value(row, TAG_HEADERS)
    if not parts:
        parts = [
            technical_segment(row['values_by_key'].get(header['key'], ''))
            for header in hierarchy_headers
            if row['values_by_key'].get(header['key'], '')
        ]
        if tag_value:
            parts.append(technical_segment(tag_value))
        ut = '-'.join(parts)
    if not parts:
        raise ValueError('sin UT.')
    equipment_is_last = _mindco_last_segment_is_equipment(row, parts, last_segment_is_equipment)
    node_parts = parts[:-1] if equipment_is_last and len(parts) > 1 else parts
    hierarchy = []
    cumulative_paths = build_cumulative_paths(node_parts)
    for index, part in enumerate(node_parts, start=1):
        header = hierarchy_headers[index - 1] if index - 1 < len(hierarchy_headers) else None
        level_name = header['label'] if header else f'Nivel {index}'
        node_name = row['values_by_key'].get(header['key'], '') if header else ''
        hierarchy.append({
            'level_order': index,
            'level_name': level_name or f'Nivel {index}',
            'node_code': part,
            'node_name': node_name or part,
            'node_path': cumulative_paths[index - 1],
        })

    tag = technical_segment(tag_value) or (parts[-1] if parts else '')
    equipment_parts = parts if equipment_is_last else parts + ([tag] if tag else [])
    nombre = _first_value(row, EQUIPMENT_NAME_HEADERS) or tag
    descripcion = _first_value(row, DESCRIPTION_HEADERS) or nombre
    return hierarchy, {
        'tag_equipo': tag,
        'nombre_equipo': nombre,
        'ut': '-'.join(equipment_parts),
        'ubicacion_tecnica': '-'.join(equipment_parts),
        'descripcion_ut': descripcion,
    }


def _sap_levels_from_headers(headers):
    level_indexes = []
    header_keys = {header['key'] for header in headers}
    for key in header_keys:
        match = re.fullmatch(r'n(\d+)', key)
        if not match:
            continue
        index = int(match.group(1))
        level_indexes.append(index)
    return sorted(level_indexes)


def build_hierarchy_from_sap_row(row, headers, last_level_is_equipment=True):
    levels = []
    for index in _sap_levels_from_headers(headers):
        path = row['values_by_key'].get(f'n{index}', '')
        if not path:
            continue
        name = row['values_by_key'].get(f'n{index} nombre', '') or path
        parts = split_ut_path(path)
        levels.append({
            'excel_index': index,
            'path': '-'.join(parts),
            'code': parts[-1] if parts else technical_segment(path),
            'name': name,
        })
    if not levels:
        raise ValueError('sin niveles N0/N1 resolubles.')

    equipment_level = levels[-1] if last_level_is_equipment and len(levels) > 1 else None
    node_levels = levels[:-1] if equipment_level else levels
    hierarchy = []
    for order, item in enumerate(node_levels, start=1):
        hierarchy.append({
            'level_order': order,
            'level_name': f'Nivel {item["excel_index"]}',
            'node_code': item['code'],
            'node_name': item['name'] or item['code'],
            'node_path': item['path'],
        })

    if equipment_level:
        equipment_path = equipment_level['path']
        equipment_name = equipment_level['name']
        equipment_tag = equipment_level['code']
    else:
        raw_ut = _first_value(row, UT_HEADERS) or levels[-1]['path']
        equipment_parts = split_ut_path(raw_ut)
        equipment_path = '-'.join(equipment_parts) if equipment_parts else levels[-1]['path']
        equipment_name = _first_value(row, DESCRIPTION_HEADERS) or levels[-1]['name']
        equipment_tag = equipment_parts[-1] if equipment_parts else levels[-1]['code']

    return hierarchy, {
        'tag_equipo': equipment_tag,
        'nombre_equipo': equipment_name or equipment_tag,
        'ut': equipment_path,
        'ubicacion_tecnica': equipment_path,
        'descripcion_ut': equipment_name or equipment_tag,
    }


def parse_mindco_simple_rows(rows, headers, empresa, servicio=None, last_segment_is_equipment=None):
    normalized = []
    errors = []
    warnings = []
    for row in rows:
        try:
            if not empresa:
                raise ValueError('sin empresa resoluble.')
            hierarchy, equipment = build_hierarchy_from_mindco_row(row, headers, last_segment_is_equipment)
            if not hierarchy:
                raise ValueError('jerarquía inválida.')
            normalized.append({
                'row_number': row['row_number'],
                'empresa': empresa,
                'servicio': servicio,
                'hierarchy': hierarchy,
                'equipment': equipment,
            })
        except ValueError as exc:
            errors.append(f'Fila {row["row_number"]}: {exc}')
    return normalized, warnings, errors


def parse_sap_uts_rows(rows, headers, empresa, servicio=None, last_level_is_equipment=True):
    normalized = []
    errors = []
    for row in rows:
        try:
            row_empresa = _resolve_empresa(row, empresa)
            if not row_empresa:
                raise ValueError('sin empresa resoluble.')
            hierarchy, equipment = build_hierarchy_from_sap_row(row, headers, last_level_is_equipment)
            if not hierarchy:
                raise ValueError('jerarquía inválida.')
            normalized.append({
                'row_number': row['row_number'],
                'empresa': row_empresa,
                'servicio': servicio,
                'hierarchy': hierarchy,
                'equipment': equipment,
            })
        except ValueError as exc:
            errors.append(f'Fila {row["row_number"]}: {exc}')
    return normalized, [], errors


def get_or_create_level(empresa, orden, nombre):
    level = models.NivelJerarquia.objects.filter(empresa=empresa, orden=orden, activo=True).first()
    if level:
        return level, False
    order = orden
    level, created = models.NivelJerarquia.objects.get_or_create(
        empresa=empresa,
        nombre=nombre or f'Nivel {orden}',
        defaults={'orden': order, 'activo': True},
    )
    return level, created


def get_or_create_node(empresa, nivel, nombre, parent, node_path=None, node_code=None):
    code = technical_segment(node_code or (split_ut_path(node_path)[-1] if node_path else nombre))
    order = (
        models.NodoJerarquia.objects.filter(empresa=empresa, parent=parent, nivel=nivel)
        .aggregate(max_order=Max('orden'))
        .get('max_order') or 0
    ) + 1
    node, created = models.NodoJerarquia.objects.get_or_create(
        empresa=empresa,
        parent=parent,
        codigo=code,
        defaults={
            'nivel': nivel,
            'nombre': (nombre or code)[:200],
            'orden': order,
            'activo': True,
        },
    )
    changed = []
    if not created:
        if node.nivel_id != nivel.pk:
            node.nivel = nivel
            changed.append('nivel')
        if nombre and node.nombre != nombre[:200]:
            node.nombre = nombre[:200]
            changed.append('nombre')
        if not node.activo:
            node.activo = True
            changed.append('activo')
        if changed:
            node.save(update_fields=changed)
    return node, created


def _find_equipment(empresa, equipment_data):
    tag = equipment_data.get('tag_equipo') or ''
    ut = equipment_data.get('ut') or ''
    qs = models.Equipo.objects.all()
    if empresa:
        qs = qs.filter(nodo__empresa=empresa)
    if tag and ut:
        item = qs.filter(tag_equipo__iexact=tag, ut__iexact=ut).first()
        if item:
            return item
    if ut:
        item = qs.filter(ut__iexact=ut).first()
        if item:
            return item
    return None


def create_or_update_equipment(empresa, servicio, equipment_data, parent_node):
    tag = technical_segment(equipment_data.get('tag_equipo') or '')
    if parent_node and tag:
        equipment_data = {
            **equipment_data,
            'tag_equipo': tag,
            'ut': f'{parent_node.ut}-{tag}',
            'ubicacion_tecnica': f'{parent_node.ut}-{tag}',
        }
    equipment = _find_equipment(empresa, equipment_data)
    created = False
    if not equipment:
        equipment = models.Equipo(
            tag_equipo=(equipment_data.get('tag_equipo') or '')[:100],
            nombre_equipo=(equipment_data.get('nombre_equipo') or '')[:200],
            ut=(equipment_data.get('ut') or '')[:200],
            descripcion_ut=(equipment_data.get('descripcion_ut') or '')[:255],
            nodo=parent_node,
        )
        equipment.save()
        created = True
    else:
        changed = []
        field_map = {
            'tag_equipo': (equipment_data.get('tag_equipo') or '')[:100],
            'nombre_equipo': (equipment_data.get('nombre_equipo') or '')[:200],
            'ut': (equipment_data.get('ut') or '')[:200],
            'descripcion_ut': (equipment_data.get('descripcion_ut') or '')[:255],
        }
        for field, value in field_map.items():
            if value and getattr(equipment, field) != value:
                setattr(equipment, field, value)
                changed.append(field)
        if parent_node and equipment.nodo_id != parent_node.pk:
            equipment.nodo = parent_node
            changed.append('nodo')
        if changed:
            equipment.save(update_fields=changed)

    service_created = False
    if servicio:
        _, service_created = models.ServicioEquipo.objects.get_or_create(servicio=servicio, equipo=equipment)
    return equipment, created, service_created


def _normalize_import_rows(file, empresa, servicio=None, sheet_name=None, format=FORMAT_AUTO, last_segment_is_equipment=None, last_level_is_equipment=True, limit=None):
    headers, rows = read_xlsx_rows(file, sheet_name=sheet_name, limit=limit)
    detected = detect_excel_format(headers) if format == FORMAT_AUTO else format
    if detected not in FORMAT_CHOICES or detected == FORMAT_AUTO:
        raise ValueError('Formato inválido. Usa auto, mindco_simple o sap_uts.')
    if detected == FORMAT_SAP:
        normalized, warnings, errors = parse_sap_uts_rows(rows, headers, empresa, servicio, last_level_is_equipment)
    else:
        normalized, warnings, errors = parse_mindco_simple_rows(rows, headers, empresa, servicio, last_segment_is_equipment)
    return headers, rows, detected, normalized, warnings, errors


def preview_equipment_import(file, empresa, servicio=None, sheet_name=None, format=FORMAT_AUTO, last_segment_is_equipment=None, last_level_is_equipment=True, limit=None):
    headers, rows, detected, normalized, warnings, errors = _normalize_import_rows(
        file,
        empresa,
        servicio=servicio,
        sheet_name=sheet_name,
        format=format,
        last_segment_is_equipment=last_segment_is_equipment,
        last_level_is_equipment=last_level_is_equipment,
        limit=limit,
    )
    report = EquipmentImportReport(rows_read=len(rows), format_detected=detected)
    report.normalized_rows = normalized
    report.warnings.extend(warnings)
    report.errors.extend(errors)
    report.skipped = len(errors)
    report.valid_rows = len(normalized)

    seen_levels = set()
    seen_nodes = set()
    simulated_nodes = {}
    seen_equipment = set()
    seen_service_equipment = set()
    for item in normalized:
        parent_key = None
        parent_node = None
        for h in item['hierarchy']:
            seen_levels.add((item['empresa'].pk, h['level_order'], h['level_name']))
            node_key = (item['empresa'].pk, parent_key, h['node_code'])
            if node_key in seen_nodes:
                parent_node = simulated_nodes.get(node_key)
                parent_key = node_key
                continue
            seen_nodes.add(node_key)
            level = models.NivelJerarquia.objects.filter(empresa=item['empresa'], orden=h['level_order'], activo=True).first()
            existing_node = None
            parent_is_new = parent_key is not None and parent_node is None
            if level and not parent_is_new:
                existing_node = models.NodoJerarquia.objects.filter(
                    empresa=item['empresa'],
                    parent=parent_node,
                    codigo__iexact=h['node_code'],
                ).first()
            if existing_node:
                report.nodes_reused += 1
                simulated_nodes[node_key] = existing_node
                parent_node = existing_node
            else:
                report.nodes_to_create += 1
                simulated_nodes[node_key] = None
                parent_node = None
            parent_key = node_key

        existing_equipment = _find_equipment(item['empresa'], item['equipment'])
        equipment_key = (item['empresa'].pk, item['equipment'].get('tag_equipo'), item['equipment'].get('ut'))
        if equipment_key not in seen_equipment:
            seen_equipment.add(equipment_key)
            if existing_equipment:
                report.equipment_to_update += 1
                report.warnings.append(f'Fila {item["row_number"]}: equipo existente será actualizado.')
            else:
                report.equipment_to_create += 1
        service = item.get('servicio')
        if service:
            service_equipment_key = (service.pk, *equipment_key)
            if service_equipment_key not in seen_service_equipment:
                seen_service_equipment.add(service_equipment_key)
                relation_exists = bool(
                    existing_equipment
                    and models.ServicioEquipo.objects.filter(servicio=service, equipo=existing_equipment).exists()
                )
                if not relation_exists:
                    report.service_equipment_to_create += 1

    report.levels_detected = [
        f'{order}. {name}' for _empresa_id, order, name in sorted(seen_levels, key=lambda value: (value[1], value[2]))
    ]
    return report


@transaction.atomic
def execute_equipment_import(preview_data, user=None):
    report = EquipmentImportReport(
        rows_read=preview_data.rows_read,
        valid_rows=preview_data.valid_rows,
        skipped=preview_data.skipped,
        format_detected=preview_data.format_detected,
        levels_detected=list(preview_data.levels_detected),
        nodes_to_create=preview_data.nodes_to_create,
        nodes_reused=preview_data.nodes_reused,
        equipment_to_create=preview_data.equipment_to_create,
        equipment_to_update=preview_data.equipment_to_update,
        service_equipment_to_create=preview_data.service_equipment_to_create,
        warnings=list(preview_data.warnings),
        errors=list(preview_data.errors),
    )
    for item in preview_data.normalized_rows:
        parent = None
        for h in item['hierarchy']:
            level, _level_created = get_or_create_level(item['empresa'], h['level_order'], h['level_name'])
            parent, created = get_or_create_node(
                item['empresa'],
                level,
                h['node_name'],
                parent,
                node_path=h.get('node_path'),
                node_code=h.get('node_code'),
            )
            if created:
                report.nodes_created += 1
        _equipment, equipment_created, service_created = create_or_update_equipment(
            item['empresa'],
            item.get('servicio'),
            item['equipment'],
            parent,
        )
        if equipment_created:
            report.equipment_created += 1
        else:
            report.equipment_updated += 1
        if service_created:
            report.service_equipment_created += 1
    return report
